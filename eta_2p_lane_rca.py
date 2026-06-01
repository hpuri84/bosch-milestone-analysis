"""
ETA 2P (port arrival, S07) Lane RCA — per-week comparison against the prior week.

Mirrors eta_2d_lane_rca.py but for ETA 2P (S07 / port arrival) instead of 2D
(S31 / delivery). SC4 ONLY — SC3 has no ETA columns.

MEASUREMENT METHOD (reconciles to the published ETA 2P headline):
  - Denominator = ARRIVED shipments = rows where ATA_DATE_TIME is present.
  - ACCEPTED  = has S07_TS_@measured present AND abs(S07_Deviation) <= 48h.
  - FAILED    = arrived but not accepted. Two sub-types:
        outside_48h : has a TS but abs(dev) > 48h
        no_t7       : arrived but NO S07 estimate ever recorded (still a failure)
  - Direction (has-TS subset only): dev < 0 = "early" (vessel arrived before ETA),
    dev > 0 = "late". no_t7 rows get direction "no_estimate".
  - Overall accuracy = accepted / arrived. no_t7 stays in the denominator so the
    overall ties to the 2P headline (75.81% CW20, 65.96% CW19, 85.41% CW17).

Output: dashboard/public/eta_2p_lane_rca.json with the SAME top-level shape as
the 2D json:
  { generated, available_weeks[], weeks:{ CWxx: {
      prior_week, current_week, weekly_stats:{prior,current},
      port_discharge_comparison[], lane_comparison[], country_origin_comparison[],
      country_dest_comparison[], service_comparison[], carrier_comparison[],
      vessel_comparison[], city_lane_comparison[], transport_comparison[] } } }
"""

import openpyxl
import json
import os
from collections import defaultdict
from datetime import datetime

BASE = os.path.dirname(os.path.abspath(__file__))
RAW_DIR = os.path.join(BASE, "Bosch Milestone raw data")
OUT_PATH = os.path.join(BASE, "dashboard", "public", "eta_2p_lane_rca.json")

WEEKS = [f"CW{i:02d}" for i in range(1, 54)]


def sc4_filename(week):
    return f"Maersk SC4_2026_{week}.xlsx"


def safe_str(val):
    return "" if val is None else str(val).strip()


def safe_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def has_value(val):
    return val is not None and str(val).strip() != ""


def find_shipments_sheet(wb):
    for sn in wb.sheetnames:
        if sn.strip().lower() == "shipments":
            return sn
    for sn in wb.sheetnames:
        if sn.strip().upper() in ("TOTAL", "ALL"):
            continue
        if sn.strip().upper().rstrip("_") in ("FCL", "BCO", "LCL"):
            continue
        ws = wb[sn]
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            for cell in row:
                if cell and "UNIQUE_SHIPMENT" in str(cell).upper():
                    return sn
    return None


def find_header_row(ws):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        for cell in row:
            if cell and "UNIQUE_SHIPMENT" in str(cell):
                return i
    return 3


def build_col_map(ws, header_row):
    headers = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
    return {str(h).strip(): i for i, h in enumerate(headers) if h}


def find_col(col_map, *patterns):
    for key in col_map:
        for pat in patterns:
            if pat in key:
                return col_map[key]
    return None


def extract_shipments(week):
    """Extract all ARRIVED SC4 shipments with ETA 2P (S07) fields for one week."""
    path = os.path.join(RAW_DIR, sc4_filename(week))
    if not os.path.exists(path):
        return None

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ship_sheet = find_shipments_sheet(wb)
    if not ship_sheet:
        print(f"  WARNING: No shipments sheet in {week}")
        wb.close()
        return []

    ws = wb[ship_sheet]
    header_row = find_header_row(ws)
    col_map = build_col_map(ws, header_row)

    # NOTE: the deviation header is "S07_Deviation_TS_@measured" and the TS list
    # header is "S07_TS_@measured_list" — both contain "S07_TS_@measured", so the
    # deviation column must be resolved with its own distinct substring first.
    c = {
        "hbl":            find_col(col_map, "HOUSE_BILL_OF_LADING"),
        "carrier":        find_col(col_map, "CARRIER_1"),
        "service":        find_col(col_map, "TRANSPORT_SERVICE_PRIORITY"),
        "transport_mode": find_col(col_map, "TRANSPORT_MODE"),
        "incoterm":       find_col(col_map, "INCOTERM"),
        "origin_city":    find_col(col_map, "CONSIGNOR_ADDRESS_CITY_NAME"),
        "origin_country": find_col(col_map, "CONSIGNOR_ADDRESS_COUNTRY"),
        "dest_country":   find_col(col_map, "CONSIGNEE_ADDRESS_COUNTRY"),
        "dest_city":      find_col(col_map, "CONSIGNEE_ADDRESS_CITY_NAME"),
        "pod":            find_col(col_map, "PORT_OF_DISCHARGE"),
        "vessel":         find_col(col_map, "VESSEL_NAME"),
        "atd":            find_col(col_map, "ATD_DATE_TIME"),
        "ata":            find_col(col_map, "ATA_DATE_TIME"),
        "eta":            find_col(col_map, "ETA_DATE_TIME"),
        "delivered":      find_col(col_map, "DELIVERED_DATE_TIME"),
        "s07_accepted":   find_col(col_map, "S07_Accepted"),
        "s07_deviation":  find_col(col_map, "S07_Deviation"),
        "s07_ts":         find_col(col_map, "S07_TS_@measured"),
        "delay_reason":   find_col(col_map, "Delay_Reason_Code_Description"),
    }

    def get(row, key):
        idx = c.get(key)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    shipments = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row = list(row)
        if not row or not row[0]:
            continue

        # Denominator = ARRIVED: ATA present.
        if not has_value(get(row, "ata")):
            continue

        has_ts = has_value(get(row, "s07_ts"))
        dev = safe_float(get(row, "s07_deviation")) if has_ts else None

        accepted = bool(has_ts and dev is not None and abs(dev) <= 48)

        if not has_ts:
            failure_type = "no_t7"
            direction = "no_estimate"
        elif accepted:
            failure_type = None
            direction = "on_time"
        else:
            failure_type = "outside_48h"
            direction = "late" if (dev is not None and dev > 0) else ("early" if (dev is not None and dev < 0) else "unknown")

        origin_country = safe_str(get(row, "origin_country"))
        dest_country = safe_str(get(row, "dest_country"))
        origin_city = safe_str(get(row, "origin_city"))
        dest_city = safe_str(get(row, "dest_city"))
        pod = safe_str(get(row, "pod"))

        shipments.append({
            "week": week,
            "hbl": safe_str(get(row, "hbl")),
            "accepted": accepted,
            "deviation_hours": dev,                       # None for no_t7
            "has_ts": has_ts,
            "failure_type": failure_type,                 # None / outside_48h / no_t7
            "direction": direction,                       # on_time/early/late/no_estimate/unknown
            "origin_country": origin_country,
            "origin_city": origin_city,
            "dest_country": dest_country,
            "dest_city": dest_city,
            "carrier": safe_str(get(row, "carrier")),
            "service": safe_str(get(row, "service")),
            "transport_mode": safe_str(get(row, "transport_mode")),
            "incoterm": safe_str(get(row, "incoterm")),
            "pod": pod,
            "vessel": safe_str(get(row, "vessel")),
            "lane": f"{origin_country} → {dest_country}",
            "city_lane": f"{origin_city} → {dest_city}",
            "delay_reason": safe_str(get(row, "delay_reason")),
        })

    wb.close()
    return shipments


def analyze_dimension(shipments, dim_key, min_count=2):
    groups = defaultdict(lambda: {"total": 0, "accepted": 0, "failed": 0,
                                  "deviations": [], "hbls": [], "no_t7": 0})
    for s in shipments:
        val = s.get(dim_key) or "Unknown"
        g = groups[val]
        g["total"] += 1
        if s["accepted"]:
            g["accepted"] += 1
        else:
            g["failed"] += 1
            if s.get("hbl"):
                g["hbls"].append(s["hbl"])
            if s["failure_type"] == "no_t7":
                g["no_t7"] += 1
        if s["deviation_hours"] is not None:
            g["deviations"].append(s["deviation_hours"])

    results = []
    for val, g in groups.items():
        if g["total"] < min_count:
            continue
        devs = g["deviations"]
        avg_dev = sum(devs) / len(devs) if devs else None
        # late_count / early_count count FAILURES only (|dev| > 48h), so that
        # late_count + early_count + no_t7 == failed. (devs includes accepted
        # within-window shipments, which must NOT be counted as late/early here.)
        late = [d for d in devs if d > 48]
        early = [d for d in devs if d < -48]
        results.append({
            "value": val,
            "total": g["total"],
            "accepted": g["accepted"],
            "failed": g["failed"],
            "no_t7": g["no_t7"],
            "accuracy": round(g["accepted"] / g["total"] * 100, 1),
            "avg_deviation_hours": round(avg_dev, 1) if avg_dev is not None else None,
            "late_count": len(late),
            "early_count": len(early),
            "sample_hbls": g["hbls"][:5],
        })
    results.sort(key=lambda x: x["failed"], reverse=True)
    return results


def deviation_buckets(shipments):
    buckets = {"≤24h": 0, "24-48h": 0, "48-72h": 0, "72-96h": 0, "96h-7d": 0, ">7d": 0, "no_estimate": 0}
    for s in shipments:
        dev = s["deviation_hours"]
        if dev is None:
            buckets["no_estimate"] += 1
            continue
        a = abs(dev)
        if a <= 24: buckets["≤24h"] += 1
        elif a <= 48: buckets["24-48h"] += 1
        elif a <= 72: buckets["48-72h"] += 1
        elif a <= 96: buckets["72-96h"] += 1
        elif a <= 168: buckets["96h-7d"] += 1
        else: buckets[">7d"] += 1
    return buckets


def what_if_windows(shipments):
    """Accuracy at broader windows over the has-TS subset (no_t7 excluded here)."""
    measurable = [s for s in shipments if s["deviation_hours"] is not None]
    total = len(measurable)
    if total == 0:
        return []
    return [
        {
            "window_hours": w,
            "window_label": f"±{w}h ({w//24}d)" if w >= 24 else f"±{w}h",
            "accepted": sum(1 for s in measurable if abs(s["deviation_hours"]) <= w),
            "total": total,
            "accuracy": round(sum(1 for s in measurable if abs(s["deviation_hours"]) <= w) / total * 100, 1),
        }
        for w in [48, 72, 96, 120, 168, 240, 336]
    ]


def direction_split(shipments):
    split = {"on_time": 0, "early": 0, "late": 0, "no_estimate": 0, "unknown": 0}
    for s in shipments:
        split[s["direction"]] = split.get(s["direction"], 0) + 1
    return split


def week_stats(shipments):
    if not shipments:
        return None
    total = len(shipments)
    acc = sum(1 for s in shipments if s["accepted"])
    failed = [s for s in shipments if not s["accepted"]]
    early = [s for s in failed if s["deviation_hours"] is not None and s["deviation_hours"] < 0]
    late = [s for s in failed if s["deviation_hours"] is not None and s["deviation_hours"] > 0]
    no_t7 = [s for s in failed if s["failure_type"] == "no_t7"]
    return {
        "total": total,
        "accepted": acc,
        "failed": total - acc,
        "accuracy": round(acc / total * 100, 1) if total > 0 else None,
        "early_failures": len(early),
        "late_failures": len(late),
        "no_t7_failures": len(no_t7),
        "avg_early_dev_hours": round(sum(s["deviation_hours"] for s in early) / len(early), 1) if early else None,
        "avg_late_dev_hours": round(sum(s["deviation_hours"] for s in late) / len(late), 1) if late else None,
        "deviation_buckets": deviation_buckets(shipments),
        "direction_split": direction_split(shipments),
        "what_if_windows": what_if_windows(shipments),
    }


def build_comparison(prior_ships, current_ships, dim_key, min_count=2):
    prior_map = {r["value"]: r for r in analyze_dimension(prior_ships, dim_key, min_count)}
    current_map = {r["value"]: r for r in analyze_dimension(current_ships, dim_key, min_count)}

    rows = []
    for val in set(prior_map) | set(current_map):
        p = prior_map.get(val)
        c = current_map.get(val)
        p_acc = p["accuracy"] if p else None
        c_acc = c["accuracy"] if c else None
        delta = round(c_acc - p_acc, 1) if (p_acc is not None and c_acc is not None) else None

        if c and p:
            if delta is not None and delta <= -10: status = "worsened"
            elif delta is not None and delta >= 10: status = "improved"
            else: status = "stable"
        elif c and not p:
            status = "new"
        else:
            status = "gone"

        rows.append({
            "value": val,
            "prior_total":    p["total"]    if p else 0,
            "prior_accepted": p["accepted"] if p else 0,
            "prior_failed":   p["failed"]   if p else 0,
            "prior_accuracy": p_acc,
            "current_total":    c["total"]    if c else 0,
            "current_accepted": c["accepted"] if c else 0,
            "current_failed":   c["failed"]   if c else 0,
            "current_accuracy": c_acc,
            "current_no_t7":    c["no_t7"]    if c else 0,
            "delta": delta,
            "status": status,
            "current_avg_dev_hours": c["avg_deviation_hours"] if c else None,
            "current_late_count":    c["late_count"]          if c else 0,
            "current_early_count":   c["early_count"]         if c else 0,
            "sample_hbls": (c["sample_hbls"] if c else []),
        })
    rows.sort(key=lambda x: x["current_failed"], reverse=True)
    return rows


def build_week_block(prior_week, current_week, prior_ships, current_ships):
    return {
        "prior_week": prior_week,
        "current_week": current_week,
        "weekly_stats": {
            "prior":   week_stats(prior_ships),
            "current": week_stats(current_ships),
        },
        "port_discharge_comparison": build_comparison(prior_ships, current_ships, "pod",            min_count=2),
        "lane_comparison":           build_comparison(prior_ships, current_ships, "lane",           min_count=2),
        "country_origin_comparison": build_comparison(prior_ships, current_ships, "origin_country", min_count=2),
        "country_dest_comparison":   build_comparison(prior_ships, current_ships, "dest_country",   min_count=2),
        "service_comparison":        build_comparison(prior_ships, current_ships, "service",        min_count=2),
        "carrier_comparison":        build_comparison(prior_ships, current_ships, "carrier",        min_count=2),
        "vessel_comparison":         build_comparison(prior_ships, current_ships, "vessel",         min_count=2)[:15],
        "city_lane_comparison":      build_comparison(prior_ships, current_ships, "city_lane",      min_count=2)[:40],
        "transport_comparison":      build_comparison(prior_ships, current_ships, "transport_mode", min_count=2),
    }


def main():
    print("=" * 80)
    print("ETA 2P (S07) Lane RCA — per-week comparison against prior week")
    print("=" * 80)

    shipments_by_week = {}
    for w in WEEKS:
        ships = extract_shipments(w)
        if ships is None:
            continue
        shipments_by_week[w] = ships
        total = len(ships)
        acc = sum(1 for s in ships if s["accepted"])
        print(f"  {w}: {acc}/{total} = {acc/total*100:.2f}%" if total else f"  {w}: no arrived rows")

    available = sorted(shipments_by_week.keys())
    week_blocks = {}
    for i, w in enumerate(available):
        if i == 0:
            continue
        prior = available[i - 1]
        week_blocks[w] = build_week_block(prior, w, shipments_by_week[prior], shipments_by_week[w])

    output = {
        "generated": datetime.now().isoformat(),
        "available_weeks": list(week_blocks.keys()),
        "weeks": week_blocks,
    }

    with open(OUT_PATH, "w") as f:
        json.dump(output, f, indent=2, default=str)

    print(f"\n  Weeks with comparison blocks: {list(week_blocks.keys())}")
    print(f"  Exported to: {OUT_PATH}")

    # Reconciliation sanity prints
    for chk in ("CW17", "CW19", "CW20"):
        if chk in week_blocks:
            cur = week_blocks[chk]["weekly_stats"]["current"]
            print(f"  RECONCILE {chk}: {cur['accepted']}/{cur['total']} = {cur['accuracy']}%  (no_t7={cur['no_t7_failures']})")
    if "CW20" in week_blocks:
        print("\n  CW20 port_discharge_comparison (top 8 by failures):")
        for r in week_blocks["CW20"]["port_discharge_comparison"][:8]:
            print(f"    {r['value']:<28} fail {r['current_failed']:>3}/{r['current_total']:<3} acc {r['current_accuracy']}%  Δ {r['delta']}")
    print("Done.")


if __name__ == "__main__":
    main()
