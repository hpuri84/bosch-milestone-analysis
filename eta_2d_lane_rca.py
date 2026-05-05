"""
ETA 2D Lane RCA — per-week comparison against the prior week.

For every available week with a prior week, extracts all measurable SC4 shipments,
computes lane-level accuracy, and produces a comparison block. Output is keyed by
current week so the dashboard can pick the block for whatever week the user selects.

Output: dashboard/public/eta_2d_lane_rca.json with shape:
  {
    "generated": "...",
    "available_weeks": ["CW02", ..., "CW14"],
    "weeks": {
       "CW14": {
         "prior_week": "CW13", "current_week": "CW14",
         "weekly_stats": { "prior": {...}, "current": {...} },
         "lane_comparison":          [ {value, prior_*, current_*, delta, status, ...}, ... ],
         "country_origin_comparison": [...],
         "country_dest_comparison":   [...],
         "carrier_comparison":        [...],
         "service_comparison":        [...],
         "city_lane_comparison":      [...],
         "transport_comparison":      [...]
       },
       ...
    }
  }
"""

import openpyxl
import json
import os
from collections import defaultdict
from datetime import datetime

BASE = os.path.dirname(os.path.abspath(__file__))
RAW_DIR = os.path.join(BASE, "Bosch Milestone raw data")
OUT_PATH = os.path.join(BASE, "dashboard", "public", "eta_2d_lane_rca.json")

# All weeks we might have data for. Script gracefully skips weeks whose files
# are missing, so callers don't need to update this list every week.
WEEKS = [f"CW{i:02d}" for i in range(1, 54)]


def sc4_filename(week):
    # CW01+ all use "Maersk SC4_2026_CWXX.xlsx"
    return f"Maersk SC4_2026_{week}.xlsx"


def safe_str(val):
    return "" if val is None else str(val).strip()


def safe_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(val).strip(), fmt)
        except ValueError:
            continue
    return None


def find_shipments_sheet(wb):
    # Preferred: sheet literally named "shipments".
    for sn in wb.sheetnames:
        if sn.strip().lower() == "shipments":
            return sn
    # Fallback: CW14 SC4 ships with Hungarian default "Munka7". Detect by content marker.
    for sn in wb.sheetnames:
        if sn.strip().upper() in ("TOTAL", "ALL"):
            continue
        if sn.strip().upper().rstrip("_") in ("FCL", "BCO", "LCL"):
            continue
        ws = wb[sn]
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            for cell in row:
                if cell and "UNIQUE_SHIPMENT_ID" in str(cell).upper():
                    return sn
    return None


def find_header_row(ws):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        for cell in row:
            if cell and "UNIQUE_SHIPMENT" in str(cell):
                return i
    return 3


def build_col_map(ws, header_row):
    headers = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
    return {str(h).strip(): i for i, h in enumerate(headers) if h}


def find_col(col_map, *patterns):
    for key in col_map:
        for pat in patterns:
            if pat in key:
                return col_map[key]
    return None


def extract_shipments(week):
    """Extract all measurable SC4 shipments with ETA 2D fields for one week."""
    path = os.path.join(RAW_DIR, sc4_filename(week))
    if not os.path.exists(path):
        return None

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ship_sheet = find_shipments_sheet(wb)
    if not ship_sheet:
        print(f"  WARNING: No shipments sheet in {week}")
        wb.close()
        return []

    ws = wb[ship_sheet]
    header_row = find_header_row(ws)
    col_map = build_col_map(ws, header_row)

    c = {
        "hbl":              find_col(col_map, "HOUSE_BILL_OF_LADING"),
        "carrier":          find_col(col_map, "CARRIER_1"),
        "service":          find_col(col_map, "TRANSPORT_SERVICE_PRIORITY"),
        "transport_mode":   find_col(col_map, "TRANSPORT_MODE"),
        "incoterm":         find_col(col_map, "INCOTERM"),
        "origin_city":      find_col(col_map, "CONSIGNOR_ADDRESS_CITY_NAME"),
        "origin_country":   find_col(col_map, "CONSIGNOR_ADDRESS_COUNTRY"),
        "dest_country":     find_col(col_map, "CONSIGNEE_ADDRESS_COUNTRY"),
        "dest_city":        find_col(col_map, "CONSIGNEE_ADDRESS_CITY_NAME"),
        "delivery_city":    find_col(col_map, "DELIVERY_ADDRESS_CITY_NAME"),
        "delivery_country": find_col(col_map, "DELIVERY_ADDRESS_COUNTRY"),
        "pod":              find_col(col_map, "PORT_OF_DISCHARGE"),
        "delivery_est":     find_col(col_map, "DELIVERY_DATE_ACT_EST_PLAN"),
        "atd":              find_col(col_map, "ATD_DATE_TIME"),
        "ata":              find_col(col_map, "ATA_DATE_TIME"),
        "delivered":        find_col(col_map, "DELIVERED_DATE_TIME"),
        "s31_accepted":     find_col(col_map, "S31_Accepted"),
        "s31_deviation":    find_col(col_map, "S31_Deviation"),
        "s31_ts":           find_col(col_map, "S31_TS_@measured"),
        "delay_reason":     find_col(col_map, "Delay_Reason_Code_Description"),
    }

    def get(row, key):
        idx = c.get(key)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    shipments = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row = list(row)
        if not row or not row[0]:
            continue

        s31_acc = get(row, "s31_accepted")
        delivered = get(row, "delivered")

        # Match rebaseline.py filter: skip unmeasured, skip in-transit non-failures.
        if s31_acc is None:
            continue
        if s31_acc == 0 and delivered is None:
            continue

        hbl = safe_str(get(row, "hbl"))
        if not hbl:
            continue

        s31_dev = safe_float(get(row, "s31_deviation"))
        delivery_est = get(row, "delivery_est")
        delivered_dt = parse_date(delivered)
        est_dt = parse_date(delivery_est)
        ata_dt = parse_date(get(row, "ata"))

        if s31_dev is None and est_dt and delivered_dt:
            s31_dev = (delivered_dt - est_dt).total_seconds() / 3600

        last_mile_days = None
        if ata_dt and delivered_dt:
            last_mile_days = round((delivered_dt - ata_dt).total_seconds() / 86400, 1)

        origin_country = safe_str(get(row, "origin_country"))
        dest_country = safe_str(get(row, "dest_country"))
        origin_city = safe_str(get(row, "origin_city"))
        delivery_city = safe_str(get(row, "delivery_city")) or safe_str(get(row, "dest_city"))
        delivery_country = safe_str(get(row, "delivery_country")) or dest_country

        shipments.append({
            "week": week,
            "hbl": hbl,
            "accepted": (s31_acc == 1),
            "deviation_hours": s31_dev,
            "deviation_days": round(s31_dev / 24, 1) if s31_dev is not None else None,
            "direction": ("late" if s31_dev and s31_dev > 0 else "early" if s31_dev and s31_dev < 0 else "unknown"),
            "origin_country": origin_country,
            "origin_city": origin_city,
            "dest_country": dest_country,
            "delivery_city": delivery_city,
            "delivery_country": delivery_country,
            "carrier": safe_str(get(row, "carrier")),
            "service": safe_str(get(row, "service")),
            "transport_mode": safe_str(get(row, "transport_mode")),
            "incoterm": safe_str(get(row, "incoterm")),
            "pod": safe_str(get(row, "pod")),
            "lane": f"{origin_country} → {dest_country}",
            "city_lane": f"{origin_city} → {delivery_city}",
            "delivery_est": str(est_dt) if est_dt else "",
            "delivered": str(delivered_dt) if delivered_dt else "",
            "last_mile_days": last_mile_days,
            "delay_reason": safe_str(get(row, "delay_reason")),
        })

    wb.close()
    return shipments


def analyze_dimension(shipments, dim_key, min_count=2):
    groups = defaultdict(lambda: {"total": 0, "accepted": 0, "failed": 0,
                                   "deviations": [], "hbls": []})
    for s in shipments:
        val = s.get(dim_key) or "Unknown"
        groups[val]["total"] += 1
        if s["accepted"]:
            groups[val]["accepted"] += 1
        else:
            groups[val]["failed"] += 1
            if s.get("hbl"):
                groups[val]["hbls"].append(s["hbl"])
        if s["deviation_hours"] is not None:
            groups[val]["deviations"].append(s["deviation_hours"])

    results = []
    for val, g in groups.items():
        if g["total"] < min_count:
            continue
        devs = g["deviations"]
        avg_dev = sum(devs) / len(devs) if devs else None
        late = [d for d in devs if d > 0]
        early = [d for d in devs if d < 0]
        results.append({
            "value": val,
            "total": g["total"],
            "accepted": g["accepted"],
            "failed": g["failed"],
            "accuracy": round(g["accepted"] / g["total"] * 100, 1),
            "avg_deviation_hours": round(avg_dev, 1) if avg_dev is not None else None,
            "late_count": len(late),
            "early_count": len(early),
            "sample_hbls": g["hbls"][:5],
        })
    results.sort(key=lambda x: x["failed"], reverse=True)
    return results


def deviation_buckets(shipments):
    buckets = {"≤24h": 0, "24-48h": 0, "48-72h": 0, "72-96h": 0, "96h-7d": 0, ">7d": 0, "no_data": 0}
    for s in shipments:
        dev = s["deviation_hours"]
        if dev is None:
            buckets["no_data"] += 1
            continue
        a = abs(dev)
        if a <= 24: buckets["≤24h"] += 1
        elif a <= 48: buckets["24-48h"] += 1
        elif a <= 72: buckets["48-72h"] += 1
        elif a <= 96: buckets["72-96h"] += 1
        elif a <= 168: buckets["96h-7d"] += 1
        else: buckets[">7d"] += 1
    return buckets


def what_if_windows(shipments):
    measurable = [s for s in shipments if s["deviation_hours"] is not None]
    total = len(measurable)
    if total == 0:
        return []
    return [
        {
            "window_hours": w,
            "window_label": f"±{w}h ({w//24}d)" if w >= 24 else f"±{w}h",
            "accepted": sum(1 for s in measurable if abs(s["deviation_hours"]) <= w),
            "total": total,
            "accuracy": round(sum(1 for s in measurable if abs(s["deviation_hours"]) <= w) / total * 100, 1),
        }
        for w in [48, 72, 96, 120, 168, 240, 336]
    ]


def week_stats(shipments):
    if not shipments:
        return None
    total = len(shipments)
    acc = sum(1 for s in shipments if s["accepted"])
    failed = [s for s in shipments if not s["accepted"]]
    early = [s for s in failed if s["deviation_hours"] is not None and s["deviation_hours"] < 0]
    late = [s for s in failed if s["deviation_hours"] is not None and s["deviation_hours"] > 0]
    return {
        "total": total,
        "accepted": acc,
        "failed": total - acc,
        "accuracy": round(acc / total * 100, 1) if total > 0 else None,
        "early_failures": len(early),
        "late_failures": len(late),
        "avg_early_dev_hours": round(sum(s["deviation_hours"] for s in early) / len(early), 1) if early else None,
        "avg_late_dev_hours": round(sum(s["deviation_hours"] for s in late) / len(late), 1) if late else None,
        "deviation_buckets": deviation_buckets(shipments),
        "what_if_windows": what_if_windows(shipments),
    }


def build_comparison(prior_ships, current_ships, dim_key, min_count=2):
    """Build a per-dimension comparison using generic prior_*/current_* fields."""
    prior_map = {r["value"]: r for r in analyze_dimension(prior_ships, dim_key, min_count)}
    current_map = {r["value"]: r for r in analyze_dimension(current_ships, dim_key, min_count)}

    rows = []
    for val in set(prior_map) | set(current_map):
        p = prior_map.get(val)
        c = current_map.get(val)
        p_acc = p["accuracy"] if p else None
        c_acc = c["accuracy"] if c else None
        delta = round(c_acc - p_acc, 1) if (p_acc is not None and c_acc is not None) else None

        if c and p:
            if delta is not None and delta <= -10: status = "worsened"
            elif delta is not None and delta >= 10: status = "improved"
            else: status = "stable"
        elif c and not p:
            status = "new"
        else:
            status = "gone"

        rows.append({
            "value": val,
            "prior_total":    p["total"]    if p else 0,
            "prior_accepted": p["accepted"] if p else 0,
            "prior_failed":   p["failed"]   if p else 0,
            "prior_accuracy": p_acc,
            "current_total":    c["total"]    if c else 0,
            "current_accepted": c["accepted"] if c else 0,
            "current_failed":   c["failed"]   if c else 0,
            "current_accuracy": c_acc,
            "delta": delta,
            "status": status,
            "current_avg_dev_hours": c["avg_deviation_hours"] if c else None,
            "current_late_count":    c["late_count"]          if c else 0,
            "current_early_count":   c["early_count"]         if c else 0,
            "sample_hbls": (c["sample_hbls"] if c else []),
        })
    rows.sort(key=lambda x: x["current_failed"], reverse=True)
    return rows


def build_week_block(prior_week, current_week, prior_ships, current_ships):
    return {
        "prior_week": prior_week,
        "current_week": current_week,
        "weekly_stats": {
            "prior":   week_stats(prior_ships),
            "current": week_stats(current_ships),
        },
        "lane_comparison":          build_comparison(prior_ships, current_ships, "lane",             min_count=2),
        "country_origin_comparison": build_comparison(prior_ships, current_ships, "origin_country",   min_count=2),
        "country_dest_comparison":   build_comparison(prior_ships, current_ships, "delivery_country", min_count=2),
        "service_comparison":        build_comparison(prior_ships, current_ships, "service",          min_count=2),
        "carrier_comparison":        build_comparison(prior_ships, current_ships, "carrier",          min_count=2),
        "city_lane_comparison":      build_comparison(prior_ships, current_ships, "city_lane",        min_count=2)[:40],
        "transport_comparison":      build_comparison(prior_ships, current_ships, "transport_mode",   min_count=2),
    }


def main():
    print("=" * 80)
    print("ETA 2D Lane RCA — per-week comparison against prior week")
    print("=" * 80)

    # Load all weeks whose files exist
    shipments_by_week = {}
    for w in WEEKS:
        ships = extract_shipments(w)
        if ships is None:
            continue  # file missing
        shipments_by_week[w] = ships
        total = len(ships)
        acc = sum(1 for s in ships if s["accepted"])
        print(f"  {w}: {acc}/{total} = {acc/total*100:.1f}%" if total else f"  {w}: no measurable rows")

    available = sorted(shipments_by_week.keys())
    week_blocks = {}
    for i, w in enumerate(available):
        if i == 0:
            continue  # no prior
        prior = available[i - 1]
        week_blocks[w] = build_week_block(prior, w, shipments_by_week[prior], shipments_by_week[w])

    output = {
        "generated": datetime.now().isoformat(),
        "available_weeks": list(week_blocks.keys()),
        "weeks": week_blocks,
    }

    with open(OUT_PATH, "w") as f:
        json.dump(output, f, indent=2, default=str)

    print(f"\n  Weeks with comparison blocks: {list(week_blocks.keys())}")
    print(f"  Exported to: {OUT_PATH}")
    print("Done.")


if __name__ == "__main__":
    main()
