"""
Deep-dive analysis: S07 (Vessel Arrival) → S31 (Delivery) transit time patterns.

Goal: Replace static POD_EST = ETA + N days with lane-specific intelligent estimates.
Focus: Door delivery shipments where POD timing matters.

Extracts per-shipment ATA and Delivered dates across CW01-CW08 from both SC3 and SC4,
groups by lane dimensions, and finds statistical patterns.
"""

import openpyxl
import os
import json
import statistics
from datetime import datetime, timedelta
from collections import defaultdict

BASE = "/Users/harsh.puri/Documents/work-maersk/Prototype Playground/Bosch Milestone Analysis"
RAW_DIR = os.path.join(BASE, "Bosch Milestone raw data")

WEEKS = [f"CW{i:02d}" for i in range(1, 10)]
SC3_FILES = {f"CW{i:02d}": f"Maersk NGTM SC3_2026_CW{i:02d}.xlsx" for i in range(1, 10)}
SC4_FILES = {f"CW{i:02d}": f"Maersk SC4_2026_CW{i:02d}.xlsx" for i in range(1, 10)}


def find_shipments_sheet(wb):
    for sn in wb.sheetnames:
        if sn.strip().lower() == "shipments":
            return sn
    return None


def valid_dt(val):
    if isinstance(val, datetime) and val.year >= 2020:
        return val
    return None


def extract_sc4_shipments(filepath, week):
    """Extract SC4 shipment-level data for ATA→Delivered analysis."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = find_shipments_sheet(wb)
    if not sheet_name:
        wb.close()
        return []

    ws = wb[sheet_name]
    header_row = 3
    headers = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
    col = {}
    for i, h in enumerate(headers):
        if h:
            col[str(h).strip()] = i

    records = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        vals = list(row)
        if not vals[0]:
            continue

        ata = valid_dt(vals[col.get("ATA_DATE_TIME", 79)])
        delivered = valid_dt(vals[col.get("DELIVERED_DATE_TIME", 81)])
        eta = valid_dt(vals[col.get("ETA_DATE_TIME", 78)])
        pod_est = valid_dt(vals[col.get("DELIVERY_DATE_ACT_EST_PLAN", 71)])

        # Need both ATA and Delivered to compute delta
        if not ata or not delivered:
            continue

        delta_days = (delivered - ata).total_seconds() / 86400

        # Skip negative deltas (data errors) and extreme outliers > 90 days
        if delta_days < 0 or delta_days > 90:
            continue

        def s(idx):
            v = vals[idx] if idx is not None and idx < len(vals) else None
            return str(v).strip() if v else ""

        hbl = s(col.get("HOUSE_BILL_OF_LADING", 1))
        pickup_country = s(col.get("PICKUP_ADDRESS_COUNTRY", 44))
        pickup_city = s(col.get("PICKUP_ADDRESS_CITY_NAME", 45))
        delivery_country = s(col.get("DELIVERY_ADDRESS_COUNTRY", 55))
        delivery_city = s(col.get("DELIVERY_ADDRESS_CITY_NAME", 53))
        consignor_country = s(col.get("CONSIGNOR_ADDRESS_COUNTRY", 35))
        consignor_city = s(col.get("CONSIGNOR_ADDRESS_CITY_NAME", 34))
        consignee_country = s(col.get("CONSIGNEE_ADDRESS_COUNTRY", 39))
        consignee_city = s(col.get("CONSIGNEE_ADDRESS_CITY_NAME", 40))
        port_discharge = s(col.get("PORT_OF_DISCHARGE_LOCATION_ADDRESS_EXT_ID", 56))
        carrier = s(col.get("CARRIER_1", 8))
        transport_priority = s(col.get("TRANSPORT_SERVICE_PRIORITY", 15))
        incoterm = s(col.get("INCOTERM", 19))
        transport_mode = s(col.get("„QUERY_NON_TMC_IFTSTA_Shipments\"[TRANSPORT_MODE]", 13))
        equipment = s(col.get("EQUIPMENT_SIZE_AND_TYPE_DESCRIPTION", 63))

        # ETA accuracy if both available
        eta_to_ata = None
        if eta and ata:
            eta_to_ata = round((ata - eta).total_seconds() / 86400, 2)

        # POD_EST accuracy
        pod_est_error = None
        if pod_est and delivered:
            pod_est_error = round((delivered - pod_est).total_seconds() / 86400, 2)

        records.append({
            "week": week,
            "scenario": "SC4",
            "hbl": hbl,
            "pickup_country": pickup_country,
            "pickup_city": pickup_city,
            "delivery_country": delivery_country,
            "delivery_city": delivery_city,
            "consignor_country": consignor_country,
            "consignor_city": consignor_city,
            "consignee_country": consignee_country,
            "consignee_city": consignee_city,
            "port_discharge": port_discharge,
            "carrier": carrier,
            "transport_priority": transport_priority,
            "incoterm": incoterm,
            "transport_mode": transport_mode,
            "equipment": equipment,
            "ata": ata.isoformat(),
            "delivered": delivered.isoformat(),
            "eta": eta.isoformat() if eta else None,
            "pod_est": pod_est.isoformat() if pod_est else None,
            "ata_to_delivered_days": round(delta_days, 2),
            "eta_to_ata_days": eta_to_ata,
            "pod_est_error_days": pod_est_error,
        })

    wb.close()
    return records


def extract_sc3_shipments(filepath, week):
    """Extract SC3 shipment-level data for ATA→Delivered analysis."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = find_shipments_sheet(wb)
    if not sheet_name:
        wb.close()
        return []

    ws = wb[sheet_name]
    header_row = 3
    headers = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
    col = {}
    for i, h in enumerate(headers):
        if h:
            col[str(h).strip()] = i

    records = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        vals = list(row)
        if not vals[0]:
            continue

        def safe(col_name, fallback):
            idx = col.get(col_name, fallback)
            if idx is not None and idx < len(vals):
                return valid_dt(vals[idx])
            return None

        ata = safe("ATA_Datetime", 92)
        delivered = safe("Delivered_Sent_DATE_TIME", 101)
        eta = safe("ETA_Datetime", 91)

        if not ata or not delivered:
            continue

        delta_days = (delivered - ata).total_seconds() / 86400
        if delta_days < 0 or delta_days > 90:
            continue

        def s(col_name, fallback):
            idx = col.get(col_name, fallback)
            v = vals[idx] if idx is not None and idx < len(vals) else None
            return str(v).strip() if v else ""

        hbl = s("House_Airway_Bill_or_House_Bill_of_Lading", 86)
        load_to = s("LOAD_TO", 2)
        consignor = s("Consignor_Name", 8)
        consignor_addr = s("Consignor_Address", 9)
        recipient = s("Recipient_Name", 11)
        recipient_addr = s("Recipient_Address", 12)
        pickup_name = s("TO_Pick_Up_Name", 14)
        delivery_name = s("TO_Delivery_Name", 17)
        delivery_addr = s("TO_Delivery_Address", 18)
        carrier = s("Main_Carrier_or_Shipping_Line", 87)
        port_loading = s("Airport_of_Departure_or_Port_of_Loading", 89)
        port_discharge = s("Airport_of_Destination_or_Port_of_Discharge", 90)
        service_to = s("Service_TO", 82)
        service_load = s("Service_Load", 83)
        pickup_country = s("Leg_Pick_up_Country", 32)
        pickup_city = s("Leg_Pick_up_City", 31)
        delivery_country = s("Leg_Delivery_Country", 37)
        delivery_city = s("Leg_Delivery_City", 36)

        eta_to_ata = None
        if eta and ata:
            eta_to_ata = round((ata - eta).total_seconds() / 86400, 2)

        records.append({
            "week": week,
            "scenario": "SC3",
            "hbl": hbl,
            "pickup_country": pickup_country,
            "pickup_city": pickup_city,
            "delivery_country": delivery_country,
            "delivery_city": delivery_city,
            "consignor_country": "",
            "consignor_city": consignor,
            "consignee_country": "",
            "consignee_city": recipient,
            "port_discharge": port_discharge,
            "carrier": carrier,
            "transport_priority": service_load,
            "incoterm": "",
            "transport_mode": "Sea",
            "equipment": "",
            "ata": ata.isoformat(),
            "delivered": delivered.isoformat(),
            "eta": eta.isoformat() if eta else None,
            "pod_est": None,
            "ata_to_delivered_days": round(delta_days, 2),
            "eta_to_ata_days": eta_to_ata,
            "pod_est_error_days": None,
        })

    wb.close()
    return records


def compute_stats(values):
    """Compute statistical summary for a list of numeric values."""
    if not values:
        return None
    n = len(values)
    vals = sorted(values)
    return {
        "count": n,
        "mean": round(statistics.mean(vals), 1),
        "median": round(statistics.median(vals), 1),
        "std": round(statistics.stdev(vals), 1) if n > 1 else 0,
        "min": round(vals[0], 1),
        "max": round(vals[-1], 1),
        "p10": round(vals[int(n * 0.10)], 1),
        "p25": round(vals[int(n * 0.25)], 1),
        "p75": round(vals[int(n * 0.75)], 1),
        "p90": round(vals[min(int(n * 0.90), n - 1)], 1),
        "recommended_offset": round(statistics.median(vals) + (statistics.stdev(vals) * 0.5 if n > 1 else 2), 1),
    }


def main():
    print("=" * 80)
    print("S07 → S31 TRANSIT PATTERN ANALYSIS")
    print("ATA (Vessel Arrival) → Delivered (Door Delivery)")
    print("=" * 80)

    all_records = []

    for week in WEEKS:
        sc4_file = os.path.join(RAW_DIR, SC4_FILES[week])
        sc3_file = os.path.join(RAW_DIR, SC3_FILES[week])

        sc4_recs = extract_sc4_shipments(sc4_file, week) if os.path.exists(sc4_file) else []
        sc3_recs = extract_sc3_shipments(sc3_file, week) if os.path.exists(sc3_file) else []

        print(f"  {week}: SC4={len(sc4_recs)}, SC3={len(sc3_recs)} shipments with ATA+Delivered")
        all_records.extend(sc4_recs)
        all_records.extend(sc3_recs)

    print(f"\nTotal records: {len(all_records)}")

    # ---- Overall Stats ----
    all_deltas = [r["ata_to_delivered_days"] for r in all_records]
    overall = compute_stats(all_deltas)
    print(f"\n{'='*60}")
    print(f"OVERALL ATA → Delivered: median={overall['median']}d, mean={overall['mean']}d, std={overall['std']}d")
    print(f"  P10={overall['p10']}d  P25={overall['p25']}d  P75={overall['p75']}d  P90={overall['p90']}d")
    print(f"  Recommended offset: {overall['recommended_offset']}d")

    # ---- By Scenario ----
    print(f"\n{'='*60}")
    print("BY SCENARIO")
    for scenario in ["SC3", "SC4"]:
        vals = [r["ata_to_delivered_days"] for r in all_records if r["scenario"] == scenario]
        if vals:
            s = compute_stats(vals)
            print(f"  {scenario}: n={s['count']}, median={s['median']}d, mean={s['mean']}d, std={s['std']}d, P90={s['p90']}d")

    # ---- By Incoterm (SC4 only, to identify door shipments) ----
    print(f"\n{'='*60}")
    print("BY INCOTERM (SC4 only)")
    incoterm_groups = defaultdict(list)
    for r in all_records:
        if r["scenario"] == "SC4" and r["incoterm"]:
            incoterm_groups[r["incoterm"]].append(r["ata_to_delivered_days"])
    for inco in sorted(incoterm_groups, key=lambda x: -len(incoterm_groups[x])):
        vals = incoterm_groups[inco]
        s = compute_stats(vals)
        print(f"  {inco}: n={s['count']}, median={s['median']}d, mean={s['mean']}d, std={s['std']}d, P90={s['p90']}d")

    # ---- By Transport Priority ----
    print(f"\n{'='*60}")
    print("BY TRANSPORT SERVICE PRIORITY")
    tsp_groups = defaultdict(list)
    for r in all_records:
        if r["transport_priority"]:
            tsp_groups[r["transport_priority"]].append(r["ata_to_delivered_days"])
    for tsp in sorted(tsp_groups, key=lambda x: -len(tsp_groups[x])):
        vals = tsp_groups[tsp]
        s = compute_stats(vals)
        print(f"  {tsp}: n={s['count']}, median={s['median']}d, mean={s['mean']}d, std={s['std']}d, P90={s['p90']}d")

    # ---- By Port of Discharge ----
    print(f"\n{'='*60}")
    print("BY PORT OF DISCHARGE")
    pod_groups = defaultdict(list)
    for r in all_records:
        if r["port_discharge"]:
            pod_groups[r["port_discharge"]].append(r["ata_to_delivered_days"])
    for port in sorted(pod_groups, key=lambda x: -len(pod_groups[x])):
        vals = pod_groups[port]
        s = compute_stats(vals)
        print(f"  {port}: n={s['count']}, median={s['median']}d, mean={s['mean']}d, std={s['std']}d, [P25={s['p25']}, P75={s['p75']}, P90={s['p90']}]")

    # ---- By Origin Country → Delivery Country ----
    print(f"\n{'='*60}")
    print("BY LANE (Origin Country → Delivery Country)")
    lane_groups = defaultdict(list)
    for r in all_records:
        orig = r["pickup_country"] or r["consignor_country"]
        dest = r["delivery_country"] or r["consignee_country"]
        if orig and dest:
            lane_groups[f"{orig} → {dest}"].append(r)
    for lane in sorted(lane_groups, key=lambda x: -len(lane_groups[x])):
        recs = lane_groups[lane]
        vals = [r["ata_to_delivered_days"] for r in recs]
        s = compute_stats(vals)
        if s["count"] >= 3:
            print(f"  {lane}: n={s['count']}, median={s['median']}d, mean={s['mean']}d, std={s['std']}d, recommended={s['recommended_offset']}d")

    # ---- By Port of Discharge → Delivery City (most granular lane) ----
    print(f"\n{'='*60}")
    print("BY LANE (Port of Discharge → Delivery City)")
    fine_lane_groups = defaultdict(list)
    for r in all_records:
        port = r["port_discharge"]
        city = r["delivery_city"] or r["consignee_city"]
        if port and city:
            fine_lane_groups[f"{port} → {city}"].append(r)
    for lane in sorted(fine_lane_groups, key=lambda x: -len(fine_lane_groups[x])):
        recs = fine_lane_groups[lane]
        vals = [r["ata_to_delivered_days"] for r in recs]
        s = compute_stats(vals)
        if s["count"] >= 3:
            print(f"  {lane}: n={s['count']}, median={s['median']}d, mean={s['mean']}d, std={s['std']}d, recommended={s['recommended_offset']}d")

    # ---- By Carrier ----
    print(f"\n{'='*60}")
    print("BY CARRIER")
    carrier_groups = defaultdict(list)
    for r in all_records:
        if r["carrier"]:
            carrier_groups[r["carrier"]].append(r["ata_to_delivered_days"])
    for car in sorted(carrier_groups, key=lambda x: -len(carrier_groups[x])):
        vals = carrier_groups[car]
        s = compute_stats(vals)
        if s["count"] >= 5:
            print(f"  {car}: n={s['count']}, median={s['median']}d, mean={s['mean']}d, std={s['std']}d")

    # ---- POD_EST accuracy analysis (SC4 only) ----
    print(f"\n{'='*60}")
    print("CURRENT POD_EST ACCURACY (SC4 — how wrong is the static estimate?)")
    pod_est_errors = [r["pod_est_error_days"] for r in all_records if r["pod_est_error_days"] is not None]
    if pod_est_errors:
        s = compute_stats([abs(e) for e in pod_est_errors])
        positive = sum(1 for e in pod_est_errors if e > 0)
        negative = sum(1 for e in pod_est_errors if e < 0)
        zero = sum(1 for e in pod_est_errors if e == 0)
        print(f"  Total with POD_EST: {len(pod_est_errors)}")
        print(f"  Delivered AFTER estimate (late): {positive} ({positive/len(pod_est_errors)*100:.1f}%)")
        print(f"  Delivered BEFORE estimate (early): {negative} ({negative/len(pod_est_errors)*100:.1f}%)")
        print(f"  Exact: {zero}")
        print(f"  Absolute error: median={s['median']}d, mean={s['mean']}d, P90={s['p90']}d")

        # Error by incoterm
        print(f"\n  POD_EST Error by Incoterm:")
        for inco in sorted(incoterm_groups, key=lambda x: -len(incoterm_groups[x])):
            errs = [r["pod_est_error_days"] for r in all_records
                    if r["scenario"] == "SC4" and r["incoterm"] == inco and r["pod_est_error_days"] is not None]
            if len(errs) >= 3:
                abs_errs = [abs(e) for e in errs]
                s = compute_stats(abs_errs)
                late = sum(1 for e in errs if e > 0)
                print(f"    {inco}: n={len(errs)}, abs_median={s['median']}d, abs_mean={s['mean']}d, late={late}/{len(errs)}")

    # ---- Build lane recommendation table ----
    print(f"\n{'='*60}")
    print("LANE RECOMMENDATION TABLE (Port → Delivery City, n≥5)")
    print(f"{'Lane':<40} {'N':>5} {'Median':>8} {'Mean':>8} {'Std':>6} {'P90':>6} {'Recommended':>12}")
    print("-" * 90)

    recommendations = []
    for lane in sorted(fine_lane_groups, key=lambda x: -len(fine_lane_groups[x])):
        recs = fine_lane_groups[lane]
        vals = [r["ata_to_delivered_days"] for r in recs]
        s = compute_stats(vals)
        if s["count"] >= 5:
            print(f"  {lane:<38} {s['count']:>5} {s['median']:>8.1f} {s['mean']:>8.1f} {s['std']:>6.1f} {s['p90']:>6.1f} {s['recommended_offset']:>12.1f}")
            recommendations.append({
                "lane": lane,
                "port_discharge": lane.split(" → ")[0],
                "delivery_city": lane.split(" → ")[1] if " → " in lane else "",
                **s,
            })

    # Also build by origin country → delivery country for fallback
    country_recommendations = []
    for lane in sorted(lane_groups, key=lambda x: -len(lane_groups[x])):
        recs = lane_groups[lane]
        vals = [r["ata_to_delivered_days"] for r in recs]
        s = compute_stats(vals)
        if s["count"] >= 5:
            country_recommendations.append({
                "lane": lane,
                "origin_country": lane.split(" → ")[0],
                "delivery_country": lane.split(" → ")[1] if " → " in lane else "",
                **s,
            })

    # ---- Export ----
    output = {
        "extraction_date": datetime.now().isoformat(),
        "total_records": len(all_records),
        "overall_stats": overall,
        "scenario_stats": {},
        "incoterm_stats": {},
        "transport_priority_stats": {},
        "port_discharge_stats": {},
        "lane_recommendations": recommendations,
        "country_lane_recommendations": country_recommendations,
        "pod_est_accuracy": {
            "total": len(pod_est_errors),
            "late_count": sum(1 for e in pod_est_errors if e > 0),
            "early_count": sum(1 for e in pod_est_errors if e < 0),
            "abs_error_stats": compute_stats([abs(e) for e in pod_est_errors]) if pod_est_errors else None,
        },
        "records": all_records,
    }

    # Fill grouped stats
    for scenario in ["SC3", "SC4"]:
        vals = [r["ata_to_delivered_days"] for r in all_records if r["scenario"] == scenario]
        if vals:
            output["scenario_stats"][scenario] = compute_stats(vals)

    for inco, vals in incoterm_groups.items():
        output["incoterm_stats"][inco] = compute_stats(vals)

    for tsp, vals in tsp_groups.items():
        output["transport_priority_stats"][tsp] = compute_stats(vals)

    for port, vals in pod_groups.items():
        if len(vals) >= 3:
            output["port_discharge_stats"][port] = compute_stats(vals)

    out_path = os.path.join(BASE, "pod_pattern_analysis.json")
    with open(out_path, "w") as f:
        json.dump(output, f, indent=2, default=str)
    print(f"\nExported to: {out_path}")

    # Also export to dashboard
    dash_path = os.path.join(BASE, "dashboard", "public", "pod_pattern_analysis.json")
    with open(dash_path, "w") as f:
        json.dump(output, f, indent=2, default=str)
    print(f"Dashboard copy: {dash_path}")

    print("\nDone.")


if __name__ == "__main__":
    main()
