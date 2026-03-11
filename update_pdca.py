"""
PDCA Excel Template Updater

Reads the Bosch PDCA template and populates a new week column with
KPI data extracted from the raw SC3/SC4 files.

Usage:
    python update_pdca.py [--week CW09] [--template path] [--output path]

If --week is omitted, auto-detects the next week from the template.
"""

import argparse
import json
import os
from datetime import datetime

import re

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

BASE = "/Users/harsh.puri/Documents/work-maersk/Prototype Playground/Bosch Milestone Analysis"
DEFAULT_TEMPLATE = os.path.join(BASE, "Bosch pdca template - milestones.xlsx")
KPI_DATA = os.path.join(BASE, "dashboard/public/kpi_data.json")
RCA_DATA = os.path.join(BASE, "dashboard/public/rca_data.json")


def load_json(path):
    with open(path) as f:
        return json.load(f)


def find_next_week_col(ws):
    """Find the next empty column in row 7 (week header row) of PDCA Main Page."""
    for col in range(7, 30):
        val = ws.cell(7, col).value
        if val is None or str(val).strip() == "":
            return col
    return None


def find_week_col(ws, week_label):
    """Find a specific week column by label in row 7."""
    for col in range(7, 30):
        val = ws.cell(7, col).value
        if val and week_label.lower().replace("cw", "w") in str(val).lower():
            return col
    return None


def update_main_page(wb, week_data, rca_data_week, week_label, col, sc4_new_col, sc3_new_col):
    """Update PDCA Main Page with header KPIs and milestone-level formulas."""
    ws = wb["PDCA Main Page"]

    # Row 7: Week header
    ws.cell(7, col, f"W{week_label.replace('CW', '')}")

    # Rows 8-14: Header KPIs (direct values)
    ws.cell(8, col, week_data["critical"]["completeness"])    # Completeness (Critical)
    ws.cell(9, col, week_data["critical"]["timeliness"])      # Timeliness (Critical)
    ws.cell(10, col, week_data["all"]["completeness"])        # Completeness (All)
    ws.cell(11, col, week_data["all"]["timeliness"])          # Timeliness (All)
    ws.cell(12, col, week_data.get("eta_2p"))                 # ETA accuracy 2P
    ws.cell(13, col, week_data.get("eta_2d"))                 # ETA Accuracy 2D
    ws.cell(14, col, week_data.get("ref_comp"))               # Reference Completeness

    # Rows 18+: Write formulas referencing the new SC Data week columns.
    # Parse the W01 formula in col 7 (e.g. "='SC4 Data'!F5") to get the sheet/row,
    # then rewrite with the new week's column letter.
    # W01 uses col F(6)=Completeness, G(7)=Timeliness. Each week block is 5 cols wide.
    # So for new week: completeness_col = sc_new_col + 3, timeliness_col = sc_new_col + 4.
    sc4_comp_letter = get_column_letter(sc4_new_col + 3)
    sc4_time_letter = get_column_letter(sc4_new_col + 4)
    sc3_comp_letter = get_column_letter(sc3_new_col + 3)
    sc3_time_letter = get_column_letter(sc3_new_col + 4)

    formula_pattern = re.compile(r"='(SC[34] Data)'!([A-Z]+)(\d+)")
    filled = 0

    for row in range(18, ws.max_row + 1):
        w01_formula = ws.cell(row, 7).value
        if not w01_formula or not str(w01_formula).startswith("="):
            continue

        m = formula_pattern.match(str(w01_formula))
        if not m:
            continue

        sheet_name = m.group(1)
        orig_col_letter = m.group(2)
        data_row = m.group(3)

        # Determine if this is a completeness or timeliness reference
        orig_col_idx = column_index_from_string(orig_col_letter)
        # W01: completeness=F(6), timeliness=G(7). Offset from week start(3): comp=+3, time=+4
        is_timeliness = (orig_col_idx % 5 == 2)  # G=7, L=12, etc: 7%5=2, 12%5=2

        if sheet_name == "SC4 Data":
            new_letter = sc4_time_letter if is_timeliness else sc4_comp_letter
        else:
            new_letter = sc3_time_letter if is_timeliness else sc3_comp_letter

        ws.cell(row, col, f"='{sheet_name}'!{new_letter}{data_row}")
        filled += 1

    print(f"  Wrote {filled} milestone formulas")


def find_sc_data_week_cols(ws):
    """Find how many weeks exist in SC data sheet and return next week start column."""
    # Headers in row 3: "Week 1" at col 3, "Week 2" at col 8, etc.
    # Each week block is 5 columns: Required, Available, In Time, Completeness, Timeliness
    max_week_col = 3
    for col in range(3, 100, 5):
        val = ws.cell(3, col).value
        if val and "week" in str(val).lower():
            max_week_col = col
        else:
            break
    return max_week_col + 5  # Next available block


def update_sc_data(wb, sheet_name, scenario, rca_data_week, week_num):
    """Update SC3 Data or SC4 Data sheet with milestone-level numbers. Returns next_col."""
    if sheet_name not in wb.sheetnames:
        print(f"  Warning: Sheet '{sheet_name}' not found, skipping")
        return None

    ws = wb[sheet_name]
    milestones = rca_data_week.get("milestones", []) if rca_data_week else []

    # Find next column block
    next_col = find_sc_data_week_cols(ws)

    # Write week header
    ws.cell(3, next_col, f"Week {week_num}")

    # Write column sub-headers
    sub_headers = ["Required", "Available", "In Time", "Completeness", "Timeliness"]
    for i, h in enumerate(sub_headers):
        ws.cell(4, next_col + i, h)

    # Build milestone lookup for this scenario
    ms_lookup = {}
    for m in milestones:
        if m["scenario"] == scenario:
            key = (m["code"], m["type"])
            ms_lookup[key] = m

    # Match rows 5+ by milestone name and type in columns A and B
    for row in range(5, ws.max_row + 1):
        ms_desc = ws.cell(row, 1).value
        ms_type = ws.cell(row, 2).value
        if not ms_desc:
            continue

        ms_desc = str(ms_desc).strip()
        ms_type = str(ms_type).strip() if ms_type else ""

        # Extract code from description like "S00 - Shipment created"
        code = None
        for c in ["S00", "S02", "S04", "S05", "S07", "S10", "S11", "S12", "S13",
                   "S16", "S17", "S18", "S31", "S45", "S46", "S50", "S51", "S52",
                   "S53", "S54", "S55", "S60"]:
            if c in ms_desc:
                code = c
                break

        if not code:
            continue

        # Determine type
        type_key = "Actual"
        if "estimated" in ms_type.lower():
            type_key = "Estimated"

        m = ms_lookup.get((code, type_key))
        if m:
            ws.cell(row, next_col, m["required"])
            ws.cell(row, next_col + 1, m["available"])
            ws.cell(row, next_col + 2, m["in_time"])
            ws.cell(row, next_col + 3, m["completeness"])
            ws.cell(row, next_col + 4, m["timeliness"])

    return next_col


def main():
    parser = argparse.ArgumentParser(description="Update PDCA template with latest KPI data")
    parser.add_argument("--week", help="Week to update (e.g. CW09). Auto-detects if omitted.")
    parser.add_argument("--template", default=DEFAULT_TEMPLATE, help="Path to PDCA template")
    parser.add_argument("--output", help="Output path. Defaults to timestamped copy.")
    args = parser.parse_args()

    print("Loading data...")
    kpi_data = load_json(KPI_DATA)
    rca_data = load_json(RCA_DATA)

    # Determine which week to update
    if args.week:
        week = args.week.upper()
    else:
        week = kpi_data[-1]["week"]

    week_kpi = next((d for d in kpi_data if d["week"] == week), None)
    week_rca = next((d for d in rca_data if d["week"] == week), None)

    if not week_kpi:
        print(f"Error: No KPI data found for {week}")
        return

    print(f"Updating PDCA template for {week}...")
    print(f"  Template: {args.template}")

    wb = openpyxl.load_workbook(args.template)

    # 1. Update PDCA Main Page
    ws_main = wb["PDCA Main Page"]
    col = find_week_col(ws_main, week)
    if not col:
        col = find_next_week_col(ws_main)
    if not col:
        print("Error: No available column in PDCA Main Page")
        return

    print(f"  Writing to column {col} ({get_column_letter(col)})")

    # 2. Update SC4 Data sheet first (Main Page formulas reference these)
    week_num = int(week.replace("CW", ""))
    print(f"  Updating SC4 Data...")
    sc4_new_col = update_sc_data(wb, "SC4 Data", "SC4", week_rca, week_num)

    # 3. Update SC3 Data sheet
    print(f"  Updating SC3 Data...")
    sc3_new_col = update_sc_data(wb, "SC3 Data", "SC3", week_rca, week_num)

    # 4. Update PDCA Main Page (header KPIs + milestone formulas)
    update_main_page(wb, week_kpi, week_rca, week, col, sc4_new_col, sc3_new_col)

    # Save
    if args.output:
        out_path = args.output
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        out_path = os.path.join(BASE, f"Bosch pdca template - milestones_{week}_{ts}.xlsx")

    wb.save(out_path)
    print(f"\nSaved: {out_path}")
    print("Done.")


if __name__ == "__main__":
    main()
