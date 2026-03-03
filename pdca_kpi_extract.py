#!/usr/bin/env python3
"""
Extract PDCA tracker KPIs from Bosch weekly raw data files.
Computes completeness & timeliness for W01, W03, W05 (both SC3 and SC4).
Also analyzes reference data, ETA accuracy, and plausibility from shipment-level sheets.

TODO: Update methodology for KPI calculations — logic to be provided later.
"""

import openpyxl
import os
import sys
from collections import defaultdict

BASE = "Bosch Milestone raw data"
WEEKS = ["CW01", "CW03", "CW05"]

# Critical milestones (Actual only) - the key transport chain milestones
CRITICAL_MILESTONES = {"S02", "S04", "S07", "S31"}

# All milestones we consider (Actual type only for completeness/timeliness base)
# We'll also track Estimated separately

def find_total_sheet(wb):
    """Find the TOTAL sheet name (may vary)."""
    for sn in wb.sheetnames:
        if "TOTAL" in sn.upper():
            return sn
    return wb.sheetnames[0]

def find_shipment_sheet(wb):
    """Find the shipment-level sheet."""
    for sn in wb.sheetnames:
        if "shipment" in sn.lower():
            return sn
    return None

def read_total_sheet(filepath):
    """Read the TOTAL summary sheet - returns list of milestone dicts."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    total_sn = find_total_sheet(wb)
    ws = wb[total_sn]
    
    milestones = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if row[0] and "Required status" in str(row[0]):
            header_found = True
            continue
        if header_found and row[0] and row[3] is not None:
            name = str(row[0]).strip()
            status_type = str(row[1]).strip() if row[1] else "Unknown"
            required = float(row[3]) if row[3] else 0
            available = float(row[4]) if row[4] else 0
            in_time = float(row[5]) if row[5] else 0
            comp_ratio = float(row[6]) if row[6] else 0
            time_ratio = float(row[7]) if row[7] else 0
            
            # Extract status code (e.g., "S02" from "S02 - Collected")
            code = name.split(" - ")[0].strip() if " - " in name else name.strip()
            
            milestones.append({
                "name": name,
                "code": code,
                "status_type": status_type,
                "required": required,
                "available": available,
                "in_time": in_time,
                "completeness": comp_ratio,
                "timeliness": time_ratio
            })
    wb.close()
    return milestones

def read_shipment_sheet(filepath, sheet_type="SC3"):
    """Read shipment-level data for reference, ETA, and plausibility checks."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ship_sn = find_shipment_sheet(wb)
    if not ship_sn:
        wb.close()
        return None, None
    
    ws = wb[ship_sn]
    headers = None
    rows_data = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        vals = list(row)
        if headers is None:
            # Look for header row
            if any(v and "LOAD_ID" in str(v) for v in vals if v):
                headers = [str(v).strip() if v else f"col_{j}" for j, v in enumerate(vals)]
                continue
            elif any(v and "CONSIGNMENT" in str(v) for v in vals if v):
                headers = [str(v).strip() if v else f"col_{j}" for j, v in enumerate(vals)]
                continue
        elif headers:
            rows_data.append(dict(zip(headers, vals)))
    
    wb.close()
    return headers, rows_data

def read_ref_sheets(filepath):
    """Read FCL REF / BCO REF / LCL ref sheets for reference data analysis."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ref_sheets = [sn for sn in wb.sheetnames if "ref" in sn.lower()]
    
    all_ref_data = []
    for sn in ref_sheets:
        ws = wb[sn]
        headers = None
        for row in ws.iter_rows(values_only=True):
            vals = list(row)
            if headers is None:
                if any(v and ("CONSIGNMENT" in str(v) or "HAWB" in str(v) or "HBL" in str(v)) for v in vals if v):
                    headers = [str(v).strip() if v else f"col_{j}" for j, v in enumerate(vals)]
                    continue
            elif headers:
                all_ref_data.append(dict(zip(headers, vals)))
    wb.close()
    return all_ref_data

def read_detail_sheets(filepath, sheet_type="SC3"):
    """Read FCL/LCL/BCO detail sheets for shipment-level milestone analysis."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    
    # For SC3: LCL, FCL, BCO sheets; For SC4: FCL, BCO, LCL sheets
    detail_sheets = []
    for sn in wb.sheetnames:
        sn_lower = sn.lower().strip()
        if sn_lower in ("fcl", "lcl", "bco") or (sn_lower.startswith("fcl") and "ref" not in sn_lower) or (sn_lower.startswith("lcl") and "ref" not in sn_lower) or (sn_lower.startswith("bco") and "ref" not in sn_lower):
            if "ref" not in sn_lower:
                detail_sheets.append(sn)
    
    all_records = []
    for sn in detail_sheets:
        ws = wb[sn]
        headers = None
        for row in ws.iter_rows(values_only=True):
            vals = list(row)
            if headers is None:
                if any(v and ("CONSIGNMENT" in str(v) or "LOAD_TO" in str(v)) for v in vals if v):
                    headers = [str(v).strip() if v else f"col_{j}" for j, v in enumerate(vals)]
                    continue
            elif headers:
                rec = dict(zip(headers, vals))
                rec["_sheet"] = sn
                all_records.append(rec)
    wb.close()
    return all_records

def compute_kpis(sc3_milestones, sc4_milestones):
    """Compute the 9 KPI metrics from milestone summary data."""
    results = {}
    
    # --- Problem 1: Critical Completeness (Actual only) ---
    # Weighted average: sum(available) / sum(required) for critical milestones
    crit_required = 0
    crit_available = 0
    for m in sc3_milestones + sc4_milestones:
        if m["code"] in CRITICAL_MILESTONES and m["status_type"] == "Actual":
            crit_required += m["required"]
            crit_available += m["available"]
    results["critical_completeness"] = (crit_available / crit_required * 100) if crit_required > 0 else 0
    
    # --- Problem 2: Critical Timeliness (Actual only) ---
    crit_in_time = 0
    for m in sc3_milestones + sc4_milestones:
        if m["code"] in CRITICAL_MILESTONES and m["status_type"] == "Actual":
            crit_in_time += m["in_time"]
    results["critical_timeliness"] = (crit_in_time / crit_required * 100) if crit_required > 0 else 0
    
    # --- Problem 8: All Milestones Completeness (weighted average) ---
    all_required = 0
    all_available = 0
    for m in sc3_milestones + sc4_milestones:
        all_required += m["required"]
        all_available += m["available"]
    results["all_completeness"] = (all_available / all_required * 100) if all_required > 0 else 0
    
    # --- Problem 9: All Milestones Timeliness (weighted average) ---
    all_in_time = 0
    for m in sc3_milestones + sc4_milestones:
        all_in_time += m["in_time"]
    results["all_timeliness"] = (all_in_time / all_required * 100) if all_required > 0 else 0
    
    # --- Also compute SC3-only and SC4-only for reference ---
    for label, ms_list in [("sc3", sc3_milestones), ("sc4", sc4_milestones)]:
        req = sum(m["required"] for m in ms_list)
        avail = sum(m["available"] for m in ms_list)
        itime = sum(m["in_time"] for m in ms_list)
        results[f"{label}_completeness"] = (avail / req * 100) if req > 0 else 0
        results[f"{label}_timeliness"] = (itime / req * 100) if req > 0 else 0
        
        # Critical only
        c_req = sum(m["required"] for m in ms_list if m["code"] in CRITICAL_MILESTONES and m["status_type"] == "Actual")
        c_avail = sum(m["available"] for m in ms_list if m["code"] in CRITICAL_MILESTONES and m["status_type"] == "Actual")
        c_itime = sum(m["in_time"] for m in ms_list if m["code"] in CRITICAL_MILESTONES and m["status_type"] == "Actual")
        results[f"{label}_crit_completeness"] = (c_avail / c_req * 100) if c_req > 0 else 0
        results[f"{label}_crit_timeliness"] = (c_itime / c_req * 100) if c_req > 0 else 0
        
        # Actual only
        a_req = sum(m["required"] for m in ms_list if m["status_type"] == "Actual")
        a_avail = sum(m["available"] for m in ms_list if m["status_type"] == "Actual")
        a_itime = sum(m["in_time"] for m in ms_list if m["status_type"] == "Actual")
        results[f"{label}_actual_completeness"] = (a_avail / a_req * 100) if a_req > 0 else 0
        results[f"{label}_actual_timeliness"] = (a_itime / a_req * 100) if a_req > 0 else 0
    
    return results

def analyze_plausibility(sc3_details, sc4_details):
    """Check if milestones follow plausible date order per shipment."""
    # Expected milestone order (by code)
    MILESTONE_ORDER = ["S00", "S60", "S52", "S02", "S46", "S16", "S17", "S10", "S50", "S53", "S04", "S07", "S54", "S51", "S18", "S12", "S13", "S45", "S05", "S31", "S55"]
    order_map = {code: idx for idx, code in enumerate(MILESTONE_ORDER)}
    
    # Group by shipment (CONSIGNMENT or LOAD_TO)
    shipment_milestones = defaultdict(list)
    for rec in (sc3_details or []) + (sc4_details or []):
        shipment_id = rec.get("CONSIGNMENT") or rec.get("LOAD_TO") or rec.get("col_0")
        status_code = rec.get("STATUS_CODE_REQUIRED") or rec.get("Status Code") or rec.get("col_1")
        avail = rec.get("STATUS_CODE_AVAILABLE összege") or rec.get("col_3") or rec.get("col_6")
        if shipment_id and status_code and avail:
            try:
                avail_val = float(avail)
            except:
                avail_val = 0
            if avail_val > 0:  # Only count available milestones
                shipment_milestones[str(shipment_id)].append(str(status_code).strip())
    
    total_shipments = len(shipment_milestones)
    plausible = 0
    for ship_id, codes in shipment_milestones.items():
        # Check order
        ordered_codes = [c for c in codes if c in order_map]
        if len(ordered_codes) <= 1:
            plausible += 1
            continue
        indices = [order_map[c] for c in ordered_codes]
        # Check if indices are monotonically non-decreasing
        is_ordered = all(indices[i] <= indices[i+1] for i in range(len(indices)-1))
        if is_ordered:
            plausible += 1
    
    return (plausible / total_shipments * 100) if total_shipments > 0 else 0, total_shipments

def analyze_eta_accuracy(sc3_shipments, sc4_shipments, sc3_milestones, sc4_milestones):
    """
    Analyze ETA accuracy: were Estimated milestones received 7 days before Actual?
    Problem 3: to Port (S04 Estimated vs S04 Actual, S07 Estimated vs S07 Actual)
    Problem 4: to Door (S31 Estimated vs S31 Actual, S05)
    Using TOTAL sheet: compare Estimated completeness ratios for relevant milestones
    """
    results = {}
    
    # ETA accuracy from TOTAL: Estimated milestones timeliness ratio
    # For "to Port": S04 Estimated + S07 Estimated
    port_required = 0
    port_in_time = 0
    door_required = 0
    door_in_time = 0
    
    for m in sc3_milestones + sc4_milestones:
        if m["status_type"] == "Estimated":
            if m["code"] in ("S04", "S07"):
                port_required += m["required"]
                port_in_time += m["in_time"]
            elif m["code"] in ("S02", "S31"):
                door_required += m["required"]
                door_in_time += m["in_time"]
    
    results["eta_2p"] = (port_in_time / port_required * 100) if port_required > 0 else 0
    results["eta_2d"] = (door_in_time / door_required * 100) if door_required > 0 else 0
    
    return results


def main():
    print("=" * 100)
    print("  PDCA TRACKER KPI EXTRACTION - Bosch Milestone Raw Data")
    print("=" * 100)
    
    for week in WEEKS:
        print(f"\n{'='*80}")
        print(f"  WEEK: {week}")
        print(f"{'='*80}")
        
        sc3_file = os.path.join(BASE, f"Maersk NGTM SC3_2026_{week}.xlsx")
        sc4_file = os.path.join(BASE, f"Maersk SC4_2026_{week}.xlsx")
        
        if not os.path.exists(sc3_file):
            print(f"  WARNING: {sc3_file} not found!")
            continue
        if not os.path.exists(sc4_file):
            print(f"  WARNING: {sc4_file} not found!")
            continue
        
        # 1. Read TOTAL sheets
        print(f"\n  Reading TOTAL sheets...")
        sc3_ms = read_total_sheet(sc3_file)
        sc4_ms = read_total_sheet(sc4_file)
        
        print(f"  SC3 milestones: {len(sc3_ms)}")
        for m in sc3_ms:
            print(f"    {m['name']:50s} | {m['status_type']:10s} | Req: {m['required']:6.0f} | Avail: {m['available']:6.0f} | InTime: {m['in_time']:6.0f} | Comp: {m['completeness']:.1%} | Time: {m['timeliness']:.1%}")
        
        print(f"\n  SC4 milestones: {len(sc4_ms)}")
        for m in sc4_ms:
            print(f"    {m['name']:50s} | {m['status_type']:10s} | Req: {m['required']:6.0f} | Avail: {m['available']:6.0f} | InTime: {m['in_time']:6.0f} | Comp: {m['completeness']:.1%} | Time: {m['timeliness']:.1%}")
        
        # 2. Compute KPIs
        kpis = compute_kpis(sc3_ms, sc4_ms)
        
        # 3. ETA accuracy
        eta = analyze_eta_accuracy(None, None, sc3_ms, sc4_ms)
        
        # 4. Read detail sheets for plausibility
        print(f"\n  Reading detail sheets for plausibility...")
        sc3_details = read_detail_sheets(sc3_file, "SC3")
        sc4_details = read_detail_sheets(sc4_file, "SC4")
        plaus_pct, plaus_total = analyze_plausibility(sc3_details, sc4_details)
        
        # 5. Reference analysis (SC4 only has REF sheets)
        ref_data = read_ref_sheets(sc4_file)
        ref_total = len(ref_data)
        ref_with_ref = 0
        ref_correct = 0
        for r in ref_data:
            # Check if any reference columns have data
            has_ref = False
            for key in r:
                if any(kw in key.lower() for kw in ["civ", "dn", "delivery_note", "invoice", "reference"]):
                    if r[key] and str(r[key]).strip():
                        has_ref = True
                        break
            if has_ref:
                ref_with_ref += 1
        
        # Print summary
        print(f"\n  {'='*60}")
        print(f"  KPI SUMMARY FOR {week}")
        print(f"  {'='*60}")
        print(f"\n  --- Problem 1: Critical Completeness ---")
        print(f"  Combined: {kpis['critical_completeness']:.1f}%")
        print(f"    SC3 Critical: {kpis['sc3_crit_completeness']:.1f}%")
        print(f"    SC4 Critical: {kpis['sc4_crit_completeness']:.1f}%")
        
        print(f"\n  --- Problem 2: Critical Timeliness ---")
        print(f"  Combined: {kpis['critical_timeliness']:.1f}%")
        print(f"    SC3 Critical: {kpis['sc3_crit_timeliness']:.1f}%")
        print(f"    SC4 Critical: {kpis['sc4_crit_timeliness']:.1f}%")
        
        print(f"\n  --- Problem 3: ETA Accuracy 2P (to Port) ---")
        print(f"  Estimated timeliness (S04+S07 Est): {eta['eta_2p']:.1f}%")
        
        print(f"\n  --- Problem 4: ETA Accuracy 2D (to Door) ---")
        print(f"  Estimated timeliness (S02+S31 Est): {eta['eta_2d']:.1f}%")
        
        print(f"\n  --- Problem 5: Reference Completeness ---")
        print(f"  SC4 REF data records: {ref_total}")
        if ref_total > 0:
            print(f"  Records with reference: {ref_with_ref} ({ref_with_ref/ref_total*100:.1f}%)")
        
        print(f"\n  --- Problem 7: Plausibility ---")
        print(f"  Shipments in plausible order: {plaus_pct:.1f}% (of {plaus_total})")
        
        print(f"\n  --- Problem 8: All Milestones Completeness ---")
        print(f"  Combined: {kpis['all_completeness']:.1f}%")
        print(f"    SC3 All: {kpis['sc3_completeness']:.1f}%")
        print(f"    SC4 All: {kpis['sc4_completeness']:.1f}%")
        print(f"    SC3 Actual-only: {kpis['sc3_actual_completeness']:.1f}%")
        print(f"    SC4 Actual-only: {kpis['sc4_actual_completeness']:.1f}%")
        
        print(f"\n  --- Problem 9: All Milestones Timeliness ---")
        print(f"  Combined: {kpis['all_timeliness']:.1f}%")
        print(f"    SC3 All: {kpis['sc3_timeliness']:.1f}%")
        print(f"    SC4 All: {kpis['sc4_timeliness']:.1f}%")
        print(f"    SC3 Actual-only: {kpis['sc3_actual_timeliness']:.1f}%")
        print(f"    SC4 Actual-only: {kpis['sc4_actual_timeliness']:.1f}%")

if __name__ == "__main__":
    main()
