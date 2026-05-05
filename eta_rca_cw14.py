"""
ETA RCA — week-vs-prior-week comparison for both ETA 2P (S07) and ETA 2D (S31).

Reads raw SC4 Excel files, extracts per-shipment S07/S31 Accepted + Deviation,
aggregates by lane (origin_country → dest_country), origin country, carrier,
service and transport mode, and prints an RCA report comparing the two weeks.

Usage:
  python3 eta_rca_cw14.py                   # auto-detect: last two CW files in raw dir
  python3 eta_rca_cw14.py --prior CW14 --current CW15
"""

import argparse
import openpyxl
import os
import re
from collections import defaultdict
from datetime import datetime

BASE = os.path.dirname(os.path.abspath(__file__))
RAW_DIR = os.path.join(BASE, "Bosch Milestone raw data")


def detect_last_two_sc4_weeks():
    weeks = set()
    for f in os.listdir(RAW_DIR):
        m = re.match(r"Maersk SC4_2026_CW(\d{2})\.xlsx$", f)
        if m:
            weeks.add(int(m.group(1)))
    ordered = sorted(weeks)
    if len(ordered) < 2:
        raise SystemExit("Need at least 2 SC4 files in the raw directory.")
    return f"CW{ordered[-2]:02d}", f"CW{ordered[-1]:02d}"


def find_shipments_sheet(wb):
    # Prefer named "shipments", fall back to scanning for UNIQUE_SHIPMENT_ID marker
    # (CW14 ships the sheet with the Hungarian default "Munka7").
    for sn in wb.sheetnames:
        if sn.strip().lower() == "shipments":
            return sn
    for sn in wb.sheetnames:
        if sn.strip().upper() in ("TOTAL", "ALL"):
            continue
        if sn.strip().upper().rstrip("_") in ("FCL", "BCO", "LCL"):
            continue
        ws = wb[sn]
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            for cell in row:
                if cell and "UNIQUE_SHIPMENT_ID" in str(cell).upper():
                    return sn
    return None


def find_header_row(ws):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        for cell in row:
            if cell and "UNIQUE_SHIPMENT" in str(cell):
                return i
    return 3


def find_col(col_map, *patterns):
    for key in col_map:
        for pat in patterns:
            if pat in key:
                return col_map[key]
    return None


def safe_str(v):
    return str(v).strip() if v is not None else ""


def safe_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def extract(week):
    path = os.path.join(RAW_DIR, f"Maersk SC4_2026_{week}.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = find_shipments_sheet(wb)
    ws = wb[sheet]
    header_row = find_header_row(ws)
    headers = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
    col_map = {str(h).strip(): i for i, h in enumerate(headers) if h}

    c = {
        "hbl":              find_col(col_map, "HOUSE_BILL_OF_LADING"),
        "carrier":          find_col(col_map, "CARRIER_1"),
        "service":          find_col(col_map, "TRANSPORT_SERVICE_PRIORITY"),
        "transport_mode":   find_col(col_map, "TRANSPORT_MODE"),
        "origin_country":   find_col(col_map, "CONSIGNOR_ADDRESS_COUNTRY"),
        "origin_city":      find_col(col_map, "CONSIGNOR_ADDRESS_CITY_NAME"),
        "dest_country":     find_col(col_map, "CONSIGNEE_ADDRESS_COUNTRY"),
        "dest_city":        find_col(col_map, "CONSIGNEE_ADDRESS_CITY_NAME"),
        "delivery_country": find_col(col_map, "DELIVERY_ADDRESS_COUNTRY"),
        "pod":              find_col(col_map, "PORT_OF_DISCHARGE"),
        "ata":              find_col(col_map, "ATA_DATE_TIME"),
        "delivered":        find_col(col_map, "DELIVERED_DATE_TIME"),
        "s07_acc":          find_col(col_map, "S07_Accepted"),
        "s07_dev":          find_col(col_map, "S07_Deviation"),
        "s31_acc":          find_col(col_map, "S31_Accepted"),
        "s31_dev":          find_col(col_map, "S31_Deviation"),
    }

    def g(row, key):
        i = c.get(key)
        if i is None or i >= len(row):
            return None
        return row[i]

    ships = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row = list(row)
        if not row or not row[0]:
            continue

        s07_acc = g(row, "s07_acc")
        s31_acc = g(row, "s31_acc")
        ata = g(row, "ata")
        delivered = g(row, "delivered")

        # Match rebaseline.py filter so lane totals reconcile with dashboard KPIs:
        # S07 measurable = accepted is 0/1, excluding in-transit (accepted=0 & no ATA).
        # S31 measurable = accepted is 0/1, excluding undelivered (accepted=0 & no delivered).
        s07_measurable = s07_acc is not None and not (s07_acc == 0 and ata is None)
        s31_measurable = s31_acc is not None and not (s31_acc == 0 and delivered is None)

        if not s07_measurable and not s31_measurable:
            continue

        origin_country = safe_str(g(row, "origin_country"))
        dest_country = safe_str(g(row, "dest_country"))
        delivery_country = safe_str(g(row, "delivery_country")) or dest_country

        ships.append({
            "week": week,
            "hbl": safe_str(g(row, "hbl")),
            "origin_country": origin_country or "?",
            "origin_city": safe_str(g(row, "origin_city")) or "?",
            "dest_country": dest_country or "?",
            "dest_city": safe_str(g(row, "dest_city")) or "?",
            "delivery_country": delivery_country or "?",
            "carrier": safe_str(g(row, "carrier")) or "?",
            "service": safe_str(g(row, "service")) or "?",
            "transport_mode": safe_str(g(row, "transport_mode")) or "?",
            "pod": safe_str(g(row, "pod")) or "?",
            "lane": f"{origin_country or '?'} -> {dest_country or '?'}",
            "city_lane": f"{safe_str(g(row, 'origin_city')) or '?'} -> {safe_str(g(row, 'dest_city')) or '?'}",
            "s07_measurable": s07_measurable,
            "s07_accepted": (s07_acc == 1) if s07_measurable else None,
            "s07_deviation_h": safe_float(g(row, "s07_dev")),
            "s31_measurable": s31_measurable,
            "s31_accepted": (s31_acc == 1) if s31_measurable else None,
            "s31_deviation_h": safe_float(g(row, "s31_dev")),
        })
    wb.close()
    return ships


def aggregate(ships, dim, kpi):
    """dim: key like 'lane'. kpi: 's07' or 's31'. Returns dict value->stats."""
    acc_flag = f"{kpi}_accepted"
    meas_flag = f"{kpi}_measurable"
    dev_key = f"{kpi}_deviation_h"
    groups = defaultdict(lambda: {"total": 0, "acc": 0, "devs": [], "hbls_failed": []})
    for s in ships:
        if not s[meas_flag]:
            continue
        val = s.get(dim) or "?"
        g = groups[val]
        g["total"] += 1
        if s[acc_flag]:
            g["acc"] += 1
        else:
            g["hbls_failed"].append(s["hbl"])
        if s[dev_key] is not None:
            g["devs"].append(s[dev_key])
    out = {}
    for val, g in groups.items():
        failed = g["total"] - g["acc"]
        rate = (g["acc"] / g["total"] * 100) if g["total"] else None
        avg_dev = (sum(g["devs"]) / len(g["devs"])) if g["devs"] else None
        out[val] = {
            "total": g["total"], "accepted": g["acc"], "failed": failed,
            "rate": round(rate, 1) if rate is not None else None,
            "avg_dev_h": round(avg_dev, 1) if avg_dev is not None else None,
            "sample_failed_hbls": g["hbls_failed"][:3],
        }
    return out


def compare(a, b, min_total=2):
    """Merge two aggregations. a = prior (CW13), b = current (CW14)."""
    keys = set(a) | set(b)
    rows = []
    for k in keys:
        pa = a.get(k)
        pb = b.get(k)
        b_total = pb["total"] if pb else 0
        a_total = pa["total"] if pa else 0
        if max(a_total, b_total) < min_total:
            continue
        delta = None
        if pa and pb and pa["rate"] is not None and pb["rate"] is not None:
            delta = round(pb["rate"] - pa["rate"], 1)
        status = "stable"
        if pa and not pb:
            status = "gone"
        elif pb and not pa:
            status = "new"
        elif delta is not None and delta <= -15:
            status = "worsened"
        elif delta is not None and delta >= 15:
            status = "improved"
        rows.append({
            "value": k,
            "a_total": a_total, "a_failed": (pa["failed"] if pa else 0), "a_rate": (pa["rate"] if pa else None),
            "b_total": b_total, "b_failed": (pb["failed"] if pb else 0), "b_rate": (pb["rate"] if pb else None),
            "a_avg_dev": (pa["avg_dev_h"] if pa else None),
            "b_avg_dev": (pb["avg_dev_h"] if pb else None),
            "delta": delta, "status": status,
            "sample_failed": (pb["sample_failed_hbls"] if pb else []),
        })
    return rows


def pct(v):
    return f"{v:.0f}%" if v is not None else "  — "


def h(v):
    return f"{v:+.0f}h" if v is not None else "  — "


def fmt_delta(v):
    if v is None:
        return "   —  "
    return f"{v:+.0f}pp"


def print_section(title):
    print("\n" + "=" * 90)
    print(title)
    print("=" * 90)


def top_failures_table(ships, kpi, week, dim, top_n=10):
    agg = aggregate(ships, dim, kpi)
    rows = sorted(agg.items(), key=lambda kv: kv[1]["failed"], reverse=True)[:top_n]
    print(f"\n  Top {top_n} {dim} by failure count — {week} {kpi.upper()}")
    print(f"  {'value':<35} {'total':>6} {'fail':>5} {'rate':>6} {'avg dev':>9}")
    print("  " + "-" * 70)
    for val, g in rows:
        print(f"  {val[:34]:<35} {g['total']:>6} {g['failed']:>5} {pct(g['rate']):>6} {h(g['avg_dev_h']):>9}")


def print_compare(rows, title, prior_label, current_label, top_n=15):
    print_section(title)
    rows_sorted = sorted(rows, key=lambda r: r["b_failed"], reverse=True)[:top_n]
    print(f"  {'value':<30} {prior_label+' n/f':>10} {prior_label+' r':>7} {current_label+' n/f':>10} {current_label+' r':>7} {'Δ':>7}  status")
    print("  " + "-" * 90)
    for r in rows_sorted:
        a_nf = f"{r['a_total']}/{r['a_failed']}"
        b_nf = f"{r['b_total']}/{r['b_failed']}"
        print(f"  {r['value'][:29]:<30} {a_nf:>10} {pct(r['a_rate']):>7} "
              f"{b_nf:>10} {pct(r['b_rate']):>7} {fmt_delta(r['delta']):>7}  {r['status']}")


def print_regressions(rows, title, prior_label, current_label, top_n=10):
    print_section(title)
    worse = [r for r in rows if r["status"] == "worsened"]
    worse.sort(key=lambda r: (r["delta"] if r["delta"] is not None else 0))
    if not worse:
        print("  (none)")
        return
    print(f"  {'value':<30} {prior_label:>8} {current_label:>8} {'Δ':>7}  {current_label} fails / total")
    print("  " + "-" * 85)
    for r in worse[:top_n]:
        print(f"  {r['value'][:29]:<30} {pct(r['a_rate']):>8} {pct(r['b_rate']):>8} {fmt_delta(r['delta']):>7}"
              f"  {r['b_failed']}/{r['b_total']}  avg_dev={h(r['b_avg_dev'])}")


def print_new(rows, title, current_label, top_n=10):
    print_section(title)
    new = [r for r in rows if r["status"] == "new"]
    new.sort(key=lambda r: r["b_failed"], reverse=True)
    if not new:
        print("  (none)")
        return
    print(f"  {'value':<30} {current_label+' fails / total':>24} {current_label+' rate':>11}")
    print("  " + "-" * 75)
    for r in new[:top_n]:
        print(f"  {r['value'][:29]:<30} {r['b_failed']}/{r['b_total']:<22} {pct(r['b_rate']):>11}  avg_dev={h(r['b_avg_dev'])}")


def print_chronic(rows, title, prior_label, current_label, top_n=10):
    print_section(title)
    chronic = [r for r in rows
               if r["a_rate"] is not None and r["b_rate"] is not None
               and r["a_rate"] < 60 and r["b_rate"] < 60
               and r["a_total"] >= 3 and r["b_total"] >= 3]
    chronic.sort(key=lambda r: r["b_failed"], reverse=True)
    if not chronic:
        print("  (none)")
        return
    print(f"  {'value':<30} {prior_label:>8} {current_label:>8}  failures ({prior_label} → {current_label})")
    print("  " + "-" * 80)
    for r in chronic[:top_n]:
        print(f"  {r['value'][:29]:<30} {pct(r['a_rate']):>8} {pct(r['b_rate']):>8}"
              f"  {r['a_failed']} -> {r['b_failed']}  avg_dev {current_label}={h(r['b_avg_dev'])}")


def week_totals(ships, kpi):
    meas = [s for s in ships if s[f"{kpi}_measurable"]]
    acc = sum(1 for s in meas if s[f"{kpi}_accepted"])
    total = len(meas)
    devs = [s[f"{kpi}_deviation_h"] for s in meas if s[f"{kpi}_deviation_h"] is not None and not s[f"{kpi}_accepted"]]
    avg_dev = sum(devs) / len(devs) if devs else None
    return {
        "total": total, "accepted": acc, "failed": total - acc,
        "rate": round(acc / total * 100, 1) if total else None,
        "avg_fail_dev_h": round(avg_dev, 1) if avg_dev is not None else None,
        "n_late": sum(1 for d in devs if d > 0),
        "n_early": sum(1 for d in devs if d < 0),
    }


def main():
    parser = argparse.ArgumentParser(description="ETA 2P/2D week-over-week RCA")
    parser.add_argument("--prior", help="Prior week, e.g. CW14")
    parser.add_argument("--current", help="Current week, e.g. CW15")
    args = parser.parse_args()

    if args.prior and args.current:
        prior_w, current_w = args.prior, args.current
    else:
        prior_w, current_w = detect_last_two_sc4_weeks()

    weeks = [prior_w, current_w]
    print("=" * 90)
    print(f"ETA RCA — {prior_w} vs {current_w} (SC4 shipments)")
    print("=" * 90)

    data = {}
    for w in weeks:
        print(f"\n  Loading {w}...")
        data[w] = extract(w)
        s07 = week_totals(data[w], "s07")
        s31 = week_totals(data[w], "s31")
        print(f"    Rows: {len(data[w])}")
        print(f"    ETA 2P (S07): {s07['accepted']}/{s07['total']} = {s07['rate']}%  "
              f"(late={s07['n_late']}, early={s07['n_early']}, avg fail dev={h(s07['avg_fail_dev_h'])})")
        print(f"    ETA 2D (S31): {s31['accepted']}/{s31['total']} = {s31['rate']}%  "
              f"(late={s31['n_late']}, early={s31['n_early']}, avg fail dev={h(s31['avg_fail_dev_h'])})")

    for kpi, label in [("s07", "ETA 2P (S07 — port arrival)"), ("s31", "ETA 2D (S31 — door delivery)")]:
        print("\n\n" + "#" * 90)
        print(f"#   {label}")
        print("#" * 90)

        for dim, _ in [
            ("lane", "country lane"),
            ("origin_country", "origin country"),
            ("delivery_country", "delivery country"),
            ("carrier", "carrier"),
            ("service", "service/priority"),
        ]:
            top_failures_table(data[current_w], kpi, current_w, dim, top_n=8)

        for dim, dim_label in [
            ("lane", "LANE"),
            ("origin_country", "ORIGIN COUNTRY"),
            ("carrier", "CARRIER"),
            ("service", "SERVICE / PRIORITY"),
        ]:
            a = aggregate(data[prior_w], dim, kpi)
            b = aggregate(data[current_w], dim, kpi)
            rows = compare(a, b, min_total=2)
            print_compare(rows, f"{label} — {dim_label}: {prior_w} vs {current_w} (sorted by {current_w} failures)",
                          prior_w, current_w, top_n=15)
            print_regressions(rows, f"{label} — {dim_label}: WORSENED week-over-week (Δ ≤ -15pp)",
                              prior_w, current_w)
            print_new(rows, f"{label} — {dim_label}: NEW in {current_w} (absent in {prior_w})", current_w)
            print_chronic(rows, f"{label} — {dim_label}: CHRONIC (< 60% both weeks)", prior_w, current_w)


if __name__ == "__main__":
    main()
