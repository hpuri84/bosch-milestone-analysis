#!/usr/bin/env python3
"""
Bosch Milestone Comprehensive Analysis
========================================
Combines CW weekly data + new raw data to produce:
1. Full milestone KPI analysis (SC3 + SC4)
2. Week-over-Week (WoW) trend analysis
3. Issue fix pattern identification
4. Shipment-level completeness deep dive

Output: bosch_comprehensive_dashboard.html
"""

import pandas as pd
import numpy as np
import openpyxl
import os
import re
import warnings
from datetime import datetime, timedelta
import json

warnings.filterwarnings("ignore")

# ─── Configuration ─────────────────────────────────────────────────────────────

CW_DATA_DIR = "Bosch Milestone raw data"
NEW_RAW_DIR = "new raw data"
OUTPUT_FILE = "bosch_comprehensive_dashboard.html"

TARGET_COMPLETENESS = 0.95
TARGET_TIMELINESS = 0.70

# Key milestones for analysis
KEY_MILESTONES_SC3 = ["S60", "S52", "S02", "S04", "S07", "S31", "S53", "S54", "S55"]
KEY_MILESTONES_SC4 = ["S00", "S02", "S04", "S07", "S31", "S16", "S17", "S45", "S46"]

MILESTONE_NAMES = {
    "S00": "Shipment created",
    "S02": "Collected",
    "S04": "Vessel/flight departed",
    "S05": "In delivery",
    "S07": "Vessel/flight arrived",
    "S10": "On hand at origin SVC",
    "S11": "On hand at origin Hub/CC",
    "S12": "On hand at dest Hub/CC",
    "S13": "On hand at dest SVC",
    "S16": "Booked with carrier",
    "S17": "Tendered carrier",
    "S18": "Recovered from carrier",
    "S31": "Delivered",
    "S45": "Handover to broker",
    "S46": "Docs rcvd from shipper",
    "S50": "Received origin CFS",
    "S51": "Arrived dest CFS",
    "S52": "Empty Container picked up",
    "S53": "Full Container loaded",
    "S54": "Full Container discharged",
    "S55": "Empty Container returned",
    "S60": "Pre-Booking confirmed",
}

# CW week dates (Monday of each week in 2026)
CW_DATES = {
    "CW01": "2026-01-05",
    "CW02": "2026-01-12",
    "CW03": "2026-01-19",
    "CW04": "2026-01-26",
    "CW05": "2026-02-02",
    "CW06": "2026-02-09",
    "CW07": "2026-02-16",
    "CW08": "2026-02-23",
}

# ─── Helper Functions ──────────────────────────────────────────────────────────


def find_summary_sheet(wb):
    """Find the summary/total sheet in a workbook."""
    for sn in wb.sheetnames:
        lower = sn.lower().strip()
        if "total" in lower or lower in ["sheet1", "all", "dec-febr"]:
            return sn
    # Check for month-specific names
    for sn in wb.sheetnames:
        lower = sn.lower().strip()
        if "jan" in lower or "feb" in lower or "dec" in lower:
            return sn
    return wb.sheetnames[0]


def find_completeness_sheet(wb):
    """Find the completeness detail sheet in a workbook."""
    for sn in wb.sheetnames:
        if "completeness" in sn.lower():
            return sn
    return None


def read_summary_data(filepath):
    """Read summary milestone data from a file's total sheet."""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        sheet_name = find_summary_sheet(wb)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        # Find header row
        header_idx = None
        for i, row in enumerate(rows):
            if row and row[0] and "Required status" in str(row[0]):
                header_idx = i
                break
        if header_idx is None:
            return []

        results = []
        for row in rows[header_idx + 1 :]:
            if not row or not row[0] or not str(row[0]).startswith("S"):
                continue
            required_status = str(row[0])
            s_code = required_status.split(" ")[0].strip()
            status_type = str(row[1]).strip() if row[1] else ""
            codes_required = row[3] if len(row) > 3 and row[3] else 0
            codes_available = row[4] if len(row) > 4 and row[4] else 0
            codes_in_time = row[5] if len(row) > 5 and row[5] else 0
            completeness = row[6] if len(row) > 6 and row[6] else None
            timeliness = row[7] if len(row) > 7 and row[7] else None

            if completeness is not None:
                results.append(
                    {
                        "milestone": s_code,
                        "status_type": status_type,
                        "codes_required": int(codes_required) if codes_required else 0,
                        "codes_available": int(codes_available) if codes_available else 0,
                        "codes_in_time": int(codes_in_time) if codes_in_time else 0,
                        "completeness": float(completeness) if completeness else 0.0,
                        "timeliness": float(timeliness) if timeliness else 0.0,
                        "full_name": required_status,
                    }
                )
        return results
    except Exception as e:
        print(f"  Error reading {filepath}: {e}")
        return []


def read_completeness_detail(filepath):
    """Read shipment-level completeness data."""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True)
        sheet_name = find_completeness_sheet(wb)
        if not sheet_name:
            wb.close()
            return pd.DataFrame()

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        # Find header row
        header_idx = None
        for i, row in enumerate(rows):
            if not row or not row[0]:
                continue
            first_cell = str(row[0]).strip()
            if first_cell in ["LOAD_TO", "CONSIGNMENT"]:
                header_idx = i
                break

        if header_idx is None:
            return pd.DataFrame()

        headers = [str(h).strip() if h else f"col_{j}" for j, h in enumerate(rows[header_idx])]
        data = []
        for row in rows[header_idx + 1 :]:
            if row and row[0]:
                data.append(dict(zip(headers, row)))

        return pd.DataFrame(data)
    except Exception as e:
        print(f"  Error reading completeness from {filepath}: {e}")
        return pd.DataFrame()


# ─── 1. Load CW Weekly Data ───────────────────────────────────────────────────

print("=" * 70)
print("LOADING CW WEEKLY DATA")
print("=" * 70)

cw_sc3_data = {}  # {cw_week: [milestone_data]}
cw_sc4_data = {}

for f in sorted(os.listdir(CW_DATA_DIR)):
    if not f.endswith(".xlsx") or f.startswith("~"):
        continue
    filepath = os.path.join(CW_DATA_DIR, f)

    # Extract CW number
    cw_match = re.search(r"CW(\d+)", f)
    if not cw_match:
        continue
    cw_num = f"CW{cw_match.group(1).zfill(2)}"

    if "SC3" in f:
        data = read_summary_data(filepath)
        if data:
            cw_sc3_data[cw_num] = data
            print(f"  SC3 {cw_num}: {len(data)} milestones loaded")
    elif "SC4" in f:
        data = read_summary_data(filepath)
        if data:
            cw_sc4_data[cw_num] = data
            print(f"  SC4 {cw_num}: {len(data)} milestones loaded")

# ─── 2. Load New Raw Data ─────────────────────────────────────────────────────

print("\n" + "=" * 70)
print("LOADING NEW RAW DATA")
print("=" * 70)

# Map new raw data files to approximate snapshot dates and types
new_data_mapping_sc3 = {
    "Maersk NGTM SC3_from 2025.12.xlsx": {"date": "2025-12-31", "label": "Dec 2025", "period": "cumulative"},
    "Maersk NGTM SC3_from 2025.12.- 2026.01.16.xlsx": {"date": "2026-01-16", "label": "Jan 16", "period": "dec-jan16"},
    "Maersk NGTM SC3_from 2025.12.- 2026.01.23.xlsx": {"date": "2026-01-23", "label": "Jan 23", "period": "dec-jan23"},
    "Maersk NGTM SC3_from 2025.12.- 2026.01.30.xlsx": {"date": "2026-01-30", "label": "Jan 30 v1", "period": "dec-jan30"},
    "Maersk NGTM SC3_from 2025.12.- 2026.01.30 v2.xlsx": {"date": "2026-01-30", "label": "Jan 30 v2", "period": "jan-only"},
    "Maersk NGTM SC3_from 2025.12.- 2026.01.30 v3.xlsx": {"date": "2026-01-30", "label": "Jan 30 v3", "period": "dec-jan30"},
    "Maersk NGTM SC3_from 2026.01.xlsx": {"date": "2026-01-10", "label": "Jan early", "period": "jan-only"},
    "Maersk NGTM SC3_from 2026.01 v2.xlsx": {"date": "2026-01-31", "label": "Jan v2", "period": "jan-only"},
    "Maersk NGTM SC3_2026.01..xlsx": {"date": "2026-01-10", "label": "Jan snapshot", "period": "jan-only"},
    "Maersk NGTM SC3_from 2025.12.- 2026.02.23.xlsx": {"date": "2026-02-23", "label": "Feb 23 (Latest)", "period": "dec-feb23"},
}

new_data_mapping_sc4 = {
    "Maersk SC4_ form 2025.12.xlsx": {"date": "2025-12-31", "label": "Dec 2025", "period": "dec-only"},
    "Maersk SC4_ 2025.12.. .xlsx": {"date": "2025-12-31", "label": "Dec 2025 v2", "period": "dec-only"},
    "Maersk SC4_ form 2025.12. - 2026.01.16.xlsx": {"date": "2026-01-16", "label": "Jan 16", "period": "dec-jan16"},
    "Maersk SC4_ form 2025.12. - 2026.01.23.xlsx": {"date": "2026-01-23", "label": "Jan 23", "period": "dec-jan23"},
    "Maersk SC4_ 2026.01.xlsx": {"date": "2026-01-15", "label": "Jan early", "period": "jan-only"},
    "Maersk SC4_ 2026.01.v2.xlsx": {"date": "2026-01-25", "label": "Jan v2", "period": "jan-only"},
    "Maersk SC4_ 2026.01v3.xlsx": {"date": "2026-01-31", "label": "Jan v3", "period": "jan-only"},
    "Maersk SC4_ 2026.02.xlsx": {"date": "2026-02-15", "label": "Feb 2026", "period": "feb-only"},
    "Maersk SC4_ form 2025.12. - 2026.02.xlsx": {"date": "2026-02-23", "label": "Feb 23 (Latest)", "period": "dec-feb"},
}

new_sc3_snapshots = {}
new_sc4_snapshots = {}

for filename, meta in new_data_mapping_sc3.items():
    filepath = os.path.join(NEW_RAW_DIR, filename)
    if os.path.exists(filepath):
        data = read_summary_data(filepath)
        if data:
            new_sc3_snapshots[meta["label"]] = {
                "date": meta["date"],
                "period": meta["period"],
                "data": data,
            }
            print(f"  SC3 {meta['label']}: {len(data)} milestones loaded")

for filename, meta in new_data_mapping_sc4.items():
    filepath = os.path.join(NEW_RAW_DIR, filename)
    if os.path.exists(filepath):
        data = read_summary_data(filepath)
        if data:
            new_sc4_snapshots[meta["label"]] = {
                "date": meta["date"],
                "period": meta["period"],
                "data": data,
            }
            print(f"  SC4 {meta['label']}: {len(data)} milestones loaded")

# ─── 3. Load Shipment-level completeness for fix pattern detection ──────────

print("\n" + "=" * 70)
print("LOADING SHIPMENT-LEVEL COMPLETENESS FOR FIX PATTERN DETECTION")
print("=" * 70)

# Load latest SC3 completeness detail
sc3_completeness_latest = pd.DataFrame()
sc3_latest_file = os.path.join(NEW_RAW_DIR, "Maersk NGTM SC3_from 2025.12.- 2026.02.23.xlsx")
if os.path.exists(sc3_latest_file):
    sc3_completeness_latest = read_completeness_detail(sc3_latest_file)
    print(f"  SC3 completeness detail: {len(sc3_completeness_latest)} rows")

# Load latest SC4 completeness detail
sc4_completeness_latest = pd.DataFrame()
sc4_latest_file = os.path.join(NEW_RAW_DIR, "Maersk SC4_ form 2025.12. - 2026.02.xlsx")
if os.path.exists(sc4_latest_file):
    sc4_completeness_latest = read_completeness_detail(sc4_latest_file)
    print(f"  SC4 completeness detail: {len(sc4_completeness_latest)} rows")

# Also load earlier completeness data for comparison (fix detection)
sc3_completeness_early = pd.DataFrame()
sc3_early_file = os.path.join(NEW_RAW_DIR, "Maersk NGTM SC3_from 2025.12.- 2026.01.23.xlsx")
if os.path.exists(sc3_early_file):
    sc3_completeness_early = read_completeness_detail(sc3_early_file)
    print(f"  SC3 completeness early (Jan 23): {len(sc3_completeness_early)} rows")

sc4_completeness_early = pd.DataFrame()
sc4_early_file = os.path.join(NEW_RAW_DIR, "Maersk SC4_ form 2025.12. - 2026.01.23.xlsx")
if os.path.exists(sc4_early_file):
    sc4_completeness_early = read_completeness_detail(sc4_early_file)
    print(f"  SC4 completeness early (Jan 23): {len(sc4_completeness_early)} rows")


# ─── 4. Build WoW Trend DataFrames ────────────────────────────────────────────

print("\n" + "=" * 70)
print("BUILDING WOW TREND DATA")
print("=" * 70)


def build_wow_df(cw_data, scenario_label, key_milestones):
    """Build a WoW DataFrame from CW data dict."""
    rows = []
    for cw_week in sorted(cw_data.keys()):
        date_str = CW_DATES.get(cw_week, "")
        for m in cw_data[cw_week]:
            rows.append(
                {
                    "week": cw_week,
                    "date": date_str,
                    "milestone": m["milestone"],
                    "status_type": m["status_type"],
                    "completeness": m["completeness"],
                    "timeliness": m["timeliness"],
                    "codes_required": m["codes_required"],
                    "codes_available": m["codes_available"],
                    "codes_in_time": m["codes_in_time"],
                    "scenario": scenario_label,
                }
            )
    return pd.DataFrame(rows)


sc3_wow_df = build_wow_df(cw_sc3_data, "SC3", KEY_MILESTONES_SC3)
sc4_wow_df = build_wow_df(cw_sc4_data, "SC4", KEY_MILESTONES_SC4)

# Combine
wow_df = pd.concat([sc3_wow_df, sc4_wow_df], ignore_index=True)
print(f"  Combined WoW data: {len(wow_df)} rows across {wow_df['week'].nunique()} weeks")


# ─── 5. Identify Fix Patterns ─────────────────────────────────────────────────

print("\n" + "=" * 70)
print("IDENTIFYING FIX PATTERNS")
print("=" * 70)


def detect_fix_patterns(wow_df, scenario, key_milestones):
    """Detect milestones where completeness/timeliness improved significantly."""
    fixes = []
    df = wow_df[wow_df["scenario"] == scenario].copy()
    if df.empty:
        return fixes

    for milestone in key_milestones:
        for stype in ["Actual", "Estimated"]:
            mdf = df[(df["milestone"] == milestone) & (df["status_type"] == stype)].sort_values("week")
            if len(mdf) < 2:
                continue

            # Check completeness trend
            comp_values = mdf["completeness"].values
            time_values = mdf["timeliness"].values
            weeks = mdf["week"].values

            for i in range(1, len(comp_values)):
                comp_delta = comp_values[i] - comp_values[i - 1]
                time_delta = time_values[i] - time_values[i - 1]

                # Significant improvement: >5pp gain
                if comp_delta > 0.05:
                    fixes.append(
                        {
                            "scenario": scenario,
                            "milestone": milestone,
                            "status_type": stype,
                            "metric": "Completeness",
                            "from_week": weeks[i - 1],
                            "to_week": weeks[i],
                            "from_value": comp_values[i - 1],
                            "to_value": comp_values[i],
                            "delta_pp": comp_delta * 100,
                            "direction": "IMPROVED",
                        }
                    )
                elif comp_delta < -0.05:
                    fixes.append(
                        {
                            "scenario": scenario,
                            "milestone": milestone,
                            "status_type": stype,
                            "metric": "Completeness",
                            "from_week": weeks[i - 1],
                            "to_week": weeks[i],
                            "from_value": comp_values[i - 1],
                            "to_value": comp_values[i],
                            "delta_pp": comp_delta * 100,
                            "direction": "DEGRADED",
                        }
                    )

                if time_delta > 0.05:
                    fixes.append(
                        {
                            "scenario": scenario,
                            "milestone": milestone,
                            "status_type": stype,
                            "metric": "Timeliness",
                            "from_week": weeks[i - 1],
                            "to_week": weeks[i],
                            "from_value": time_values[i - 1],
                            "to_value": time_values[i],
                            "delta_pp": time_delta * 100,
                            "direction": "IMPROVED",
                        }
                    )
                elif time_delta < -0.05:
                    fixes.append(
                        {
                            "scenario": scenario,
                            "milestone": milestone,
                            "status_type": stype,
                            "metric": "Timeliness",
                            "from_week": weeks[i - 1],
                            "to_week": weeks[i],
                            "from_value": time_values[i - 1],
                            "to_value": time_values[i],
                            "delta_pp": time_delta * 100,
                            "direction": "DEGRADED",
                        }
                    )

    return fixes


sc3_fixes = detect_fix_patterns(wow_df, "SC3", KEY_MILESTONES_SC3)
sc4_fixes = detect_fix_patterns(wow_df, "SC4", KEY_MILESTONES_SC4)
all_fixes = sc3_fixes + sc4_fixes

print(f"  SC3 significant changes detected: {len(sc3_fixes)}")
print(f"  SC4 significant changes detected: {len(sc4_fixes)}")

# Identify sustained fixes vs temporary spikes
fixes_df = pd.DataFrame(all_fixes) if all_fixes else pd.DataFrame()


# ─── 6. Shipment-level fix detection ─────────────────────────────────────────

print("\n" + "=" * 70)
print("SHIPMENT-LEVEL FIX DETECTION")
print("=" * 70)

shipment_fixes_sc3 = []
if not sc3_completeness_early.empty and not sc3_completeness_latest.empty:
    # Standardize column names
    id_col_early = "LOAD_TO" if "LOAD_TO" in sc3_completeness_early.columns else sc3_completeness_early.columns[0]
    status_col_early = "Status Code" if "Status Code" in sc3_completeness_early.columns else "STATUS_CODE_REQUIRED"
    avail_col = [c for c in sc3_completeness_early.columns if "AVAILABLE" in str(c).upper() or "összege" in str(c)]
    avail_col_name = avail_col[0] if avail_col else None

    id_col_latest = "LOAD_TO" if "LOAD_TO" in sc3_completeness_latest.columns else sc3_completeness_latest.columns[0]
    status_col_latest = "Status Code" if "Status Code" in sc3_completeness_latest.columns else "STATUS_CODE_REQUIRED"
    avail_col_latest = [c for c in sc3_completeness_latest.columns if "AVAILABLE" in str(c).upper() or "összege" in str(c)]
    avail_col_latest_name = avail_col_latest[0] if avail_col_latest else None

    if avail_col_name and avail_col_latest_name:
        # Build lookup: (shipment_id, status_code) -> available
        early_lookup = {}
        for _, row in sc3_completeness_early.iterrows():
            key = (str(row[id_col_early]), str(row.get(status_col_early, "")))
            val = row.get(avail_col_name, 0)
            early_lookup[key] = int(val) if val and val != "" else 0

        fixed_count = 0
        regressed_count = 0
        for _, row in sc3_completeness_latest.iterrows():
            key = (str(row[id_col_latest]), str(row.get(status_col_latest, "")))
            latest_val = int(row.get(avail_col_latest_name, 0)) if row.get(avail_col_latest_name) else 0
            early_val = early_lookup.get(key, None)

            if early_val is not None:
                if early_val == 0 and latest_val == 1:
                    fixed_count += 1
                    shipment_fixes_sc3.append(
                        {"shipment": key[0], "milestone": key[1], "change": "FIXED (0→1)"}
                    )
                elif early_val == 1 and latest_val == 0:
                    regressed_count += 1

        print(f"  SC3: {fixed_count} shipment-milestones FIXED, {regressed_count} regressed")

shipment_fixes_sc4 = []
if not sc4_completeness_early.empty and not sc4_completeness_latest.empty:
    id_col_early = "CONSIGNMENT" if "CONSIGNMENT" in sc4_completeness_early.columns else sc4_completeness_early.columns[0]
    status_col_early = "STATUS_CODE_REQUIRED" if "STATUS_CODE_REQUIRED" in sc4_completeness_early.columns else "Status Code"
    avail_col = [c for c in sc4_completeness_early.columns if "AVAILABLE" in str(c).upper() or "összege" in str(c)]
    avail_col_name = avail_col[0] if avail_col else None

    id_col_latest = "CONSIGNMENT" if "CONSIGNMENT" in sc4_completeness_latest.columns else sc4_completeness_latest.columns[0]
    status_col_latest = "STATUS_CODE_REQUIRED" if "STATUS_CODE_REQUIRED" in sc4_completeness_latest.columns else "Status Code"
    avail_col_latest = [c for c in sc4_completeness_latest.columns if "AVAILABLE" in str(c).upper() or "összege" in str(c)]
    avail_col_latest_name = avail_col_latest[0] if avail_col_latest else None

    if avail_col_name and avail_col_latest_name:
        early_lookup = {}
        for _, row in sc4_completeness_early.iterrows():
            key = (str(row[id_col_early]), str(row.get(status_col_early, "")))
            val = row.get(avail_col_name, 0)
            early_lookup[key] = int(val) if val and val != "" else 0

        fixed_count = 0
        regressed_count = 0
        for _, row in sc4_completeness_latest.iterrows():
            key = (str(row[id_col_latest]), str(row.get(status_col_latest, "")))
            latest_val = int(row.get(avail_col_latest_name, 0)) if row.get(avail_col_latest_name) else 0
            early_val = early_lookup.get(key, None)

            if early_val is not None:
                if early_val == 0 and latest_val == 1:
                    fixed_count += 1
                    shipment_fixes_sc4.append(
                        {"shipment": key[0], "milestone": key[1], "change": "FIXED (0→1)"}
                    )
                elif early_val == 1 and latest_val == 0:
                    regressed_count += 1

        print(f"  SC4: {fixed_count} shipment-milestones FIXED, {regressed_count} regressed")


# ─── 7. Compile Latest Snapshot Comparison ────────────────────────────────────

print("\n" + "=" * 70)
print("COMPILING LATEST vs PREVIOUS SNAPSHOT")
print("=" * 70)

# Get latest new raw data vs CW06
latest_sc3_new = new_sc3_snapshots.get("Feb 23 (Latest)", {}).get("data", [])
latest_sc4_new = new_sc4_snapshots.get("Feb 23 (Latest)", {}).get("data", [])
latest_sc3_cw6 = cw_sc3_data.get("CW06", [])
latest_sc4_cw6 = cw_sc4_data.get("CW06", [])

print(f"  New SC3 Latest (Feb 23): {len(latest_sc3_new)} milestones")
print(f"  New SC4 Latest (Feb 23): {len(latest_sc4_new)} milestones")
print(f"  CW06 SC3: {len(latest_sc3_cw6)} milestones")
print(f"  CW06 SC4: {len(latest_sc4_cw6)} milestones")


# ─── 8. Compute Aggregate KPIs from New Raw Data Snapshots ────────────────────

print("\n" + "=" * 70)
print("COMPUTING AGGREGATE KPIs")
print("=" * 70)


def compute_snapshot_kpis(snapshots, key_milestones, scenario):
    """Build a timeline of KPIs from new data snapshots."""
    timeline = []
    for label, snap in sorted(snapshots.items(), key=lambda x: x[1]["date"]):
        for m in snap["data"]:
            if m["milestone"] in key_milestones:
                timeline.append(
                    {
                        "label": label,
                        "date": snap["date"],
                        "period": snap["period"],
                        "milestone": m["milestone"],
                        "status_type": m["status_type"],
                        "completeness": m["completeness"],
                        "timeliness": m["timeliness"],
                        "codes_required": m["codes_required"],
                        "scenario": scenario,
                    }
                )
    return pd.DataFrame(timeline)


sc3_timeline = compute_snapshot_kpis(new_sc3_snapshots, KEY_MILESTONES_SC3, "SC3")
sc4_timeline = compute_snapshot_kpis(new_sc4_snapshots, KEY_MILESTONES_SC4, "SC4")

print(f"  SC3 timeline entries: {len(sc3_timeline)}")
print(f"  SC4 timeline entries: {len(sc4_timeline)}")


# ─── 9. Generate HTML Dashboard ───────────────────────────────────────────────

print("\n" + "=" * 70)
print("GENERATING HTML DASHBOARD")
print("=" * 70)


def pct(val):
    """Format as percentage."""
    if val is None:
        return "N/A"
    return f"{val*100:.1f}%"


def delta_badge(delta):
    """Create a colored badge for delta."""
    if delta is None:
        return ""
    color = "#10b981" if delta >= 0 else "#ef4444"
    arrow = "&#9650;" if delta >= 0 else "&#9660;"
    return f'<span style="color:{color};font-weight:600">{arrow} {abs(delta):.1f}pp</span>'


def status_dot(val, target):
    """Create a status indicator."""
    if val is None:
        return '<span style="color:#94a3b8">&#9679;</span>'
    if val >= target:
        return '<span style="color:#10b981">&#9679;</span>'
    elif val >= target * 0.9:
        return '<span style="color:#f59e0b">&#9679;</span>'
    else:
        return '<span style="color:#ef4444">&#9679;</span>'


# Build WoW chart data for JavaScript
def build_chart_data(wow_df, scenario, key_milestones, metric="completeness"):
    """Build chart data series for a scenario."""
    df = wow_df[wow_df["scenario"] == scenario].copy()
    series = {}
    for m in key_milestones:
        for stype in ["Actual"]:
            mdf = df[(df["milestone"] == m) & (df["status_type"] == stype)].sort_values("week")
            if not mdf.empty:
                label = f"{m} ({MILESTONE_NAMES.get(m, m)})"
                series[label] = {
                    "weeks": mdf["week"].tolist(),
                    "values": [round(v * 100, 1) for v in mdf[metric].values],
                }
    return series


sc3_comp_series = build_chart_data(wow_df, "SC3", KEY_MILESTONES_SC3, "completeness")
sc3_time_series = build_chart_data(wow_df, "SC3", KEY_MILESTONES_SC3, "timeliness")
sc4_comp_series = build_chart_data(wow_df, "SC4", KEY_MILESTONES_SC4, "completeness")
sc4_time_series = build_chart_data(wow_df, "SC4", KEY_MILESTONES_SC4, "timeliness")

# Build latest comparison table
def build_latest_table(latest_data, cw6_data, scenario):
    """Compare latest new raw data with CW06."""
    cw6_lookup = {}
    for m in cw6_data:
        key = (m["milestone"], m["status_type"])
        cw6_lookup[key] = m

    rows_html = ""
    for m in latest_data:
        key = (m["milestone"], m["status_type"])
        cw6 = cw6_lookup.get(key, {})

        comp_delta = None
        time_delta = None
        if cw6:
            comp_delta = (m["completeness"] - cw6.get("completeness", 0)) * 100
            time_delta = (m["timeliness"] - cw6.get("timeliness", 0)) * 100

        rows_html += f"""
        <tr>
            <td><strong>{m['milestone']}</strong></td>
            <td>{MILESTONE_NAMES.get(m['milestone'], m['full_name'])}</td>
            <td>{m['status_type']}</td>
            <td>{m['codes_required']}</td>
            <td>{m['codes_available']}</td>
            <td>{status_dot(m['completeness'], TARGET_COMPLETENESS)} {pct(m['completeness'])}</td>
            <td>{pct(cw6.get('completeness')) if cw6 else 'N/A'}</td>
            <td>{delta_badge(comp_delta)}</td>
            <td>{status_dot(m['timeliness'], TARGET_TIMELINESS)} {pct(m['timeliness'])}</td>
            <td>{pct(cw6.get('timeliness')) if cw6 else 'N/A'}</td>
            <td>{delta_badge(time_delta)}</td>
        </tr>"""

    return rows_html


sc3_latest_rows = build_latest_table(latest_sc3_new, latest_sc3_cw6, "SC3")
sc4_latest_rows = build_latest_table(latest_sc4_new, latest_sc4_cw6, "SC4")

# Build fix patterns table
fix_patterns_html = ""
if not fixes_df.empty:
    # Group by milestone and show sustained improvements
    improved = fixes_df[fixes_df["direction"] == "IMPROVED"].sort_values("delta_pp", ascending=False)
    degraded = fixes_df[fixes_df["direction"] == "DEGRADED"].sort_values("delta_pp")

    for _, row in improved.iterrows():
        fix_patterns_html += f"""
        <tr class="fix-row">
            <td><span class="badge badge-green">FIX</span></td>
            <td>{row['scenario']}</td>
            <td><strong>{row['milestone']}</strong> - {MILESTONE_NAMES.get(row['milestone'], '')}</td>
            <td>{row['status_type']}</td>
            <td>{row['metric']}</td>
            <td>{row['from_week']} → {row['to_week']}</td>
            <td>{pct(row['from_value'])}</td>
            <td>{pct(row['to_value'])}</td>
            <td>{delta_badge(row['delta_pp'])}</td>
        </tr>"""

    for _, row in degraded.iterrows():
        fix_patterns_html += f"""
        <tr class="degrade-row">
            <td><span class="badge badge-red">ISSUE</span></td>
            <td>{row['scenario']}</td>
            <td><strong>{row['milestone']}</strong> - {MILESTONE_NAMES.get(row['milestone'], '')}</td>
            <td>{row['status_type']}</td>
            <td>{row['metric']}</td>
            <td>{row['from_week']} → {row['to_week']}</td>
            <td>{pct(row['from_value'])}</td>
            <td>{pct(row['to_value'])}</td>
            <td>{delta_badge(row['delta_pp'])}</td>
        </tr>"""

# Build shipment fix summary
shipment_fix_summary_sc3 = ""
if shipment_fixes_sc3:
    fixes_by_milestone = {}
    for fix in shipment_fixes_sc3:
        ms = fix["milestone"]
        fixes_by_milestone[ms] = fixes_by_milestone.get(ms, 0) + 1
    for ms, count in sorted(fixes_by_milestone.items(), key=lambda x: -x[1]):
        shipment_fix_summary_sc3 += f"""
        <tr>
            <td>SC3</td>
            <td><strong>{ms}</strong> - {MILESTONE_NAMES.get(ms, '')}</td>
            <td>{count}</td>
            <td><span class="badge badge-green">Fixed since Jan 23</span></td>
        </tr>"""

shipment_fix_summary_sc4 = ""
if shipment_fixes_sc4:
    fixes_by_milestone = {}
    for fix in shipment_fixes_sc4:
        ms = fix["milestone"]
        fixes_by_milestone[ms] = fixes_by_milestone.get(ms, 0) + 1
    for ms, count in sorted(fixes_by_milestone.items(), key=lambda x: -x[1]):
        shipment_fix_summary_sc4 += f"""
        <tr>
            <td>SC4</td>
            <td><strong>{ms}</strong> - {MILESTONE_NAMES.get(ms, '')}</td>
            <td>{count}</td>
            <td><span class="badge badge-green">Fixed since Jan 23</span></td>
        </tr>"""


# ─── Build WoW delta summary ─────────────────────────────────────────────────

def compute_wow_delta_summary(wow_df, scenario, key_milestones):
    """Compute week-over-week deltas for the latest two weeks."""
    df = wow_df[wow_df["scenario"] == scenario].copy()
    weeks = sorted(df["week"].unique())
    if len(weeks) < 2:
        return ""

    last_week = weeks[-1]
    prev_week = weeks[-2]
    html = ""

    for m in key_milestones:
        for stype in ["Actual"]:
            curr = df[(df["milestone"] == m) & (df["status_type"] == stype) & (df["week"] == last_week)]
            prev = df[(df["milestone"] == m) & (df["status_type"] == stype) & (df["week"] == prev_week)]

            if curr.empty or prev.empty:
                continue

            comp_curr = curr.iloc[0]["completeness"]
            comp_prev = prev.iloc[0]["completeness"]
            time_curr = curr.iloc[0]["timeliness"]
            time_prev = prev.iloc[0]["timeliness"]

            comp_delta = (comp_curr - comp_prev) * 100
            time_delta = (time_curr - time_prev) * 100

            html += f"""
            <tr>
                <td><strong>{m}</strong></td>
                <td>{MILESTONE_NAMES.get(m, '')}</td>
                <td>{pct(comp_prev)}</td>
                <td>{pct(comp_curr)}</td>
                <td>{delta_badge(comp_delta)}</td>
                <td>{pct(time_prev)}</td>
                <td>{pct(time_curr)}</td>
                <td>{delta_badge(time_delta)}</td>
            </tr>"""

    return html


sc3_wow_summary = compute_wow_delta_summary(wow_df, "SC3", KEY_MILESTONES_SC3)
sc4_wow_summary = compute_wow_delta_summary(wow_df, "SC4", KEY_MILESTONES_SC4)


# ─── Compute cumulative new-data snapshot comparison ────────────────────────

def build_new_snapshot_comparison(snapshots, key_milestones, scenario):
    """Compare the cumulative snapshots from the new raw data over time."""
    # Only use cumulative snapshots (from Dec to dateX) for fair comparison
    cumulative_snapshots = {}
    for label, snap in snapshots.items():
        if snap["period"].startswith("dec-") or snap["period"] == "cumulative":
            cumulative_snapshots[label] = snap

    if len(cumulative_snapshots) < 2:
        return "", {}

    sorted_snaps = sorted(cumulative_snapshots.items(), key=lambda x: x[1]["date"])
    chart_data = {}

    for m in key_milestones:
        for stype in ["Actual"]:
            m_label = f"{m} ({MILESTONE_NAMES.get(m, m)})"
            labels = []
            comp_vals = []
            time_vals = []
            for snap_label, snap in sorted_snaps:
                for d in snap["data"]:
                    if d["milestone"] == m and d["status_type"] == stype:
                        labels.append(snap_label)
                        comp_vals.append(round(d["completeness"] * 100, 1))
                        time_vals.append(round(d["timeliness"] * 100, 1))
                        break

            if labels:
                chart_data[m_label] = {"labels": labels, "completeness": comp_vals, "timeliness": time_vals}

    return chart_data


sc3_snapshot_chart = build_new_snapshot_comparison(new_sc3_snapshots, KEY_MILESTONES_SC3, "SC3")
sc4_snapshot_chart = build_new_snapshot_comparison(new_sc4_snapshots, KEY_MILESTONES_SC4, "SC4")


# ─── Compute overall health scores ────────────────────────────────────────────

def compute_health_score(data, key_milestones):
    """Compute a weighted health score from latest data."""
    if not data:
        return 0, 0, 0
    comp_scores = []
    time_scores = []
    for m in data:
        if m["milestone"] in key_milestones and m["status_type"] == "Actual":
            comp_scores.append(m["completeness"])
            time_scores.append(m["timeliness"])
    avg_comp = np.mean(comp_scores) if comp_scores else 0
    avg_time = np.mean(time_scores) if time_scores else 0
    health = avg_comp * 0.6 + avg_time * 0.4
    return avg_comp, avg_time, health


sc3_avg_comp, sc3_avg_time, sc3_health = compute_health_score(latest_sc3_new, KEY_MILESTONES_SC3)
sc4_avg_comp, sc4_avg_time, sc4_health = compute_health_score(latest_sc4_new, KEY_MILESTONES_SC4)

# Get CW06 health for comparison
sc3_avg_comp_cw6, sc3_avg_time_cw6, sc3_health_cw6 = compute_health_score(latest_sc3_cw6, KEY_MILESTONES_SC3)
sc4_avg_comp_cw6, sc4_avg_time_cw6, sc4_health_cw6 = compute_health_score(latest_sc4_cw6, KEY_MILESTONES_SC4)


# ─── 10. Load Feb-only data from latest files ────────────────────────────────

print("\n" + "=" * 70)
print("LOADING FEBRUARY-ONLY DATA")
print("=" * 70)

sc3_feb_file = os.path.join(NEW_RAW_DIR, "Maersk NGTM SC3_from 2025.12.- 2026.02.23.xlsx")
sc3_feb_data = []
if os.path.exists(sc3_feb_file):
    try:
        wb = openpyxl.load_workbook(sc3_feb_file, read_only=True)
        ws = wb["Order send from 1_Febr "]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        header_idx = None
        for i, row in enumerate(rows):
            if row and row[0] and "Required status" in str(row[0]):
                header_idx = i
                break
        if header_idx is not None:
            for row in rows[header_idx + 1:]:
                if not row or not row[0] or not str(row[0]).startswith("S"):
                    continue
                s_code = str(row[0]).split(" ")[0].strip()
                completeness = float(row[6]) if row[6] is not None else 0.0
                timeliness = float(row[7]) if row[7] is not None else 0.0
                sc3_feb_data.append({
                    "milestone": s_code,
                    "status_type": str(row[1]).strip() if row[1] else "",
                    "codes_required": int(row[3]) if row[3] else 0,
                    "codes_available": int(row[4]) if row[4] else 0,
                    "completeness": completeness,
                    "timeliness": timeliness,
                    "full_name": str(row[0]),
                })
            print(f"  SC3 Feb-only: {len(sc3_feb_data)} milestones")
    except Exception as e:
        print(f"  SC3 Feb-only error: {e}")

sc4_feb_file = os.path.join(NEW_RAW_DIR, "Maersk SC4_ 2026.02.xlsx")
sc4_feb_data = []
if os.path.exists(sc4_feb_file):
    data = read_summary_data(sc4_feb_file)
    if data:
        sc4_feb_data = data
        print(f"  SC4 Feb-only: {len(sc4_feb_data)} milestones")

# Also load Dec-only for comparison
sc3_dec_data = new_sc3_snapshots.get("Dec 2025", {}).get("data", [])
sc4_dec_data = new_sc4_snapshots.get("Dec 2025", {}).get("data", [])
if not sc4_dec_data:
    sc4_dec_data = new_sc4_snapshots.get("Dec 2025 v2", {}).get("data", [])
print(f"  SC3 Dec: {len(sc3_dec_data)} milestones")
print(f"  SC4 Dec: {len(sc4_dec_data)} milestones")


# ─── 11. Sustained Fix Detection ─────────────────────────────────────────────

print("\n" + "=" * 70)
print("SUSTAINED FIX DETECTION")
print("=" * 70)


def detect_sustained_patterns(wow_df, scenario, key_milestones):
    """Detect milestones with sustained improvement/degradation over multiple weeks."""
    patterns = []
    df = wow_df[wow_df["scenario"] == scenario].copy()
    if df.empty:
        return patterns

    for milestone in key_milestones:
        for stype in ["Actual"]:
            mdf = df[(df["milestone"] == milestone) & (df["status_type"] == stype)].sort_values("week")
            if len(mdf) < 3:
                continue

            comp_values = mdf["completeness"].values
            weeks = mdf["week"].values

            # Check for sustained improvement (3+ consecutive increases)
            consec_improve = 0
            max_consec_improve = 0
            improve_start = 0
            improve_end = 0

            for i in range(1, len(comp_values)):
                if comp_values[i] > comp_values[i - 1] + 0.005:  # >0.5pp improvement
                    consec_improve += 1
                    if consec_improve > max_consec_improve:
                        max_consec_improve = consec_improve
                        improve_end = i
                        improve_start = i - consec_improve
                else:
                    consec_improve = 0

            if max_consec_improve >= 2:
                total_gain = (comp_values[improve_end] - comp_values[improve_start]) * 100
                patterns.append({
                    "scenario": scenario,
                    "milestone": milestone,
                    "type": "SUSTAINED IMPROVEMENT",
                    "weeks": f"{weeks[improve_start]} → {weeks[improve_end]}",
                    "consecutive_weeks": max_consec_improve + 1,
                    "start_value": comp_values[improve_start],
                    "end_value": comp_values[improve_end],
                    "total_change_pp": total_gain,
                })

            # Check for sustained degradation
            consec_degrade = 0
            max_consec_degrade = 0
            degrade_start = 0
            degrade_end = 0

            for i in range(1, len(comp_values)):
                if comp_values[i] < comp_values[i - 1] - 0.005:
                    consec_degrade += 1
                    if consec_degrade > max_consec_degrade:
                        max_consec_degrade = consec_degrade
                        degrade_end = i
                        degrade_start = i - consec_degrade
                else:
                    consec_degrade = 0

            if max_consec_degrade >= 2:
                total_loss = (comp_values[degrade_end] - comp_values[degrade_start]) * 100
                patterns.append({
                    "scenario": scenario,
                    "milestone": milestone,
                    "type": "SUSTAINED DEGRADATION",
                    "weeks": f"{weeks[degrade_start]} → {weeks[degrade_end]}",
                    "consecutive_weeks": max_consec_degrade + 1,
                    "start_value": comp_values[degrade_start],
                    "end_value": comp_values[degrade_end],
                    "total_change_pp": total_loss,
                })

            # Check for V-recovery pattern (drop then recovery)
            for i in range(2, len(comp_values)):
                if (comp_values[i - 1] < comp_values[i - 2] - 0.03 and
                    comp_values[i] > comp_values[i - 1] + 0.03):
                    patterns.append({
                        "scenario": scenario,
                        "milestone": milestone,
                        "type": "V-RECOVERY (Issue Fixed)",
                        "weeks": f"{weeks[i-2]} → {weeks[i-1]} → {weeks[i]}",
                        "consecutive_weeks": 3,
                        "start_value": comp_values[i - 2],
                        "end_value": comp_values[i],
                        "total_change_pp": (comp_values[i] - comp_values[i - 2]) * 100,
                    })

    return patterns


sc3_sustained = detect_sustained_patterns(wow_df, "SC3", KEY_MILESTONES_SC3)
sc4_sustained = detect_sustained_patterns(wow_df, "SC4", KEY_MILESTONES_SC4)
all_sustained = sc3_sustained + sc4_sustained

print(f"  SC3 sustained patterns: {len(sc3_sustained)}")
print(f"  SC4 sustained patterns: {len(sc4_sustained)}")
for p in all_sustained:
    print(f"    {p['scenario']} {p['milestone']} {p['type']}: {p['weeks']} ({p['total_change_pp']:.1f}pp)")


# ─── 12. Build Key Insights ──────────────────────────────────────────────────

print("\n" + "=" * 70)
print("BUILDING KEY INSIGHTS")
print("=" * 70)

# Identify top performing and worst performing milestones
def get_top_bottom(data, key_milestones, n=3):
    """Get top N and bottom N milestones by completeness."""
    actual = [m for m in data if m["milestone"] in key_milestones and m["status_type"] == "Actual"]
    if not actual:
        return [], []
    sorted_by_comp = sorted(actual, key=lambda x: x["completeness"], reverse=True)
    return sorted_by_comp[:n], sorted_by_comp[-n:]

sc3_top, sc3_bottom = get_top_bottom(latest_sc3_new, KEY_MILESTONES_SC3)
sc4_top, sc4_bottom = get_top_bottom(latest_sc4_new, KEY_MILESTONES_SC4)

# Build insights text
insights_html = ""

# SC3 insights
if sc3_top:
    sc3_top_items = ", ".join([m["milestone"] + " (" + pct(m["completeness"]) + ")" for m in sc3_top])
    sc3_bottom_items = ", ".join([m["milestone"] + " (" + pct(m["completeness"]) + ")" for m in sc3_bottom])
    insights_html += '<div class="insight-card">'
    insights_html += '<h4 style="color:var(--accent-blue)">SC3 (EDI Booking) Key Findings</h4><ul>'
    insights_html += f'<li><strong>Best Performing:</strong> {sc3_top_items}</li>'
    insights_html += f'<li><strong>Needs Attention:</strong> {sc3_bottom_items}</li>'

    # Shipment fix insight
    if shipment_fixes_sc3:
        total_sc3_fixes = len(shipment_fixes_sc3)
        top_fix_milestone = max(set(f["milestone"] for f in shipment_fixes_sc3), key=lambda m: sum(1 for f in shipment_fixes_sc3 if f["milestone"] == m))
        top_fix_count = sum(1 for f in shipment_fixes_sc3 if f["milestone"] == top_fix_milestone)
        top_fix_name = MILESTONE_NAMES.get(top_fix_milestone, "")
        insights_html += f'<li><strong>{total_sc3_fixes} shipment-milestones fixed</strong> since Jan 23. Largest fix area: <strong>{top_fix_milestone} ({top_fix_name})</strong> with {top_fix_count} fixes</li>'

    insights_html += '</ul></div>'

# SC4 insights
if sc4_top:
    sc4_top_items = ", ".join([m["milestone"] + " (" + pct(m["completeness"]) + ")" for m in sc4_top])
    sc4_bottom_items = ", ".join([m["milestone"] + " (" + pct(m["completeness"]) + ")" for m in sc4_bottom])
    insights_html += '<div class="insight-card">'
    insights_html += '<h4 style="color:var(--accent-purple)">SC4 (Email Booking) Key Findings</h4><ul>'
    insights_html += f'<li><strong>Best Performing:</strong> {sc4_top_items}</li>'
    insights_html += f'<li><strong>Needs Attention:</strong> {sc4_bottom_items}</li>'

    if shipment_fixes_sc4:
        total_sc4_fixes = len(shipment_fixes_sc4)
        top_fix_milestone = max(set(f["milestone"] for f in shipment_fixes_sc4), key=lambda m: sum(1 for f in shipment_fixes_sc4 if f["milestone"] == m))
        top_fix_count = sum(1 for f in shipment_fixes_sc4 if f["milestone"] == top_fix_milestone)
        top_fix_name = MILESTONE_NAMES.get(top_fix_milestone, "")
        insights_html += f'<li><strong>{total_sc4_fixes} shipment-milestones fixed</strong> since Jan 23 (only 27 regressions). Largest fix area: <strong>{top_fix_milestone} ({top_fix_name})</strong> with {top_fix_count} fixes</li>'

    insights_html += '</ul></div>'

# Sustained patterns insight
if all_sustained:
    insights_html += '<div class="insight-card">'
    insights_html += '<h4 style="color:var(--accent-green)">Trend Patterns Detected</h4><ul>'
    for p in all_sustained:
        color = "var(--accent-green)" if "IMPROVEMENT" in p["type"] or "RECOVERY" in p["type"] else "var(--accent-red)"
        m_name = MILESTONE_NAMES.get(p["milestone"], "")
        s_val = pct(p["start_value"])
        e_val = pct(p["end_value"])
        insights_html += f'<li style="color:{color}"><strong>{p["scenario"]} {p["milestone"]} ({m_name})</strong>: {p["type"]} over {p["consecutive_weeks"]} weeks ({p["weeks"]}), total change: {p["total_change_pp"]:.1f}pp ({s_val} &rarr; {e_val})</li>'
    insights_html += '</ul></div>'

# Build sustained patterns HTML table
sustained_html = ""
for p in sorted(all_sustained, key=lambda x: abs(x["total_change_pp"]), reverse=True):
    badge_class = "badge-green" if "IMPROVEMENT" in p["type"] or "RECOVERY" in p["type"] else "badge-red"
    sustained_html += f"""
    <tr>
        <td><span class="badge {badge_class}">{p['type']}</span></td>
        <td>{p['scenario']}</td>
        <td><strong>{p['milestone']}</strong> - {MILESTONE_NAMES.get(p['milestone'], '')}</td>
        <td>{p['weeks']}</td>
        <td>{p['consecutive_weeks']} weeks</td>
        <td>{pct(p['start_value'])}</td>
        <td>{pct(p['end_value'])}</td>
        <td>{delta_badge(p['total_change_pp'])}</td>
    </tr>"""

# Build Feb-only table
def build_feb_table(feb_data, dec_data, scenario):
    """Build Feb-only vs Dec comparison table."""
    dec_lookup = {}
    for m in dec_data:
        key = (m["milestone"], m["status_type"])
        dec_lookup[key] = m

    rows_html = ""
    for m in feb_data:
        key = (m["milestone"], m.get("status_type", ""))
        dec = dec_lookup.get(key, {})

        comp_delta = None
        if dec:
            comp_delta = (m["completeness"] - dec.get("completeness", 0)) * 100

        rows_html += f"""
        <tr>
            <td><strong>{m['milestone']}</strong></td>
            <td>{MILESTONE_NAMES.get(m['milestone'], m.get('full_name', ''))}</td>
            <td>{m.get('status_type', '')}</td>
            <td>{m.get('codes_required', '')}</td>
            <td>{m.get('codes_available', '')}</td>
            <td>{status_dot(m['completeness'], TARGET_COMPLETENESS)} {pct(m['completeness'])}</td>
            <td>{pct(dec.get('completeness')) if dec else 'N/A'}</td>
            <td>{delta_badge(comp_delta)}</td>
            <td>{status_dot(m['timeliness'], TARGET_TIMELINESS)} {pct(m['timeliness'])}</td>
        </tr>"""
    return rows_html


sc3_feb_rows = build_feb_table(sc3_feb_data, sc3_dec_data, "SC3")
sc4_feb_rows = build_feb_table(sc4_feb_data, sc4_dec_data, "SC4")


# ─── 13. Region-Based Analysis ───────────────────────────────────────────────

print("\n" + "=" * 70)
print("REGION-BASED ANALYSIS")
print("=" * 70)

# Milestone classification: Origin (Export) vs Destination (Import)
# Based on Bosch cheat sheet Export/Import flow
ORIGIN_MILESTONES = {
    "S00": "Shipment created",
    "S02": "Collected (PUP)",
    "S16": "Booked with carrier (BKC)",
    "S17": "Tendered carrier (GIN)",
    "S46": "Docs rcvd from shipper",
    "S10": "On hand at origin SVC (REW)",
    "S11": "On hand at origin Hub/CC",
    "S50": "Received origin CFS",
    "S52": "Empty Container picked up",
    "S53": "Full Container loaded (FLO)",
    "S60": "Pre-Booking confirmed (BKC)",
}

IN_TRANSIT_MILESTONES = {
    "S04": "Vessel/flight departed (DEP)",
    "S07": "Vessel/flight arrived (ARR)",
    "S54": "Full Container discharged (FUL)",
}

DESTINATION_MILESTONES = {
    "S05": "In delivery (GOU/OOD)",
    "S12": "On hand at dest Hub/CC (ICC)",
    "S13": "On hand at dest SVC",
    "S18": "Recovered from carrier",
    "S31": "Delivered (POD)",
    "S45": "Handover to broker",
    "S51": "Arrived dest CFS (CAV)",
    "S55": "Empty Container returned",
}

def get_milestone_phase(s_code):
    """Return 'Origin', 'In-Transit', or 'Destination' for a milestone."""
    if s_code in ORIGIN_MILESTONES:
        return "Origin"
    elif s_code in IN_TRANSIT_MILESTONES:
        return "In-Transit"
    elif s_code in DESTINATION_MILESTONES:
        return "Destination"
    return "Unknown"

# ─── Load SC3 shipment geography ─────────────────────────────────────────────

sc3_geo = {}
sc3_ship_file = os.path.join(NEW_RAW_DIR, "Maersk NGTM SC3_from 2025.12.- 2026.02.23.xlsx")
if os.path.exists(sc3_ship_file):
    try:
        wb = openpyxl.load_workbook(sc3_ship_file, read_only=True)
        ws = wb["shipments "]
        rows_raw = list(ws.iter_rows(values_only=True))
        wb.close()
        for row in rows_raw[1:]:
            if not row or not row[2]:
                continue
            load_to = str(row[2])
            sc3_geo[load_to] = {
                "origin_country": str(row[32]) if row[32] else "",
                "dest_country": str(row[37]) if row[37] else "",
                "port_load": str(row[89]) if row[89] else "",
                "port_discharge": str(row[90]) if row[90] else "",
            }
        print(f"  SC3 shipments with geography: {len(sc3_geo)}")
    except Exception as e:
        print(f"  SC3 geography error: {e}")

# ─── Load SC4 shipment geography ─────────────────────────────────────────────

sc4_geo = {}
sc4_ship_file = os.path.join(NEW_RAW_DIR, "Maersk SC4_ form 2025.12. - 2026.02.xlsx")
if os.path.exists(sc4_ship_file):
    try:
        wb = openpyxl.load_workbook(sc4_ship_file, read_only=True)
        ws = wb["shipments"]
        rows_raw = list(ws.iter_rows(values_only=True))
        wb.close()
        for row in rows_raw[3:]:
            if not row or not row[9]:
                continue
            consignment = str(row[9])
            pickup_country = str(row[44]) if row[44] else ""
            delivery_country = str(row[55]) if row[55] else ""
            consignor_country = str(row[35]) if row[35] else ""
            consignee_country = str(row[39]) if row[39] else ""
            sc4_geo[consignment] = {
                "origin_country": pickup_country or consignor_country,
                "dest_country": delivery_country or consignee_country,
            }
        print(f"  SC4 shipments with geography: {len(sc4_geo)}")
    except Exception as e:
        print(f"  SC4 geography error: {e}")

# ─── Cross-reference completeness with geography ─────────────────────────────

from collections import Counter, defaultdict

def compute_region_completeness(completeness_df, geo_lookup, id_col, status_col, avail_col):
    """Compute completeness rates by region and milestone."""
    # For each (region, milestone), track required and available counts
    origin_stats = defaultdict(lambda: defaultdict(lambda: {"required": 0, "available": 0}))
    dest_stats = defaultdict(lambda: defaultdict(lambda: {"required": 0, "available": 0}))

    matched = 0
    unmatched = 0
    for _, row in completeness_df.iterrows():
        shipment_id = str(row[id_col])
        milestone = str(row.get(status_col, ""))
        avail = int(row.get(avail_col, 0)) if row.get(avail_col) is not None and row.get(avail_col) != "" else 0

        geo = geo_lookup.get(shipment_id)
        if not geo:
            unmatched += 1
            continue
        matched += 1

        phase = get_milestone_phase(milestone)
        origin = geo.get("origin_country", "Unknown")
        dest = geo.get("dest_country", "Unknown")

        if not origin:
            origin = "Unknown"
        if not dest:
            dest = "Unknown"

        # For origin milestones, group by origin country
        # For destination milestones, group by destination country
        # For in-transit, use origin (departure side issue) or dest (arrival side)
        if phase == "Origin" or phase == "In-Transit":
            origin_stats[origin][milestone]["required"] += 1
            origin_stats[origin][milestone]["available"] += avail
        if phase == "Destination" or phase == "In-Transit":
            dest_stats[dest][milestone]["required"] += 1
            dest_stats[dest][milestone]["available"] += avail

    return origin_stats, dest_stats, matched, unmatched


# SC3 region analysis
sc3_origin_stats = {}
sc3_dest_stats = {}
if not sc3_completeness_latest.empty and sc3_geo:
    id_col = "LOAD_TO" if "LOAD_TO" in sc3_completeness_latest.columns else sc3_completeness_latest.columns[0]
    status_col = "Status Code" if "Status Code" in sc3_completeness_latest.columns else "STATUS_CODE_REQUIRED"
    avail_cols = [c for c in sc3_completeness_latest.columns if "AVAILABLE" in str(c).upper() or "összege" in str(c)]
    avail_col = avail_cols[0] if avail_cols else None

    if avail_col:
        sc3_origin_stats, sc3_dest_stats, m, u = compute_region_completeness(
            sc3_completeness_latest, sc3_geo, id_col, status_col, avail_col
        )
        print(f"  SC3 region analysis: {m} matched, {u} unmatched")

# SC4 region analysis
sc4_origin_stats = {}
sc4_dest_stats = {}
if not sc4_completeness_latest.empty and sc4_geo:
    id_col = "CONSIGNMENT" if "CONSIGNMENT" in sc4_completeness_latest.columns else sc4_completeness_latest.columns[0]
    status_col = "STATUS_CODE_REQUIRED" if "STATUS_CODE_REQUIRED" in sc4_completeness_latest.columns else "Status Code"
    avail_cols = [c for c in sc4_completeness_latest.columns if "AVAILABLE" in str(c).upper() or "összege" in str(c)]
    avail_col = avail_cols[0] if avail_cols else None

    if avail_col:
        sc4_origin_stats, sc4_dest_stats, m, u = compute_region_completeness(
            sc4_completeness_latest, sc4_geo, id_col, status_col, avail_col
        )
        print(f"  SC4 region analysis: {m} matched, {u} unmatched")


# ─── Build region summary tables ─────────────────────────────────────────────

COUNTRY_NAMES = {
    "CN": "China", "DE": "Germany", "HU": "Hungary", "HR": "Croatia",
    "HK": "Hong Kong", "MY": "Malaysia", "VN": "Vietnam", "TW": "Taiwan",
    "TH": "Thailand", "CZ": "Czech Rep.", "CH": "Switzerland", "RO": "Romania",
    "RS": "Serbia", "PL": "Poland", "SI": "Slovenia", "JP": "Japan",
    "PT": "Portugal", "IT": "Italy", "ES": "Spain", "FR": "France",
    "SG": "Singapore", "IN": "India", "ID": "Indonesia", "TR": "Turkey",
    "KR": "South Korea", "US": "United States", "BR": "Brazil", "MX": "Mexico",
    "AT": "Austria", "BE": "Belgium", "NL": "Netherlands", "SE": "Sweden",
    "GB": "United Kingdom",
}


def build_region_summary(region_stats, side_label, top_n=15):
    """Build a summary table of worst-performing regions."""
    # Aggregate across all milestones per region
    region_agg = {}
    for country, milestones in region_stats.items():
        total_req = sum(m["required"] for m in milestones.values())
        total_avail = sum(m["available"] for m in milestones.values())
        if total_req > 0:
            overall_comp = total_avail / total_req
            # Find worst milestone for this region
            worst_milestone = None
            worst_comp = 1.0
            for ms, data in milestones.items():
                if data["required"] >= 5:  # minimum sample
                    ms_comp = data["available"] / data["required"] if data["required"] > 0 else 0
                    if ms_comp < worst_comp:
                        worst_comp = ms_comp
                        worst_milestone = ms

            region_agg[country] = {
                "total_required": total_req,
                "total_available": total_avail,
                "completeness": overall_comp,
                "worst_milestone": worst_milestone,
                "worst_milestone_comp": worst_comp,
                "milestone_detail": milestones,
            }

    # Sort by total missing (required - available), descending
    sorted_regions = sorted(
        region_agg.items(),
        key=lambda x: x[1]["total_required"] - x[1]["total_available"],
        reverse=True,
    )[:top_n]

    return sorted_regions


sc3_origin_summary = build_region_summary(sc3_origin_stats, "Origin")
sc3_dest_summary = build_region_summary(sc3_dest_stats, "Destination")
sc4_origin_summary = build_region_summary(sc4_origin_stats, "Origin")
sc4_dest_summary = build_region_summary(sc4_dest_stats, "Destination")

print(f"  SC3 origin regions: {len(sc3_origin_summary)}, dest: {len(sc3_dest_summary)}")
print(f"  SC4 origin regions: {len(sc4_origin_summary)}, dest: {len(sc4_dest_summary)}")

# Print top issues
for label, summary in [("SC3 Origin", sc3_origin_summary), ("SC3 Dest", sc3_dest_summary),
                        ("SC4 Origin", sc4_origin_summary), ("SC4 Dest", sc4_dest_summary)]:
    if summary:
        top = summary[0]
        c_name = COUNTRY_NAMES.get(top[0], top[0])
        missing = top[1]["total_required"] - top[1]["total_available"]
        print(f"    {label} worst: {c_name} ({top[0]}) - {missing} missing, comp={top[1]['completeness']:.1%}")


def build_region_html_table(summary, side_label):
    """Build HTML table rows for a region summary."""
    rows_html = ""
    for country_code, data in summary:
        c_name = COUNTRY_NAMES.get(country_code, country_code)
        missing = data["total_required"] - data["total_available"]
        worst_ms = data["worst_milestone"] or "N/A"
        worst_ms_name = MILESTONE_NAMES.get(worst_ms, "")
        worst_ms_phase = get_milestone_phase(worst_ms)

        rows_html += f"""
        <tr>
            <td><strong>{country_code}</strong></td>
            <td>{c_name}</td>
            <td>{data['total_required']}</td>
            <td>{missing}</td>
            <td>{status_dot(data['completeness'], TARGET_COMPLETENESS)} {pct(data['completeness'])}</td>
            <td><strong>{worst_ms}</strong> - {worst_ms_name}</td>
            <td>{pct(data['worst_milestone_comp'])}</td>
            <td><span class="badge badge-{'blue' if worst_ms_phase == 'Origin' else 'yellow' if worst_ms_phase == 'In-Transit' else 'purple'}">{worst_ms_phase}</span></td>
        </tr>"""
    return rows_html


sc3_origin_html = build_region_html_table(sc3_origin_summary, "Origin")
sc3_dest_html = build_region_html_table(sc3_dest_summary, "Destination")
sc4_origin_html = build_region_html_table(sc4_origin_summary, "Origin")
sc4_dest_html = build_region_html_table(sc4_dest_summary, "Destination")

# Build region chart data (top 10 origins by missing count for bar chart)
def build_region_chart_data(summary, top_n=10):
    """Build chart data for region analysis."""
    labels = []
    comp_values = []
    missing_values = []
    for country_code, data in summary[:top_n]:
        c_name = COUNTRY_NAMES.get(country_code, country_code)
        labels.append(f"{country_code} ({c_name})")
        comp_values.append(round(data["completeness"] * 100, 1))
        missing_values.append(data["total_required"] - data["total_available"])
    return {"labels": labels, "completeness": comp_values, "missing": missing_values}


sc3_origin_chart = build_region_chart_data(sc3_origin_summary)
sc3_dest_chart = build_region_chart_data(sc3_dest_summary)
sc4_origin_chart = build_region_chart_data(sc4_origin_summary)
sc4_dest_chart = build_region_chart_data(sc4_dest_summary)

# Build per-milestone region heatmap data (top 5 regions x key milestones)
def build_milestone_region_detail(region_stats, key_milestones, top_regions, side):
    """Build detailed milestone x region breakdown."""
    rows_html = ""
    for country_code, _ in top_regions[:8]:
        if country_code not in region_stats:
            continue
        c_name = COUNTRY_NAMES.get(country_code, country_code)
        milestones = region_stats[country_code]

        for ms in key_milestones:
            if ms not in milestones:
                continue
            phase = get_milestone_phase(ms)
            if side == "Origin" and phase not in ["Origin", "In-Transit"]:
                continue
            if side == "Destination" and phase not in ["Destination", "In-Transit"]:
                continue

            data = milestones[ms]
            if data["required"] < 3:
                continue
            comp = data["available"] / data["required"] if data["required"] > 0 else 0
            missing = data["required"] - data["available"]

            rows_html += f"""
            <tr>
                <td><strong>{country_code}</strong> ({c_name})</td>
                <td><strong>{ms}</strong> - {MILESTONE_NAMES.get(ms, '')}</td>
                <td><span class="badge badge-{'blue' if phase == 'Origin' else 'yellow' if phase == 'In-Transit' else 'purple'}">{phase}</span></td>
                <td>{data['required']}</td>
                <td>{missing}</td>
                <td>{status_dot(comp, TARGET_COMPLETENESS)} {pct(comp)}</td>
            </tr>"""
    return rows_html


sc3_origin_detail = build_milestone_region_detail(sc3_origin_stats, KEY_MILESTONES_SC3, sc3_origin_summary, "Origin")
sc3_dest_detail = build_milestone_region_detail(sc3_dest_stats, KEY_MILESTONES_SC3, sc3_dest_summary, "Destination")
sc4_origin_detail = build_milestone_region_detail(sc4_origin_stats, KEY_MILESTONES_SC4, sc4_origin_summary, "Origin")
sc4_dest_detail = build_milestone_region_detail(sc4_dest_stats, KEY_MILESTONES_SC4, sc4_dest_summary, "Destination")


# ─── Generate the HTML ────────────────────────────────────────────────────────

# Color palette
CHART_COLORS = [
    "#3b82f6", "#ef4444", "#10b981", "#f59e0b", "#8b5cf6",
    "#ec4899", "#06b6d4", "#84cc16", "#f97316", "#6366f1",
    "#14b8a6", "#e11d48", "#a855f7", "#0ea5e9", "#22c55e",
]

html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Bosch Milestone Comprehensive Analysis</title>
<style>
  :root {{
    --bg-primary: #0f172a;
    --bg-secondary: #1e293b;
    --bg-card: #1e293b;
    --bg-card-hover: #334155;
    --text-primary: #f1f5f9;
    --text-secondary: #94a3b8;
    --text-muted: #64748b;
    --border-color: #334155;
    --accent-blue: #3b82f6;
    --accent-green: #10b981;
    --accent-red: #ef4444;
    --accent-yellow: #f59e0b;
    --accent-purple: #8b5cf6;
  }}

  * {{ margin: 0; padding: 0; box-sizing: border-box; }}

  body {{
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
    background: var(--bg-primary);
    color: var(--text-primary);
    line-height: 1.6;
  }}

  .container {{ max-width: 1600px; margin: 0 auto; padding: 24px; }}

  .header {{
    background: linear-gradient(135deg, #1e3a5f 0%, #0f172a 100%);
    border: 1px solid var(--border-color);
    border-radius: 16px;
    padding: 32px;
    margin-bottom: 24px;
  }}
  .header h1 {{ font-size: 28px; font-weight: 700; margin-bottom: 8px; }}
  .header .subtitle {{ color: var(--text-secondary); font-size: 14px; }}
  .header .date {{ color: var(--accent-blue); font-size: 13px; margin-top: 4px; }}

  .kpi-grid {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
    gap: 16px;
    margin-bottom: 24px;
  }}

  .kpi-card {{
    background: var(--bg-card);
    border: 1px solid var(--border-color);
    border-radius: 12px;
    padding: 24px;
    transition: all 0.2s;
  }}
  .kpi-card:hover {{ border-color: var(--accent-blue); }}
  .kpi-card .kpi-label {{ font-size: 12px; text-transform: uppercase; letter-spacing: 0.5px; color: var(--text-secondary); margin-bottom: 8px; }}
  .kpi-card .kpi-value {{ font-size: 32px; font-weight: 700; }}
  .kpi-card .kpi-delta {{ font-size: 13px; margin-top: 4px; }}
  .kpi-card .kpi-sub {{ font-size: 12px; color: var(--text-muted); margin-top: 4px; }}

  .section {{
    background: var(--bg-card);
    border: 1px solid var(--border-color);
    border-radius: 12px;
    padding: 24px;
    margin-bottom: 24px;
  }}
  .section-title {{
    font-size: 18px;
    font-weight: 600;
    margin-bottom: 16px;
    padding-bottom: 12px;
    border-bottom: 1px solid var(--border-color);
  }}

  .tabs {{
    display: flex;
    gap: 4px;
    margin-bottom: 16px;
    background: var(--bg-primary);
    border-radius: 8px;
    padding: 4px;
  }}
  .tab {{
    padding: 8px 16px;
    border-radius: 6px;
    cursor: pointer;
    font-size: 13px;
    font-weight: 500;
    color: var(--text-secondary);
    border: none;
    background: transparent;
    transition: all 0.2s;
  }}
  .tab.active {{ background: var(--accent-blue); color: white; }}
  .tab:hover:not(.active) {{ background: var(--bg-card-hover); color: var(--text-primary); }}

  .tab-content {{ display: none; }}
  .tab-content.active {{ display: block; }}

  table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 13px;
  }}
  th {{
    text-align: left;
    padding: 10px 12px;
    background: var(--bg-primary);
    color: var(--text-secondary);
    font-weight: 600;
    font-size: 11px;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    border-bottom: 1px solid var(--border-color);
    position: sticky;
    top: 0;
    cursor: pointer;
    user-select: none;
    white-space: nowrap;
  }}
  th:hover {{ color: var(--text-primary); }}
  th .sort-arrow {{
    display: inline-block;
    margin-left: 4px;
    font-size: 10px;
    color: var(--text-muted);
    opacity: 0.4;
    transition: opacity 0.2s;
  }}
  th.sort-asc .sort-arrow, th.sort-desc .sort-arrow {{
    opacity: 1;
    color: var(--accent-blue);
  }}
  td {{
    padding: 10px 12px;
    border-bottom: 1px solid var(--border-color);
    color: var(--text-primary);
  }}
  tr:hover {{ background: rgba(59, 130, 246, 0.05); }}

  .badge {{
    padding: 2px 8px;
    border-radius: 4px;
    font-size: 11px;
    font-weight: 600;
    text-transform: uppercase;
  }}
  .badge-green {{ background: rgba(16, 185, 129, 0.15); color: #10b981; }}
  .badge-red {{ background: rgba(239, 68, 68, 0.15); color: #ef4444; }}
  .badge-yellow {{ background: rgba(245, 158, 11, 0.15); color: #f59e0b; }}
  .badge-blue {{ background: rgba(59, 130, 246, 0.15); color: #3b82f6; }}
  .badge-purple {{ background: rgba(139, 92, 246, 0.15); color: #8b5cf6; }}

  .region-bar-container {{
    width: 100%;
    height: 340px;
    position: relative;
  }}

  .region-kpi-grid {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
    gap: 12px;
    margin-bottom: 16px;
  }}
  .region-kpi {{
    background: var(--bg-primary);
    border: 1px solid var(--border-color);
    border-radius: 8px;
    padding: 16px;
    text-align: center;
  }}
  .region-kpi .rk-label {{ font-size: 11px; text-transform: uppercase; color: var(--text-secondary); margin-bottom: 4px; }}
  .region-kpi .rk-value {{ font-size: 24px; font-weight: 700; }}

  .chart-container {{
    width: 100%;
    height: 420px;
    position: relative;
  }}

  .grid-2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 24px; }}
  @media (max-width: 1200px) {{ .grid-2 {{ grid-template-columns: 1fr; }} }}

  .fix-row td {{ background: rgba(16, 185, 129, 0.03); }}
  .degrade-row td {{ background: rgba(239, 68, 68, 0.03); }}

  .legend {{
    display: flex;
    flex-wrap: wrap;
    gap: 12px;
    margin-top: 12px;
    font-size: 12px;
  }}
  .legend-item {{
    display: flex;
    align-items: center;
    gap: 6px;
    cursor: pointer;
    opacity: 1;
    transition: opacity 0.2s;
  }}
  .legend-item.hidden {{ opacity: 0.3; }}
  .legend-dot {{
    width: 10px;
    height: 10px;
    border-radius: 50%;
  }}

  .target-line {{
    stroke-dasharray: 6 4;
    stroke-width: 1.5;
  }}

  .nav {{
    display: flex;
    flex-wrap: wrap;
    gap: 8px;
    margin-bottom: 24px;
  }}
  .nav a {{
    color: var(--text-secondary);
    text-decoration: none;
    padding: 6px 12px;
    border-radius: 6px;
    background: var(--bg-card);
    font-size: 13px;
    border: 1px solid var(--border-color);
    transition: all 0.2s;
  }}
  .nav a:hover {{ border-color: var(--accent-blue); color: var(--text-primary); }}

  canvas {{ background: transparent; }}

  .summary-box {{
    background: var(--bg-primary);
    border: 1px solid var(--border-color);
    border-radius: 8px;
    padding: 16px;
    margin-bottom: 16px;
  }}
  .summary-box h4 {{ font-size: 14px; margin-bottom: 8px; color: var(--text-secondary); }}
  .summary-box p {{ font-size: 13px; color: var(--text-primary); line-height: 1.7; }}

  .scroll-table {{
    max-height: 500px;
    overflow-y: auto;
  }}

  .insight-card {{
    background: var(--bg-primary);
    border: 1px solid var(--border-color);
    border-radius: 8px;
    padding: 16px;
    margin-bottom: 12px;
  }}
  .insight-card h4 {{ font-size: 14px; margin-bottom: 8px; }}
  .insight-card ul {{ padding-left: 20px; }}
  .insight-card li {{ font-size: 13px; color: var(--text-primary); line-height: 1.8; }}
</style>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
</head>
<body>
<div class="container">

<!-- Header -->
<div class="header">
  <h1>Bosch Milestone Comprehensive Analysis</h1>
  <div class="subtitle">SC3 (EDI Booking) + SC4 (Email Booking) | Combined CW Weekly + Raw Data Snapshots</div>
  <div class="date">Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Data range: Dec 2025 - Feb 23, 2026 | Weeks: CW01-CW06 + Latest Raw</div>
</div>

<!-- Navigation -->
<div class="nav">
  <a href="#health">Health Overview</a>
  <a href="#insights">Key Insights</a>
  <a href="#wow-trends">WoW Trends</a>
  <a href="#feb-analysis">Feb-Only Analysis</a>
  <a href="#latest-snapshot">Latest Snapshot</a>
  <a href="#sustained-patterns">Sustained Patterns</a>
  <a href="#fix-patterns">Fix Patterns</a>
  <a href="#shipment-fixes">Shipment Fixes</a>
  <a href="#wow-summary">WoW Delta Summary</a>
  <a href="#raw-snapshot-trends">Raw Data Trends</a>
  <a href="#region-analysis">Region Analysis</a>
</div>

<!-- Health Overview -->
<div id="health" class="kpi-grid">
  <div class="kpi-card">
    <div class="kpi-label">SC3 Avg Completeness (Latest Raw)</div>
    <div class="kpi-value" style="color:{'#10b981' if sc3_avg_comp >= TARGET_COMPLETENESS else '#ef4444'}">{pct(sc3_avg_comp)}</div>
    <div class="kpi-delta">{delta_badge((sc3_avg_comp - sc3_avg_comp_cw6)*100)} vs CW06 ({pct(sc3_avg_comp_cw6)})</div>
    <div class="kpi-sub">Target: {pct(TARGET_COMPLETENESS)}</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-label">SC3 Avg Timeliness (Latest Raw)</div>
    <div class="kpi-value" style="color:{'#10b981' if sc3_avg_time >= TARGET_TIMELINESS else '#ef4444'}">{pct(sc3_avg_time)}</div>
    <div class="kpi-delta">{delta_badge((sc3_avg_time - sc3_avg_time_cw6)*100)} vs CW06 ({pct(sc3_avg_time_cw6)})</div>
    <div class="kpi-sub">Target: {pct(TARGET_TIMELINESS)}</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-label">SC4 Avg Completeness (Latest Raw)</div>
    <div class="kpi-value" style="color:{'#10b981' if sc4_avg_comp >= TARGET_COMPLETENESS else '#ef4444'}">{pct(sc4_avg_comp)}</div>
    <div class="kpi-delta">{delta_badge((sc4_avg_comp - sc4_avg_comp_cw6)*100)} vs CW06 ({pct(sc4_avg_comp_cw6)})</div>
    <div class="kpi-sub">Target: {pct(TARGET_COMPLETENESS)}</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-label">SC4 Avg Timeliness (Latest Raw)</div>
    <div class="kpi-value" style="color:{'#10b981' if sc4_avg_time >= TARGET_TIMELINESS else '#ef4444'}">{pct(sc4_avg_time)}</div>
    <div class="kpi-delta">{delta_badge((sc4_avg_time - sc4_avg_time_cw6)*100)} vs CW06 ({pct(sc4_avg_time_cw6)})</div>
    <div class="kpi-sub">Target: {pct(TARGET_TIMELINESS)}</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-label">SC3 Health Score</div>
    <div class="kpi-value" style="color:{'#10b981' if sc3_health >= 0.8 else '#f59e0b' if sc3_health >= 0.6 else '#ef4444'}">{pct(sc3_health)}</div>
    <div class="kpi-delta">{delta_badge((sc3_health - sc3_health_cw6)*100)} vs CW06</div>
    <div class="kpi-sub">60% completeness + 40% timeliness</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-label">SC4 Health Score</div>
    <div class="kpi-value" style="color:{'#10b981' if sc4_health >= 0.8 else '#f59e0b' if sc4_health >= 0.6 else '#ef4444'}">{pct(sc4_health)}</div>
    <div class="kpi-delta">{delta_badge((sc4_health - sc4_health_cw6)*100)} vs CW06</div>
    <div class="kpi-sub">60% completeness + 40% timeliness</div>
  </div>
</div>

<!-- Key Insights -->
<div id="insights" class="section">
  <div class="section-title">Key Insights &amp; Executive Summary</div>
  {insights_html}
</div>

<!-- WoW Trend Charts -->
<div id="wow-trends" class="section">
  <div class="section-title">Week-over-Week (WoW) Milestone Trends (CW01-CW06)</div>

  <div class="tabs">
    <button class="tab active" onclick="switchTab('wow', 'sc3-comp', this)">SC3 Completeness</button>
    <button class="tab" onclick="switchTab('wow', 'sc3-time', this)">SC3 Timeliness</button>
    <button class="tab" onclick="switchTab('wow', 'sc4-comp', this)">SC4 Completeness</button>
    <button class="tab" onclick="switchTab('wow', 'sc4-time', this)">SC4 Timeliness</button>
  </div>

  <div id="wow-sc3-comp" class="tab-content active">
    <div class="chart-container"><canvas id="chart-sc3-comp"></canvas></div>
  </div>
  <div id="wow-sc3-time" class="tab-content">
    <div class="chart-container"><canvas id="chart-sc3-time"></canvas></div>
  </div>
  <div id="wow-sc4-comp" class="tab-content">
    <div class="chart-container"><canvas id="chart-sc4-comp"></canvas></div>
  </div>
  <div id="wow-sc4-time" class="tab-content">
    <div class="chart-container"><canvas id="chart-sc4-time"></canvas></div>
  </div>
</div>

<!-- February-Only Analysis -->
<div id="feb-analysis" class="section">
  <div class="section-title">February 2026 - Month-Specific Performance</div>
  <div class="summary-box">
    <h4>About February Data</h4>
    <p>
      February-only data isolates recent shipments (ordered from Feb 1 onward). Later-stage milestones
      (S07 Arrived, S31 Delivered, S54/S55 Container return) show low rates because these shipments are
      still in transit - this is expected lifecycle behavior, not a performance issue.
      Focus on early-stage milestones (S60, S02, S04) for actionable February performance assessment.
    </p>
  </div>

  <div class="tabs">
    <button class="tab active" onclick="switchTab('feb', 'sc3-feb', this)">SC3 Feb-Only</button>
    <button class="tab" onclick="switchTab('feb', 'sc4-feb', this)">SC4 Feb-Only</button>
  </div>

  <div id="feb-sc3-feb" class="tab-content active">
    <div class="scroll-table">
    <table>
      <thead>
        <tr>
          <th>Code</th><th>Milestone</th><th>Type</th>
          <th>Required</th><th>Available</th>
          <th>Completeness</th><th>Dec Comp</th><th>vs Dec</th>
          <th>Timeliness</th>
        </tr>
      </thead>
      <tbody>{sc3_feb_rows}</tbody>
    </table>
    </div>
  </div>

  <div id="feb-sc4-feb" class="tab-content">
    <div class="scroll-table">
    <table>
      <thead>
        <tr>
          <th>Code</th><th>Milestone</th><th>Type</th>
          <th>Required</th><th>Available</th>
          <th>Completeness</th><th>Dec Comp</th><th>vs Dec</th>
          <th>Timeliness</th>
        </tr>
      </thead>
      <tbody>{sc4_feb_rows}</tbody>
    </table>
    </div>
  </div>
</div>

<!-- WoW Delta Summary -->
<div id="wow-summary" class="section">
  <div class="section-title">WoW Delta Summary (CW05 → CW06)</div>
  <div class="grid-2">
    <div>
      <h3 style="font-size:15px;margin-bottom:12px;color:var(--accent-blue)">SC3 (EDI Booking) - Actual</h3>
      <table>
        <thead>
          <tr>
            <th>Code</th><th>Milestone</th>
            <th>Comp (Prev)</th><th>Comp (Curr)</th><th>Comp Delta</th>
            <th>Time (Prev)</th><th>Time (Curr)</th><th>Time Delta</th>
          </tr>
        </thead>
        <tbody>{sc3_wow_summary}</tbody>
      </table>
    </div>
    <div>
      <h3 style="font-size:15px;margin-bottom:12px;color:var(--accent-purple)">SC4 (Email Booking) - Actual</h3>
      <table>
        <thead>
          <tr>
            <th>Code</th><th>Milestone</th>
            <th>Comp (Prev)</th><th>Comp (Curr)</th><th>Comp Delta</th>
            <th>Time (Prev)</th><th>Time (Curr)</th><th>Time Delta</th>
          </tr>
        </thead>
        <tbody>{sc4_wow_summary}</tbody>
      </table>
    </div>
  </div>
</div>

<!-- Latest Snapshot vs CW06 -->
<div id="latest-snapshot" class="section">
  <div class="section-title">Latest Raw Data Snapshot (Feb 23) vs CW06 (Feb 9-15)</div>

  <div class="tabs">
    <button class="tab active" onclick="switchTab('latest', 'sc3-latest', this)">SC3 (EDI Booking)</button>
    <button class="tab" onclick="switchTab('latest', 'sc4-latest', this)">SC4 (Email Booking)</button>
  </div>

  <div id="latest-sc3-latest" class="tab-content active">
    <div class="scroll-table">
    <table>
      <thead>
        <tr>
          <th>Code</th><th>Milestone</th><th>Type</th>
          <th>Required</th><th>Available</th>
          <th>Completeness</th><th>CW06 Comp</th><th>Delta</th>
          <th>Timeliness</th><th>CW06 Time</th><th>Delta</th>
        </tr>
      </thead>
      <tbody>{sc3_latest_rows}</tbody>
    </table>
    </div>
  </div>

  <div id="latest-sc4-latest" class="tab-content">
    <div class="scroll-table">
    <table>
      <thead>
        <tr>
          <th>Code</th><th>Milestone</th><th>Type</th>
          <th>Required</th><th>Available</th>
          <th>Completeness</th><th>CW06 Comp</th><th>Delta</th>
          <th>Timeliness</th><th>CW06 Time</th><th>Delta</th>
        </tr>
      </thead>
      <tbody>{sc4_latest_rows}</tbody>
    </table>
    </div>
  </div>
</div>

<!-- Sustained Patterns -->
<div id="sustained-patterns" class="section">
  <div class="section-title">Sustained Trend Patterns (Multi-Week Analysis)</div>
  <div class="summary-box">
    <h4>Pattern Types Detected</h4>
    <p>
      <strong>Sustained Improvement:</strong> 3+ consecutive weeks of improvement - indicates successful process or integration fixes.<br>
      <strong>Sustained Degradation:</strong> 3+ consecutive weeks of decline - indicates ongoing or worsening issues.<br>
      <strong>V-Recovery:</strong> Drop followed by recovery - indicates a temporary issue that was identified and corrected.
    </p>
  </div>
  <div class="scroll-table">
  <table>
    <thead>
      <tr>
        <th>Pattern</th><th>Scenario</th><th>Milestone</th>
        <th>Period</th><th>Duration</th>
        <th>Start</th><th>End</th><th>Total Change</th>
      </tr>
    </thead>
    <tbody>
      {sustained_html}
      {'<tr><td colspan="8" style="color:var(--text-muted);text-align:center">No sustained patterns detected with current thresholds</td></tr>' if not sustained_html else ''}
    </tbody>
  </table>
  </div>
</div>

<!-- Fix Patterns -->
<div id="fix-patterns" class="section">
  <div class="section-title">Fix &amp; Issue Patterns (>5pp WoW change detected)</div>
  <div class="summary-box">
    <h4>Pattern Analysis Summary</h4>
    <p>
      Analyzed {len(wow_df['week'].unique())} weeks of CW data.
      Detected <strong>{len([f for f in all_fixes if f['direction']=='IMPROVED'])}</strong> improvement events and
      <strong>{len([f for f in all_fixes if f['direction']=='DEGRADED'])}</strong> degradation events (>5 percentage point change).
      {'<br>Key insight: Milestones showing sustained improvement indicate successful process/integration fixes.' if all_fixes else ''}
    </p>
  </div>
  <div class="scroll-table">
  <table>
    <thead>
      <tr>
        <th>Type</th><th>Scenario</th><th>Milestone</th><th>Status</th>
        <th>Metric</th><th>Period</th>
        <th>From</th><th>To</th><th>Change</th>
      </tr>
    </thead>
    <tbody>{fix_patterns_html}</tbody>
  </table>
  </div>
</div>

<!-- Shipment-level Fixes -->
<div id="shipment-fixes" class="section">
  <div class="section-title">Shipment-Level Fix Detection (Jan 23 → Feb 23)</div>
  <div class="summary-box">
    <h4>How this works</h4>
    <p>
      Compares individual shipment milestone availability between Jan 23 snapshot and Feb 23 snapshot.
      A "fix" means a milestone that was missing (0) in Jan 23 is now present (1) in Feb 23, indicating
      the status update was sent/received after the initial report.
    </p>
  </div>
  <table>
    <thead>
      <tr>
        <th>Scenario</th><th>Milestone</th><th>Shipments Fixed</th><th>Status</th>
      </tr>
    </thead>
    <tbody>
      {shipment_fix_summary_sc3}
      {shipment_fix_summary_sc4}
      {'<tr><td colspan="4" style="color:var(--text-muted);text-align:center">No shipment-level comparison data available</td></tr>' if not shipment_fix_summary_sc3 and not shipment_fix_summary_sc4 else ''}
    </tbody>
  </table>
</div>

<!-- Raw Data Snapshot Trends -->
<div id="raw-snapshot-trends" class="section">
  <div class="section-title">Raw Data Cumulative Snapshot Trends</div>
  <div class="summary-box">
    <h4>About these charts</h4>
    <p>
      These charts show milestone metrics from the new raw data files, using only cumulative snapshots
      (Dec onward) for fair comparison. Each point represents a Bosch BI report extracted at that date.
    </p>
  </div>

  <div class="tabs">
    <button class="tab active" onclick="switchTab('raw', 'raw-sc3', this)">SC3 Completeness</button>
    <button class="tab" onclick="switchTab('raw', 'raw-sc4', this)">SC4 Completeness</button>
  </div>

  <div id="raw-raw-sc3" class="tab-content active">
    <div class="chart-container"><canvas id="chart-raw-sc3"></canvas></div>
  </div>
  <div id="raw-raw-sc4" class="tab-content">
    <div class="chart-container"><canvas id="chart-raw-sc4"></canvas></div>
  </div>
</div>

</div>

<!-- Region Analysis -->
<div id="region-analysis" class="section">
  <div class="section-title">Region-Based Milestone Analysis (Origin vs Destination)</div>
  <div class="summary-box">
    <h4>How to read this section</h4>
    <p>
      Milestones are classified as <span class="badge badge-blue">Origin</span> (pickup, loading, departure-side),
      <span class="badge badge-yellow">In-Transit</span> (departure & arrival events), or
      <span class="badge badge-purple">Destination</span> (arrival, delivery, customs-side).
      Origin milestones are grouped by <strong>origin country</strong>; Destination milestones by <strong>destination country</strong>.
      In-Transit milestones appear on both sides. Sorted by total missing milestone events (= gap to close).
    </p>
  </div>

  <div class="region-kpi-grid">
    <div class="region-kpi">
      <div class="rk-label">SC3 Origin Countries</div>
      <div class="rk-value">{len(sc3_origin_summary)}</div>
    </div>
    <div class="region-kpi">
      <div class="rk-label">SC3 Dest Countries</div>
      <div class="rk-value">{len(sc3_dest_summary)}</div>
    </div>
    <div class="region-kpi">
      <div class="rk-label">SC4 Origin Countries</div>
      <div class="rk-value">{len(sc4_origin_summary)}</div>
    </div>
    <div class="region-kpi">
      <div class="rk-label">SC4 Dest Countries</div>
      <div class="rk-value">{len(sc4_dest_summary)}</div>
    </div>
  </div>

  <div class="tabs">
    <button class="tab active" onclick="switchTab('region', 'sc3-origin', this)">SC3 (EDI) Origin</button>
    <button class="tab" onclick="switchTab('region', 'sc3-dest', this)">SC3 (EDI) Destination</button>
    <button class="tab" onclick="switchTab('region', 'sc4-origin', this)">SC4 (Email) Origin</button>
    <button class="tab" onclick="switchTab('region', 'sc4-dest', this)">SC4 (Email) Destination</button>
  </div>

  <!-- SC3 Origin -->
  <div id="region-sc3-origin" class="tab-content active">
    <h4 style="margin-bottom:12px; color: var(--text-secondary);">SC3 (EDI Booking) &mdash; Worst-Performing Origin Countries</h4>
    <div class="region-bar-container"><canvas id="chart-region-sc3-origin"></canvas></div>
    <div style="max-height:400px; overflow-y:auto; margin-top:16px;">
    <table>
      <thead>
        <tr><th>Code</th><th>Country</th><th>Required</th><th>Missing</th><th>Completeness</th><th>Worst Milestone</th><th>Ms Comp%</th><th>Phase</th></tr>
      </thead>
      <tbody>{sc3_origin_html}</tbody>
    </table>
    </div>
    <h4 style="margin:20px 0 12px; color: var(--text-secondary);">Milestone x Region Breakdown (Origin Side)</h4>
    <div style="max-height:400px; overflow-y:auto;">
    <table>
      <thead>
        <tr><th>Region</th><th>Milestone</th><th>Phase</th><th>Required</th><th>Missing</th><th>Completeness</th></tr>
      </thead>
      <tbody>{sc3_origin_detail}</tbody>
    </table>
    </div>
  </div>

  <!-- SC3 Destination -->
  <div id="region-sc3-dest" class="tab-content">
    <h4 style="margin-bottom:12px; color: var(--text-secondary);">SC3 (EDI Booking) &mdash; Worst-Performing Destination Countries</h4>
    <div class="region-bar-container"><canvas id="chart-region-sc3-dest"></canvas></div>
    <div style="max-height:400px; overflow-y:auto; margin-top:16px;">
    <table>
      <thead>
        <tr><th>Code</th><th>Country</th><th>Required</th><th>Missing</th><th>Completeness</th><th>Worst Milestone</th><th>Ms Comp%</th><th>Phase</th></tr>
      </thead>
      <tbody>{sc3_dest_html}</tbody>
    </table>
    </div>
    <h4 style="margin:20px 0 12px; color: var(--text-secondary);">Milestone x Region Breakdown (Destination Side)</h4>
    <div style="max-height:400px; overflow-y:auto;">
    <table>
      <thead>
        <tr><th>Region</th><th>Milestone</th><th>Phase</th><th>Required</th><th>Missing</th><th>Completeness</th></tr>
      </thead>
      <tbody>{sc3_dest_detail}</tbody>
    </table>
    </div>
  </div>

  <!-- SC4 Origin -->
  <div id="region-sc4-origin" class="tab-content">
    <h4 style="margin-bottom:12px; color: var(--text-secondary);">SC4 (Email Booking) &mdash; Worst-Performing Origin Countries</h4>
    <div class="region-bar-container"><canvas id="chart-region-sc4-origin"></canvas></div>
    <div style="max-height:400px; overflow-y:auto; margin-top:16px;">
    <table>
      <thead>
        <tr><th>Code</th><th>Country</th><th>Required</th><th>Missing</th><th>Completeness</th><th>Worst Milestone</th><th>Ms Comp%</th><th>Phase</th></tr>
      </thead>
      <tbody>{sc4_origin_html}</tbody>
    </table>
    </div>
    <h4 style="margin:20px 0 12px; color: var(--text-secondary);">Milestone x Region Breakdown (Origin Side)</h4>
    <div style="max-height:400px; overflow-y:auto;">
    <table>
      <thead>
        <tr><th>Region</th><th>Milestone</th><th>Phase</th><th>Required</th><th>Missing</th><th>Completeness</th></tr>
      </thead>
      <tbody>{sc4_origin_detail}</tbody>
    </table>
    </div>
  </div>

  <!-- SC4 Destination -->
  <div id="region-sc4-dest" class="tab-content">
    <h4 style="margin-bottom:12px; color: var(--text-secondary);">SC4 (Email Booking) &mdash; Worst-Performing Destination Countries</h4>
    <div class="region-bar-container"><canvas id="chart-region-sc4-dest"></canvas></div>
    <div style="max-height:400px; overflow-y:auto; margin-top:16px;">
    <table>
      <thead>
        <tr><th>Code</th><th>Country</th><th>Required</th><th>Missing</th><th>Completeness</th><th>Worst Milestone</th><th>Ms Comp%</th><th>Phase</th></tr>
      </thead>
      <tbody>{sc4_dest_html}</tbody>
    </table>
    </div>
    <h4 style="margin:20px 0 12px; color: var(--text-secondary);">Milestone x Region Breakdown (Destination Side)</h4>
    <div style="max-height:400px; overflow-y:auto;">
    <table>
      <thead>
        <tr><th>Region</th><th>Milestone</th><th>Phase</th><th>Required</th><th>Missing</th><th>Completeness</th></tr>
      </thead>
      <tbody>{sc4_dest_detail}</tbody>
    </table>
    </div>
  </div>
</div>

</div><!-- container -->

<script>
const COLORS = {json.dumps(CHART_COLORS)};

function switchTab(group, tabId, btn) {{
  const parent = btn.closest('.section');
  if (!parent) return;

  // Deactivate all tabs in this group
  parent.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
  btn.classList.add('active');

  // Hide all tab content in this section, show the target
  parent.querySelectorAll('.tab-content').forEach(tc => tc.classList.remove('active'));
  const targetId = group + '-' + tabId;
  const targetEl = document.getElementById(targetId);
  if (targetEl) targetEl.classList.add('active');

  // Trigger chart resize so hidden charts render correctly
  setTimeout(() => window.dispatchEvent(new Event('resize')), 50);
}}

function createLineChart(canvasId, chartData, targetValue, yLabel) {{
  const ctx = document.getElementById(canvasId);
  if (!ctx) return;

  const datasets = [];
  const labels = Object.keys(chartData);
  let allWeeks = new Set();

  labels.forEach(label => {{
    chartData[label].weeks.forEach(w => allWeeks.add(w));
  }});

  const sortedWeeks = Array.from(allWeeks).sort();

  labels.forEach((label, idx) => {{
    const data = sortedWeeks.map(w => {{
      const wIdx = chartData[label].weeks.indexOf(w);
      return wIdx >= 0 ? chartData[label].values[wIdx] : null;
    }});

    datasets.push({{
      label: label,
      data: data,
      borderColor: COLORS[idx % COLORS.length],
      backgroundColor: COLORS[idx % COLORS.length] + '20',
      borderWidth: 2,
      pointRadius: 4,
      pointHoverRadius: 6,
      tension: 0.3,
      fill: false,
    }});
  }});

  // Add target line
  datasets.push({{
    label: `Target (${{targetValue}}%)`,
    data: sortedWeeks.map(() => targetValue),
    borderColor: '#94a3b8',
    borderWidth: 1.5,
    borderDash: [6, 4],
    pointRadius: 0,
    fill: false,
  }});

  new Chart(ctx, {{
    type: 'line',
    data: {{ labels: sortedWeeks, datasets: datasets }},
    options: {{
      responsive: true,
      maintainAspectRatio: false,
      plugins: {{
        legend: {{
          position: 'bottom',
          labels: {{ color: '#94a3b8', font: {{ size: 11 }}, padding: 12, usePointStyle: true }}
        }},
        tooltip: {{
          backgroundColor: '#1e293b',
          titleColor: '#f1f5f9',
          bodyColor: '#94a3b8',
          borderColor: '#334155',
          borderWidth: 1,
          callbacks: {{
            label: function(context) {{
              return context.dataset.label + ': ' + context.parsed.y + '%';
            }}
          }}
        }}
      }},
      scales: {{
        x: {{
          grid: {{ color: '#334155' }},
          ticks: {{ color: '#94a3b8', font: {{ size: 11 }} }}
        }},
        y: {{
          min: 0,
          max: 100,
          grid: {{ color: '#334155' }},
          ticks: {{ color: '#94a3b8', font: {{ size: 11 }}, callback: v => v + '%' }},
          title: {{ display: true, text: yLabel, color: '#94a3b8' }}
        }}
      }}
    }}
  }});
}}

function createSnapshotChart(canvasId, chartData) {{
  const ctx = document.getElementById(canvasId);
  if (!ctx) return;

  const datasets = [];
  const labels = Object.keys(chartData);
  let allSnapLabels = new Set();

  labels.forEach(label => {{
    chartData[label].labels.forEach(l => allSnapLabels.add(l));
  }});

  const sortedLabels = Array.from(allSnapLabels);

  labels.forEach((label, idx) => {{
    const data = sortedLabels.map(sl => {{
      const sIdx = chartData[label].labels.indexOf(sl);
      return sIdx >= 0 ? chartData[label].completeness[sIdx] : null;
    }});

    datasets.push({{
      label: label,
      data: data,
      borderColor: COLORS[idx % COLORS.length],
      backgroundColor: COLORS[idx % COLORS.length] + '20',
      borderWidth: 2,
      pointRadius: 5,
      pointHoverRadius: 7,
      tension: 0.3,
      fill: false,
    }});
  }});

  // Add target
  datasets.push({{
    label: 'Target (95%)',
    data: sortedLabels.map(() => 95),
    borderColor: '#94a3b8',
    borderWidth: 1.5,
    borderDash: [6, 4],
    pointRadius: 0,
    fill: false,
  }});

  new Chart(ctx, {{
    type: 'line',
    data: {{ labels: sortedLabels, datasets: datasets }},
    options: {{
      responsive: true,
      maintainAspectRatio: false,
      plugins: {{
        legend: {{
          position: 'bottom',
          labels: {{ color: '#94a3b8', font: {{ size: 11 }}, padding: 12, usePointStyle: true }}
        }},
        tooltip: {{
          backgroundColor: '#1e293b',
          titleColor: '#f1f5f9',
          bodyColor: '#94a3b8',
          borderColor: '#334155',
          borderWidth: 1,
        }}
      }},
      scales: {{
        x: {{
          grid: {{ color: '#334155' }},
          ticks: {{ color: '#94a3b8', font: {{ size: 11 }} }}
        }},
        y: {{
          min: 0,
          max: 100,
          grid: {{ color: '#334155' }},
          ticks: {{ color: '#94a3b8', font: {{ size: 11 }}, callback: v => v + '%' }},
          title: {{ display: true, text: 'Completeness %', color: '#94a3b8' }}
        }}
      }}
    }}
  }});
}}

function createRegionBarChart(canvasId, chartData, title) {{
  const ctx = document.getElementById(canvasId);
  if (!ctx || !chartData || !chartData.labels || chartData.labels.length === 0) return;

  new Chart(ctx, {{
    type: 'bar',
    data: {{
      labels: chartData.labels,
      datasets: [
        {{
          label: 'Missing Milestones',
          data: chartData.missing,
          backgroundColor: 'rgba(239, 68, 68, 0.7)',
          borderColor: '#ef4444',
          borderWidth: 1,
          borderRadius: 4,
          yAxisID: 'y',
        }},
        {{
          label: 'Completeness %',
          data: chartData.completeness,
          type: 'line',
          borderColor: '#3b82f6',
          backgroundColor: 'rgba(59, 130, 246, 0.1)',
          borderWidth: 2,
          pointRadius: 4,
          pointHoverRadius: 6,
          tension: 0.3,
          fill: false,
          yAxisID: 'y1',
        }}
      ]
    }},
    options: {{
      responsive: true,
      maintainAspectRatio: false,
      plugins: {{
        legend: {{
          position: 'bottom',
          labels: {{ color: '#94a3b8', font: {{ size: 11 }}, padding: 12, usePointStyle: true }}
        }},
        title: {{
          display: true,
          text: title,
          color: '#f1f5f9',
          font: {{ size: 14 }}
        }},
        tooltip: {{
          backgroundColor: '#1e293b',
          titleColor: '#f1f5f9',
          bodyColor: '#94a3b8',
          borderColor: '#334155',
          borderWidth: 1,
        }}
      }},
      scales: {{
        x: {{
          grid: {{ color: '#334155' }},
          ticks: {{ color: '#94a3b8', font: {{ size: 10 }}, maxRotation: 45 }}
        }},
        y: {{
          position: 'left',
          grid: {{ color: '#334155' }},
          ticks: {{ color: '#ef4444', font: {{ size: 11 }} }},
          title: {{ display: true, text: 'Missing Count', color: '#ef4444' }}
        }},
        y1: {{
          position: 'right',
          min: 0,
          max: 100,
          grid: {{ drawOnChartArea: false }},
          ticks: {{ color: '#3b82f6', font: {{ size: 11 }}, callback: v => v + '%' }},
          title: {{ display: true, text: 'Completeness %', color: '#3b82f6' }}
        }}
      }}
    }}
  }});
}}

// Initialize charts
document.addEventListener('DOMContentLoaded', function() {{
  createLineChart('chart-sc3-comp', {json.dumps(sc3_comp_series)}, {TARGET_COMPLETENESS*100}, 'Completeness %');
  createLineChart('chart-sc3-time', {json.dumps(sc3_time_series)}, {TARGET_TIMELINESS*100}, 'Timeliness %');
  createLineChart('chart-sc4-comp', {json.dumps(sc4_comp_series)}, {TARGET_COMPLETENESS*100}, 'Completeness %');
  createLineChart('chart-sc4-time', {json.dumps(sc4_time_series)}, {TARGET_TIMELINESS*100}, 'Timeliness %');

  createSnapshotChart('chart-raw-sc3', {json.dumps(sc3_snapshot_chart)});
  createSnapshotChart('chart-raw-sc4', {json.dumps(sc4_snapshot_chart)});

  // Region bar charts
  createRegionBarChart('chart-region-sc3-origin', {json.dumps(sc3_origin_chart)}, 'SC3 Origin - Missing Milestones by Country');
  createRegionBarChart('chart-region-sc3-dest', {json.dumps(sc3_dest_chart)}, 'SC3 Destination - Missing Milestones by Country');
  createRegionBarChart('chart-region-sc4-origin', {json.dumps(sc4_origin_chart)}, 'SC4 Origin - Missing Milestones by Country');
  createRegionBarChart('chart-region-sc4-dest', {json.dumps(sc4_dest_chart)}, 'SC4 Destination - Missing Milestones by Country');

  // Make all tables sortable
  initSortableTables();
}});

function initSortableTables() {{
  document.querySelectorAll('table').forEach(table => {{
    const headers = table.querySelectorAll('th');
    if (headers.length === 0) return;

    headers.forEach((th, colIdx) => {{
      // Add sort arrow indicator
      const arrow = document.createElement('span');
      arrow.className = 'sort-arrow';
      arrow.textContent = '\u2195';
      th.appendChild(arrow);

      th.addEventListener('click', function() {{
        sortTable(table, colIdx, th);
      }});
    }});
  }});
}}

function parseSortValue(text) {{
  // Strip HTML tags to get raw text
  const clean = text.replace(/<[^>]*>/g, '').trim();

  // Try percentage (e.g. "85.3%" or "  85.3%")
  const pctMatch = clean.match(/([\d.]+)%/);
  if (pctMatch) return parseFloat(pctMatch[1]);

  // Try signed number with pp suffix (e.g. "+5.2pp" or "-3.1pp")
  const ppMatch = clean.match(/([+-]?[\d.]+)\s*pp/);
  if (ppMatch) return parseFloat(ppMatch[1]);

  // Try plain number / comma-separated number
  const numClean = clean.replace(/,/g, '');
  const numMatch = numClean.match(/^-?[\d.]+$/);
  if (numMatch) return parseFloat(numClean);

  // Try number anywhere in the string (for cells like "3 weeks")
  const anyNum = clean.match(/([\d,]+)/);
  if (anyNum) return parseFloat(anyNum[1].replace(/,/g, ''));

  // Fall back to lowercase string
  return clean.toLowerCase();
}}

function sortTable(table, colIdx, clickedTh) {{
  const tbody = table.querySelector('tbody') || table;
  const rows = Array.from(tbody.querySelectorAll('tr'));
  if (rows.length === 0) return;

  // Determine sort direction
  const isAsc = clickedTh.classList.contains('sort-asc');
  const dir = isAsc ? -1 : 1;

  // Clear sort classes from all headers in this table
  table.querySelectorAll('th').forEach(th => {{
    th.classList.remove('sort-asc', 'sort-desc');
  }});
  clickedTh.classList.add(isAsc ? 'sort-desc' : 'sort-asc');

  rows.sort((a, b) => {{
    const cellA = a.cells[colIdx];
    const cellB = b.cells[colIdx];
    if (!cellA || !cellB) return 0;

    const valA = parseSortValue(cellA.innerHTML);
    const valB = parseSortValue(cellB.innerHTML);

    if (typeof valA === 'number' && typeof valB === 'number') {{
      return (valA - valB) * dir;
    }}
    return String(valA).localeCompare(String(valB)) * dir;
  }});

  rows.forEach(row => tbody.appendChild(row));
}}
</script>

</body>
</html>
"""

with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
    f.write(html)

print(f"\n{'='*70}")
print(f"Dashboard saved to: {OUTPUT_FILE}")
print(f"{'='*70}")
print(f"\nSummary:")
print(f"  SC3 Health: {pct(sc3_health)} (Comp: {pct(sc3_avg_comp)}, Time: {pct(sc3_avg_time)})")
print(f"  SC4 Health: {pct(sc4_health)} (Comp: {pct(sc4_avg_comp)}, Time: {pct(sc4_avg_time)})")
print(f"  Total fix events detected: {len(all_fixes)}")
print(f"  Improvements: {len([f for f in all_fixes if f['direction']=='IMPROVED'])}")
print(f"  Degradations: {len([f for f in all_fixes if f['direction']=='DEGRADED'])}")
