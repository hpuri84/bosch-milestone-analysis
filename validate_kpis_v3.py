"""
Validate KPI calculations against PDCA template values for CW05-CW07.
v3: Uses exact PDCA formulas decoded from the template:
  - Completeness = sum(Available) / sum(Required)
  - Timeliness = sum(In_Time) / sum(Available)  [NOT /Required]
  - Critical = S00(SC4), S02, S04, S07, S45, S31 (both Actual + Estimated)
  - All (SC4) = rows 8-25 (S46..S31, excludes S00+S02), All (SC3) = all rows
"""

import openpyxl
import os

BASE = "/Users/harsh.puri/Documents/work-maersk/Prototype Playground/Bosch Milestone Analysis"
RAW_DIR = os.path.join(BASE, "Bosch Milestone raw data")

SC3_FILES = {
    "CW01": "Maersk NGTM SC3_2026_CW01.xlsx",
    "CW02": "Maersk NGTM SC3_2026_CW02.xlsx",
    "CW03": "Maersk NGTM SC3_2026_CW03.xlsx",
    "CW04": "Maersk NGTM SC3_2026_CW04.xlsx",
    "CW05": "Maersk NGTM SC3_2026_CW05.xlsx",
    "CW06": "Maersk NGTM SC3_2026_CW06.xlsx",
    "CW07": "Maersk NGTM SC3_2026_CW07.xlsx",
}

SC4_FILES = {
    "CW01": "Maersk SC4_2026_CW01.xlsx",
    "CW02": "Maersk SC4_2026_CW02v2.xlsx",
    "CW03": "Maersk SC4_2026_CW03.xlsx",
    "CW04": "Maersk SC4_2026_CW04.xlsx",
    "CW05": "Maersk SC4_2026_CW05.xlsx",
    "CW06": "Maersk SC4_2026_CW06.xlsx",
    "CW07": "Maersk SC4_2026_CW07.xlsx",
}

# Exact milestone+type combos from PDCA formula
# Critical (SC4 Data rows): 5(S00-Act), 6(S02-Est), 7(S02-Act), 14(S04-Est), 15(S04-Act),
#   16(S07-Est), 17(S07-Act), 22(S45-Act), 24(S31-Est), 25(S31-Act)
# Critical (SC3 Data rows): 7(S02-Act), 8(S02-Est), 12(S04-Act), 13(S04-Est),
#   14(S07-Act), 15(S07-Est), 19(S45-Act), 20(S45-Est), 22(S31-Act), 23(S31-Est)
SC4_CRITICAL_CODES = {
    ("S00", "Actual"), ("S02", "Estimated"), ("S02", "Actual"),
    ("S04", "Estimated"), ("S04", "Actual"), ("S07", "Estimated"), ("S07", "Actual"),
    ("S45", "Actual"), ("S31", "Estimated"), ("S31", "Actual"),
}
SC3_CRITICAL_CODES = {
    ("S02", "Actual"), ("S02", "Estimated"),
    ("S04", "Actual"), ("S04", "Estimated"), ("S07", "Actual"), ("S07", "Estimated"),
    ("S45", "Actual"), ("S45", "Estimated"), ("S31", "Actual"), ("S31", "Estimated"),
}

# "All" formula:
# SC4 rows 8-25 = everything EXCEPT S00(row5), S02-Est(row6), S02-Act(row7)
# SC3 rows 5-24 = everything
SC4_ALL_EXCLUDE = {("S00", "Actual"), ("S02", "Estimated"), ("S02", "Actual")}


def find_total_sheet(wb):
    for sn in wb.sheetnames:
        s = sn.strip().upper()
        if s.startswith("TOTAL") or s == "ALL":
            return sn
    return None


def find_shipments_sheet(wb):
    for sn in wb.sheetnames:
        if sn.strip().lower() == "shipments":
            return sn
    return None


def parse_milestone_code(name):
    if name and " - " in str(name):
        return str(name).split(" - ")[0].strip()
    return str(name).strip() if name else None


def read_summary_sheet(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = find_total_sheet(wb)
    if not sheet_name:
        print(f"  WARNING: No TOTAL sheet in {os.path.basename(filepath)}")
        wb.close()
        return []

    ws = wb[sheet_name]
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        for cell in row:
            if cell and "required" in str(cell).lower() and "status" in str(cell).lower():
                header_row = i
                break
        if header_row:
            break
    if not header_row:
        header_row = 3

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        milestone_name = row[0]
        if not milestone_name or "required" in str(milestone_name).lower():
            continue
        code = parse_milestone_code(milestone_name)
        status_type = str(row[1]).strip() if row[1] else ""
        required = row[3] if row[3] and isinstance(row[3], (int, float)) else 0
        available = row[4] if row[4] and isinstance(row[4], (int, float)) else 0
        in_time = row[5] if row[5] and isinstance(row[5], (int, float)) else 0

        rows.append({
            "code": code,
            "type": status_type,
            "required": required,
            "available": available,
            "in_time": in_time,
        })
    wb.close()
    return rows


def read_sc4_shipments(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = find_shipments_sheet(wb)
    if not sheet_name:
        wb.close()
        return []

    ws = wb[sheet_name]
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        for cell in row:
            if cell and "UNIQUE_SHIPMENT" in str(cell):
                header_row = i
                break
        if header_row:
            break
    if not header_row:
        header_row = 3

    headers = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            col_map[str(h).strip()] = i

    shipments = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        vals = list(row)
        if not vals[0]:
            continue
        shipment = {}
        for name, idx in col_map.items():
            shipment[name] = vals[idx] if idx < len(vals) else None
        shipments.append(shipment)

    wb.close()
    return shipments


# PDCA reference values
PDCA = {
    "Comp_Crit": {"W5": 0.8277, "W6": 0.8168, "W7": 0.8315},
    "Time_Crit": {"W5": 0.5885, "W6": 0.6012, "W7": 0.6415},
    "Comp_All":  {"W5": 0.7481, "W6": 0.7585, "W7": 0.7838},
    "Time_All":  {"W5": 0.5996, "W6": 0.6140, "W7": 0.6257},
    "ETA_2P":    {"W5": 0.44,   "W6": 0.62,   "W7": 0.39},
    "ETA_2D":    {"W5": 0.13,   "W6": 0.17,   "W7": 0.17},
    "Ref_Comp":  {"W5": 0.71,   "W6": 0.74,   "W7": 0.76},
}


def match_check(computed, pdca, tol=0.005):
    diff = abs(computed - pdca)
    return "MATCH" if diff < tol else f"MISS (d={diff:.4f})"


def main():
    print("=" * 100)
    print("BOSCH MILESTONE KPI VALIDATION v3 — Using Exact PDCA Formulas")
    print("  Completeness = sum(Available) / sum(Required)")
    print("  Timeliness   = sum(In_Time) / sum(Available)")
    print("  Critical     = S00(SC4), S02, S04, S07, S45, S31 (Act+Est)")
    print("  All(SC4)     = All milestones EXCEPT S00 & S02")
    print("  All(SC3)     = All milestones")
    print("=" * 100)

    for week in ["CW05", "CW06", "CW07"]:
        wkey = week.replace("CW0", "W")
        print(f"\n{'─'*100}")
        print(f"  {week} ({wkey})")
        print(f"{'─'*100}")

        sc3_file = os.path.join(RAW_DIR, SC3_FILES[week])
        sc4_file = os.path.join(RAW_DIR, SC4_FILES[week])

        sc3_data = read_summary_sheet(sc3_file)
        sc4_data = read_summary_sheet(sc4_file)

        # === CRITICAL ===
        sc3_crit = [r for r in sc3_data if (r["code"], r["type"]) in SC3_CRITICAL_CODES]
        sc4_crit = [r for r in sc4_data if (r["code"], r["type"]) in SC4_CRITICAL_CODES]
        crit = sc3_crit + sc4_crit

        crit_req = sum(r["required"] for r in crit)
        crit_avl = sum(r["available"] for r in crit)
        crit_it = sum(r["in_time"] for r in crit)
        crit_comp = crit_avl / crit_req if crit_req > 0 else 0
        crit_time = crit_it / crit_avl if crit_avl > 0 else 0

        pc = PDCA["Comp_Crit"][wkey]
        pt = PDCA["Time_Crit"][wkey]

        print(f"\n  CRITICAL MILESTONES (SC3: {len(sc3_crit)} rows, SC4: {len(sc4_crit)} rows)")
        print(f"    Required={crit_req}  Available={crit_avl}  InTime={crit_it}")
        print(f"    Completeness: {crit_comp:.4f}  PDCA={pc:.4f}  [{match_check(crit_comp, pc)}]")
        print(f"    Timeliness:   {crit_time:.4f}  PDCA={pt:.4f}  [{match_check(crit_time, pt)}]")

        # Print which milestones were included
        for label, rows in [("SC3", sc3_crit), ("SC4", sc4_crit)]:
            for r in rows:
                print(f"      {label} {r['code']:4s} {r['type']:10s} | req={r['required']:4} avl={r['available']:4} it={r['in_time']:4}")

        # === ALL ===
        sc3_all = sc3_data  # All SC3 milestones
        sc4_all = [r for r in sc4_data if (r["code"], r["type"]) not in SC4_ALL_EXCLUDE]

        all_rows = sc3_all + sc4_all
        all_req = sum(r["required"] for r in all_rows)
        all_avl = sum(r["available"] for r in all_rows)
        all_it = sum(r["in_time"] for r in all_rows)
        all_comp = all_avl / all_req if all_req > 0 else 0
        all_time = all_it / all_avl if all_avl > 0 else 0

        pac = PDCA["Comp_All"][wkey]
        pat = PDCA["Time_All"][wkey]

        print(f"\n  ALL MILESTONES (SC3: {len(sc3_all)} rows, SC4: {len(sc4_all)} rows)")
        print(f"    Required={all_req}  Available={all_avl}  InTime={all_it}")
        print(f"    Completeness: {all_comp:.4f}  PDCA={pac:.4f}  [{match_check(all_comp, pac)}]")
        print(f"    Timeliness:   {all_time:.4f}  PDCA={pat:.4f}  [{match_check(all_time, pat)}]")

        # === ETA ACCURACY ===
        shipments = read_sc4_shipments(sc4_file)

        # ETA 2P (S07_Accepted)
        s07_col = None
        s31_col = None
        civ_col = dn_col = None
        for key in (shipments[0].keys() if shipments else []):
            if "S07_Accepted" in key:
                s07_col = key
            if "S31_Accepted" in key:
                s31_col = key
            if "COMMERCIAL_INVOICE" in key:
                civ_col = key
            if "DELIVERY_NOTE" in key:
                dn_col = key

        s07_total = sum(1 for s in shipments if s.get(s07_col) is not None)
        s07_acc = sum(1 for s in shipments if s.get(s07_col) == 1)
        eta_2p = s07_acc / s07_total if s07_total > 0 else 0

        s31_total = sum(1 for s in shipments if s.get(s31_col) is not None)
        s31_acc = sum(1 for s in shipments if s.get(s31_col) == 1)
        eta_2d = s31_acc / s31_total if s31_total > 0 else 0

        ref_total = len(shipments)
        ref_complete = 0
        for s in shipments:
            civ = s.get(civ_col)
            dn = s.get(dn_col)
            has_civ = civ is not None and str(civ).strip() not in ("", "None")
            has_dn = dn is not None and str(dn).strip() not in ("", "None")
            if has_civ or has_dn:
                ref_complete += 1
        ref_comp = ref_complete / ref_total if ref_total > 0 else 0

        p2p = PDCA["ETA_2P"][wkey]
        p2d = PDCA["ETA_2D"][wkey]
        pref = PDCA["Ref_Comp"][wkey]

        print(f"\n  ETA ACCURACY & REFERENCE (SC4)")
        print(f"    ETA 2P (S07): {s07_acc:3d}/{s07_total:3d} = {eta_2p:.4f}  PDCA={p2p:.2f}  [{match_check(eta_2p, p2p, 0.02)}]")
        print(f"    ETA 2D (S31): {s31_acc:3d}/{s31_total:3d} = {eta_2d:.4f}  PDCA={p2d:.2f}  [{match_check(eta_2d, p2d, 0.02)}]")
        print(f"    Ref Complete: {ref_complete:3d}/{ref_total:3d} = {ref_comp:.4f}  PDCA={pref:.2f}  [{match_check(ref_comp, pref, 0.02)}]")


if __name__ == "__main__":
    main()
