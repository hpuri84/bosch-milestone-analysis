"""Build the ETA 2D pushback evidence workbook from the reconciled rollups + shipments.
Static analytical export (no formulas -> zero formula errors). Optional enrichment
from eta_2d_findings.json (the multi-agent synthesis) if present.
"""
import json, os
from collections import defaultdict
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

ROLL = json.load(open("eta_2d_rollups.json"))
SHIPS = json.load(open("eta_2d_shipments.json"))
FIND = json.load(open("eta_2d_findings.json")) if os.path.exists("eta_2d_findings.json") else None

ORANGE = "C15F3C"; INK = "2B2B28"; CREAM = "FAF9F5"; BAND = "F2F1EC"; BORDER = "E5E4DF"
TEXT = "Calibri"; MONO = "Consolas"
hdr_font = Font(name=TEXT, bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill("solid", fgColor=ORANGE)
title_font = Font(name=TEXT, bold=True, color=INK, size=16)
sub_font = Font(name=TEXT, color="6B6862", size=10)
thin = Side(style="thin", color=BORDER)
box = Border(left=thin, right=thin, top=thin, bottom=thin)


def mode(g):
    if (g.get("no_t7_rate") or 0) >= 0.5: return "no-T7 transmission"
    if (g.get("early_gt48") or 0) > (g.get("late_gt48") or 0): return "early-bias (measurement)"
    if (g.get("avg_dev_days") or 0) > 14: return "gross-late"
    return "late"


def sheet(wb, name, headers, rows, mono_cols=(), pct_cols=(), dev_cols=(), heat_col=None, widths=None, freeze="A2"):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c); cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = box
    for r in rows:
        ws.append(r)
    nrow = len(rows) + 1
    for ri in range(2, nrow + 1):
        band = PatternFill("solid", fgColor=BAND) if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for ci in range(1, len(headers) + 1):
            cell = ws.cell(ri, ci); cell.border = box; cell.fill = band
            cell.font = Font(name=MONO if ci in mono_cols else TEXT, size=10, color=INK)
            if ci in pct_cols:
                cell.number_format = "0.0%"; cell.alignment = Alignment(horizontal="center")
            elif ci in dev_cols:
                cell.number_format = '0.0" d";(0.0)" d";-'; cell.alignment = Alignment(horizontal="center")
            elif isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"; cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{nrow}"
    for ci, w in enumerate((widths or [16] * len(headers)), 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    if heat_col and nrow > 1:
        L = get_column_letter(heat_col)
        ws.conditional_formatting.add(f"{L}2:{L}{nrow}", ColorScaleRule(
            start_type="num", start_value=0, start_color="F4C7C3",
            mid_type="num", mid_value=0.5, mid_color="FFF4C7",
            end_type="num", end_value=0.9, end_color="C6E0B4"))
    return ws


def hotspot_rows(groups, min_vol=1):
    out = []
    for g in groups:
        if g["total"] < min_vol: continue
        out.append([g["label"] if "label" in g else g["key"], g["total"], g["acc"], g["failed"],
                    g["rate"], g["no_t7"], g.get("no_t7_rate") or 0, g["late_gt48"], g["early_gt48"],
                    g["avg_dev_days"] if g["avg_dev_days"] is not None else "", g["recent_rate"] if g["recent_rate"] is not None else "",
                    g["gap_vs_target"] if g["gap_vs_target"] is not None else "", mode(g)])
    return out


HOT_HDR = ["Area", "Delivered (n)", "Accepted", "Failed", "Acc rate", "No-T7", "No-T7 rate",
           "Late >48h", "Early >48h", "Avg dev", "Recent rate (CW18-21)", "Gap vs 90%", "Dominant mode"]
HOT_W = [30, 13, 11, 9, 10, 9, 11, 11, 11, 11, 20, 12, 24]
HOT_MONO = (2, 3, 4, 6, 8, 9); HOT_PCT = (5, 7, 11, 12); HOT_DEV = (10,)

wb = Workbook(); wb.remove(wb.active)

# ---- Overview ----
ov = wb.create_sheet("Overview")
ov.sheet_view.showGridLines = False
ov.column_dimensions["A"].width = 2
for col in "BCDEFGH": ov.column_dimensions[col].width = 17
ov["B2"] = "Bosch ETA Accuracy 2D — Root-Cause & Pushback Evidence"; ov["B2"].font = title_font
ov["B3"] = "Door delivery (S31): T-7 ETA snapshot vs actual, accepted if |deviation| ≤ 48h. SC4, CW01–CW21. Reconciled to live pipeline (kpi_data.json)."
ov["B3"].font = sub_font
R = ROLL["responsibility"]; tot = R["total_delivered"]; b = R["buckets"]
def pct(x): return f"{x/tot*100:.1f}%"
cards = [
    ("Delivered shipments", f"{tot:,}", "denominator (all 21 weeks)"),
    ("Accepted (±48h)", f"{b.get('accepted',0):,}  ({pct(b.get('accepted',0))})", "pass the KPI"),
    ("No T-7 snapshot", f"{b.get('no_t7',0):,}  ({pct(b.get('no_t7',0))})", "never transmitted — not a delivery miss"),
    ("Delivered >48h EARLY", f"{b.get('early_gt48',0):,}  ({pct(b.get('early_gt48',0))})", "arrived ahead of estimate — penalised"),
    ("Delivered >48h late", f"{b.get('late_gt48',0):,}  ({pct(b.get('late_gt48',0))})", "real lateness — attributed in tabs"),
    ("Not a late-delivery failure", f"{b.get('no_t7',0)+b.get('early_gt48',0):,}  ({(b.get('no_t7',0)+b.get('early_gt48',0))/tot*100:.0f}%)", "no-T7 + early combined"),
]
r = 5
for label, val, note in cards:
    ov.cell(r, 2, label).font = Font(name=TEXT, bold=True, color=INK, size=11)
    ov.cell(r, 4, val).font = Font(name=MONO, bold=True, color=ORANGE, size=12)
    ov.cell(r, 6, note).font = sub_font
    ov.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
    r += 1
r += 1
if FIND:
    ov.cell(r, 2, "Executive summary").font = Font(name=TEXT, bold=True, color=INK, size=12); r += 1
    for line in [FIND["executive_summary"]]:
        ov.cell(r, 2, line).font = Font(name=TEXT, color=INK, size=10)
        ov.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ov.row_dimensions[r].height = 70; ov.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top"); r += 2
    rs = FIND["responsibility_split"]
    ov.cell(r, 2, "Gap ownership split").font = Font(name=TEXT, bold=True, color=INK, size=12); r += 1
    for k, v in [("Measurement-design (no-T7 + early)", rs.get("measurement_design_pct")),
                 ("Estimate-staleness / structural", rs.get("estimate_staleness_structural_pct")),
                 ("Genuine Maersk last-mile exposure", rs.get("genuine_maersk_lastmile_pct"))]:
        ov.cell(r, 2, k).font = Font(name=TEXT, color=INK, size=10)
        ov.cell(r, 4, f"{v}%" if v is not None else "—").font = Font(name=MONO, bold=True, color=ORANGE, size=11); r += 1
    r += 1
ov.cell(r, 2, "Data-quality caveats").font = Font(name=TEXT, bold=True, color=INK, size=11); r += 1
caveats = (FIND["data_quality_caveats"] if FIND else [
    "Delay-reason fields are sparsely populated; attribution from them is directional, not exhaustive.",
    "City spellings vary (e.g. 'Ceske' vs 'Cheske Budejovice'); grouped on a best-effort basis.",
    "No literal 'Budapest' in the data — see Dest City tab for the Hungarian network + České Budějovice.",
])
for cc in caveats:
    ov.cell(r, 2, "• " + cc).font = sub_font
    ov.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8); r += 1

# ---- Weekly trend ----
trend_rows = [[t["week"], t["total"], t["acc"], t["rate"], t["no_t7"], t["t7_comp"],
               t["total"] - t["acc"] - t["no_t7"]] for t in ROLL["trend"]]
sheet(wb, "Weekly Trend",
      ["Week", "Delivered", "Accepted", "Acc rate", "No-T7", "T7 completeness", "Miss w/ snapshot"],
      trend_rows, mono_cols=(1, 2, 3, 5, 7), pct_cols=(4, 6), heat_col=4,
      widths=[10, 12, 11, 11, 9, 16, 16])

# ---- Hotspot tabs ----
sheet(wb, "Dest City", HOT_HDR, hotspot_rows(ROLL["by_dest_city"], 5), HOT_MONO, HOT_PCT, HOT_DEV, heat_col=5, widths=HOT_W)
sheet(wb, "Dest Country", HOT_HDR, hotspot_rows(ROLL["by_dest_country"], 1), HOT_MONO, HOT_PCT, HOT_DEV, heat_col=5, widths=HOT_W)
sheet(wb, "Lane", HOT_HDR, hotspot_rows(ROLL["by_lane"], 5), HOT_MONO, HOT_PCT, HOT_DEV, heat_col=5, widths=HOT_W)
sheet(wb, "Origin Country", HOT_HDR, hotspot_rows(ROLL["by_origin_country"], 1), HOT_MONO, HOT_PCT, HOT_DEV, heat_col=5, widths=HOT_W)
sheet(wb, "Carrier", HOT_HDR, hotspot_rows(ROLL["by_carrier"], 1), HOT_MONO, HOT_PCT, HOT_DEV, heat_col=5, widths=HOT_W)
sheet(wb, "Service", HOT_HDR, hotspot_rows(ROLL["by_service"], 1), HOT_MONO, HOT_PCT, HOT_DEV, heat_col=5, widths=HOT_W)

# ---- City x Week heatmap (top cities by volume) ----
weeks = ROLL["meta"]["weeks"]
top_cities = sorted([g for g in ROLL["by_dest_city"] if g["total"] >= 30], key=lambda g: -g["total"])[:25]
hm_hdr = ["Dest city", "Total n"] + weeks
hm_rows = []
for g in top_cities:
    sm = {s["week"]: s["rate"] for s in g["series"]}
    hm_rows.append([g["label"], g["total"]] + [sm.get(w, "") for w in weeks])
hm = sheet(wb, "City x Week (acc rate)", hm_hdr, hm_rows, mono_cols=(2,), pct_cols=tuple(range(3, 3 + len(weeks))),
           widths=[26, 9] + [7] * len(weeks), freeze="C2")
L0, L1 = get_column_letter(3), get_column_letter(2 + len(weeks))
hm.conditional_formatting.add(f"{L0}2:{L1}{len(hm_rows)+1}", ColorScaleRule(
    start_type="num", start_value=0, start_color="F4C7C3", mid_type="num", mid_value=0.5,
    mid_color="FFF4C7", end_type="num", end_value=0.9, end_color="C6E0B4"))

# ---- Country x Month (acc rate, ranked by failure volume) ----
def cw_month(week):
    return date.fromisocalendar(2026, int(week[2:]), 4).strftime("%b")
months = []
for w in ROLL["meta"]["weeks"]:
    mlab = cw_month(w)
    if mlab not in months:
        months.append(mlab)
cmagg = defaultdict(lambda: defaultdict(lambda: [0, 0]))   # country -> month -> [acc, total]
for s in SHIPS:
    c = s["dest_country"]
    if not c:
        continue
    cell = cmagg[c][cw_month(s["week"])]
    cell[1] += 1
    cell[0] += 1 if s["accepted"] else 0
cm_order = sorted([g for g in ROLL["by_dest_country"] if g["total"] >= 20], key=lambda g: -g["failed"])
cm_rows = []
for g in cm_order:
    c = g["key"]
    row = [c, g["total"], g["failed"]]
    for mlab in months:
        a, t = cmagg[c].get(mlab, [0, 0])
        row.append(round(a / t, 4) if t else "")
    cm_rows.append(row)
cmsheet = sheet(wb, "Country x Month", ["Dest country", "Total n", "Failed"] + months, cm_rows,
                mono_cols=(2, 3), pct_cols=tuple(range(4, 4 + len(months))),
                widths=[16, 9, 8] + [8] * len(months), freeze="D2")
cmL0, cmL1 = get_column_letter(4), get_column_letter(3 + len(months))
cmsheet.conditional_formatting.add(f"{cmL0}2:{cmL1}{len(cm_rows)+1}", ColorScaleRule(
    start_type="num", start_value=0, start_color="F4C7C3", mid_type="num", mid_value=0.5,
    mid_color="FFF4C7", end_type="num", end_value=0.9, end_color="C6E0B4"))

# ---- Responsibility ----
resp_rows = [["Accepted (±48h)", b.get("accepted", 0), b.get("accepted", 0) / tot],
             ["No T-7 snapshot (transmission gap)", b.get("no_t7", 0), b.get("no_t7", 0) / tot],
             ["Delivered >48h EARLY (measurement)", b.get("early_gt48", 0), b.get("early_gt48", 0) / tot],
             ["Delivered >48h late", b.get("late_gt48", 0), b.get("late_gt48", 0) / tot],
             ["Measured, no deviation value", b.get("measured_no_dev", 0), b.get("measured_no_dev", 0) / tot]]
sheet(wb, "Responsibility", ["Bucket", "Shipments", "Share"], resp_rows, mono_cols=(2,), pct_cols=(3,), widths=[40, 13, 11])
rsh = wb["Responsibility"]
rr = len(resp_rows) + 3
rsh.cell(rr, 1, "Late-miss port attribution (where port data exists)").font = Font(name=TEXT, bold=True, color=INK)
att = R["late_miss_attribution"]; rr += 1
for k, v in [("Late misses with port data", att["coverage"]),
             ("…arrived >48h late at PORT (upstream)", att["upstream_port"]),
             ("…arrived ~on-time at port, late at door", att["last_mile"]),
             ("Upstream share", f"{(att['upstream_share'] or 0)*100:.1f}%")]:
    rsh.cell(rr, 1, k).font = Font(name=TEXT, size=10, color=INK)
    rsh.cell(rr, 2, v).font = Font(name=MONO, size=10, color=INK); rr += 1
rr += 1
rsh.cell(rr, 1, "Populated delay reasons (late misses)").font = Font(name=TEXT, bold=True, color=INK); rr += 1
for reason, n in list(R["delay_reasons"].items()):
    rsh.cell(rr, 1, reason).font = Font(name=TEXT, size=10, color=INK)
    rsh.cell(rr, 2, n).font = Font(name=MONO, size=10, color=INK); rr += 1
rsh.cell(rr, 1, f"Supplier-responsible = Y").font = Font(name=TEXT, size=10, color=INK)
rsh.cell(rr, 2, R["supplier_responsible_Y"]).font = Font(name=MONO, size=10, color=INK)

# ---- Shipment detail (all delivered) ----
det_hdr = ["Week", "HBL", "Service", "Carrier", "Origin city", "Orig ctry", "Dest city", "Dest ctry",
           "Lane", "T7?", "Dev days", "Dir", "Bucket", "Accepted", "Delay reason", "Supplier resp",
           "Port late (h)", "Last-mile (d)", "ETA", "ATA", "Delivered"]
det_rows = [[s["week"], s["hbl"], s["service"], s["carrier"], s["origin_city"], s["origin_country"],
             s["dest_city"], s["dest_country"], s["lane"], "Y" if s["t7_present"] else "N",
             s["dev_days"] if s["dev_days"] is not None else "", s["direction"] or "", s["bucket"],
             "Y" if s["accepted"] else "N", s["delay_reason"], s["supplier_responsible"],
             s["port_late_h"] if s["port_late_h"] is not None else "", s["last_mile_d"] if s["last_mile_d"] is not None else "",
             s["eta"] or "", s["ata"] or "", s["delivered"] or ""] for s in SHIPS]
sheet(wb, "Shipment detail", det_hdr, det_rows, mono_cols=(1, 2, 4, 6, 8, 9, 11, 17, 18, 19, 20, 21),
      dev_cols=(11,), widths=[8, 16, 9, 10, 16, 8, 16, 8, 10, 5, 9, 7, 14, 9, 22, 10, 11, 11, 16, 16, 16])

out = "Bosch_ETA2D_Pushback_Evidence.xlsx"
wb.save(out)
print(f"Wrote {out} with {len(wb.sheetnames)} tabs: {wb.sheetnames}")
print(f"Shipment detail rows: {len(det_rows)}  |  findings enrichment: {'yes' if FIND else 'no (run after workflow)'}")
