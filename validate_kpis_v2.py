"""
Validate KPI calculations against PDCA template values for CW01-CW07.
v2: Fixed sheet name handling and trying multiple calculation methods.
"""

import openpyxl
import os

BASE = "/Users/harsh.puri/Documents/work-maersk/Prototype Playground/Bosch Milestone Analysis"
RAW_DIR = os.path.join(BASE, "Bosch Milestone raw data")

SC3_FILES = {
    "CW01": "Maersk NGTM SC3_2026_CW01.xlsx",
    "CW02": "Maersk NGTM SC3_2026_CW02.xlsx",
    "CW03": "Maersk NGTM SC3_2026_CW03.xlsx",
    "CW04": "Maersk NGTM SC3_2026_CW04.xlsx",
    "CW05": "Maersk NGTM SC3_2026_CW05.xlsx",
    "CW06": "Maersk NGTM SC3_2026_CW06.xlsx",
    "CW07": "Maersk NGTM SC3_2026_CW07.xlsx",
}

SC4_FILES = {
    "CW01": "Maersk SC4_2026_CW01.xlsx",
    "CW02": "Maersk SC4_2026_CW02v2.xlsx",
    "CW03": "Maersk SC4_2026_CW03.xlsx",
    "CW04": "Maersk SC4_2026_CW04.xlsx",
    "CW05": "Maersk SC4_2026_CW05.xlsx",
    "CW06": "Maersk SC4_2026_CW06.xlsx",
    "CW07": "Maersk SC4_2026_CW07.xlsx",
}

SC3_CRITICAL = ["S02", "S04", "S07", "S31"]
SC4_CRITICAL = ["S00", "S02", "S04", "S07", "S31"]


def find_total_sheet(wb):
    """Find the TOTAL summary sheet regardless of naming convention."""
    for sn in wb.sheetnames:
        s = sn.strip().upper()
        if s.startswith("TOTAL") or s == "ALL":
            return sn
    return None


def find_shipments_sheet(wb):
    """Find the shipments sheet regardless of naming convention."""
    for sn in wb.sheetnames:
        if sn.strip().lower() == "shipments":
            return sn
    return None


def parse_milestone_code(name):
    if name and " - " in str(name):
        return str(name).split(" - ")[0].strip()
    return str(name).strip() if name else None


def read_summary_sheet(filepath):
    """Read TOTAL/ALL summary sheet and return milestone data."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = find_total_sheet(wb)
    if not sheet_name:
        print(f"  WARNING: No TOTAL sheet found in {filepath}")
        wb.close()
        return []

    ws = wb[sheet_name]

    # Find header row (contains "Required status" or "Status codes required")
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        for cell in row:
            if cell and "required" in str(cell).lower() and "status" in str(cell).lower():
                header_row = i
                break
        if header_row:
            break

    if not header_row:
        # Default: assume header at row 3
        header_row = 3

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        milestone_name = row[0]
        if not milestone_name or "required" in str(milestone_name).lower():
            continue
        code = parse_milestone_code(milestone_name)
        status_type = str(row[1]).strip() if row[1] else ""
        required = row[3] if row[3] and isinstance(row[3], (int, float)) else 0
        available = row[4] if row[4] and isinstance(row[4], (int, float)) else 0
        in_time = row[5] if row[5] and isinstance(row[5], (int, float)) else 0
        completeness = row[6] if row[6] and isinstance(row[6], (int, float)) else 0
        timeliness = row[7] if row[7] and isinstance(row[7], (int, float)) else 0

        rows.append({
            "name": str(milestone_name),
            "code": code,
            "type": status_type,
            "required": required,
            "available": available,
            "in_time": in_time,
            "completeness": completeness,
            "timeliness": timeliness,
        })
    wb.close()
    return rows


def read_sc4_shipments(filepath):
    """Read SC4 shipment-level data."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = find_shipments_sheet(wb)
    if not sheet_name:
        wb.close()
        return []

    ws = wb[sheet_name]

    # Find header row
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        for cell in row:
            if cell and "UNIQUE_SHIPMENT" in str(cell):
                header_row = i
                break
        if header_row:
            break

    if not header_row:
        header_row = 3

    headers = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            col_map[str(h).strip()] = i

    shipments = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        vals = list(row)
        if not vals[0]:
            continue
        shipment = {}
        for name, idx in col_map.items():
            shipment[name] = vals[idx] if idx < len(vals) else None
        shipments.append(shipment)

    wb.close()
    return shipments


def compute_eta_2p(shipments):
    """ETA accuracy 2P: S07_Accepted / total with S07_Accepted not None."""
    col = None
    for key in (shipments[0].keys() if shipments else []):
        if "S07_Accepted" in key:
            col = key
            break
    if not col:
        return None, 0, 0
    total = sum(1 for s in shipments if s.get(col) is not None)
    accepted = sum(1 for s in shipments if s.get(col) == 1)
    return (accepted / total if total > 0 else 0), accepted, total


def compute_eta_2d(shipments):
    """ETA accuracy 2D: S31_Accepted / total with S31_Accepted not None."""
    col = None
    for key in (shipments[0].keys() if shipments else []):
        if "S31_Accepted" in key:
            col = key
            break
    if not col:
        return None, 0, 0
    total = sum(1 for s in shipments if s.get(col) is not None)
    accepted = sum(1 for s in shipments if s.get(col) == 1)
    return (accepted / total if total > 0 else 0), accepted, total


def compute_ref_completeness(shipments):
    """SC4 Reference Completeness: has CIV or DN."""
    civ_col = dn_col = None
    for key in (shipments[0].keys() if shipments else []):
        if "COMMERCIAL_INVOICE" in key:
            civ_col = key
        if "DELIVERY_NOTE" in key:
            dn_col = key
    total = len(shipments)
    complete = 0
    for s in shipments:
        civ = s.get(civ_col)
        dn = s.get(dn_col)
        has_civ = civ is not None and str(civ).strip() not in ("", "None")
        has_dn = dn is not None and str(dn).strip() not in ("", "None")
        if has_civ or has_dn:
            complete += 1
    return (complete / total if total > 0 else 0), complete, total


# PDCA reference values
PDCA = {
    "Comp_Crit": {"W5": 0.8277, "W6": 0.8168, "W7": 0.8315},
    "Time_Crit": {"W5": 0.5885, "W6": 0.6012, "W7": 0.6415},
    "Comp_All":  {"W5": 0.7481, "W6": 0.7585, "W7": 0.7838},
    "Time_All":  {"W5": 0.5996, "W6": 0.6140, "W7": 0.6257},
    "ETA_2P":    {"W5": 0.44,   "W6": 0.62,   "W7": 0.39},
    "ETA_2D":    {"W5": 0.13,   "W6": 0.17,   "W7": 0.17},
    "Ref_Comp":  {"W5": 0.71,   "W6": 0.74,   "W7": 0.76},
}


def match_str(computed, pdca, tol=0.005):
    if pdca == "?":
        return "N/A"
    diff = abs(computed - pdca)
    return f"MATCH (d={diff:.4f})" if diff < tol else f"MISS  (d={diff:.4f})"


def main():
    print("=" * 110)
    print("BOSCH MILESTONE KPI VALIDATION v2")
    print("=" * 110)

    for week in ["CW05", "CW06", "CW07"]:
        wkey = week.replace("CW0", "W")
        print(f"\n{'='*110}")
        print(f"  WEEK: {week} ({wkey})")
        print(f"{'='*110}")

        sc3_file = os.path.join(RAW_DIR, SC3_FILES[week])
        sc4_file = os.path.join(RAW_DIR, SC4_FILES[week])

        sc3_data = read_summary_sheet(sc3_file)
        sc4_data = read_summary_sheet(sc4_file)

        # ----- Print raw data -----
        print(f"\n  SC3 milestones ({len(sc3_data)} rows):")
        for r in sc3_data:
            print(f"    {r['code']:4s} {r['type']:10s} | req={r['required']:4} avl={r['available']:4} it={r['in_time']:4} | C={r['completeness']:.3f} T={r['timeliness']:.3f}")

        print(f"\n  SC4 milestones ({len(sc4_data)} rows):")
        for r in sc4_data:
            print(f"    {r['code']:4s} {r['type']:10s} | req={r['required']:4} avl={r['available']:4} it={r['in_time']:4} | C={r['completeness']:.3f} T={r['timeliness']:.3f}")

        # ----- Method 1: Weighted sum (sum_avl / sum_req) -----
        # ----- Method 2: Simple average of per-milestone completeness ratios -----
        # ----- Method 3: Average of per-milestone ratios, one entry per milestone code (avg Actual+Estimated) -----

        # Define milestone+type combos to try
        # The PDCA problems list references specific milestone+type combos:
        # SC4: S00-Actual, S02-Estimated, S02-Actual, S04-Estimated, S04-Actual, S07-Estimated, S07-Actual, S31-Estimated, S31-Actual
        # SC3: S02-Actual, S02-Estimated, S04-Actual, S04-Estimated, S07-Actual, S07-Estimated, S31-Actual, S31-Estimated

        # So each milestone has both Actual and Estimated rows.
        # "Critical" means specific milestones. The question is how they aggregate.

        print("\n  ---- CRITICAL MILESTONES ----")

        # Try: average of individual milestone-type completeness values
        sc3_crit = [r for r in sc3_data if r["code"] in SC3_CRITICAL]
        sc4_crit = [r for r in sc4_data if r["code"] in SC4_CRITICAL]
        all_crit = sc3_crit + sc4_crit

        # Method A: Weighted sum, Actual+Estimated
        total_req = sum(r["required"] for r in all_crit)
        total_avl = sum(r["available"] for r in all_crit)
        total_it = sum(r["in_time"] for r in all_crit)
        ws_comp = total_avl / total_req if total_req > 0 else 0
        ws_time = total_it / total_req if total_req > 0 else 0

        # Method B: Simple average of completeness ratios (all rows)
        if all_crit:
            avg_comp = sum(r["completeness"] for r in all_crit) / len(all_crit)
            avg_time = sum(r["timeliness"] for r in all_crit) / len(all_crit)
        else:
            avg_comp = avg_time = 0

        # Method C: Weighted sum, Actual only
        actual_crit = [r for r in all_crit if r["type"] == "Actual"]
        if actual_crit:
            ws_a_comp = sum(r["available"] for r in actual_crit) / sum(r["required"] for r in actual_crit)
            ws_a_time = sum(r["in_time"] for r in actual_crit) / sum(r["required"] for r in actual_crit)
        else:
            ws_a_comp = ws_a_time = 0

        # Method D: Average of per-milestone-type completeness, Actual only
        if actual_crit:
            avg_a_comp = sum(r["completeness"] for r in actual_crit) / len(actual_crit)
            avg_a_time = sum(r["timeliness"] for r in actual_crit) / len(actual_crit)
        else:
            avg_a_comp = avg_a_time = 0

        # Method E: Average of per-milestone ratios, one per code, using highest type (Actual preferred)
        # For each milestone code, use Actual if available, else Estimated
        by_code = {}
        for r in all_crit:
            key = r["code"]
            if key not in by_code or r["type"] == "Actual":
                by_code[key] = r
        method_e_rows = list(by_code.values())
        if method_e_rows:
            avg_e_comp = sum(r["completeness"] for r in method_e_rows) / len(method_e_rows)
            avg_e_time = sum(r["timeliness"] for r in method_e_rows) / len(method_e_rows)
        else:
            avg_e_comp = avg_e_time = 0

        pdca_comp = PDCA["Comp_Crit"].get(wkey, "?")
        pdca_time = PDCA["Time_Crit"].get(wkey, "?")

        print(f"  PDCA Completeness: {pdca_comp}")
        print(f"  PDCA Timeliness:   {pdca_time}")
        print(f"  Method A (weighted, A+E):  C={ws_comp:.4f} [{match_str(ws_comp, pdca_comp)}]  T={ws_time:.4f} [{match_str(ws_time, pdca_time)}]")
        print(f"  Method B (avg ratio, A+E): C={avg_comp:.4f} [{match_str(avg_comp, pdca_comp)}]  T={avg_time:.4f} [{match_str(avg_time, pdca_time)}]")
        print(f"  Method C (weighted, Act):  C={ws_a_comp:.4f} [{match_str(ws_a_comp, pdca_comp)}]  T={ws_a_time:.4f} [{match_str(ws_a_time, pdca_time)}]")
        print(f"  Method D (avg ratio, Act): C={avg_a_comp:.4f} [{match_str(avg_a_comp, pdca_comp)}]  T={avg_a_time:.4f} [{match_str(avg_a_time, pdca_time)}]")
        print(f"  Method E (avg, Actual per code): C={avg_e_comp:.4f} [{match_str(avg_e_comp, pdca_comp)}]  T={avg_e_time:.4f} [{match_str(avg_e_time, pdca_time)}]")

        # Method F: For each milestone, use specific type: S02=Est, S04=Act, S07=Act, S31=Est, S00=Act
        # This is based on what Bosch cares about per milestone
        type_map = {
            "S00": "Actual",
            "S02": "Estimated",
            "S04": "Actual",
            "S07": "Actual",
            "S31": "Estimated",
        }
        method_f_rows = []
        for r in all_crit:
            expected_type = type_map.get(r["code"])
            if expected_type and r["type"] == expected_type:
                method_f_rows.append(r)

        if method_f_rows:
            ws_f_comp = sum(r["available"] for r in method_f_rows) / sum(r["required"] for r in method_f_rows)
            ws_f_time = sum(r["in_time"] for r in method_f_rows) / sum(r["required"] for r in method_f_rows)
            avg_f_comp = sum(r["completeness"] for r in method_f_rows) / len(method_f_rows)
            avg_f_time = sum(r["timeliness"] for r in method_f_rows) / len(method_f_rows)
        else:
            ws_f_comp = ws_f_time = avg_f_comp = avg_f_time = 0

        print(f"  Method F (weighted, specific types): C={ws_f_comp:.4f} [{match_str(ws_f_comp, pdca_comp)}]  T={ws_f_time:.4f} [{match_str(ws_f_time, pdca_time)}]")
        print(f"  Method F (avg, specific types):      C={avg_f_comp:.4f} [{match_str(avg_f_comp, pdca_comp)}]  T={avg_f_time:.4f} [{match_str(avg_f_time, pdca_time)}]")

        # Method G: average completeness across all milestone-type combos for critical milestones
        # But weight SC3 and SC4 separately, then average
        if sc3_crit and sc4_crit:
            sc3_comp = sum(r["completeness"] for r in sc3_crit) / len(sc3_crit)
            sc4_comp = sum(r["completeness"] for r in sc4_crit) / len(sc4_crit)
            sc3_time = sum(r["timeliness"] for r in sc3_crit) / len(sc3_crit)
            sc4_time = sum(r["timeliness"] for r in sc4_crit) / len(sc4_crit)
            avg_g_comp = (sc3_comp + sc4_comp) / 2
            avg_g_time = (sc3_time + sc4_time) / 2
            print(f"  Method G (avg of SC3+SC4 avgs):  C={avg_g_comp:.4f} [{match_str(avg_g_comp, pdca_comp)}]  T={avg_g_time:.4f} [{match_str(avg_g_time, pdca_time)}]")

        # ----- ALL MILESTONES -----
        print("\n  ---- ALL MILESTONES ----")
        all_rows = sc3_data + sc4_data

        pdca_comp_all = PDCA["Comp_All"].get(wkey, "?")
        pdca_time_all = PDCA["Time_All"].get(wkey, "?")

        # Method A: weighted, A+E
        t_req = sum(r["required"] for r in all_rows)
        t_avl = sum(r["available"] for r in all_rows)
        t_it = sum(r["in_time"] for r in all_rows)
        a_comp = t_avl / t_req if t_req > 0 else 0
        a_time = t_it / t_req if t_req > 0 else 0

        # Method B: avg of ratios, A+E
        b_comp = sum(r["completeness"] for r in all_rows) / len(all_rows) if all_rows else 0
        b_time = sum(r["timeliness"] for r in all_rows) / len(all_rows) if all_rows else 0

        # Method C: weighted, Actual only
        actual_all = [r for r in all_rows if r["type"] == "Actual"]
        c_comp = sum(r["available"] for r in actual_all) / sum(r["required"] for r in actual_all) if actual_all else 0
        c_time = sum(r["in_time"] for r in actual_all) / sum(r["required"] for r in actual_all) if actual_all else 0

        # Method D: avg ratio, Actual only
        d_comp = sum(r["completeness"] for r in actual_all) / len(actual_all) if actual_all else 0
        d_time = sum(r["timeliness"] for r in actual_all) / len(actual_all) if actual_all else 0

        print(f"  PDCA Completeness: {pdca_comp_all}")
        print(f"  PDCA Timeliness:   {pdca_time_all}")
        print(f"  Method A (weighted, A+E):  C={a_comp:.4f} [{match_str(a_comp, pdca_comp_all)}]  T={a_time:.4f} [{match_str(a_time, pdca_time_all)}]")
        print(f"  Method B (avg ratio, A+E): C={b_comp:.4f} [{match_str(b_comp, pdca_comp_all)}]  T={b_time:.4f} [{match_str(b_time, pdca_time_all)}]")
        print(f"  Method C (weighted, Act):  C={c_comp:.4f} [{match_str(c_comp, pdca_comp_all)}]  T={c_time:.4f} [{match_str(c_time, pdca_time_all)}]")
        print(f"  Method D (avg ratio, Act): C={d_comp:.4f} [{match_str(d_comp, pdca_comp_all)}]  T={d_time:.4f} [{match_str(d_time, pdca_time_all)}]")

        # ----- ETA ACCURACY -----
        print("\n  ---- ETA ACCURACY & REFERENCE (SC4 only) ----")
        shipments = read_sc4_shipments(sc4_file)

        eta_2p, s07_a, s07_t = compute_eta_2p(shipments)
        eta_2d, s31_a, s31_t = compute_eta_2d(shipments)
        ref, ref_c, ref_t = compute_ref_completeness(shipments)

        p2p = PDCA["ETA_2P"].get(wkey, "?")
        p2d = PDCA["ETA_2D"].get(wkey, "?")
        pref = PDCA["Ref_Comp"].get(wkey, "?")

        print(f"  ETA 2P: {s07_a:3d}/{s07_t:3d} = {eta_2p:.4f}  PDCA={p2p}  [{match_str(eta_2p, p2p, 0.02)}]")
        print(f"  ETA 2D: {s31_a:3d}/{s31_t:3d} = {eta_2d:.4f}  PDCA={p2d}  [{match_str(eta_2d, p2d, 0.02)}]")
        print(f"  RefComp: {ref_c:3d}/{ref_t:3d} = {ref:.4f}  PDCA={pref}  [{match_str(ref, pref, 0.02)}]")


if __name__ == "__main__":
    main()
