"""
Validate KPI calculations against PDCA template values for CW01-CW07.
This script reads raw SC3/SC4 files and computes all 7 KPIs, then compares
against known PDCA values.
"""

import openpyxl
import os
from collections import defaultdict

BASE = "/Users/harsh.puri/Documents/work-maersk/Prototype Playground/Bosch Milestone Analysis"
RAW_DIR = os.path.join(BASE, "Bosch Milestone raw data")

# File mapping
SC3_FILES = {
    "CW01": "Maersk NGTM SC3_2026_CW01.xlsx",
    "CW02": "Maersk NGTM SC3_2026_CW02.xlsx",
    "CW03": "Maersk NGTM SC3_2026_CW03.xlsx",
    "CW04": "Maersk NGTM SC3_2026_CW04.xlsx",
    "CW05": "Maersk NGTM SC3_2026_CW05.xlsx",
    "CW06": "Maersk NGTM SC3_2026_CW06.xlsx",
    "CW07": "Maersk NGTM SC3_2026_CW07.xlsx",
}

SC4_FILES = {
    "CW01": "Maersk SC4_2026_CW01.xlsx",
    "CW02": "Maersk SC4_2026_CW02v2.xlsx",
    "CW03": "Maersk SC4_2026_CW03.xlsx",
    "CW04": "Maersk SC4_2026_CW04.xlsx",
    "CW05": "Maersk SC4_2026_CW05.xlsx",
    "CW06": "Maersk SC4_2026_CW06.xlsx",
    "CW07": "Maersk SC4_2026_CW07.xlsx",
}

# Critical milestones
SC3_CRITICAL = ["S02", "S04", "S07", "S31"]
SC4_CRITICAL = ["S00", "S02", "S04", "S07", "S31"]


def parse_milestone_code(name):
    """Extract milestone code like 'S02' from 'S02 - Collected'."""
    if name and " - " in str(name):
        return str(name).split(" - ")[0].strip()
    return str(name).strip() if name else None


def read_summary_sheet(filepath, sheet_name):
    """Read a summary sheet (TOTAL, BCO, FCL, LCL) and return milestone data."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    # SC3 has trailing space in "TOTAL " sheet name
    actual_sheet = None
    for sn in wb.sheetnames:
        if sn.strip() == sheet_name.strip():
            actual_sheet = sn
            break
    if not actual_sheet:
        return []

    ws = wb[actual_sheet]
    rows = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        milestone_name = row[0]
        if not milestone_name:
            continue
        code = parse_milestone_code(milestone_name)
        status_type = str(row[1]).strip() if row[1] else ""
        required = row[3] if row[3] else 0
        available = row[4] if row[4] else 0
        in_time = row[5] if row[5] else 0
        completeness = row[6] if row[6] else 0
        timeliness = row[7] if row[7] else 0

        rows.append({
            "name": str(milestone_name),
            "code": code,
            "type": status_type,
            "required": required,
            "available": available,
            "in_time": in_time,
            "completeness": completeness,
            "timeliness": timeliness,
        })
    wb.close()
    return rows


def compute_completeness_timeliness(sc3_data, sc4_data, milestone_codes, include_types=None):
    """
    Compute weighted average completeness and timeliness for given milestones.
    Uses sum(available)/sum(required) for completeness,
    sum(in_time)/sum(required) for timeliness.
    """
    total_required = 0
    total_available = 0
    total_in_time = 0

    for row in sc3_data + sc4_data:
        if row["code"] not in milestone_codes:
            continue
        if include_types and row["type"] not in include_types:
            continue
        total_required += row["required"]
        total_available += row["available"]
        total_in_time += row["in_time"]

    completeness = total_available / total_required if total_required > 0 else 0
    timeliness = total_in_time / total_required if total_required > 0 else 0
    return completeness, timeliness, total_required, total_available, total_in_time


def read_sc4_shipments(filepath):
    """Read SC4 shipment-level data for ETA accuracy and reference completeness."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["shipments"]

    # Get headers from row 3
    headers = list(next(ws.iter_rows(min_row=3, max_row=3, values_only=True)))

    # Build column index
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            col_map[str(h).strip()] = i

    shipments = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        vals = list(row)
        if not vals[0]:
            continue

        shipment = {}
        for name, idx in col_map.items():
            shipment[name] = vals[idx] if idx < len(vals) else None
        shipments.append(shipment)

    wb.close()
    return shipments


def compute_eta_accuracy_2p(shipments):
    """
    ETA accuracy 2P (to Port): S07 accuracy.
    S07_Accepted = 1 means ETA was within ±48h of actual vessel arrival.
    """
    # Find the right column name
    s07_acc_col = None
    for key in shipments[0].keys() if shipments else []:
        if "S07_Accepted" in key:
            s07_acc_col = key
            break

    if not s07_acc_col:
        return None, 0, 0

    total = 0
    accepted = 0
    for s in shipments:
        val = s.get(s07_acc_col)
        if val is not None:
            total += 1
            if val == 1:
                accepted += 1

    return accepted / total if total > 0 else 0, accepted, total


def compute_eta_accuracy_2d(shipments):
    """
    ETA accuracy 2D (to Door): S31 accuracy.
    S31_Accepted = 1 means delivery estimate was within ±48h of actual delivery.
    """
    s31_acc_col = None
    for key in shipments[0].keys() if shipments else []:
        if "S31_Accepted" in key:
            s31_acc_col = key
            break

    if not s31_acc_col:
        return None, 0, 0

    total = 0
    accepted = 0
    for s in shipments:
        val = s.get(s31_acc_col)
        if val is not None:
            total += 1
            if val == 1:
                accepted += 1

    return accepted / total if total > 0 else 0, accepted, total


def compute_reference_completeness(shipments):
    """
    SC4 Reference Completeness: % of shipments with CIV or DN number present.
    """
    total = len(shipments)
    complete = 0

    # Find column names
    civ_col = None
    dn_col = None
    for key in shipments[0].keys() if shipments else []:
        if "COMMERCIAL_INVOICE" in key:
            civ_col = key
        if "DELIVERY_NOTE" in key:
            dn_col = key

    for s in shipments:
        civ = s.get(civ_col)
        dn = s.get(dn_col)
        has_civ = civ is not None and str(civ).strip() not in ("", "None", "NONE")
        has_dn = dn is not None and str(dn).strip() not in ("", "None", "NONE")
        if has_civ or has_dn:
            complete += 1

    return complete / total if total > 0 else 0, complete, total


# PDCA reference values (from the template)
PDCA_VALUES = {
    # Col positions: Nov'25=6, Dec'25=7, Jan'26=8, W5=9, W6=10, W7=11
    "Completeness (Critical)": {
        "Nov'25": 0.82, "Dec'25": 0.84, "Jan'26": 0.84,
        "W5": 0.8277, "W6": 0.8168, "W7": 0.8315,
    },
    "Timeliness (Critical)": {
        "Nov'25": 0.10, "Dec'25": 0.22, "Jan'26": 0.48,
        "W5": 0.5885, "W6": 0.6012, "W7": 0.6415,
    },
    "Completeness (All)": {
        "Jan'26": 0.71, "W5": 0.7481, "W6": 0.7585, "W7": 0.7838,
    },
    "Timeliness (All)": {
        "Jan'26": 0.43, "W5": 0.5996, "W6": 0.6140, "W7": 0.6257,
    },
    "ETA accuracy 2P": {
        "Jan'26": 0.38, "W5": 0.44, "W6": 0.62, "W7": 0.39,
    },
    "ETA accuracy 2D": {
        "Jan'26": 0.11, "W5": 0.13, "W6": 0.17, "W7": 0.17,
    },
    "SC4 Reference Completeness": {
        "Jan'26": 0.69, "W5": 0.71, "W6": 0.74, "W7": 0.76,
    },
}


def main():
    print("=" * 100)
    print("BOSCH MILESTONE KPI VALIDATION - CW01 to CW07")
    print("=" * 100)

    # =========================================================================
    # KPI 1 & 2: Completeness (Critical) and Timeliness (Critical)
    # =========================================================================
    print("\n" + "=" * 100)
    print("KPI 1: COMPLETENESS (CRITICAL)  &  KPI 2: TIMELINESS (CRITICAL)")
    print("Critical milestones: SC3={}, SC4={}".format(SC3_CRITICAL, SC4_CRITICAL))
    print("=" * 100)

    # Try different type combinations to find what matches PDCA
    type_combos = [
        ("Actual only", ["Actual"]),
        ("Estimated only", ["Estimated"]),
        ("Actual + Estimated", ["Actual", "Estimated"]),
    ]

    for week in ["CW05", "CW06", "CW07"]:
        print(f"\n--- {week} ---")

        sc3_file = os.path.join(RAW_DIR, SC3_FILES[week])
        sc4_file = os.path.join(RAW_DIR, SC4_FILES[week])

        sc3_total = read_summary_sheet(sc3_file, "TOTAL")
        sc4_total = read_summary_sheet(sc4_file, "TOTAL")

        # Print raw milestone data for critical milestones
        print(f"  SC3 Critical milestones:")
        for row in sc3_total:
            if row["code"] in SC3_CRITICAL:
                print(f"    {row['code']} {row['type']:10s} | req={row['required']:4d} avl={row['available']:4d} intime={row['in_time']:4d} | comp={row['completeness']:.3f} time={row['timeliness']:.3f}")

        print(f"  SC4 Critical milestones:")
        for row in sc4_total:
            if row["code"] in SC4_CRITICAL:
                print(f"    {row['code']} {row['type']:10s} | req={row['required']:4d} avl={row['available']:4d} intime={row['in_time']:4d} | comp={row['completeness']:.3f} time={row['timeliness']:.3f}")

        # Try different combinations
        for label, types in type_combos:
            # Combined SC3+SC4 with respective critical milestones
            sc3_filtered = [r for r in sc3_total if r["code"] in SC3_CRITICAL and r["type"] in types]
            sc4_filtered = [r for r in sc4_total if r["code"] in SC4_CRITICAL and r["type"] in types]

            total_req = sum(r["required"] for r in sc3_filtered + sc4_filtered)
            total_avl = sum(r["available"] for r in sc3_filtered + sc4_filtered)
            total_it = sum(r["in_time"] for r in sc3_filtered + sc4_filtered)

            comp = total_avl / total_req if total_req > 0 else 0
            timel = total_it / total_req if total_req > 0 else 0

            pdca_key = week.replace("CW0", "W")
            pdca_comp = PDCA_VALUES["Completeness (Critical)"].get(pdca_key, "?")
            pdca_time = PDCA_VALUES["Timeliness (Critical)"].get(pdca_key, "?")

            comp_match = "MATCH" if isinstance(pdca_comp, float) and abs(comp - pdca_comp) < 0.005 else "MISS"
            time_match = "MATCH" if isinstance(pdca_time, float) and abs(timel - pdca_time) < 0.005 else "MISS"

            print(f"  [{label}] Comp={comp:.4f} (PDCA={pdca_comp}) [{comp_match}] | Time={timel:.4f} (PDCA={pdca_time}) [{time_match}]")

    # =========================================================================
    # KPI 3 & 4: Completeness (All) and Timeliness (All)
    # =========================================================================
    print("\n" + "=" * 100)
    print("KPI 3: COMPLETENESS (ALL)  &  KPI 4: TIMELINESS (ALL)")
    print("=" * 100)

    for week in ["CW05", "CW06", "CW07"]:
        print(f"\n--- {week} ---")

        sc3_file = os.path.join(RAW_DIR, SC3_FILES[week])
        sc4_file = os.path.join(RAW_DIR, SC4_FILES[week])

        sc3_total = read_summary_sheet(sc3_file, "TOTAL")
        sc4_total = read_summary_sheet(sc4_file, "TOTAL")

        for label, types in type_combos:
            all_rows = [r for r in sc3_total + sc4_total if r["type"] in types]
            total_req = sum(r["required"] for r in all_rows)
            total_avl = sum(r["available"] for r in all_rows)
            total_it = sum(r["in_time"] for r in all_rows)

            comp = total_avl / total_req if total_req > 0 else 0
            timel = total_it / total_req if total_req > 0 else 0

            pdca_key = week.replace("CW0", "W")
            pdca_comp = PDCA_VALUES["Completeness (All)"].get(pdca_key, "?")
            pdca_time = PDCA_VALUES["Timeliness (All)"].get(pdca_key, "?")

            comp_match = "MATCH" if isinstance(pdca_comp, float) and abs(comp - pdca_comp) < 0.005 else "MISS"
            time_match = "MATCH" if isinstance(pdca_time, float) and abs(timel - pdca_time) < 0.005 else "MISS"

            print(f"  [{label}] Comp={comp:.4f} (PDCA={pdca_comp}) [{comp_match}] | Time={timel:.4f} (PDCA={pdca_time}) [{time_match}]")

    # =========================================================================
    # KPI 5 & 6: ETA Accuracy 2P and 2D
    # =========================================================================
    print("\n" + "=" * 100)
    print("KPI 5: ETA ACCURACY 2P (SC4)  &  KPI 6: ETA ACCURACY 2D (SC4)")
    print("=" * 100)

    for week in ["CW05", "CW06", "CW07"]:
        print(f"\n--- {week} ---")
        sc4_file = os.path.join(RAW_DIR, SC4_FILES[week])
        shipments = read_sc4_shipments(sc4_file)

        eta_2p, s07_acc, s07_tot = compute_eta_accuracy_2p(shipments)
        eta_2d, s31_acc, s31_tot = compute_eta_accuracy_2d(shipments)

        pdca_key = week.replace("CW0", "W")
        pdca_2p = PDCA_VALUES["ETA accuracy 2P"].get(pdca_key, "?")
        pdca_2d = PDCA_VALUES["ETA accuracy 2D"].get(pdca_key, "?")

        p2_match = "MATCH" if isinstance(pdca_2p, float) and abs(eta_2p - pdca_2p) < 0.02 else "MISS"
        d2_match = "MATCH" if isinstance(pdca_2d, float) and abs(eta_2d - pdca_2d) < 0.02 else "MISS"

        print(f"  ETA 2P: {s07_acc}/{s07_tot} = {eta_2p:.4f} (PDCA={pdca_2p}) [{p2_match}]")
        print(f"  ETA 2D: {s31_acc}/{s31_tot} = {eta_2d:.4f} (PDCA={pdca_2d}) [{d2_match}]")

    # =========================================================================
    # KPI 7: SC4 Reference Completeness
    # =========================================================================
    print("\n" + "=" * 100)
    print("KPI 7: SC4 REFERENCE COMPLETENESS (CIV or DN)")
    print("=" * 100)

    for week in ["CW05", "CW06", "CW07"]:
        print(f"\n--- {week} ---")
        sc4_file = os.path.join(RAW_DIR, SC4_FILES[week])
        shipments = read_sc4_shipments(sc4_file)

        ref_comp, ref_complete, ref_total = compute_reference_completeness(shipments)

        pdca_key = week.replace("CW0", "W")
        pdca_ref = PDCA_VALUES["SC4 Reference Completeness"].get(pdca_key, "?")
        ref_match = "MATCH" if isinstance(pdca_ref, float) and abs(ref_comp - pdca_ref) < 0.02 else "MISS"

        print(f"  Ref Completeness: {ref_complete}/{ref_total} = {ref_comp:.4f} (PDCA={pdca_ref}) [{ref_match}]")


if __name__ == "__main__":
    main()
