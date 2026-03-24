"""
Extract detailed milestone-level RCA (Root Cause Analysis) data.

For each week, produces:
  - Per-milestone summary: required, available, in_time, missing count, gaps
  - Shipment-level detail: which HBLs are missing each milestone or late
  - SC4: HBL directly from detail sheets
  - SC3: LOAD_TO joined with shipments sheet to get HBL

Output: rca_data.json consumed by the dashboard.
"""

import openpyxl
import os
import json
from datetime import timedelta

BASE = "/Users/harsh.puri/Documents/work-maersk/Prototype Playground/Bosch Milestone Analysis"
RAW_DIR = os.path.join(BASE, "Bosch Milestone raw data")

WEEKS = ["CW01", "CW02", "CW03", "CW04", "CW05", "CW06", "CW07", "CW08", "CW09", "CW10"]
SC3_FILES = {f"CW{i:02d}": f"Maersk NGTM SC3_2026_CW{i:02d}.xlsx" for i in range(1, 10)}
SC3_FILES["CW10"] = "Maersk SC3_2026_CW10.xlsx"  # CW10 has different naming
SC4_FILES = {f"CW{i:02d}": f"Maersk SC4_2026_CW{i:02d}.xlsx" for i in range(1, 11)}

SC3_CRITICAL_CODES = {"S02", "S04", "S07", "S31"}
SC4_CRITICAL_CODES = {"S00", "S02", "S04", "S07", "S31"}

MILESTONE_NAMES = {
    "S00": "Shipment created",
    "S02": "Collected",
    "S04": "Vessel/flight departed",
    "S07": "Vessel/flight arrived",
    "S10": "On hand at origin SVC",
    "S13": "On hand at destination SVC",
    "S16": "Customs initiated",
    "S31": "Delivered",
    "S45": "Handover to broker",
    "S46": "Full Container loaded on vessel",
    "S50": "Received origin CFS",
    "S51": "Arrived destination CFS",
    "S52": "Empty Container picked up",
    "S53": "Full Container loaded on vessel",
    "S54": "Full Container discharge from vessel",
    "S55": "Empty Container returned",
    "S05": "In delivery",
    "S60": "Pre-Booking confirmed",
}


def find_sheet(wb, pattern_fn):
    for sn in wb.sheetnames:
        if pattern_fn(sn):
            return sn
    return None


def find_shipments_sheet(wb):
    return find_sheet(wb, lambda sn: sn.strip().lower() == "shipments")


def find_detail_sheet(wb, service_type):
    """Find the detail sheet for a service type.
    CW08+ has paired sheets: FCL (summary) + FCL_ (detail).
    CW01-CW07 has just FCL/BCO/LCL with detail data directly.
    """
    # First try the _ variant (CW08+)
    detail_name = find_sheet(wb, lambda sn: sn.strip() == f"{service_type}_")
    if detail_name:
        return detail_name
    # Fall back to direct sheet (CW01-CW07)
    return find_sheet(wb, lambda sn: sn.strip().upper() == service_type.upper())


def parse_milestone_code(name):
    if name and " - " in str(name):
        return str(name).split(" - ")[0].strip()
    return str(name).strip() if name else None


def read_total_summary(filepath):
    """Read TOTAL sheet summary rows."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = find_sheet(wb, lambda sn: sn.strip().upper().startswith("TOTAL") or sn.strip().upper() == "ALL")
    if not sheet_name:
        wb.close()
        return []

    ws = wb[sheet_name]
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        for cell in row:
            if cell and str(cell).strip().lower().startswith("required status"):
                header_row = i
                break
        if header_row:
            break
    if not header_row:
        header_row = 3

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        milestone_name = row[0]
        if not milestone_name or str(milestone_name).strip().lower().startswith("required status"):
            continue
        code = parse_milestone_code(milestone_name)
        status_type = str(row[1]).strip() if row[1] else ""
        required = row[3] if row[3] and isinstance(row[3], (int, float)) else 0
        available = row[4] if row[4] and isinstance(row[4], (int, float)) else 0
        in_time = row[5] if row[5] and isinstance(row[5], (int, float)) else 0

        rows.append({
            "code": code,
            "name": MILESTONE_NAMES.get(code, code),
            "type": status_type,
            "required": int(required),
            "available": int(available),
            "in_time": int(in_time),
            "missing": int(required - available),
            "late": int(available - in_time),
            "completeness": round(available / required, 4) if required > 0 else 0,
            "timeliness": round(in_time / required, 4) if required > 0 else 0,
        })
    wb.close()
    return rows


def build_sc3_load_to_hbl_map(filepath):
    """Build mapping from LOAD_ID_TO_ID -> HBL from SC3 shipments sheet."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = find_shipments_sheet(wb)
    if not sheet_name:
        wb.close()
        return {}

    ws = wb[sheet_name]
    # Header at row 3
    header_row = 3
    headers = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))

    # Find LOAD_ID (col 0), TO_ID (col 1), HBL (col 86), MBL (col 85)
    load_col = 0
    to_col = 1
    hbl_col = None
    mbl_col = None
    for idx, h in enumerate(headers):
        if h and "House_Airway_Bill" in str(h):
            hbl_col = idx
        if h and "Airway_Bill_or_Bill_of_lading" == str(h).strip():
            mbl_col = idx

    if hbl_col is None:
        # Fallback to col 86
        hbl_col = 86 if len(headers) > 86 else None

    mapping = {}
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        vals = list(row)
        if not vals[load_col]:
            continue
        load_id = str(vals[load_col]).strip()
        to_id = str(vals[to_col]).strip() if vals[to_col] else ""
        key = f"{load_id}_{to_id}"
        hbl = str(vals[hbl_col]).strip() if hbl_col and vals[hbl_col] else ""
        mbl = str(vals[mbl_col]).strip() if mbl_col and vals[mbl_col] else ""
        mapping[key] = {"hbl": hbl, "mbl": mbl, "load_id": load_id, "to_id": to_id}
        # Also map just load_id for cases where detail doesn't have full LOAD_TO
        if load_id not in mapping:
            mapping[load_id] = {"hbl": hbl, "mbl": mbl, "load_id": load_id, "to_id": to_id}

    wb.close()
    return mapping


def extract_sc4_detail(filepath):
    """Extract SC4 shipment-level milestone detail with HBLs."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    all_details = []

    for svc in ["FCL", "BCO", "LCL"]:
        sheet_name = find_detail_sheet(wb, svc)
        if not sheet_name:
            continue

        ws = wb[sheet_name]
        # Find header row with CONSIGNMENT
        header_row = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            row_strs = [str(c).strip().upper() if c else "" for c in row]
            if "CONSIGNMENT" in row_strs:
                header_row = i
                break
        if not header_row:
            continue

        headers = [str(c).strip().upper() if c else "" for c in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
        col_map = {h: i for i, h in enumerate(headers) if h}

        consignment_col = col_map.get("CONSIGNMENT", 0)
        hbl_col = col_map.get("HBL", 2)
        mbl_col = col_map.get("MBL", 3)
        code_col = col_map.get("STATUS_CODE_REQUIRED", 4)
        type_col = col_map.get("STATUS_TYPE", 5)
        avl_col = None
        for k, v in col_map.items():
            if "STATUS_CODE_AVAILABLE" in k:
                avl_col = v
                break

        if avl_col is None:
            continue

        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
            vals = list(row)
            if not vals[consignment_col]:
                continue

            available = int(vals[avl_col]) if isinstance(vals[avl_col], (int, float)) else 0
            code = str(vals[code_col]).strip() if vals[code_col] else ""
            status_type = str(vals[type_col]).strip() if vals[type_col] else ""

            all_details.append({
                "consignment": str(vals[consignment_col]).strip(),
                "hbl": str(vals[hbl_col]).strip() if vals[hbl_col] else "",
                "mbl": str(vals[mbl_col]).strip() if vals[mbl_col] else "",
                "code": code,
                "type": status_type,
                "available": available,
                "service": svc,
            })

    wb.close()
    return all_details


def extract_sc3_detail(filepath):
    """Extract SC3 shipment-level milestone detail."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    all_details = []

    for svc in ["FCL", "BCO", "LCL"]:
        sheet_name = find_detail_sheet(wb, svc)
        if not sheet_name:
            continue

        ws = wb[sheet_name]
        header_row = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            row_strs = [str(c).strip().upper() if c else "" for c in row]
            if any("STATUS_CODE_AVAILABLE" in s for s in row_strs):
                header_row = i
                break
        if not header_row:
            continue

        headers = [str(c).strip().upper() if c else "" for c in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]

        load_to_col = 0
        code_col = None
        type_col = None
        avl_col = None
        for idx, h in enumerate(headers):
            if h in ("STATUS CODE", "STATUS_CODE_REQUIRED"):
                code_col = idx
            if h == "STATUS_TYPE":
                type_col = idx
            if "STATUS_CODE_AVAILABLE" in h:
                avl_col = idx

        if avl_col is None or code_col is None:
            continue

        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
            vals = list(row)
            if not vals[load_to_col]:
                continue

            available = int(vals[avl_col]) if isinstance(vals[avl_col], (int, float)) else 0
            code = str(vals[code_col]).strip() if vals[code_col] else ""
            status_type = str(vals[type_col]).strip() if type_col is not None and vals[type_col] else ""

            all_details.append({
                "load_to": str(vals[load_to_col]).strip(),
                "code": code,
                "type": status_type,
                "available": available,
                "service": svc,
            })

    wb.close()
    return all_details


def extract_eta_ref_rca(filepath):
    """Extract ETA accuracy and Reference Completeness shipment-level RCA from SC4 shipments sheet.

    Returns dict with:
      - eta_2p: {total, accepted, failed, rate, failed_shipments[]}
      - eta_2d: {total, accepted, failed, rate, failed_shipments[]}
      - ref: {total, complete, incomplete, rate, incomplete_shipments[]}
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = find_shipments_sheet(wb)
    if not sheet_name:
        wb.close()
        return None

    ws = wb[sheet_name]
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        for cell in row:
            if cell and "UNIQUE_SHIPMENT" in str(cell):
                header_row = i
                break
        if header_row:
            break
    if not header_row:
        header_row = 3

    headers = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            col_map[str(h).strip()] = i

    # Find relevant columns
    hbl_col = col_map.get("HOUSE_BILL_OF_LADING", 1)
    mbl_col = col_map.get("MASTER_BILL_OF_LADING", 2)
    consignment_col = col_map.get("CONSIGNMENT_ID_1", 9)
    transport_col = col_map.get("TRANSPORT_SERVICE_PRIORITY", None)
    origin_col = col_map.get("CONSIGNOR_ADDRESS_CITY_NAME", None)
    dest_col = col_map.get("CONSIGNEE_ADDRESS_CITY_NAME", None)
    carrier_col = col_map.get("CARRIER_1", None)

    s07_col = s31_col = s07_dev_col = s31_dev_col = None
    civ_col = dn_col = None
    eta_col = ata_col = del_est_col = delivered_col = None
    s07_eta_measured_col = s31_eta_measured_col = None
    for key, idx in col_map.items():
        if "S07_Accepted" in key:
            s07_col = idx
        if "S31_Accepted" in key:
            s31_col = idx
        if "S07_Deviation" in key:
            s07_dev_col = idx
        if "S31_Deviation" in key:
            s31_dev_col = idx
        if "S07_TS_@measured" in key and "list" in key:
            s07_eta_measured_col = idx
        if "S31_TS_@measured" in key and "list" in key:
            s31_eta_measured_col = idx
        if "COMMERCIAL_INVOICE" in key:
            civ_col = idx
        if "DELIVERY_NOTE" in key:
            dn_col = idx
        if key == "ETA_DATE_TIME":
            eta_col = idx
        if key == "ATA_DATE_TIME":
            ata_col = idx
        if key == "DELIVERY_DATE_ACT_EST_PLAN":
            del_est_col = idx
        if key == "DELIVERED_DATE_TIME":
            delivered_col = idx

    eta_2p_failed = []
    eta_2p_accepted = 0
    eta_2p_total = 0
    eta_2d_failed = []
    eta_2d_accepted = 0
    eta_2d_total = 0
    ref_incomplete = []
    ref_complete_count = 0
    ref_total = 0

    DEBUG = False  # Set to True to enable debug output
    for row_num, row in enumerate(ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True), header_row + 1):
        vals = list(row)
        if not vals[0]:
            continue

        hbl = str(vals[hbl_col]).strip() if vals[hbl_col] else ""
        mbl = str(vals[mbl_col]).strip() if vals[mbl_col] else ""
        consignment = str(vals[consignment_col]).strip() if vals[consignment_col] else ""
        transport = str(vals[transport_col]).strip() if transport_col is not None and vals[transport_col] else ""
        origin = str(vals[origin_col]).strip() if origin_col is not None and vals[origin_col] else ""
        dest = str(vals[dest_col]).strip() if dest_col is not None and vals[dest_col] else ""
        carrier = str(vals[carrier_col]).strip() if carrier_col is not None and vals[carrier_col] else ""

        base_info = {
            "hbl": hbl,
            "mbl": mbl,
            "consignment": consignment,
            "transport": transport,
            "origin": origin,
            "dest": dest,
            "carrier": carrier,
        }

        # ETA 2P (S07) — only count if vessel has actually arrived (ATA exists)
        ata_val = vals[ata_col] if ata_col is not None and ata_col < len(vals) else None
        if DEBUG and row_num == 17 and '25HA30007606A' in hbl:
            print(f'DEBUG row {row_num}: HBL={hbl}, s07_col={s07_col}, len(vals)={len(vals)}')
            if s07_col is not None and s07_col < len(vals):
                print(f'  vals[s07_col]={vals[s07_col]}, ata_col={ata_col}, ata_val={ata_val}')
        if s07_col is not None and s07_col < len(vals) and vals[s07_col] is not None:
            s07_val = vals[s07_col]
            if DEBUG and '25HA30007606A' in hbl:
                print(f'  S07 check: s07_val={s07_val}, ata_val={ata_val}')
            # Skip in-transit shipments: S07_Accepted=0 but no ATA means not yet measurable
            if s07_val == 0 and ata_val is None:
                if DEBUG and '25HA30007606A' in hbl:
                    print(f'  SKIPPED: in-transit (S07=0, no ATA)')
                pass  # in-transit, exclude from measurement
            else:
                eta_2p_total += 1
                if s07_val == 1:
                    eta_2p_accepted += 1
                    if DEBUG and '25HA30007606A' in hbl:
                        print(f'  ACCEPTED')
                else:
                    if DEBUG and '25HA30007606A' in hbl:
                        print(f'  FAILED - adding to list')
                    deviation = vals[s07_dev_col] if s07_dev_col and s07_dev_col < len(vals) and vals[s07_dev_col] else None
                    # Use S07 measured ETA if available (the ETA at time of measurement), fallback to current ETA
                    eta_val = vals[s07_eta_measured_col] if s07_eta_measured_col and s07_eta_measured_col < len(vals) and vals[s07_eta_measured_col] else (vals[eta_col] if eta_col is not None and eta_col < len(vals) else None)
                    if DEBUG and '25HA30007606A' in hbl:
                        print(f'    deviation={deviation}, eta_val={eta_val}, s07_eta_measured_col={s07_eta_measured_col}')
                    dev_hours = round(float(deviation), 1) if deviation and isinstance(deviation, (int, float)) else None
                    eta_2p_failed.append({
                        **base_info,
                        "deviation_hours": dev_hours,
                        "deviation_days": round(dev_hours / 24, 1) if dev_hours else None,
                        "direction": "late" if dev_hours and dev_hours > 0 else "early" if dev_hours and dev_hours < 0 else None,
                        "eta_baseline": str(eta_val)[:16] if eta_val else None,
                        "estimated": str(eta_val)[:16] if eta_val else None,
                        "actual": str(ata_val)[:16] if ata_val else None,
                        "window_start": str(eta_val - timedelta(hours=48))[:16] if eta_val and hasattr(eta_val, '__sub__') else None,
                        "window_end": str(eta_val + timedelta(hours=48))[:16] if eta_val and hasattr(eta_val, '__add__') else None,
                    })
                    if DEBUG and '25HA30007606A' in hbl:
                        print(f'    Added to eta_2p_failed. List size now: {len(eta_2p_failed)}')

        # ETA 2D (S31) — only count if actually delivered
        delivered_val = vals[delivered_col] if delivered_col is not None else None
        if s31_col is not None and vals[s31_col] is not None:
            s31_val = vals[s31_col]
            # Skip undelivered shipments: S31_Accepted=0 but no delivered date means not yet measurable
            if s31_val == 0 and delivered_val is None:
                pass  # not yet delivered, exclude from measurement
            else:
                eta_2d_total += 1
                if s31_val == 1:
                    eta_2d_accepted += 1
                else:
                    deviation = vals[s31_dev_col] if s31_dev_col and vals[s31_dev_col] else None
                    # Use S31 measured ETA if available (the ETA at time of measurement), fallback to current delivery estimate
                    del_est_val = vals[s31_eta_measured_col] if s31_eta_measured_col and vals[s31_eta_measured_col] else (vals[del_est_col] if del_est_col is not None else None)
                    dev_hours = round(float(deviation), 1) if deviation and isinstance(deviation, (int, float)) else None
                    eta_2d_failed.append({
                        **base_info,
                        "deviation_hours": dev_hours,
                        "deviation_days": round(dev_hours / 24, 1) if dev_hours else None,
                        "direction": "late" if dev_hours and dev_hours > 0 else "early" if dev_hours and dev_hours < 0 else None,
                        "eta_baseline": str(del_est_val)[:16] if del_est_val else None,
                        "estimated": str(del_est_val)[:16] if del_est_val else None,
                        "actual": str(delivered_val)[:16] if delivered_val else None,
                        "window_start": str(del_est_val - timedelta(hours=48))[:16] if del_est_val and hasattr(del_est_val, '__sub__') else None,
                        "window_end": str(del_est_val + timedelta(hours=48))[:16] if del_est_val and hasattr(del_est_val, '__add__') else None,
                    })

        # Reference Completeness
        ref_total += 1
        civ_val = vals[civ_col] if civ_col else None
        dn_val = vals[dn_col] if dn_col else None
        has_civ = civ_val is not None and str(civ_val).strip() not in ("", "None", "NA")
        has_dn = dn_val is not None and str(dn_val).strip() not in ("", "None", "NA")
        if has_civ or has_dn:
            ref_complete_count += 1
        else:
            ref_incomplete.append({
                **base_info,
                "has_civ": has_civ,
                "has_dn": has_dn,
            })

    wb.close()

    return {
        "eta_2p": {
            "total": eta_2p_total,
            "accepted": eta_2p_accepted,
            "failed": eta_2p_total - eta_2p_accepted,
            "rate": round(eta_2p_accepted / eta_2p_total, 4) if eta_2p_total > 0 else None,
            "failed_shipments": eta_2p_failed,
            "total_failed_shipments": len(eta_2p_failed),
        },
        "eta_2d": {
            "total": eta_2d_total,
            "accepted": eta_2d_accepted,
            "failed": eta_2d_total - eta_2d_accepted,
            "rate": round(eta_2d_accepted / eta_2d_total, 4) if eta_2d_total > 0 else None,
            "failed_shipments": eta_2d_failed,
            "total_failed_shipments": len(eta_2d_failed),
        },
        "ref": {
            "total": ref_total,
            "complete": ref_complete_count,
            "incomplete": ref_total - ref_complete_count,
            "rate": round(ref_complete_count / ref_total, 4) if ref_total > 0 else None,
            "incomplete_shipments": ref_incomplete,
            "total_incomplete_shipments": len(ref_incomplete),
        },
    }


"""
Milestone Accuracy Check — Plausibility Rules

Expected milestone sequence:
  PUP_EST (S02 est) ≤ PUP (S02 act) ≤ ETD/ATD (S04) ≤ ETA/ATA (S07) ≤ POD_EST/POD (S31)

24 pairwise rules enforce this ordering:
  - S02 cannot be after S04, S07, S31
  - S04 cannot be before S02, cannot be after S07, S31
  - S07 cannot be before S02/S04, cannot be after S31
  - S31 cannot be before S02/S04/S07
"""

# Milestone date field names (logical)
#   PUP_EST = Planned pickup (S02 est)
#   PUP     = Collected (S02 act)
#   ETD     = Est. departure (S04 est)
#   ATD     = Act. departure (S04 act — VD for LCL, DEP for BCN/FCL)
#   ETA     = Est. arrival (S07 est)
#   ATA     = Act. arrival (S07 act — VA for LCL, ARR for BCN/FCL)
#   POD_EST = Planned delivery (S31 est)
#   POD     = Delivered (S31 act)

PLAUSIBILITY_RULES = [
    # S02 (est/act) cannot be after S04/S07/S31
    ("PUP_EST", "ETD",     "S02 est > S04 est",  "Planned pickup after est. departure"),
    ("PUP_EST", "ATD",     "S02 est > S04 act",  "Planned pickup after act. departure"),
    ("PUP_EST", "ETA",     "S02 est > S07 est",  "Planned pickup after est. arrival"),
    ("PUP_EST", "ATA",     "S02 est > S07 act",  "Planned pickup after act. arrival"),
    ("PUP_EST", "POD_EST", "S02 est > S31 est",  "Planned pickup after planned delivery"),
    ("PUP_EST", "POD",     "S02 est > S31 act",  "Planned pickup after delivery"),
    ("PUP",     "ETD",     "S02 act > S04 est",  "Collected after est. departure"),
    ("PUP",     "ATD",     "S02 act > S04 act",  "Collected after act. departure"),
    ("PUP",     "ETA",     "S02 act > S07 est",  "Collected after est. arrival"),
    ("PUP",     "ATA",     "S02 act > S07 act",  "Collected after act. arrival"),
    ("PUP",     "POD_EST", "S02 act > S31 est",  "Collected after planned delivery"),
    ("PUP",     "POD",     "S02 act > S31 act",  "Collected after delivery"),
    # S04 cannot be after S07/S31
    ("ETD",     "ETA",     "S04 est > S07 est",  "Est. departure after est. arrival"),
    ("ETD",     "ATA",     "S04 est > S07 act",  "Est. departure after act. arrival"),
    ("ETD",     "POD_EST", "S04 est > S31 est",  "Est. departure after planned delivery"),
    ("ETD",     "POD",     "S04 est > S31 act",  "Est. departure after delivery"),
    ("ATD",     "ETA",     "S04 act > S07 est",  "Act. departure after est. arrival"),
    ("ATD",     "ATA",     "S04 act > S07 act",  "Act. departure after act. arrival"),
    ("ATD",     "POD_EST", "S04 act > S31 est",  "Act. departure after planned delivery"),
    ("ATD",     "POD",     "S04 act > S31 act",  "Act. departure after delivery"),
    # S07 cannot be after S31
    ("ETA",     "POD_EST", "S07 est > S31 est",  "Est. arrival after planned delivery"),
    ("ETA",     "POD",     "S07 est > S31 act",  "Est. arrival after delivery"),
    ("ATA",     "POD_EST", "S07 act > S31 est",  "Act. arrival after planned delivery"),
    ("ATA",     "POD",     "S07 act > S31 act",  "Act. arrival after delivery"),
]

# Milestone group for summary: which milestone pair category
MILESTONE_GROUPS = {
    "S02 est": "S02", "S02 act": "S02",
    "S04 est": "S04", "S04 act": "S04",
    "S07 est": "S07", "S07 act": "S07",
    "S31 est": "S31", "S31 act": "S31",
}

# Minimum valid date — SC3 has bogus dates like 1900-01-02
MIN_VALID_DATE = None  # set in function


def _valid_dt(val):
    """Return datetime if valid (post-2020), else None."""
    from datetime import datetime
    if isinstance(val, datetime) and val.year >= 2020:
        return val
    return None


def _safe_col(vals, col_idx):
    """Safely get a column value."""
    if col_idx is not None and col_idx < len(vals):
        return _valid_dt(vals[col_idx])
    return None


def extract_plausibility_rca_sc4(filepath):
    """Extract milestone accuracy violations from SC4 shipments sheet."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = find_shipments_sheet(wb)
    if not sheet_name:
        wb.close()
        return [], 0, set()

    ws = wb[sheet_name]
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        for cell in row:
            if cell and "UNIQUE_SHIPMENT" in str(cell):
                header_row = i
                break
        if header_row:
            break
    if not header_row:
        header_row = 3

    headers = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            col_map[str(h).strip()] = i

    hbl_col = col_map.get("HOUSE_BILL_OF_LADING", 1)
    mbl_col = col_map.get("MASTER_BILL_OF_LADING", 2)
    consignment_col = col_map.get("CONSIGNMENT_ID_1", 9)
    carrier_col = col_map.get("CARRIER_1", None)
    origin_col = col_map.get("CONSIGNOR_ADDRESS_CITY_NAME", None)
    dest_col = col_map.get("CONSIGNEE_ADDRESS_CITY_NAME", None)
    transport_col = col_map.get("TRANSPORT_SERVICE_PRIORITY", None)
    dataset_col = col_map.get("DATASET", None)

    # SC4 date column mapping
    date_cols = {
        "PUP_EST": col_map.get("TARGET_COLLECTION", 73),
        "PUP":     col_map.get("COLLECTED_DATE_TIME", 74),
        "ETD":     col_map.get("ETD_DATE_TIME", 76),
        "ATD":     col_map.get("ATD_DATE_TIME", 77),
        "ETA":     col_map.get("ETA_DATE_TIME", 78),
        "ATA":     col_map.get("ATA_DATE_TIME", 79),
        "POD_EST": col_map.get("DELIVERY_DATE_ACT_EST_PLAN", 71),
        "POD":     col_map.get("DELIVERED_DATE_TIME", 81),
    }

    violations = []
    total_shipments = 0
    affected_hbls = set()

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        vals = list(row)
        if not vals[0]:
            continue
        total_shipments += 1

        hbl = str(vals[hbl_col]).strip() if vals[hbl_col] else ""
        mbl = str(vals[mbl_col]).strip() if vals[mbl_col] else ""
        consignment = str(vals[consignment_col]).strip() if vals[consignment_col] else ""
        carrier = str(vals[carrier_col]).strip() if carrier_col is not None and vals[carrier_col] else ""
        origin = str(vals[origin_col]).strip() if origin_col is not None and vals[origin_col] else ""
        dest = str(vals[dest_col]).strip() if dest_col is not None and vals[dest_col] else ""
        transport = str(vals[transport_col]).strip() if transport_col is not None and vals[transport_col] else ""
        dataset = str(vals[dataset_col]).strip() if dataset_col is not None and vals[dataset_col] else ""

        base_info = {
            "hbl": hbl, "mbl": mbl, "consignment": consignment,
            "carrier": carrier, "origin": origin, "dest": dest,
            "transport": transport, "dataset": dataset, "scenario": "SC4",
        }

        # Get all 8 date values
        dates = {k: _safe_col(vals, c) for k, c in date_cols.items()}

        for field_a, field_b, rule_id, desc in PLAUSIBILITY_RULES:
            da = dates.get(field_a)
            db = dates.get(field_b)
            if da and db and da > db:
                gap_hours = (da - db).total_seconds() / 3600
                affected_hbls.add(hbl)
                violations.append({
                    **base_info,
                    "rule": rule_id,
                    "rule_label": f"{field_a} > {field_b}",
                    "description": desc,
                    "field_a": field_a,
                    "field_b": field_b,
                    "date_a": da.isoformat(),
                    "date_b": db.isoformat(),
                    "gap_hours": round(gap_hours, 1),
                })

    wb.close()
    return violations, total_shipments, affected_hbls


def extract_plausibility_rca_sc3(filepath):
    """Extract milestone accuracy violations from SC3 shipments sheet."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = find_shipments_sheet(wb)
    if not sheet_name:
        wb.close()
        return [], 0, set()

    ws = wb[sheet_name]
    header_row = 3

    headers = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            col_map[str(h).strip()] = i

    hbl_col = col_map.get("House_Airway_Bill_or_House_Bill_of_Lading", 86)
    mbl_col = col_map.get("Airway_Bill_or_Bill_of_lading", 85)
    load_to_col = col_map.get("LOAD_TO", 2)
    origin_col = col_map.get("Consignor_Name", None)
    dest_col = col_map.get("Recipient_Name", None)
    carrier_col = col_map.get("Main_Carrier_or_Shipping_Line", None)

    # SC3 date column mapping
    date_cols = {
        "PUP_EST": col_map.get("Pick_up_date", 51),
        "PUP":     col_map.get("Collected_Sent_DATE_TIME", 99),
        "ETD":     col_map.get("ETD_Datetime", 93),
        "ATD":     col_map.get("ATD_Datetime", 94),
        "ETA":     col_map.get("ETA_Datetime", 91),
        "ATA":     col_map.get("ATA_Datetime", 92),
        "POD_EST": col_map.get("Arrival_date", 55),
        "POD":     col_map.get("Delivered_Sent_DATE_TIME", 101),
    }

    violations = []
    total_shipments = 0
    affected_hbls = set()

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        vals = list(row)
        if not vals[0]:
            continue
        total_shipments += 1

        hbl = str(vals[hbl_col]).strip() if hbl_col < len(vals) and vals[hbl_col] else ""
        mbl = str(vals[mbl_col]).strip() if mbl_col < len(vals) and vals[mbl_col] else ""
        load_to = str(vals[load_to_col]).strip() if load_to_col < len(vals) and vals[load_to_col] else ""
        origin = str(vals[origin_col]).strip() if origin_col is not None and origin_col < len(vals) and vals[origin_col] else ""
        dest = str(vals[dest_col]).strip() if dest_col is not None and dest_col < len(vals) and vals[dest_col] else ""
        carrier = str(vals[carrier_col]).strip() if carrier_col is not None and carrier_col < len(vals) and vals[carrier_col] else ""

        base_info = {
            "hbl": hbl, "mbl": mbl, "consignment": load_to,
            "carrier": carrier, "origin": origin, "dest": dest,
            "transport": "", "dataset": "", "scenario": "SC3",
        }

        dates = {k: _safe_col(vals, c) for k, c in date_cols.items()}

        for field_a, field_b, rule_id, desc in PLAUSIBILITY_RULES:
            da = dates.get(field_a)
            db = dates.get(field_b)
            if da and db and da > db:
                gap_hours = (da - db).total_seconds() / 3600
                affected_hbls.add(hbl or load_to)
                violations.append({
                    **base_info,
                    "rule": rule_id,
                    "rule_label": f"{field_a} > {field_b}",
                    "description": desc,
                    "field_a": field_a,
                    "field_b": field_b,
                    "date_a": da.isoformat(),
                    "date_b": db.isoformat(),
                    "gap_hours": round(gap_hours, 1),
                })

    wb.close()
    return violations, total_shipments, affected_hbls


def extract_plausibility_rca(sc4_filepath, sc3_filepath):
    """Extract combined SC3+SC4 milestone accuracy violations."""
    from collections import Counter

    sc4_violations, sc4_total, sc4_affected = extract_plausibility_rca_sc4(sc4_filepath)
    sc3_violations, sc3_total, sc3_affected = extract_plausibility_rca_sc3(sc3_filepath)

    all_violations = sc4_violations + sc3_violations
    total_shipments = sc4_total + sc3_total
    all_affected = sc4_affected | sc3_affected

    rule_counts = Counter(v["rule"] for v in all_violations)

    # Group violations by milestone pair category
    milestone_pair_counts = Counter()
    for v in all_violations:
        pair = f"{v['field_a']} > {v['field_b']}"
        milestone_pair_counts[pair] += 1

    return {
        "total_shipments": total_shipments,
        "sc4_shipments": sc4_total,
        "sc3_shipments": sc3_total,
        "affected_hbls": len(all_affected),
        "sc4_affected": len(sc4_affected),
        "sc3_affected": len(sc3_affected),
        "total_violations": len(all_violations),
        "rule_counts": dict(rule_counts),
        "milestone_pair_counts": dict(milestone_pair_counts),
        "violations": all_violations,
        "total_violation_records": len(all_violations),
    }


def build_missing_shipments(sc4_details, sc3_details, sc3_hbl_map):
    """Build lists of missing/late shipments per milestone."""
    missing = {}

    # SC4 missing (available=0)
    for d in sc4_details:
        if d["available"] == 0:
            key = (d["code"], d["type"], "SC4")
            if key not in missing:
                missing[key] = []
            missing[key].append({
                "hbl": d["hbl"],
                "mbl": d["mbl"],
                "consignment": d["consignment"],
                "service": d["service"],
            })

    # SC3 missing (available=0)
    for d in sc3_details:
        if d["available"] == 0:
            key = (d["code"], d["type"], "SC3")
            if key not in missing:
                missing[key] = []
            # Resolve HBL from mapping
            load_to = d["load_to"]
            ref = sc3_hbl_map.get(load_to, {})
            missing[key].append({
                "hbl": ref.get("hbl", ""),
                "mbl": ref.get("mbl", ""),
                "load_to": load_to,
                "service": d["service"],
            })

    return missing


def process_week(week):
    """Process one week and return RCA data."""
    print(f"  Processing {week}...")

    sc3_file = os.path.join(RAW_DIR, SC3_FILES[week])
    sc4_file = os.path.join(RAW_DIR, SC4_FILES[week])

    if not os.path.exists(sc3_file) or not os.path.exists(sc4_file):
        print(f"    Files missing for {week}")
        return None

    # Summary data
    sc3_summary = read_total_summary(sc3_file)
    sc4_summary = read_total_summary(sc4_file)

    # Detail data
    sc4_details = extract_sc4_detail(sc4_file)
    sc3_details = extract_sc3_detail(sc3_file)

    # SC3 HBL mapping
    sc3_hbl_map = build_sc3_load_to_hbl_map(sc3_file)

    # Build missing shipment lists
    missing_map = build_missing_shipments(sc4_details, sc3_details, sc3_hbl_map)

    # Build milestone-level RCA
    milestones = []

    for scenario, summary, crit_codes in [
        ("SC3", sc3_summary, SC3_CRITICAL_CODES),
        ("SC4", sc4_summary, SC4_CRITICAL_CODES),
    ]:
        for row in summary:
            code = row["code"]
            stype = row["type"]
            is_critical = code in crit_codes
            missing_key = (code, stype, scenario)
            missing_list = missing_map.get(missing_key, [])

            missing_hbls = missing_list
            total_missing_shipments = len(missing_list)

            milestones.append({
                "scenario": scenario,
                "code": code,
                "name": row["name"],
                "type": stype,
                "is_critical": is_critical,
                "required": row["required"],
                "available": row["available"],
                "in_time": row["in_time"],
                "missing": row["missing"],
                "late": row["late"],
                "completeness": row["completeness"],
                "timeliness": row["timeliness"],
                "comp_gap": round(0.95 - row["completeness"], 4) if row["completeness"] < 0.95 else 0,
                "time_gap": round(0.70 - row["timeliness"], 4) if row["timeliness"] < 0.70 else 0,
                "severity": "critical" if row["completeness"] < 0.7 else "warning" if row["completeness"] < 0.9 else "ok",
                "missing_shipments": missing_hbls,
                "total_missing_shipments": total_missing_shipments,
            })

    # Sort by severity then by missing count
    severity_order = {"critical": 0, "warning": 1, "ok": 2}
    milestones.sort(key=lambda m: (severity_order.get(m["severity"], 3), -m["missing"]))

    # Aggregate stats
    total_missing = sum(m["missing"] for m in milestones)
    critical_issues = sum(1 for m in milestones if m["severity"] == "critical")
    warning_issues = sum(1 for m in milestones if m["severity"] == "warning")

    # ETA & Reference RCA (SC4 only)
    eta_ref_rca = extract_eta_ref_rca(sc4_file)

    # Plausibility RCA (SC3 + SC4 combined)
    plausibility_rca = extract_plausibility_rca(sc4_file, sc3_file)

    return {
        "week": week,
        "milestones": milestones,
        "total_missing": total_missing,
        "critical_issues": critical_issues,
        "warning_issues": warning_issues,
        "ok_issues": sum(1 for m in milestones if m["severity"] == "ok"),
        "eta_ref_rca": eta_ref_rca,
        "plausibility_rca": plausibility_rca,
    }


def main():
    print("=" * 80)
    print("BOSCH MILESTONE RCA DATA EXTRACTION")
    print("=" * 80)

    all_rca = []
    for week in WEEKS:
        result = process_week(week)
        if result:
            all_rca.append(result)

    # Export
    output_path = os.path.join(BASE, "rca_data.json")
    with open(output_path, "w") as f:
        json.dump(all_rca, f, indent=2, default=str)
    print(f"\nExported RCA data to: {output_path}")

    # Quick summary
    for r in all_rca:
        print(f"\n  {r['week']}: {r['critical_issues']} critical, {r['warning_issues']} warnings, {r['total_missing']} missing statuses")
        # Top 3 worst milestones
        for m in r["milestones"][:3]:
            print(f"    {m['scenario']} {m['code']} {m['type']:10s} — missing={m['missing']:4d}, comp={m['completeness']:.1%}, time={m['timeliness']:.1%}")

    print("\nDone.")


if __name__ == "__main__":
    main()
