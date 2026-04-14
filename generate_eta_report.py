"""Generate ETA 2D RCA Excel report with detailed calculations."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
from collections import defaultdict

BASE = "/Users/harsh.puri/Documents/work-maersk/Prototype Playground/Bosch Milestone Analysis"

with open(f"{BASE}/eta_2d_analysis.json") as f:
    data = json.load(f)

wb = openpyxl.Workbook()

# Styles
header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
data_font = Font(name='Calibri', size=10)
bold_font = Font(name='Calibri', bold=True, size=10)
title_font = Font(name='Calibri', bold=True, size=14, color='003366')
subtitle_font = Font(name='Calibri', bold=True, size=12, color='003366')
pct_fmt = '0.0%'
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
light_blue = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
blue_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
orange_fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = thin_border


def style_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.font = data_font
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')


def color_acc(cell, val):
    if val is None:
        return
    if val >= 0.9:
        cell.fill = green_fill
    elif val >= 0.5:
        cell.fill = yellow_fill
    else:
        cell.fill = red_fill


def auto_width(ws, max_col, min_w=12):
    for col in range(1, max_col + 1):
        mx = min_w
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    mx = max(mx, len(str(cell.value)) + 2)
        ws.column_dimensions[get_column_letter(col)].width = min(mx, 35)


summary = data['summary']
total = summary['cw12_total']
accepted = summary['cw12_accepted']
failed_total = total - accepted

# ===================== SHEET 1: Executive Summary =====================
ws = wb.active
ws.title = 'Executive Summary'
ws.sheet_properties.tabColor = '003366'

r = 1
ws.cell(r, 1, 'ETA 2D Delivery Accuracy — Root Cause Analysis').font = title_font
ws.merge_cells('A1:H1')
r = 2
ws.cell(r, 1, 'Bosch Milestone Monitoring | CW09-CW12 Analysis | April 2026').font = Font(name='Calibri', italic=True, size=10, color='666666')
ws.merge_cells('A2:H2')

r = 4
ws.cell(r, 1, 'CURRENT STATE').font = subtitle_font
r = 5
headers = ['Metric', 'CW09', 'CW10', 'CW11', 'CW12', 'Trend', 'Target']
for c, h in enumerate(headers, 1):
    ws.cell(r, c, h)
style_header(ws, r, len(headers))

weekly = data['weekly_trend']
r = 6
ws.cell(r, 1, 'ETA 2D Accuracy (±48h)').font = bold_font
for i, w in enumerate(weekly):
    cell = ws.cell(r, i + 2, w['rate'] / 100)
    cell.number_format = pct_fmt
    color_acc(cell, w['rate'] / 100)
ws.cell(r, 6, 'Flat ~23-29%').font = data_font
ws.cell(r, 7, '90%+').font = bold_font
ws.cell(r, 7).fill = green_fill
style_row(ws, r, len(headers))

r = 7
ws.cell(r, 1, 'Measured Shipments').font = bold_font
for i, w in enumerate(weekly):
    ws.cell(r, i + 2, w['total'])
style_row(ws, r, len(headers))

r = 8
ws.cell(r, 1, 'Failed Shipments').font = bold_font
for i, w in enumerate(weekly):
    ws.cell(r, i + 2, w['failed'])
    ws.cell(r, i + 2).fill = red_fill
style_row(ws, r, len(headers))

r = 10
ws.cell(r, 1, 'GAP TO 90% (CW12)').font = subtitle_font
r = 11
gap_data = [
    ('Current Accuracy', f"{summary['cw12_accuracy']}%"),
    ('Total Measured', summary['cw12_total']),
    ('Currently Accepted (within ±48h)', summary['cw12_accepted']),
    ('Currently Failed', failed_total),
    ('Additional Needed for 90%', summary['needed_for_90pct']),
    ('% of Failures to Fix', f"{round(summary['needed_for_90pct'] / failed_total * 100, 1)}%"),
]
for label, val in gap_data:
    ws.cell(r, 1, label).font = bold_font
    ws.cell(r, 2, val).font = data_font
    ws.cell(r, 1).border = thin_border
    ws.cell(r, 2).border = thin_border
    r += 1

r += 1
ws.cell(r, 1, 'ROOT CAUSE SUMMARY').font = subtitle_font
r += 1
headers2 = ['#', 'Root Cause', 'Description', 'Impact']
for c, h in enumerate(headers2, 1):
    ws.cell(r, c, h)
style_header(ws, r, 4)

causes = [
    ('RC1', 'Stale ETA Baseline', 'S31 measured ETA captured too early, not updated after ATA. 83/85 early failures show delivery AFTER estimate date but measurement snapshot is outdated.', '~30% of failures'),
    ('RC2', 'CN->DE LCL Consolidation Delays', 'China to Germany LCL shipments avg +15.8 days late. Last-mile after ATA averages 12-14 days, far exceeding ±48h window.', '~28% of failures (80 ships)'),
    ('RC3', 'Systematic Bias on HU/PT Lanes', 'Hungary (Miskolc, Hatvan) and Portugal (Braga) lanes show consistent patterns — static ETAs not calibrated to actual transit.', '~23% of failures (67 ships)'),
    ('RC4', 'No ETA Update After Port Arrival', 'Last mile: 11.9 days (accepted) vs 13.7 days (failed). ETA not recalculated post-ATA for customs + inland transport.', 'Structural issue'),
]
for cause in causes:
    r += 1
    for c, v in enumerate(cause, 1):
        ws.cell(r, c, v).font = data_font
        ws.cell(r, c).border = thin_border
        ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical='top')

auto_width(ws, 8)
ws.column_dimensions['C'].width = 60
ws.column_dimensions['D'].width = 24

# ===================== SHEET 2: Top 10 Lanes =====================
ws2 = wb.create_sheet('Top 10 Failure Lanes')
ws2.sheet_properties.tabColor = 'C00000'

r = 1
ws2.cell(r, 1, 'Top 10 Failure Lanes — Cumulative Benefit Analysis').font = title_font
ws2.merge_cells('A1:K1')

r = 3
headers = ['#', 'Lane', 'Total', 'Accepted', 'Failed', 'Lane Acc', 'Cum Fixed', 'New Overall Acc', 'Gain (pp)', 'Avg Dev (h)', 'Direction']
for c, h in enumerate(headers, 1):
    ws2.cell(r, c, h)
style_header(ws2, r, len(headers))

lanes = data['by_country_lane'][:10]
cum = 0
for i, lane in enumerate(lanes):
    r += 1
    cum += lane['failed']
    new_acc = (accepted + cum) / total
    gain = new_acc - accepted / total
    avg_dev = lane['avg_deviation_hours']
    direction = 'LATE' if avg_dev > 48 else ('EARLY' if avg_dev < -48 else 'MIXED')
    ws2.cell(r, 1, i + 1)
    ws2.cell(r, 2, lane['value']).font = bold_font
    ws2.cell(r, 3, lane['total'])
    ws2.cell(r, 4, lane['accepted'])
    ws2.cell(r, 5, lane['failed']).fill = red_fill
    cell_acc = ws2.cell(r, 6, lane['accuracy'] / 100)
    cell_acc.number_format = pct_fmt
    color_acc(cell_acc, lane['accuracy'] / 100)
    ws2.cell(r, 7, cum)
    cell_new = ws2.cell(r, 8, new_acc)
    cell_new.number_format = pct_fmt
    color_acc(cell_new, new_acc)
    ws2.cell(r, 9, round(gain * 100, 1))
    ws2.cell(r, 10, round(avg_dev, 1))
    ws2.cell(r, 11, direction)
    style_row(ws2, r, len(headers))

r += 1
ws2.cell(r, 2, 'TOTAL (Top 10)').font = bold_font
ws2.cell(r, 5, cum).font = bold_font
ws2.cell(r, 5).fill = red_fill
ws2.cell(r, 7, cum).font = bold_font
cell_final = ws2.cell(r, 8, (accepted + cum) / total)
cell_final.number_format = pct_fmt
cell_final.font = bold_font
color_acc(cell_final, (accepted + cum) / total)
ws2.cell(r, 9, round(((accepted + cum) / total - accepted / total) * 100, 1)).font = bold_font
style_row(ws2, r, len(headers))
for c in range(1, len(headers) + 1):
    ws2.cell(r, c).fill = light_blue

r += 2
ws2.cell(r, 1, 'These 10 lanes = 72.2% of all CW12 failures. Fixing all 10 reaches 78.6%. Remaining 80 failures in long-tail lanes needed for 90%+.').font = Font(name='Calibri', italic=True, size=10)
ws2.merge_cells(f'A{r}:K{r}')
auto_width(ws2, len(headers))

# ===================== SHEET 3: What-If Window =====================
ws3 = wb.create_sheet('What-If Window Analysis')
ws3.sheet_properties.tabColor = '4472C4'

r = 1
ws3.cell(r, 1, 'What-If: Accuracy at Different Measurement Windows').font = title_font
ws3.merge_cells('A1:G1')

r = 3
headers = ['Window (±hours)', 'Window (±days)', 'Accepted', 'Total', 'Accuracy', 'Gap to 90%', 'Note']
for c, h in enumerate(headers, 1):
    ws3.cell(r, c, h)
style_header(ws3, r, len(headers))

for w in data['what_if_windows']:
    r += 1
    ws3.cell(r, 1, w['window_hours'])
    ws3.cell(r, 2, w['window_days'])
    ws3.cell(r, 3, w['accepted'])
    ws3.cell(r, 4, w['total'])
    cell = ws3.cell(r, 5, w['accuracy'] / 100)
    cell.number_format = pct_fmt
    color_acc(cell, w['accuracy'] / 100)
    ws3.cell(r, 6, round(90 - w['accuracy'], 1))
    if w['window_hours'] == 48:
        ws3.cell(r, 7, 'CURRENT WINDOW').font = bold_font
        ws3.cell(r, 7).fill = yellow_fill
    style_row(ws3, r, len(headers))

r += 2
ws3.cell(r, 1, 'Even at ±14 days (336h), accuracy = 86.8%. The ETAs require fundamental recalibration to reach 90%.').font = Font(name='Calibri', italic=True, size=10, color='C00000')
ws3.merge_cells(f'A{r}:G{r}')
auto_width(ws3, len(headers))

# ===================== SHEET 4: Early vs Late =====================
ws4 = wb.create_sheet('Early vs Late Analysis')
ws4.sheet_properties.tabColor = 'ED7D31'

failed_ships = data['failed_shipments_detail']
early = [s for s in failed_ships if s.get('deviation_hours') and s['deviation_hours'] < -48]
late = [s for s in failed_ships if s.get('deviation_hours') and s['deviation_hours'] > 48]

r = 1
ws4.cell(r, 1, 'Early vs Late Delivery Failure Breakdown (CW12)').font = title_font
ws4.merge_cells('A1:H1')

# Early by lane
r = 3
ws4.cell(r, 1, f'EARLY FAILURES ({len(early)} shipments, 29.6% of failures)').font = subtitle_font
ws4.merge_cells(f'A{r}:F{r}')
r = 4
ws4.cell(r, 1, 'Root cause: Stale ETA baseline not updated after faster-than-expected transit').font = Font(name='Calibri', italic=True, size=10)
ws4.merge_cells(f'A{r}:F{r}')

r = 6
headers = ['Lane', 'Count', 'Avg Dev (h)', 'Avg Dev (d)', '% of Early']
for c, h in enumerate(headers, 1):
    ws4.cell(r, c, h)
style_header(ws4, r, len(headers))

early_lanes = defaultdict(list)
for s in early:
    early_lanes[s['lane']].append(s)
for lane in sorted(early_lanes, key=lambda x: len(early_lanes[x]), reverse=True):
    ships = early_lanes[lane]
    r += 1
    avg = sum(s['deviation_hours'] for s in ships) / len(ships)
    ws4.cell(r, 1, lane).font = bold_font
    ws4.cell(r, 2, len(ships))
    ws4.cell(r, 3, round(avg, 1))
    ws4.cell(r, 4, round(avg / 24, 1))
    cell = ws4.cell(r, 5, len(ships) / len(early))
    cell.number_format = pct_fmt
    style_row(ws4, r, len(headers))

# Early by delivery city
r += 2
ws4.cell(r, 1, 'Early Failures by Delivery City (Top 10)').font = subtitle_font
r += 1
headers = ['Delivery City', 'Count', 'Avg Dev (d)', '% of Early']
for c, h in enumerate(headers, 1):
    ws4.cell(r, c, h)
style_header(ws4, r, len(headers))

early_cities = defaultdict(list)
for s in early:
    early_cities[s['delivery_city']].append(s)
for city in sorted(early_cities, key=lambda x: len(early_cities[x]), reverse=True)[:10]:
    ships = early_cities[city]
    r += 1
    avg = sum(s['deviation_hours'] for s in ships) / len(ships)
    ws4.cell(r, 1, city).font = bold_font
    ws4.cell(r, 2, len(ships))
    ws4.cell(r, 3, round(avg / 24, 1))
    cell = ws4.cell(r, 4, len(ships) / len(early))
    cell.number_format = pct_fmt
    style_row(ws4, r, len(headers))

# Late by lane
r += 2
ws4.cell(r, 1, f'LATE FAILURES ({len(late)} shipments, 48.4% of failures)').font = subtitle_font
ws4.merge_cells(f'A{r}:F{r}')
r += 1
ws4.cell(r, 1, 'Root cause: Last-mile transit + customs not factored into ETA, consolidation delays on CN->DE').font = Font(name='Calibri', italic=True, size=10)
ws4.merge_cells(f'A{r}:F{r}')

r += 2
headers = ['Lane', 'Count', 'Avg Dev (h)', 'Avg Dev (d)', '% of Late']
for c, h in enumerate(headers, 1):
    ws4.cell(r, c, h)
style_header(ws4, r, len(headers))

late_lanes = defaultdict(list)
for s in late:
    late_lanes[s['lane']].append(s)
for lane in sorted(late_lanes, key=lambda x: len(late_lanes[x]), reverse=True)[:15]:
    ships = late_lanes[lane]
    r += 1
    avg = sum(s['deviation_hours'] for s in ships) / len(ships)
    ws4.cell(r, 1, lane).font = bold_font
    ws4.cell(r, 2, len(ships))
    ws4.cell(r, 3, round(avg, 1))
    ws4.cell(r, 4, round(avg / 24, 1))
    cell = ws4.cell(r, 5, len(ships) / len(late))
    cell.number_format = pct_fmt
    style_row(ws4, r, len(headers))

auto_width(ws4, 8)

# ===================== SHEET 5: By Destination =====================
ws5 = wb.create_sheet('By Destination')
ws5.sheet_properties.tabColor = '7030A0'

r = 1
ws5.cell(r, 1, 'ETA 2D Accuracy by Destination (CW12)').font = title_font
ws5.merge_cells('A1:G1')

r = 3
ws5.cell(r, 1, 'By Destination Country').font = subtitle_font
r = 4
headers = ['Country', 'Total', 'Accepted', 'Failed', 'Accuracy', 'Avg Dev (h)']
for c, h in enumerate(headers, 1):
    ws5.cell(r, c, h)
style_header(ws5, r, len(headers))

for dest in data['by_dest_country']:
    r += 1
    ws5.cell(r, 1, dest['value']).font = bold_font
    ws5.cell(r, 2, dest['total'])
    ws5.cell(r, 3, dest['accepted'])
    ws5.cell(r, 4, dest['failed'])
    cell = ws5.cell(r, 5, dest['accuracy'] / 100)
    cell.number_format = pct_fmt
    color_acc(cell, dest['accuracy'] / 100)
    ws5.cell(r, 6, dest['avg_deviation_hours'])
    style_row(ws5, r, len(headers))

r += 2
ws5.cell(r, 1, 'By Delivery City (Top 20)').font = subtitle_font
r += 1
for c, h in enumerate(headers, 1):
    ws5.cell(r, c, h)
style_header(ws5, r, len(headers))

for city in data['by_delivery_city']:
    r += 1
    ws5.cell(r, 1, city['value']).font = bold_font
    ws5.cell(r, 2, city['total'])
    ws5.cell(r, 3, city['accepted'])
    ws5.cell(r, 4, city['failed'])
    cell = ws5.cell(r, 5, city['accuracy'] / 100)
    cell.number_format = pct_fmt
    color_acc(cell, city['accuracy'] / 100)
    ws5.cell(r, 6, city['avg_deviation_hours'])
    style_row(ws5, r, len(headers))

auto_width(ws5, len(headers))

# ===================== SHEET 6: By Service Type =====================
ws6 = wb.create_sheet('By Service Type')
ws6.sheet_properties.tabColor = '70AD47'

r = 1
ws6.cell(r, 1, 'ETA 2D Accuracy by Service Type (CW12)').font = title_font
ws6.merge_cells('A1:G1')
r = 3
headers = ['Service', 'Total', 'Accepted', 'Failed', 'Accuracy', 'Avg Dev (h)', 'Med Abs Dev (h)']
for c, h in enumerate(headers, 1):
    ws6.cell(r, c, h)
style_header(ws6, r, len(headers))

for svc in data['by_service']:
    r += 1
    ws6.cell(r, 1, svc['value']).font = bold_font
    ws6.cell(r, 2, svc['total'])
    ws6.cell(r, 3, svc['accepted'])
    ws6.cell(r, 4, svc['failed']).fill = red_fill
    cell = ws6.cell(r, 5, svc['accuracy'] / 100)
    cell.number_format = pct_fmt
    color_acc(cell, svc['accuracy'] / 100)
    ws6.cell(r, 6, svc['avg_deviation_hours'])
    ws6.cell(r, 7, svc['median_abs_deviation_hours'])
    style_row(ws6, r, len(headers))

auto_width(ws6, len(headers))

# ===================== SHEET 7: Failed Shipments Detail =====================
ws7 = wb.create_sheet('Failed Shipments Detail')
ws7.sheet_properties.tabColor = 'FF0000'

r = 1
ws7.cell(r, 1, f'CW12 — All Failed ETA 2D Shipments ({failed_total} outside ±48h)').font = title_font
ws7.merge_cells('A1:N1')

r = 3
headers = ['HBL', 'MBL', 'Service', 'Origin Country', 'Origin City', 'Dest Country',
           'Dest City', 'Delivery City', 'Carrier', 'ETA Baseline', 'Actual Delivered',
           'Deviation (h)', 'Deviation (d)', 'Direction']
for c, h in enumerate(headers, 1):
    ws7.cell(r, c, h)
style_header(ws7, r, len(headers))

for s in sorted(failed_ships, key=lambda x: abs(x.get('deviation_hours') or 0), reverse=True):
    r += 1
    vals = [
        s.get('hbl', ''), s.get('mbl', ''), s.get('service', ''),
        s.get('origin_country', ''), s.get('origin_city', ''),
        s.get('dest_country', ''), s.get('dest_city', ''),
        s.get('delivery_city', ''), s.get('carrier', ''),
        s.get('delivery_est', ''), s.get('delivered', ''),
        round(s['deviation_hours'], 1) if s.get('deviation_hours') else '',
        round(s['deviation_hours'] / 24, 1) if s.get('deviation_hours') else '',
        s.get('direction', ''),
    ]
    for c, v in enumerate(vals, 1):
        ws7.cell(r, c, v)
    style_row(ws7, r, len(headers))
    if s.get('direction') == 'early':
        ws7.cell(r, 14).fill = blue_fill
    elif s.get('direction') == 'late':
        ws7.cell(r, 14).fill = orange_fill

auto_width(ws7, len(headers))

# ===================== SHEET 8: Preventive Actions =====================
ws8 = wb.create_sheet('Preventive Actions')
ws8.sheet_properties.tabColor = '00B050'

r = 1
ws8.cell(r, 1, 'Preventive Actions & Roadmap to 90%+ ETA 2D Accuracy').font = title_font
ws8.merge_cells('A1:F1')

r = 3
headers = ['Priority', 'Action', 'Root Cause Addressed', 'Expected Impact', 'Owner', 'Timeline']
for c, h in enumerate(headers, 1):
    ws8.cell(r, c, h)
style_header(ws8, r, len(headers))

actions = [
    ('P1 - Critical',
     'Implement dynamic ETA update post-ATA:\nRecalculate S31 ETA after vessel arrival (S07) incorporating customs clearance and inland transit estimates per destination.',
     'RC#1: Stale ETA baseline\nRC#4: No post-ATA update',
     '+15-20pp accuracy\n(fixes ~85 early + ~50 near-miss late)',
     'Bosch IT / EDI Team', 'CW16-CW20'),
    ('P1 - Critical',
     'Calibrate last-mile transit buffers per destination cluster:\n- Germany: avg 12d last mile\n- Hungary (Miskolc/Hatvan): avg 11d\n- Portugal (Braga): avg 10d',
     'RC#3: Systematic estimate bias on HU/PT/DE lanes',
     '+10-15pp accuracy\n(fixes ~67 lane-specific failures)',
     'Maersk Operations', 'CW14-CW18'),
    ('P2 - High',
     'Investigate CN->DE LCL consolidation delays:\n80 failures with avg +15.8 days late. Root cause likely at origin CFS or transshipment.',
     'RC#2: CN->DE consolidation delays',
     '+8-10pp accuracy\n(80 shipments, largest single lane)',
     'Origin Operations / CFS', 'CW14-CW16'),
    ('P2 - High',
     'Add ETA measurement checkpoint at customs clearance (S16):\nCapture updated delivery ETA after customs release for more accurate last-mile estimate.',
     'RC#1 + RC#4',
     '+5-8pp accuracy',
     'Bosch IT / Maersk IT', 'CW18-CW22'),
    ('P3 - Medium',
     'Lane-specific ETA models:\nBuild historical transit time distributions per origin-dest pair to replace static ETAs with data-driven estimates.',
     'All root causes',
     'Long-term: 85-90%+ achievable',
     'Data & Analytics', 'CW20-CW30'),
    ('P3 - Medium',
     'Negotiate measurement window review:\nPresent data showing even ±7d = 73.4%, ±14d = 86.8%. Discuss ±72h or phased targets with Bosch.',
     'Structural: ±48h too tight for ocean freight door delivery',
     'Resets achievable target',
     'Account Management', 'CW14'),
    ('P4 - Quick Win',
     'Fix Immenstadt (DE) deliveries:\n9 shipments avg +58 days deviation — specific warehouse or customs issue.',
     'RC#2: Specific consignee',
     '+2pp (9 ships)',
     'DE Operations', 'CW13-CW14'),
    ('P4 - Quick Win',
     'Fix DE->TH / DE->JP reverse lanes:\n17 shipments at 0% accuracy — investigate if different ETA methodology.',
     'Possible: Different ETA source',
     '+4pp (17 ships)',
     'Operations', 'CW13-CW14'),
]

for action in actions:
    r += 1
    for c, v in enumerate(action, 1):
        ws8.cell(r, c, v).font = data_font
        ws8.cell(r, c).border = thin_border
        ws8.cell(r, c).alignment = Alignment(wrap_text=True, vertical='top')
    prio = action[0]
    if 'P1' in prio:
        ws8.cell(r, 1).fill = red_fill
    elif 'P2' in prio:
        ws8.cell(r, 1).fill = yellow_fill
    elif 'P3' in prio:
        ws8.cell(r, 1).fill = blue_fill
    else:
        ws8.cell(r, 1).fill = green_fill

ws8.column_dimensions['A'].width = 16
ws8.column_dimensions['B'].width = 55
ws8.column_dimensions['C'].width = 30
ws8.column_dimensions['D'].width = 24
ws8.column_dimensions['E'].width = 22
ws8.column_dimensions['F'].width = 14
for row in ws8.iter_rows(min_row=4, max_row=r):
    ws8.row_dimensions[row[0].row].height = 50

out_path = f"{BASE}/ETA_2D_RCA_Analysis_CW12.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
print("8 sheets: Executive Summary | Top 10 Failure Lanes | What-If Window | Early vs Late | By Destination | By Service Type | Failed Shipments Detail | Preventive Actions")
