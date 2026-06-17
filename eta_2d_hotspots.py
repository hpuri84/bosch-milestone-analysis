"""
ETA 2D (Door, T-7 ETA snapshot +/-48h) hotspot extraction.

Produces a per-shipment dataset across all weeks plus dimensional rollups with
rolling averages, for the Bosch ETA-2D pushback analysis.

Canonical rule (mirrors rebaseline.py compute_eta_and_ref):
  - Denominator = delivered shipments (DELIVERED_DATE_TIME present)
  - Accepted    = S31 T-7 snapshot present AND |S31 deviation| <= 48h
  - no_t7       = delivered but no S31 T-7 snapshot (counts as a miss)
Self-reconciles per week against kpi_data.json (s31_acc / s31_total / s31_no_t7).
"""
import openpyxl, json
from collections import defaultdict
from datetime import datetime

WEEKS = [f"CW{i:02d}" for i in range(1, 23)]
SC4_FILES = {f"CW{i:02d}": f"Bosch Milestone raw data/Maersk SC4_2026_CW{i:02d}.xlsx" for i in range(1, 23)}
WINDOW_H = 48
ROLL = 4
TARGET = 0.90          # assumed contractual ETA-2D target (per prior analysis; confirm vs contract)
RECENT = ["CW19", "CW20", "CW21", "CW22"]   # current-state window for hotspot ranking
MIN_VOL = 15           # min pooled delivered volume for a group to rank as a hotspot


def find_shipments_sheet(wb):
    for sn in wb.sheetnames:
        if sn.strip().lower() == "shipments":
            return sn
    for sn in wb.sheetnames:
        if sn.strip().upper() in ("TOTAL", "ALL") or sn.strip().upper().rstrip("_") in ("FCL", "BCO", "LCL"):
            continue
        ws = wb[sn]
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            if row and any(c and "UNIQUE_SHIPMENT_ID" in str(c) for c in row):
                return sn
    return None


def detect_header_row(ws):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), 1):
        if row and any(c and "UNIQUE_SHIPMENT_ID" in str(c) for c in row):
            return i
    return 3


def dt(v):
    return v if isinstance(v, datetime) else None


def parse_week(path, week):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sn = find_shipments_sheet(wb)
    if not sn:
        wb.close()
        return []
    ws = wb[sn]
    hr = detect_header_row(ws)
    headers = list(next(ws.iter_rows(min_row=hr, max_row=hr, values_only=True)))
    col = {}
    for i, h in enumerate(headers):
        if h is not None and str(h).strip():
            col[str(h).strip()] = i

    def gi(name):
        return col.get(name)

    c = dict(
        hbl=gi("HOUSE_BILL_OF_LADING"), mbl=gi("MASTER_BILL_OF_LADING"), uid=gi("UNIQUE_SHIPMENT_ID"),
        carrier=gi("CARRIER_1"), svc=gi("TRANSPORT_SERVICE_PRIORITY"),
        ocity=gi("CONSIGNOR_ADDRESS_CITY_NAME"), oco=gi("CONSIGNOR_ADDRESS_COUNTRY"),
        dcity=gi("CONSIGNEE_ADDRESS_CITY_NAME"), dco=gi("CONSIGNEE_ADDRESS_COUNTRY"),
        delcity=gi("DELIVERY_ADDRESS_CITY_NAME"), delco=gi("DELIVERY_ADDRESS_COUNTRY"),
        eta=gi("ETA_DATE_TIME"), ata=gi("ATA_DATE_TIME"), deliv=gi("DELIVERED_DATE_TIME"),
        coll=gi("COLLECTED_DATE_TIME"), del_est=gi("DELIVERY_DATE_ACT_EST_PLAN"),
    )
    c_mode = c_dreason = c_dsupp = c_s31_ts = c_s31_dev = None
    for k, idx in col.items():
        if "TRANSPORT_MODE" in k:
            c_mode = idx
        if "Delay_Reason_Code_Description" in k:
            c_dreason = idx
        if "Reason_Supplier_Responsible" in k:
            c_dsupp = idx
        if "S31_TS_@measured" in k:
            c_s31_ts = idx
        elif "S31_Deviation_TS_@measured" in k:
            c_s31_dev = idx

    out = []
    for row in ws.iter_rows(min_row=hr + 1, values_only=True):
        if not row or row[0] is None:
            continue

        def g(ci):
            return row[ci] if ci is not None and ci < len(row) else None

        def s(ci):
            v = g(ci)
            return str(v).strip() if v is not None and str(v).strip() else ""

        delivered = g(c["deliv"])
        if delivered is None:
            continue                      # 2D denominator = delivered shipments only

        ts = g(c_s31_ts)
        dev = g(c_s31_dev)
        t7 = ts not in (None, "")
        dev_h = round(float(dev), 1) if isinstance(dev, (int, float)) else None
        if not t7:
            bucket, accepted = "no_t7", False
        elif dev_h is None:
            bucket, accepted = "measured_no_dev", False
        elif abs(dev_h) <= WINDOW_H:
            bucket, accepted = "accepted", True
        else:
            bucket, accepted = ("late_gt48" if dev_h > 0 else "early_gt48"), False

        eta, ata, dl = dt(g(c["eta"])), dt(g(c["ata"])), dt(delivered)
        port_late_h = round((ata - eta).total_seconds() / 3600, 1) if eta and ata else None
        last_mile_d = round((dl - ata).total_seconds() / 86400, 1) if ata and dl else None
        oco, dco = s(c["oco"]).upper(), s(c["dco"]).upper()

        out.append({
            "week": week, "hbl": s(c["hbl"]), "mbl": s(c["mbl"]),
            "service": s(c["svc"]), "carrier": s(c["carrier"]),
            "origin_city": s(c["ocity"]), "origin_country": oco,
            "dest_city": s(c["dcity"]), "dest_country": dco,
            "delivery_city": s(c["delcity"]), "delivery_country": s(c["delco"]).upper(),
            "lane": f"{oco}->{dco}" if oco and dco else "",
            "t7_present": t7, "dev_hours": dev_h,
            "dev_days": round(dev_h / 24, 1) if dev_h is not None else None,
            "direction": ("late" if dev_h and dev_h > 0 else "early" if dev_h and dev_h < 0 else None),
            "accepted": accepted, "bucket": bucket,
            "delay_reason": s(c_dreason), "supplier_responsible": s(c_dsupp),
            "port_late_h": port_late_h, "last_mile_d": last_mile_d,
            "delivered": str(dl)[:16] if dl else None,
            "eta": str(eta)[:16] if eta else None, "ata": str(ata)[:16] if ata else None,
        })
    wb.close()
    return out


def label_city(r):
    city = (r["dest_city"] or "(blank)").title()
    return f"{city} ({r['dest_country'] or '?'})"


def rate(acc, tot):
    return round(acc / tot, 4) if tot else None


def rollup(ships, keyfn, label=None):
    by = defaultdict(lambda: defaultdict(lambda: {"total": 0, "acc": 0, "no_t7": 0,
                                                  "late": 0, "early": 0, "mnd": 0, "dev_sum": 0.0, "dev_n": 0}))
    disp = {}
    for r in ships:
        k = keyfn(r)
        if not k or k in ("", "?", "(blank) (?)", "-> "):
            continue
        if label:
            disp[k] = label(r)
        cell = by[k][r["week"]]
        cell["total"] += 1
        cell["acc"] += 1 if r["accepted"] else 0
        if r["bucket"] == "no_t7":
            cell["no_t7"] += 1
        elif r["bucket"] == "late_gt48":
            cell["late"] += 1
        elif r["bucket"] == "early_gt48":
            cell["early"] += 1
        elif r["bucket"] == "measured_no_dev":
            cell["mnd"] += 1
        if r["dev_days"] is not None:
            cell["dev_sum"] += r["dev_days"]
            cell["dev_n"] += 1

    groups = []
    for k, weeks in by.items():
        tot = sum(w["total"] for w in weeks.values())
        acc = sum(w["acc"] for w in weeks.values())
        no_t7 = sum(w["no_t7"] for w in weeks.values())
        late = sum(w["late"] for w in weeks.values())
        early = sum(w["early"] for w in weeks.values())
        dev_sum = sum(w["dev_sum"] for w in weeks.values())
        dev_n = sum(w["dev_n"] for w in weeks.values())
        rec = [weeks[w] for w in RECENT if w in weeks]
        rtot = sum(w["total"] for w in rec)
        racc = sum(w["acc"] for w in rec)
        # per-week series + rolling avg of acc_rate
        series = []
        for w in WEEKS:
            if w in weeks:
                wc = weeks[w]
                series.append({"week": w, "total": wc["total"], "acc": wc["acc"],
                               "rate": rate(wc["acc"], wc["total"]), "no_t7": wc["no_t7"]})
        roll = []
        for i in range(len(series)):
            window = series[max(0, i - ROLL + 1): i + 1]
            wt = sum(x["total"] for x in window)
            wa = sum(x["acc"] for x in window)
            roll.append({"week": series[i]["week"], "roll_rate": rate(wa, wt), "roll_vol": wt})
        groups.append({
            "key": k, "label": disp.get(k, k),
            "total": tot, "acc": acc, "failed": tot - acc, "rate": rate(acc, tot),
            "no_t7": no_t7, "no_t7_rate": rate(no_t7, tot),
            "late_gt48": late, "early_gt48": early,
            "snapshot_but_miss": tot - acc - no_t7,
            "avg_dev_days": round(dev_sum / dev_n, 1) if dev_n else None,
            "recent_total": rtot, "recent_acc": racc, "recent_rate": rate(racc, rtot),
            "gap_vs_target": round(TARGET - rate(acc, tot), 4) if tot else None,
            "series": series, "rolling": roll,
        })
    groups.sort(key=lambda g: (g["rate"] if g["rate"] is not None else 1, -g["total"]))
    return groups


def responsibility(ships):
    """Decompose every delivered shipment + the >48h misses into attribution buckets."""
    tot = len(ships)
    b = defaultdict(int)
    for r in ships:
        b[r["bucket"]] += 1
    # late-miss attribution: arrived late at port (upstream) vs on-time-port-but-late-door (last mile)
    late_miss = [r for r in ships if r["bucket"] in ("late_gt48",) and r["port_late_h"] is not None]
    upstream = sum(1 for r in late_miss if r["port_late_h"] > WINDOW_H)
    lastmile = sum(1 for r in late_miss if r["port_late_h"] <= WINDOW_H)
    # delay-reason tally (populated rows only)
    reasons = defaultdict(int)
    supp_y = 0
    for r in ships:
        if r["delay_reason"]:
            reasons[r["delay_reason"]] += 1
        if (r["supplier_responsible"] or "").upper().startswith("Y"):
            supp_y += 1
    return {
        "total_delivered": tot,
        "buckets": dict(b),
        "no_t7_share": rate(b["no_t7"], tot),
        "accepted_share": rate(b["accepted"], tot),
        "miss_with_snapshot": tot - b["accepted"] - b["no_t7"],
        "late_miss_attribution": {
            "coverage": len(late_miss), "upstream_port": upstream, "last_mile": lastmile,
            "upstream_share": rate(upstream, len(late_miss)),
        },
        "delay_reasons": dict(sorted(reasons.items(), key=lambda x: -x[1])),
        "supplier_responsible_Y": supp_y,
    }


def main():
    all_ships = []
    per_week = {}
    for wk in WEEKS:
        path = SC4_FILES[wk]
        ships = parse_week(path, wk)
        all_ships.extend(ships)
        acc = sum(1 for r in ships if r["accepted"])
        no_t7 = sum(1 for r in ships if r["bucket"] == "no_t7")
        per_week[wk] = {"total": len(ships), "acc": acc, "no_t7": no_t7,
                        "rate": rate(acc, len(ships)),
                        "t7_comp": rate(len(ships) - no_t7, len(ships))}

    # ---- reconcile against kpi_data.json ----
    kpi = {r["week"]: r for r in json.load(open("kpi_data.json"))}
    print(f"{'week':6} {'mine acc/tot/no_t7':22} {'kpi acc/tot/no_t7':22} {'match'}")
    ok = True
    for wk in WEEKS:
        m, k = per_week[wk], kpi.get(wk, {})
        match = (m["total"] == k.get("s31_total") and m["acc"] == k.get("s31_acc") and m["no_t7"] == k.get("s31_no_t7"))
        ok = ok and match
        print(f"{wk:6} {m['acc']}/{m['total']}/{m['no_t7']:<16} "
              f"{k.get('s31_acc')}/{k.get('s31_total')}/{k.get('s31_no_t7'):<16} {'OK' if match else 'MISMATCH'}")
    print("RECONCILIATION:", "PASS" if ok else "FAIL")

    rollups = {
        "meta": {"weeks": WEEKS, "window_h": WINDOW_H, "roll": ROLL, "target": TARGET,
                 "recent_weeks": RECENT, "min_vol": MIN_VOL, "reconciled": ok,
                 "total_delivered": len(all_ships)},
        "trend": [{"week": wk, **per_week[wk]} for wk in WEEKS],
        "by_dest_city": rollup(all_ships, label_city, label_city),
        "by_dest_country": rollup(all_ships, lambda r: r["dest_country"]),
        "by_lane": rollup(all_ships, lambda r: r["lane"]),
        "by_origin_country": rollup(all_ships, lambda r: r["origin_country"]),
        "by_carrier": rollup(all_ships, lambda r: r["carrier"]),
        "by_service": rollup(all_ships, lambda r: r["service"]),
        "responsibility": responsibility(all_ships),
    }
    json.dump(rollups, open("eta_2d_rollups.json", "w"), indent=1, default=str)
    json.dump(all_ships, open("eta_2d_shipments.json", "w"), indent=1, default=str)
    print(f"\nWrote eta_2d_rollups.json ({len(all_ships)} delivered shipments) + eta_2d_shipments.json")
    # quick headline
    print("\nWorst dest-city hotspots (pooled, vol>=%d):" % MIN_VOL)
    for g in [x for x in rollups["by_dest_city"] if x["total"] >= MIN_VOL][:12]:
        print(f"  {g['label']:28} n={g['total']:4} acc={g['rate']*100:5.1f}%  no_t7={g['no_t7_rate']*100:4.0f}%  avgdev={g['avg_dev_days']}d")


if __name__ == "__main__":
    main()
