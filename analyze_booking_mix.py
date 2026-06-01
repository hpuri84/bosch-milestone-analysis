#!/usr/bin/env python3
"""Booking Mix analysis: SC3 vs SC4 shipment-count split per week vs 80/20 target.

SC4 = ocean (Asia->Europe). SC3 = NGTM road/inland legs. Different datasets.
Headline mix = share of shipment COUNT that is SC3 vs SC4 per week.

Origin/lane columns:
  SC3 origin  -> Leg_Pick_up_Country  ; SC3 dest -> Leg_Delivery_Country
  SC4 origin  -> CONSIGNOR_ADDRESS_COUNTRY ; SC4 dest -> CONSIGNEE_ADDRESS_COUNTRY
"""
import json
import os
from datetime import datetime, timezone

import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
RAW = os.path.join(BASE, "Bosch Milestone raw data")
OUT = os.path.join(BASE, "dashboard", "public", "booking_mix.json")

WEEKS = [f"CW{i:02d}" for i in range(1, 21)]
TARGET_SC3 = 0.80
TARGET_SC4 = 0.20


def sc3_path(week):
    n = int(week[2:])
    name = f"Maersk NGTM SC3_2026_{week}.xlsx" if n <= 9 else f"Maersk SC3_2026_{week}.xlsx"
    return os.path.join(RAW, name)


def sc4_path(week):
    return os.path.join(RAW, f"Maersk SC4_2026_{week}.xlsx")


def find_shipments_sheet(wb):
    # 1. by name (case-insensitive, strip trailing spaces)
    for ws in wb.worksheets:
        if ws.title.strip().lower() == "shipments":
            return ws
    # 2. scan first 5 rows of each sheet for a marker
    for ws in wb.worksheets:
        for r in range(1, 6):
            row = next(ws.iter_rows(min_row=r, max_row=r, values_only=True), ())
            for v in row:
                if isinstance(v, str):
                    u = v.upper()
                    if "UNIQUE_SHIPMENT" in u or u == "LOAD_ID":
                        return ws
    return wb.worksheets[0]


def detect_header_row(ws):
    # SC4 has UNIQUE_SHIPMENT_ID; SC3 starts with LOAD_ID. Fallback row 3.
    for r in range(1, 6):
        row = next(ws.iter_rows(min_row=r, max_row=r, values_only=True), ())
        for v in row:
            if isinstance(v, str):
                u = v.upper()
                if "UNIQUE_SHIPMENT" in u or u == "LOAD_ID":
                    return r
    return 3


def col_index(headers, substr):
    for i, h in enumerate(headers):
        if isinstance(h, str) and substr.lower() in h.lower():
            return i
    return None


def analyze_file(path, origin_substr, dest_substr):
    """Return (row_count, origin_counts, dest_counts, lane_counts)."""
    if not os.path.exists(path):
        return None, {}, {}, {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = find_shipments_sheet(wb)
        hdr_row = detect_header_row(ws)
        headers = list(next(ws.iter_rows(min_row=hdr_row, max_row=hdr_row, values_only=True), ()))
        o_idx = col_index(headers, origin_substr)
        d_idx = col_index(headers, dest_substr)
        count = 0
        origins = {}
        dests = {}
        lanes = {}
        for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
            if not row:
                continue
            first = row[0]
            if first in (None, ""):
                continue
            count += 1
            # origin
            orig = None
            if o_idx is not None and o_idx < len(row) and row[o_idx] not in (None, ""):
                orig = str(row[o_idx]).strip().upper()
                origins[orig] = origins.get(orig, 0) + 1
            # dest
            dest = None
            if d_idx is not None and d_idx < len(row) and row[d_idx] not in (None, ""):
                dest = str(row[d_idx]).strip().upper()
                dests[dest] = dests.get(dest, 0) + 1
            # lane = origin → dest (only when both known)
            if orig and dest:
                lane = f"{orig} → {dest}"
                lanes[lane] = lanes.get(lane, 0) + 1
        return count, origins, dests, lanes
    finally:
        wb.close()


def merge_country_rows(sc3_c, sc4_c):
    all_c = set(sc3_c) | set(sc4_c)
    rows = []
    for c in all_c:
        a = sc3_c.get(c, 0)
        b = sc4_c.get(c, 0)
        t = a + b
        rows.append({
            "country": c,
            "sc3": a,
            "sc4": b,
            "total": t,
            "sc3_share": round(a / t, 4) if t else 0,
        })
    rows.sort(key=lambda x: x["total"], reverse=True)
    return rows


def merge_lane_rows(sc3_lanes, sc4_lanes, top_n=30):
    all_l = set(sc3_lanes) | set(sc4_lanes)
    rows = []
    for lane in all_l:
        a = sc3_lanes.get(lane, 0)
        b = sc4_lanes.get(lane, 0)
        t = a + b
        rows.append({
            "lane": lane,
            "sc3": a,
            "sc4": b,
            "total": t,
            "sc3_share": round(a / t, 4) if t else 0,
        })
    rows.sort(key=lambda x: x["total"], reverse=True)
    return rows[:top_n]


def main():
    weekly = []
    by_dest = {}
    by_origin = {}
    by_lane = {}
    diag = {}

    for week in WEEKS:
        p3, p4 = sc3_path(week), sc4_path(week)
        if not (os.path.exists(p3) or os.path.exists(p4)):
            continue

        sc3, sc3_orig, sc3_dest, sc3_lanes = analyze_file(
            p3, "Leg_Pick_up_Country", "Leg_Delivery_Country"
        )
        sc4, sc4_orig, sc4_dest, sc4_lanes = analyze_file(
            p4, "CONSIGNOR_ADDRESS_COUNTRY", "CONSIGNEE_ADDRESS_COUNTRY"
        )
        sc3 = sc3 or 0
        sc4 = sc4 or 0
        total = sc3 + sc4
        if total == 0:
            continue

        sc3_share = sc3 / total
        sc4_share = sc4 / total
        weekly.append({
            "week": week,
            "sc3": sc3,
            "sc4": sc4,
            "total": total,
            "sc3_share": round(sc3_share, 4),
            "sc4_share": round(sc4_share, 4),
            "gap_pp": round(sc3_share * 100 - 80, 1),
        })
        diag[week] = (sc3, sum(sc3_orig.values()))

        by_dest[week] = merge_country_rows(sc3_dest, sc4_dest)
        by_origin[week] = merge_country_rows(sc3_orig, sc4_orig)
        by_lane[week] = merge_lane_rows(sc3_lanes, sc4_lanes)

    # trailing 4-week avg sc3_share
    trailing4 = {}
    for i, w in enumerate(weekly):
        window = weekly[max(0, i - 3): i + 1]
        avg = sum(x["sc3_share"] for x in window) / len(window)
        trailing4[w["week"]] = round(avg, 4)

    latest_week = weekly[-1]["week"] if weekly else None

    notes = [
        "Booking mix is measured by shipment COUNT, not container or volume.",
        "Weekly files are activity snapshots, not cumulative bookings — mix is volatile week to week.",
        "CW20 SC3 (200 rows) and CW18 SC4 (208 rows) appear to be PARTIAL extracts vs typical ~400-650; treat single-week values as directional. Trailing-average is more reliable.",
        "SC3 origin = Leg_Pick_up_Country; SC4 origin = CONSIGNOR_ADDRESS_COUNTRY. These use different sourcing logic — lane-level comparisons are directional.",
        "SC4 = ocean freight (Asia origin -> Europe). SC3 = NGTM road/inland legs (mostly within Europe or from CN/KR via road).",
    ]

    payload = {
        "generated": datetime.now(timezone.utc).isoformat(),
        "target_sc3": TARGET_SC3,
        "target_sc4": TARGET_SC4,
        "weekly": weekly,
        "trailing4_sc3_share": trailing4,
        "by_dest_country": by_dest,
        "by_origin_country": by_origin,
        "by_lane": by_lane,
        "latest_week": latest_week,
        "notes": notes,
    }
    with open(OUT, "w") as f:
        json.dump(payload, f, indent=2)

    print(f"Wrote {OUT}\n")
    print(f"{'WEEK':6} {'SC3':>6} {'SC4':>6} {'TOTAL':>7} {'SC3%':>7} {'gap_pp':>7} {'trail4%':>8}")
    for w in weekly:
        print(f"{w['week']:6} {w['sc3']:>6} {w['sc4']:>6} {w['total']:>7} "
              f"{w['sc3_share']*100:>6.1f}% {w['gap_pp']:>7} "
              f"{trailing4[w['week']]*100:>7.1f}%")

    if latest_week:
        print(f"\nTop 10 origin countries for {latest_week}:")
        for r in by_origin[latest_week][:10]:
            print(f"  {r['country']:4} sc3={r['sc3']:>4} sc4={r['sc4']:>4} "
                  f"total={r['total']:>4} sc3%={r['sc3_share']*100:>5.1f}%")
        print(f"\nTop 10 lanes for {latest_week}:")
        for r in by_lane[latest_week][:10]:
            print(f"  {r['lane']:20} sc3={r['sc3']:>4} sc4={r['sc4']:>4} "
                  f"total={r['total']:>4} sc3%={r['sc3_share']*100:>5.1f}%")


if __name__ == "__main__":
    main()
