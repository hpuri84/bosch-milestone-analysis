"""
Validate KPI calculations — v4 with corrected business logic:
  - Critical = S00(SC4 only), S02, S04, S07, S31 (Act+Est) — NO S45
  - All = ALL milestones from both SC3 and SC4 (including S00, S02)
  - Completeness = sum(Available) / sum(Required)
  - Timeliness = sum(In_Time) / sum(Available)
  - ETA 2P = S07_Accepted / count(S07 measured)
  - ETA 2D = S31_Accepted / count(S31 measured)
"""

import openpyxl
import os

BASE = "/Users/harsh.puri/Documents/work-maersk/Prototype Playground/Bosch Milestone Analysis"
RAW_DIR = os.path.join(BASE, "Bosch Milestone raw data")

SC3_FILES = {
    "CW01": "Maersk NGTM SC3_2026_CW01.xlsx",
    "CW02": "Maersk NGTM SC3_2026_CW02.xlsx",
    "CW03": "Maersk NGTM SC3_2026_CW03.xlsx",
    "CW04": "Maersk NGTM SC3_2026_CW04.xlsx",
    "CW05": "Maersk NGTM SC3_2026_CW05.xlsx",
    "CW06": "Maersk NGTM SC3_2026_CW06.xlsx",
    "CW07": "Maersk NGTM SC3_2026_CW07.xlsx",
}

SC4_FILES = {
    "CW01": "Maersk SC4_2026_CW01.xlsx",
    "CW02": "Maersk SC4_2026_CW02v2.xlsx",
    "CW03": "Maersk SC4_2026_CW03.xlsx",
    "CW04": "Maersk SC4_2026_CW04.xlsx",
    "CW05": "Maersk SC4_2026_CW05.xlsx",
    "CW06": "Maersk SC4_2026_CW06.xlsx",
    "CW07": "Maersk SC4_2026_CW07.xlsx",
}

SC3_CRITICAL_CODES = {"S02", "S04", "S07", "S31"}
SC4_CRITICAL_CODES = {"S00", "S02", "S04", "S07", "S31"}


def find_total_sheet(wb):
    for sn in wb.sheetnames:
        s = sn.strip().upper()
        if s.startswith("TOTAL") or s == "ALL":
            return sn
    return None


def find_shipments_sheet(wb):
    for sn in wb.sheetnames:
        if sn.strip().lower() == "shipments":
            return sn
    return None


def parse_milestone_code(name):
    if name and " - " in str(name):
        return str(name).split(" - ")[0].strip()
    return str(name).strip() if name else None


def read_summary_sheet(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = find_total_sheet(wb)
    if not sheet_name:
        print(f"  WARNING: No TOTAL sheet in {os.path.basename(filepath)}")
        wb.close()
        return []

    ws = wb[sheet_name]
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        for cell in row:
            if cell and "required" in str(cell).lower() and "status" in str(cell).lower():
                header_row = i
                break
        if header_row:
            break
    if not header_row:
        header_row = 3

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        milestone_name = row[0]
        if not milestone_name or "required" in str(milestone_name).lower():
            continue
        code = parse_milestone_code(milestone_name)
        status_type = str(row[1]).strip() if row[1] else ""
        required = row[3] if row[3] and isinstance(row[3], (int, float)) else 0
        available = row[4] if row[4] and isinstance(row[4], (int, float)) else 0
        in_time = row[5] if row[5] and isinstance(row[5], (int, float)) else 0

        rows.append({
            "code": code,
            "type": status_type,
            "required": required,
            "available": available,
            "in_time": in_time,
        })
    wb.close()
    return rows


def read_sc4_shipments(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = find_shipments_sheet(wb)
    if not sheet_name:
        wb.close()
        return []

    ws = wb[sheet_name]
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        for cell in row:
            if cell and "UNIQUE_SHIPMENT" in str(cell):
                header_row = i
                break
        if header_row:
            break
    if not header_row:
        header_row = 3

    headers = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            col_map[str(h).strip()] = i

    shipments = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        vals = list(row)
        if not vals[0]:
            continue
        shipment = {}
        for name, idx in col_map.items():
            shipment[name] = vals[idx] if idx < len(vals) else None
        shipments.append(shipment)

    wb.close()
    return shipments


# PDCA reference values (old, with S45 included — shown for comparison)
PDCA_OLD = {
    "Comp_Crit": {"W5": 0.8277, "W6": 0.8168, "W7": 0.8315},
    "Time_Crit": {"W5": 0.5885, "W6": 0.6012, "W7": 0.6415},
    "Comp_All":  {"W5": 0.7481, "W6": 0.7585, "W7": 0.7838},
    "Time_All":  {"W5": 0.5996, "W6": 0.6140, "W7": 0.6257},
    "ETA_2P":    {"W5": 0.44,   "W6": 0.62,   "W7": 0.39},
    "ETA_2D":    {"W5": 0.13,   "W6": 0.17,   "W7": 0.17},
    "Ref_Comp":  {"W5": 0.71,   "W6": 0.74,   "W7": 0.76},
}


def fmt_pct(val):
    return f"{val:.2%}"


def main():
    print("=" * 110)
    print("BOSCH MILESTONE KPI VALIDATION v4 — Corrected Business Logic")
    print("  Completeness = sum(Available) / sum(Required)")
    print("  Timeliness   = sum(In_Time) / sum(Available)")
    print("  Critical     = S00(SC4 only), S02, S04, S07, S31 (Act+Est) — NO S45")
    print("  All          = ALL milestones from SC3 + SC4 (including S00, S02)")
    print("=" * 110)

    results = {}

    for week in ["CW01", "CW02", "CW03", "CW04", "CW05", "CW06", "CW07"]:
        wkey = week.replace("CW0", "W")

        sc3_file = os.path.join(RAW_DIR, SC3_FILES[week])
        sc4_file = os.path.join(RAW_DIR, SC4_FILES[week])

        sc3_data = read_summary_sheet(sc3_file)
        sc4_data = read_summary_sheet(sc4_file)

        # === CRITICAL (no S45) ===
        sc3_crit = [r for r in sc3_data if r["code"] in SC3_CRITICAL_CODES]
        sc4_crit = [r for r in sc4_data if r["code"] in SC4_CRITICAL_CODES]
        crit = sc3_crit + sc4_crit

        crit_req = sum(r["required"] for r in crit)
        crit_avl = sum(r["available"] for r in crit)
        crit_it = sum(r["in_time"] for r in crit)
        crit_comp = crit_avl / crit_req if crit_req > 0 else 0
        crit_time = crit_it / crit_avl if crit_avl > 0 else 0

        # === ALL (everything) ===
        all_rows = sc3_data + sc4_data
        all_req = sum(r["required"] for r in all_rows)
        all_avl = sum(r["available"] for r in all_rows)
        all_it = sum(r["in_time"] for r in all_rows)
        all_comp = all_avl / all_req if all_req > 0 else 0
        all_time = all_it / all_avl if all_avl > 0 else 0

        # === ETA & Reference ===
        shipments = read_sc4_shipments(sc4_file)

        s07_col = s31_col = civ_col = dn_col = None
        for key in (shipments[0].keys() if shipments else []):
            if "S07_Accepted" in key: s07_col = key
            if "S31_Accepted" in key: s31_col = key
            if "COMMERCIAL_INVOICE" in key: civ_col = key
            if "DELIVERY_NOTE" in key: dn_col = key

        s07_total = sum(1 for s in shipments if s.get(s07_col) is not None) if s07_col else 0
        s07_acc = sum(1 for s in shipments if s.get(s07_col) == 1) if s07_col else 0
        eta_2p = s07_acc / s07_total if s07_total > 0 else None

        s31_total = sum(1 for s in shipments if s.get(s31_col) is not None) if s31_col else 0
        s31_acc = sum(1 for s in shipments if s.get(s31_col) == 1) if s31_col else 0
        eta_2d = s31_acc / s31_total if s31_total > 0 else None

        ref_total = len(shipments)
        ref_complete = 0
        for s in shipments:
            civ = s.get(civ_col) if civ_col else None
            dn = s.get(dn_col) if dn_col else None
            has_civ = civ is not None and str(civ).strip() not in ("", "None", "NA")
            has_dn = dn is not None and str(dn).strip() not in ("", "None", "NA")
            if has_civ or has_dn:
                ref_complete += 1
        ref_comp = ref_complete / ref_total if ref_total > 0 else None

        results[week] = {
            "crit_comp": crit_comp, "crit_time": crit_time,
            "crit_req": crit_req, "crit_avl": crit_avl, "crit_it": crit_it,
            "all_comp": all_comp, "all_time": all_time,
            "all_req": all_req, "all_avl": all_avl, "all_it": all_it,
            "eta_2p": eta_2p, "s07_acc": s07_acc, "s07_total": s07_total,
            "eta_2d": eta_2d, "s31_acc": s31_acc, "s31_total": s31_total,
            "ref_comp": ref_comp, "ref_complete": ref_complete, "ref_total": ref_total,
            "sc3_crit_count": len(sc3_crit), "sc4_crit_count": len(sc4_crit),
        }

    # Print summary table
    weeks = ["CW01", "CW02", "CW03", "CW04", "CW05", "CW06", "CW07"]
    wkeys = [w.replace("CW0", "W") for w in weeks]

    print(f"\n{'KPI':<40s} | {'CW01':>8s} | {'CW02':>8s} | {'CW03':>8s} | {'CW04':>8s} | {'CW05':>8s} | {'CW06':>8s} | {'CW07':>8s}")
    print("─" * 110)

    # Completeness Critical
    vals = [fmt_pct(results[w]["crit_comp"]) for w in weeks]
    print(f"{'Completeness (Critical)':<40s} | {'  |  '.join(vals)}")

    # vs PDCA old
    pdca_vals = ["" if w not in ["CW05","CW06","CW07"] else fmt_pct(PDCA_OLD["Comp_Crit"][w.replace("CW0","W")]) for w in weeks]
    print(f"{'  └─ PDCA (old, with S45)':<40s} | {'  |  '.join(v.rjust(8) for v in pdca_vals)}")

    # Timeliness Critical
    vals = [fmt_pct(results[w]["crit_time"]) for w in weeks]
    print(f"{'Timeliness (Critical)':<40s} | {'  |  '.join(vals)}")

    pdca_vals = ["" if w not in ["CW05","CW06","CW07"] else fmt_pct(PDCA_OLD["Time_Crit"][w.replace("CW0","W")]) for w in weeks]
    print(f"{'  └─ PDCA (old, with S45)':<40s} | {'  |  '.join(v.rjust(8) for v in pdca_vals)}")

    print("─" * 110)

    # Completeness All
    vals = [fmt_pct(results[w]["all_comp"]) for w in weeks]
    print(f"{'Completeness (All)':<40s} | {'  |  '.join(vals)}")

    pdca_vals = ["" if w not in ["CW05","CW06","CW07"] else fmt_pct(PDCA_OLD["Comp_All"][w.replace("CW0","W")]) for w in weeks]
    print(f"{'  └─ PDCA (old, excl S00/S02 SC4)':<40s} | {'  |  '.join(v.rjust(8) for v in pdca_vals)}")

    # Timeliness All
    vals = [fmt_pct(results[w]["all_time"]) for w in weeks]
    print(f"{'Timeliness (All)':<40s} | {'  |  '.join(vals)}")

    pdca_vals = ["" if w not in ["CW05","CW06","CW07"] else fmt_pct(PDCA_OLD["Time_All"][w.replace("CW0","W")]) for w in weeks]
    print(f"{'  └─ PDCA (old, excl S00/S02 SC4)':<40s} | {'  |  '.join(v.rjust(8) for v in pdca_vals)}")

    print("─" * 110)

    # ETA 2P
    vals = [f"{results[w]['eta_2p']:.2%}" if results[w]["eta_2p"] is not None else "N/A" for w in weeks]
    print(f"{'ETA accuracy 2P (SC4)':<40s} | {'  |  '.join(v.rjust(8) for v in vals)}")

    pdca_vals = ["" if w not in ["CW05","CW06","CW07"] else fmt_pct(PDCA_OLD["ETA_2P"][w.replace("CW0","W")]) for w in weeks]
    print(f"{'  └─ PDCA':<40s} | {'  |  '.join(v.rjust(8) for v in pdca_vals)}")

    # ETA 2D
    vals = [f"{results[w]['eta_2d']:.2%}" if results[w]["eta_2d"] is not None else "N/A" for w in weeks]
    print(f"{'ETA accuracy 2D (SC4)':<40s} | {'  |  '.join(v.rjust(8) for v in vals)}")

    pdca_vals = ["" if w not in ["CW05","CW06","CW07"] else fmt_pct(PDCA_OLD["ETA_2D"][w.replace("CW0","W")]) for w in weeks]
    print(f"{'  └─ PDCA':<40s} | {'  |  '.join(v.rjust(8) for v in pdca_vals)}")

    print("─" * 110)

    # Reference Completeness
    vals = [f"{results[w]['ref_comp']:.2%}" if results[w]["ref_comp"] is not None else "N/A" for w in weeks]
    print(f"{'SC4 Ref Completeness (CIV or DN)':<40s} | {'  |  '.join(v.rjust(8) for v in vals)}")

    pdca_vals = ["" if w not in ["CW05","CW06","CW07"] else fmt_pct(PDCA_OLD["Ref_Comp"][w.replace("CW0","W")]) for w in weeks]
    print(f"{'  └─ PDCA':<40s} | {'  |  '.join(v.rjust(8) for v in pdca_vals)}")

    # Detail breakdown
    print(f"\n\n{'='*110}")
    print("DETAIL: Critical milestone rows per week")
    print(f"{'='*110}")
    for week in weeks:
        r = results[week]
        print(f"\n  {week}: SC3={r['sc3_crit_count']} rows, SC4={r['sc4_crit_count']} rows | Req={r['crit_req']} Avl={r['crit_avl']} IT={r['crit_it']}")

    print(f"\n\n{'='*110}")
    print("DELTA: New (corrected) vs Old PDCA values for W5-W7")
    print(f"{'='*110}")
    for w in ["CW05", "CW06", "CW07"]:
        wk = w.replace("CW0", "W")
        r = results[w]
        print(f"\n  {w}:")
        print(f"    Comp Crit: {r['crit_comp']:.4f} (new) vs {PDCA_OLD['Comp_Crit'][wk]:.4f} (old PDCA) | delta={r['crit_comp']-PDCA_OLD['Comp_Crit'][wk]:+.4f}")
        print(f"    Time Crit: {r['crit_time']:.4f} (new) vs {PDCA_OLD['Time_Crit'][wk]:.4f} (old PDCA) | delta={r['crit_time']-PDCA_OLD['Time_Crit'][wk]:+.4f}")
        print(f"    Comp All:  {r['all_comp']:.4f} (new) vs {PDCA_OLD['Comp_All'][wk]:.4f} (old PDCA)  | delta={r['all_comp']-PDCA_OLD['Comp_All'][wk]:+.4f}")
        print(f"    Time All:  {r['all_time']:.4f} (new) vs {PDCA_OLD['Time_All'][wk]:.4f} (old PDCA)  | delta={r['all_time']-PDCA_OLD['Time_All'][wk]:+.4f}")


if __name__ == "__main__":
    main()
