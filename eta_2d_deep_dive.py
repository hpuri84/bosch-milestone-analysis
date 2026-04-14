"""
ETA 2D Deep Dive Analysis
Analyzes delivery accuracy patterns across origin, destination, service type,
carrier, lanes, and deviation distribution to identify what drives the ~30% accuracy
and what it takes to reach 90%+.
"""

import openpyxl
import json
from collections import defaultdict
from datetime import datetime, timedelta

BASE = "/Users/harsh.puri/Documents/work-maersk/Prototype Playground/Bosch Milestone Analysis"
RAW_DIR = BASE + "/Bosch Milestone raw data"

# Analyze recent weeks with enough data
WEEKS_TO_ANALYZE = ["CW09", "CW10", "CW11", "CW12"]
SC4_FILES = {f"CW{i:02d}": f"Maersk SC4_2026_CW{i:02d}.xlsx" for i in range(1, 13)}

# Column indices (0-based) from Shipments sheet
COL = {
    "hbl": 1,
    "mbl": 2,
    "carrier": 8,
    "consignment": 9,
    "transport_mode": 13,
    "service_priority": 15,
    "carriage_condition": 17,
    "incoterm": 19,
    "vessel": 21,
    "container_id": 22,
    "consignor_name": 33,
    "consignor_city": 34,
    "consignor_country": 35,
    "consignee_country": 39,
    "consignee_city": 40,
    "pickup_city": 45,
    "pickup_country": 44,
    "delivery_name": 50,
    "delivery_city": 53,
    "delivery_postcode": 54,
    "delivery_country": 55,
    "port_of_discharge": 56,
    "planned_departure": 67,
    "planned_arrival": 68,
    "delivery_est": 71,
    "delivery_est_type": 72,
    "collected_date": 74,
    "etd": 76,
    "atd": 77,
    "eta": 78,
    "ata": 79,
    "delivered": 81,
    "planned_transit": 82,
    "total_transit": 88,
    "delay_reason_code": 94,
    "delay_reason_desc": 95,
    "reason_supplier": 96,
    "s07_accepted": 105,
    "s07_deviation": 106,
    "s07_measured_eta": 107,
    "s31_accepted": 108,
    "s31_deviation": 109,
    "s31_measured_eta": 110,
}


def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()


def safe_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    try:
        return datetime.strptime(str(val).strip(), "%Y-%m-%d %H:%M:%S")
    except:
        try:
            return datetime.strptime(str(val).strip(), "%Y-%m-%d %H:%M")
        except:
            return None


def find_header_row(ws):
    """Find the header row dynamically (same logic as rebaseline.py)."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        for cell in row:
            if cell and "UNIQUE_SHIPMENT" in str(cell):
                return i
    return 3  # default fallback


def build_col_map(ws, header_row):
    """Build column name -> index map from header row."""
    headers = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            col_map[str(h).strip()] = i
    return col_map


def find_col(col_map, *patterns):
    """Find a column by matching any of the given patterns in column names."""
    for key in col_map:
        for pat in patterns:
            if pat in key:
                return col_map[key]
    return None


def extract_shipments(week):
    """Extract all SC4 shipments with ETA 2D data."""
    path = f"{RAW_DIR}/{SC4_FILES[week]}"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Find Shipments sheet
    ship_sheet = None
    for sn in wb.sheetnames:
        if sn.strip().lower() == "shipments":
            ship_sheet = sn
            break
    if not ship_sheet:
        print(f"  WARNING: No Shipments sheet in {week}")
        return []

    ws = wb[ship_sheet]

    # Dynamic header detection
    header_row = find_header_row(ws)
    col_map = build_col_map(ws, header_row)

    # Map columns dynamically by matching known patterns
    c_hbl = find_col(col_map, "HOUSE_BILL_OF_LADING")
    c_mbl = find_col(col_map, "MASTER_BILL_OF_LADING")
    c_carrier = find_col(col_map, "CARRIER_1")
    c_consignment = find_col(col_map, "CONSIGNMENT_ID_1")
    c_service = find_col(col_map, "TRANSPORT_SERVICE_PRIORITY")
    c_transport_mode = find_col(col_map, "TRANSPORT_MODE")
    c_incoterm = find_col(col_map, "INCOTERM")
    c_consignor_city = find_col(col_map, "CONSIGNOR_ADDRESS_CITY_NAME")
    c_consignor_country = find_col(col_map, "CONSIGNOR_ADDRESS_COUNTRY")
    c_consignee_country = find_col(col_map, "CONSIGNEE_ADDRESS_COUNTRY")
    c_consignee_city = find_col(col_map, "CONSIGNEE_ADDRESS_CITY_NAME")
    c_delivery_city = find_col(col_map, "DELIVERY_ADDRESS_CITY_NAME")
    c_delivery_country = find_col(col_map, "DELIVERY_ADDRESS_COUNTRY")
    c_port_discharge = find_col(col_map, "PORT_OF_DISCHARGE")
    c_delivery_est = find_col(col_map, "DELIVERY_DATE_ACT_EST_PLAN")
    c_collected = find_col(col_map, "COLLECTED_DATE_TIME")
    c_etd = find_col(col_map, "ETD_DATE_TIME")
    c_atd = find_col(col_map, "ATD_DATE_TIME")
    c_eta = find_col(col_map, "ETA_DATE_TIME")
    c_ata = find_col(col_map, "ATA_DATE_TIME")
    c_delivered = find_col(col_map, "DELIVERED_DATE_TIME")
    c_delay_reason = find_col(col_map, "Delay_Reason_Code_Description")
    c_s31_accepted = find_col(col_map, "S31_Accepted")
    c_s31_deviation = find_col(col_map, "S31_Deviation")
    c_s31_measured = find_col(col_map, "S31_TS_@measured")

    def get(row, idx):
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    shipments = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row = list(row)
        if not row[0]:
            continue

        s31_acc = get(row, c_s31_accepted)
        delivered = get(row, c_delivered)
        delivery_est = get(row, c_delivery_est)
        s31_deviation = safe_float(get(row, c_s31_deviation))

        # Only include measurable shipments (same filter as rebaseline.py)
        # Skip when s31_accepted is None (no measurement)
        if s31_acc is None:
            continue
        # Skip undelivered: s31_accepted=0 AND no delivered date
        if s31_acc == 0 and delivered is None:
            continue

        hbl = safe_str(get(row, c_hbl))
        if not hbl:
            continue

        accepted = (s31_acc == 1)

        # Compute deviation if not provided
        if s31_deviation is None and delivery_est and delivered:
            est_dt = parse_date(delivery_est)
            act_dt = parse_date(delivered)
            if est_dt and act_dt:
                s31_deviation = (act_dt - est_dt).total_seconds() / 3600

        origin_country = safe_str(get(row, c_consignor_country))
        origin_city = safe_str(get(row, c_consignor_city))
        dest_country = safe_str(get(row, c_consignee_country))
        dest_city = safe_str(get(row, c_consignee_city))
        delivery_city = safe_str(get(row, c_delivery_city))
        delivery_country = safe_str(get(row, c_delivery_country))
        carrier = safe_str(get(row, c_carrier))
        service = safe_str(get(row, c_service))
        transport_mode = safe_str(get(row, c_transport_mode))
        incoterm = safe_str(get(row, c_incoterm))
        port_discharge = safe_str(get(row, c_port_discharge))

        # Transit phase analysis
        etd = parse_date(get(row, c_etd))
        atd = parse_date(get(row, c_atd))
        eta_port = parse_date(get(row, c_eta))
        ata = parse_date(get(row, c_ata))
        collected = parse_date(get(row, c_collected))
        delivered_dt = parse_date(delivered)
        est_dt = parse_date(delivery_est)

        # Phase durations
        last_mile_days = None
        if ata and delivered_dt:
            last_mile_days = (delivered_dt - ata).total_seconds() / 86400

        transit_days = None
        if atd and ata:
            transit_days = (ata - atd).total_seconds() / 86400

        delay_reason = safe_str(get(row, c_delay_reason))
        s31_measured = safe_str(get(row, c_s31_measured))

        shipments.append({
            "week": week,
            "hbl": hbl,
            "mbl": safe_str(get(row, c_mbl)),
            "accepted": accepted,
            "deviation_hours": s31_deviation,
            "deviation_days": round(s31_deviation / 24, 1) if s31_deviation else None,
            "direction": "late" if (s31_deviation and s31_deviation > 0) else ("early" if (s31_deviation and s31_deviation < 0) else "unknown"),
            "origin_country": origin_country,
            "origin_city": origin_city,
            "dest_country": dest_country,
            "dest_city": dest_city,
            "delivery_city": delivery_city,
            "delivery_country": delivery_country,
            "carrier": carrier,
            "service": service,
            "transport_mode": transport_mode,
            "incoterm": incoterm,
            "port_discharge": port_discharge,
            "lane": f"{origin_country}->{dest_country}",
            "city_lane": f"{origin_city}->{dest_city}",
            "delivery_est": str(est_dt) if est_dt else "",
            "delivered": str(delivered_dt) if delivered_dt else "",
            "last_mile_days": round(last_mile_days, 1) if last_mile_days else None,
            "transit_days": round(transit_days, 1) if transit_days else None,
            "delay_reason": delay_reason,
            "s31_measured_eta": s31_measured,
        })

    wb.close()
    return shipments


def analyze_dimension(shipments, dim_key, label, min_count=3):
    """Analyze accuracy by a dimension, return sorted by impact."""
    groups = defaultdict(lambda: {"total": 0, "accepted": 0, "failed": 0, "deviations": []})
    for s in shipments:
        val = s.get(dim_key, "Unknown") or "Unknown"
        groups[val]["total"] += 1
        if s["accepted"]:
            groups[val]["accepted"] += 1
        else:
            groups[val]["failed"] += 1
        if s["deviation_hours"] is not None:
            groups[val]["deviations"].append(s["deviation_hours"])

    results = []
    for val, g in groups.items():
        if g["total"] < min_count:
            continue
        rate = g["accepted"] / g["total"] if g["total"] > 0 else 0
        devs = g["deviations"]
        avg_dev = sum(devs) / len(devs) if devs else 0
        abs_devs = [abs(d) for d in devs]
        median_abs = sorted(abs_devs)[len(abs_devs)//2] if abs_devs else 0
        results.append({
            "value": val,
            "total": g["total"],
            "accepted": g["accepted"],
            "failed": g["failed"],
            "accuracy": round(rate * 100, 1),
            "avg_deviation_hours": round(avg_dev, 1),
            "median_abs_deviation_hours": round(median_abs, 1),
            "contribution_to_failures": g["failed"],
        })

    # Sort by number of failures (impact)
    results.sort(key=lambda x: x["failed"], reverse=True)
    return results


def deviation_distribution(shipments):
    """Bucket deviations to understand the spread."""
    buckets = {
        "within_24h": 0,
        "24h_to_48h": 0,
        "48h_to_72h": 0,
        "72h_to_96h": 0,
        "96h_to_1week": 0,
        "over_1week": 0,
        "no_data": 0,
    }
    early_late = {"early": 0, "late": 0, "on_time": 0}

    for s in shipments:
        dev = s["deviation_hours"]
        if dev is None:
            buckets["no_data"] += 1
            continue
        abs_dev = abs(dev)
        if abs_dev <= 24:
            buckets["within_24h"] += 1
        elif abs_dev <= 48:
            buckets["24h_to_48h"] += 1
        elif abs_dev <= 72:
            buckets["48h_to_72h"] += 1
        elif abs_dev <= 96:
            buckets["96h_to_1week"] += 1
        elif abs_dev <= 168:
            buckets["96h_to_1week"] += 1
        else:
            buckets["over_1week"] += 1

        if dev > 48:
            early_late["late"] += 1
        elif dev < -48:
            early_late["early"] += 1
        else:
            early_late["on_time"] += 1

    return buckets, early_late


def what_if_analysis(shipments):
    """Calculate what accuracy would be at different windows."""
    windows = [48, 72, 96, 120, 168, 240, 336]
    results = []
    measurable = [s for s in shipments if s["deviation_hours"] is not None]
    total = len(measurable)
    if total == 0:
        return results

    for w in windows:
        accepted = sum(1 for s in measurable if abs(s["deviation_hours"]) <= w)
        results.append({
            "window_hours": w,
            "window_days": round(w / 24, 1),
            "accepted": accepted,
            "total": total,
            "accuracy": round(accepted / total * 100, 1),
        })
    return results


def top_failing_lanes(shipments, min_failures=2):
    """Find origin->dest lanes with worst accuracy."""
    lanes = defaultdict(lambda: {"total": 0, "failed": 0, "deviations": [], "shipments": []})
    for s in shipments:
        if not s["accepted"]:
            key = s["city_lane"]
            lanes[key]["total"] += 1
            lanes[key]["failed"] += 1
            if s["deviation_hours"]:
                lanes[key]["deviations"].append(s["deviation_hours"])
            lanes[key]["shipments"].append(s["hbl"])

    # Also count accepted for each lane
    for s in shipments:
        key = s["city_lane"]
        if key in lanes:
            if s["accepted"]:
                lanes[key]["total"] += 1

    results = []
    for lane, data in lanes.items():
        if data["failed"] < min_failures:
            continue
        devs = data["deviations"]
        avg = sum(devs) / len(devs) if devs else 0
        results.append({
            "lane": lane,
            "total": data["total"],
            "failed": data["failed"],
            "accuracy": round((1 - data["failed"] / data["total"]) * 100, 1) if data["total"] > 0 else 0,
            "avg_deviation_hours": round(avg, 1),
            "avg_deviation_days": round(avg / 24, 1),
            "sample_hbls": data["shipments"][:5],
        })
    results.sort(key=lambda x: x["failed"], reverse=True)
    return results


def main():
    all_shipments = []
    weekly_stats = []

    for week in WEEKS_TO_ANALYZE:
        print(f"  Extracting {week}...")
        shipments = extract_shipments(week)
        total = len(shipments)
        accepted = sum(1 for s in shipments if s["accepted"])
        failed = total - accepted
        rate = round(accepted / total * 100, 1) if total > 0 else 0
        print(f"    {week}: {accepted}/{total} = {rate}% accuracy ({failed} failed)")
        weekly_stats.append({"week": week, "total": total, "accepted": accepted, "failed": failed, "rate": rate})
        all_shipments.extend(shipments)

    # Focus analysis on CW12 but include trends
    cw12 = [s for s in all_shipments if s["week"] == "CW12"]
    recent = all_shipments  # CW09-CW12 combined

    print(f"\n{'='*100}")
    print(f"ETA 2D DEEP DIVE — CW12 ({len(cw12)} shipments) + Recent Trend ({len(recent)} shipments)")
    print(f"{'='*100}")

    # 1. Deviation distribution
    print(f"\n--- DEVIATION DISTRIBUTION (CW12) ---")
    buckets, early_late = deviation_distribution(cw12)
    print(f"  Direction: Early={early_late['early']}, On-time={early_late['on_time']}, Late={early_late['late']}")
    for k, v in buckets.items():
        pct = round(v / len(cw12) * 100, 1) if cw12 else 0
        print(f"  {k:20s}: {v:4d} ({pct}%)")

    # 2. What-if analysis (window expansion)
    print(f"\n--- WHAT-IF: ACCURACY AT DIFFERENT WINDOWS (CW12) ---")
    what_if = what_if_analysis(cw12)
    for w in what_if:
        marker = " <-- CURRENT" if w["window_hours"] == 48 else ""
        marker = " <-- 90% TARGET?" if w["accuracy"] >= 90 and not any(
            x["accuracy"] >= 90 for x in what_if if x["window_hours"] < w["window_hours"]
        ) else marker
        print(f"  ±{w['window_hours']:3d}h ({w['window_days']:4.1f}d): {w['accepted']:3d}/{w['total']:3d} = {w['accuracy']:5.1f}%{marker}")

    # 3. By origin country
    print(f"\n--- BY ORIGIN COUNTRY (CW12, sorted by failures) ---")
    by_origin = analyze_dimension(cw12, "origin_country", "Origin Country")
    print(f"  {'Country':<12} {'Total':>6} {'OK':>5} {'Fail':>5} {'Acc%':>7} {'AvgDev(h)':>10} {'MedAbsDev(h)':>13}")
    for r in by_origin[:15]:
        print(f"  {r['value']:<12} {r['total']:>6} {r['accepted']:>5} {r['failed']:>5} {r['accuracy']:>6.1f}% {r['avg_deviation_hours']:>10.1f} {r['median_abs_deviation_hours']:>13.1f}")

    # 4. By destination country
    print(f"\n--- BY DESTINATION COUNTRY (CW12) ---")
    by_dest = analyze_dimension(cw12, "dest_country", "Dest Country")
    print(f"  {'Country':<12} {'Total':>6} {'OK':>5} {'Fail':>5} {'Acc%':>7} {'AvgDev(h)':>10}")
    for r in by_dest[:15]:
        print(f"  {r['value']:<12} {r['total']:>6} {r['accepted']:>5} {r['failed']:>5} {r['accuracy']:>6.1f}% {r['avg_deviation_hours']:>10.1f}")

    # 5. By service type
    print(f"\n--- BY SERVICE TYPE (CW12) ---")
    by_service = analyze_dimension(cw12, "service", "Service")
    print(f"  {'Service':<25} {'Total':>6} {'OK':>5} {'Fail':>5} {'Acc%':>7} {'AvgDev(h)':>10}")
    for r in by_service:
        print(f"  {r['value']:<25} {r['total']:>6} {r['accepted']:>5} {r['failed']:>5} {r['accuracy']:>6.1f}% {r['avg_deviation_hours']:>10.1f}")

    # 6. By carrier
    print(f"\n--- BY CARRIER (CW12) ---")
    by_carrier = analyze_dimension(cw12, "carrier", "Carrier")
    print(f"  {'Carrier':<25} {'Total':>6} {'OK':>5} {'Fail':>5} {'Acc%':>7} {'AvgDev(h)':>10}")
    for r in by_carrier[:10]:
        print(f"  {r['value']:<25} {r['total']:>6} {r['accepted']:>5} {r['failed']:>5} {r['accuracy']:>6.1f}% {r['avg_deviation_hours']:>10.1f}")

    # 7. By country lane (origin->dest)
    print(f"\n--- BY COUNTRY LANE (CW12) ---")
    by_lane = analyze_dimension(cw12, "lane", "Lane")
    print(f"  {'Lane':<25} {'Total':>6} {'OK':>5} {'Fail':>5} {'Acc%':>7} {'AvgDev(h)':>10}")
    for r in by_lane[:15]:
        print(f"  {r['value']:<25} {r['total']:>6} {r['accepted']:>5} {r['failed']:>5} {r['accuracy']:>6.1f}% {r['avg_deviation_hours']:>10.1f}")

    # 8. By delivery city (last mile)
    print(f"\n--- BY DELIVERY CITY (CW12, top failing) ---")
    by_del_city = analyze_dimension(cw12, "delivery_city", "Delivery City", min_count=3)
    print(f"  {'City':<30} {'Total':>6} {'OK':>5} {'Fail':>5} {'Acc%':>7} {'AvgDev(h)':>10}")
    for r in by_del_city[:15]:
        print(f"  {r['value']:<30} {r['total']:>6} {r['accepted']:>5} {r['failed']:>5} {r['accuracy']:>6.1f}% {r['avg_deviation_hours']:>10.1f}")

    # 9. Top failing city lanes
    print(f"\n--- TOP FAILING CITY LANES (CW12) ---")
    failing_lanes = top_failing_lanes(cw12)
    print(f"  {'Lane':<50} {'Total':>6} {'Fail':>5} {'Acc%':>7} {'AvgDev(d)':>10}")
    for r in failing_lanes[:20]:
        print(f"  {r['lane']:<50} {r['total']:>6} {r['failed']:>5} {r['accuracy']:>6.1f}% {r['avg_deviation_days']:>10.1f}")

    # 10. Early vs Late analysis
    print(f"\n--- EARLY vs LATE BREAKDOWN (CW12 failures only) ---")
    failed = [s for s in cw12 if not s["accepted"] and s["deviation_hours"] is not None]
    early = [s for s in failed if s["deviation_hours"] < -48]
    late = [s for s in failed if s["deviation_hours"] > 48]
    print(f"  Total failures: {len(failed)}")
    print(f"  Early (delivered too soon): {len(early)} ({round(len(early)/len(failed)*100,1) if failed else 0}%)")
    if early:
        avg_early = sum(s["deviation_hours"] for s in early) / len(early)
        print(f"    Avg deviation: {avg_early:.1f}h ({avg_early/24:.1f} days)")
    print(f"  Late (delivered too late):  {len(late)} ({round(len(late)/len(failed)*100,1) if failed else 0}%)")
    if late:
        avg_late = sum(s["deviation_hours"] for s in late) / len(late)
        print(f"    Avg deviation: +{avg_late:.1f}h (+{avg_late/24:.1f} days)")

    # 11. Last mile analysis
    print(f"\n--- LAST MILE DURATION (ATA -> Delivered, CW12) ---")
    with_lastmile = [s for s in cw12 if s["last_mile_days"] is not None]
    if with_lastmile:
        acc_lm = [s["last_mile_days"] for s in with_lastmile if s["accepted"]]
        fail_lm = [s["last_mile_days"] for s in with_lastmile if not s["accepted"]]
        print(f"  Accepted shipments avg last mile: {sum(acc_lm)/len(acc_lm):.1f} days" if acc_lm else "  No accepted data")
        print(f"  Failed shipments avg last mile:   {sum(fail_lm)/len(fail_lm):.1f} days" if fail_lm else "  No failed data")

    # 12. Weekly trend
    print(f"\n--- WEEKLY TREND ---")
    print(f"  {'Week':<8} {'Total':>6} {'OK':>5} {'Fail':>5} {'Accuracy':>9}")
    for ws in weekly_stats:
        print(f"  {ws['week']:<8} {ws['total']:>6} {ws['accepted']:>5} {ws['failed']:>5} {ws['rate']:>8.1f}%")

    # 13. Path to 90%: identify fixable segments
    print(f"\n{'='*100}")
    print(f"PATH TO 90% — ANALYSIS")
    print(f"{'='*100}")
    total_cw12 = len(cw12)
    accepted_cw12 = sum(1 for s in cw12 if s["accepted"])
    current_rate = round(accepted_cw12 / total_cw12 * 100, 1)
    needed_for_90 = int(0.9 * total_cw12) - accepted_cw12
    print(f"  Current: {accepted_cw12}/{total_cw12} = {current_rate}%")
    print(f"  Need {needed_for_90} more shipments within ±48h to reach 90%")
    print(f"  That means fixing {needed_for_90} out of {total_cw12 - accepted_cw12} current failures")

    # Near-misses (just outside window)
    near_miss_72 = [s for s in cw12 if not s["accepted"] and s["deviation_hours"] is not None and abs(s["deviation_hours"]) <= 72]
    near_miss_96 = [s for s in cw12 if not s["accepted"] and s["deviation_hours"] is not None and abs(s["deviation_hours"]) <= 96]
    print(f"\n  Near-misses (48-72h deviation): {len(near_miss_72)} shipments")
    print(f"  Near-misses (48-96h deviation): {len(near_miss_96)} shipments")
    print(f"  If all 48-72h near-misses were fixed: {round((accepted_cw12 + len(near_miss_72)) / total_cw12 * 100, 1)}%")
    print(f"  If all 48-96h near-misses were fixed: {round((accepted_cw12 + len(near_miss_96)) / total_cw12 * 100, 1)}%")

    # Top actionable segments
    print(f"\n  TOP SEGMENTS TO FIX (by failure volume, CW12):")
    for i, r in enumerate(by_lane[:5]):
        print(f"    {i+1}. Lane {r['value']}: {r['failed']} failures, {r['accuracy']}% accuracy, avg dev {r['avg_deviation_hours']}h")
    for i, r in enumerate(by_service):
        print(f"    Service {r['value']}: {r['failed']} failures out of {r['total']}, {r['accuracy']}% accuracy")

    # Export detailed data for dashboard
    output = {
        "generated": datetime.now().isoformat(),
        "weeks_analyzed": WEEKS_TO_ANALYZE,
        "summary": {
            "cw12_total": total_cw12,
            "cw12_accepted": accepted_cw12,
            "cw12_accuracy": current_rate,
            "needed_for_90pct": needed_for_90,
        },
        "weekly_trend": weekly_stats,
        "deviation_distribution": buckets,
        "early_late_split": early_late,
        "what_if_windows": what_if,
        "by_origin_country": by_origin,
        "by_dest_country": by_dest,
        "by_service": by_service,
        "by_carrier": by_carrier,
        "by_country_lane": by_lane,
        "by_delivery_city": by_del_city[:20],
        "top_failing_city_lanes": failing_lanes[:30],
        "near_misses": {
            "within_72h": len(near_miss_72),
            "within_96h": len(near_miss_96),
            "accuracy_if_72h_fixed": round((accepted_cw12 + len(near_miss_72)) / total_cw12 * 100, 1),
            "accuracy_if_96h_fixed": round((accepted_cw12 + len(near_miss_96)) / total_cw12 * 100, 1),
        },
        "failed_shipments_detail": [s for s in cw12 if not s["accepted"]],
    }

    out_path = f"{BASE}/eta_2d_analysis.json"
    with open(out_path, "w") as f:
        json.dump(output, f, indent=2, default=str)
    print(f"\n  Exported detailed analysis to: {out_path}")


if __name__ == "__main__":
    main()
