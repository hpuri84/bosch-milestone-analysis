"""
Extract shipment-level ETA/POD detail for CW11-CW13.
Produces a single Excel sheet with estimated vs actual for both port arrival (2P) and delivery (2D).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

BASE = "/Users/harsh.puri/Documents/work-maersk/Prototype Playground/Bosch Milestone Analysis"
RAW_DIR = BASE + "/Bosch Milestone raw data"

WEEKS = ["CW11", "CW12", "CW13"]
SC4_FILES = {f"CW{i:02d}": f"Maersk SC4_2026_CW{i:02d}.xlsx" for i in range(1, 14)}

# Column patterns to search for in headers
COL_PATTERNS = {
    "unique_id": "UNIQUE_SHIPMENT_ID",
    "hbl": "HOUSE_BILL_OF_LADING",
    "mbl": "MASTER_BILL_OF_LADING",
    "carrier": "CARRIER_1",
    "consignment": "CONSIGNMENT_ID_1",
    "transport_mode": "TRANSPORT_MODE",
    "service": "TRANSPORT_SERVICE_PRIORITY",
    "incoterm": "INCOTERM",
    "vessel": "VESSEL_NAME",
    "container": "CONTAINER_ID",
    "consignor_name": "CONSIGNOR_ADDRESS_NAME1",
    "consignor_city": "CONSIGNOR_ADDRESS_CITY_NAME",
    "consignor_country": "CONSIGNOR_ADDRESS_COUNTRY",
    "consignee_name": "CONSIGNEE_ADDRESS_NAME1",
    "consignee_country": "CONSIGNEE_ADDRESS_COUNTRY",
    "consignee_city": "CONSIGNEE_ADDRESS_CITY_NAME",
    "delivery_name": "DELIVERY_ADDRESS_NAME1",
    "delivery_city": "DELIVERY_ADDRESS_CITY_NAME",
    "delivery_country": "DELIVERY_ADDRESS_COUNTRY",
    "port_of_discharge": "PORT_OF_DISCHARGE",
    "planned_departure": "PLANNED_DEPARTURE_DATE_TIME",
    "planned_arrival": "PLANNED_ARRIVAL_DATE_TIME",
    "delivery_est": "DELIVERY_DATE_ACT_EST_PLAN",
    "delivery_est_type": "DELIVERY_DATE_ACT_EST_PLAN_date_type",
    "collected": "COLLECTED_DATE_TIME",
    "etd": "ETD_DATE_TIME",
    "atd": "ATD_DATE_TIME",
    "eta": "ETA_DATE_TIME",
    "ata": "ATA_DATE_TIME",
    "delivered": "DELIVERED_DATE_TIME",
    "planned_transit": "PLANNED_TRANSIT_TIME",
    "total_transit": "TOTAL_TRANSIT_TIME",
    "s07_accepted": "S07_Accepted",
    "s07_deviation": "S07_Deviation",
    "s07_measured": "S07_TS_@measured",
    "s31_accepted": "S31_Accepted",
    "s31_deviation": "S31_Deviation",
    "s31_measured": "S31_TS_@measured",
}


def find_header_row(ws):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        for cell in row:
            if cell and "UNIQUE_SHIPMENT" in str(cell):
                return i
    return 3


def build_col_map(ws, header_row):
    headers = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            col_map[str(h).strip()] = i
    return col_map


def find_col(col_map, pattern):
    for key in col_map:
        if pattern in key:
            return col_map[key]
    return None


def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()


def fmt_date(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M")
    s = str(val).strip()
    return s if s and s != "None" else ""


def safe_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def extract_week(week):
    path = f"{RAW_DIR}/{SC4_FILES[week]}"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    ship_sheet = None
    for sn in wb.sheetnames:
        if sn.strip().lower() == "shipments":
            ship_sheet = sn
            break
    if not ship_sheet:
        print(f"  WARNING: No Shipments sheet in {week}")
        wb.close()
        return []

    ws = wb[ship_sheet]
    header_row = find_header_row(ws)
    col_map = build_col_map(ws, header_row)

    # Map all columns
    cols = {}
    for key, pattern in COL_PATTERNS.items():
        cols[key] = find_col(col_map, pattern)

    def get(row, key):
        idx = cols.get(key)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    records = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row = list(row)
        if not row[0]:
            continue

        hbl = safe_str(get(row, "hbl"))
        if not hbl:
            continue

        # ETA 2P fields
        eta_port = get(row, "eta")
        ata_port = get(row, "ata")
        s07_acc = get(row, "s07_accepted")
        s07_dev = safe_float(get(row, "s07_deviation"))
        s07_measured = get(row, "s07_measured")

        # ETA 2D fields
        pod_est = get(row, "delivery_est")
        pod_act = get(row, "delivered")
        s31_acc = get(row, "s31_accepted")
        s31_dev = safe_float(get(row, "s31_deviation"))
        s31_measured = get(row, "s31_measured")

        # Compute deviations manually if not available
        if s07_dev is None and eta_port and ata_port:
            try:
                e = eta_port if isinstance(eta_port, datetime) else datetime.strptime(str(eta_port).strip(), "%Y-%m-%d %H:%M:%S")
                a = ata_port if isinstance(ata_port, datetime) else datetime.strptime(str(ata_port).strip(), "%Y-%m-%d %H:%M:%S")
                s07_dev = (a - e).total_seconds() / 3600
            except:
                pass

        if s31_dev is None and pod_est and pod_act:
            try:
                e = pod_est if isinstance(pod_est, datetime) else datetime.strptime(str(pod_est).strip(), "%Y-%m-%d %H:%M:%S")
                a = pod_act if isinstance(pod_act, datetime) else datetime.strptime(str(pod_act).strip(), "%Y-%m-%d %H:%M:%S")
                s31_dev = (a - e).total_seconds() / 3600
            except:
                pass

        # 2P status
        if s07_acc == 1:
            status_2p = "Within ±48h"
        elif s07_acc == 0 and ata_port:
            status_2p = "Outside ±48h"
        elif s07_acc == 0 and not ata_port:
            status_2p = "In Transit"
        else:
            status_2p = "Not Measured"

        # 2D status
        if s31_acc == 1:
            status_2d = "Within ±48h"
        elif s31_acc == 0 and pod_act:
            status_2d = "Outside ±48h"
        elif s31_acc == 0 and not pod_act:
            status_2d = "Not Delivered"
        else:
            status_2d = "Not Measured"

        records.append({
            "week": week,
            "hbl": hbl,
            "mbl": safe_str(get(row, "mbl")),
            "carrier": safe_str(get(row, "carrier")),
            "service": safe_str(get(row, "service")),
            "vessel": safe_str(get(row, "vessel")),
            "container": safe_str(get(row, "container")),
            "incoterm": safe_str(get(row, "incoterm")),
            # Route
            "origin_name": safe_str(get(row, "consignor_name")),
            "origin_city": safe_str(get(row, "consignor_city")),
            "origin_country": safe_str(get(row, "consignor_country")),
            "dest_name": safe_str(get(row, "consignee_name")),
            "dest_city": safe_str(get(row, "consignee_city")),
            "dest_country": safe_str(get(row, "consignee_country")),
            "delivery_name": safe_str(get(row, "delivery_name")),
            "delivery_city": safe_str(get(row, "delivery_city")),
            "delivery_country": safe_str(get(row, "delivery_country")),
            "port_of_discharge": safe_str(get(row, "port_of_discharge")),
            "lane": f"{safe_str(get(row, 'consignor_country'))}->{safe_str(get(row, 'consignee_country'))}",
            # Departure
            "etd": fmt_date(get(row, "etd")),
            "atd": fmt_date(get(row, "atd")),
            # 2P - Port Arrival
            "eta_port": fmt_date(eta_port),
            "ata_port": fmt_date(ata_port),
            "eta_2p_baseline": fmt_date(s07_measured),
            "s07_accepted": s07_acc,
            "s07_deviation_hours": round(s07_dev, 1) if s07_dev is not None else None,
            "s07_deviation_days": round(s07_dev / 24, 1) if s07_dev is not None else None,
            "status_2p": status_2p,
            # 2D - Door Delivery
            "pod_estimated": fmt_date(pod_est),
            "pod_actual": fmt_date(pod_act),
            "eta_2d_baseline": fmt_date(s31_measured),
            "s31_accepted": s31_acc,
            "s31_deviation_hours": round(s31_dev, 1) if s31_dev is not None else None,
            "s31_deviation_days": round(s31_dev / 24, 1) if s31_dev is not None else None,
            "status_2d": status_2d,
        })

    wb.close()
    return records


# ===================== Generate Excel =====================
def generate_excel(all_records):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ETA Detail CW11-CW13"
    ws.sheet_properties.tabColor = "003366"

    header_font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    section_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    data_font = Font(name='Calibri', size=9)
    green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    grey = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Section headers (row 1)
    sections = [
        (1, 2, "Shipment ID"),
        (3, 11, "Route Details"),
        (12, 13, "Departure"),
        (14, 21, "ETA 2P — Port Arrival (±48h)"),
        (22, 29, "ETA 2D — Door Delivery / POD (±48h)"),
    ]

    for start, end, label in sections:
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        cell = ws.cell(1, start, label)
        cell.font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal='center')
        for c in range(start, end + 1):
            ws.cell(1, c).fill = section_fill
            ws.cell(1, c).border = thin_border

    # Column headers (row 2)
    headers = [
        # ID
        "Week", "HBL",
        # Route
        "Origin Country", "Origin City", "Dest Country", "Dest City",
        "Delivery City", "Lane", "Service", "Carrier", "Vessel",
        # Departure
        "ETD", "ATD",
        # 2P
        "ETA (Port)", "ATA (Port)", "2P Baseline ETA", "2P Accepted",
        "2P Dev (hours)", "2P Dev (days)", "2P Direction", "2P Status",
        # 2D
        "POD Estimated", "POD Actual", "2D Baseline ETA", "2D Accepted",
        "2D Dev (hours)", "2D Dev (days)", "2D Direction", "2D Status",
    ]

    for c, h in enumerate(headers, 1):
        cell = ws.cell(2, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    # Data rows
    for i, r in enumerate(all_records, 3):
        s07_dir = ""
        if r["s07_deviation_hours"] is not None:
            s07_dir = "Late" if r["s07_deviation_hours"] > 0 else ("Early" if r["s07_deviation_hours"] < 0 else "On Time")

        s31_dir = ""
        if r["s31_deviation_hours"] is not None:
            s31_dir = "Late" if r["s31_deviation_hours"] > 0 else ("Early" if r["s31_deviation_hours"] < 0 else "On Time")

        values = [
            r["week"], r["hbl"],
            r["origin_country"], r["origin_city"], r["dest_country"], r["dest_city"],
            r["delivery_city"], r["lane"], r["service"], r["carrier"], r["vessel"],
            r["etd"], r["atd"],
            r["eta_port"], r["ata_port"], r["eta_2p_baseline"], r["s07_accepted"],
            r["s07_deviation_hours"], r["s07_deviation_days"], s07_dir, r["status_2p"],
            r["pod_estimated"], r["pod_actual"], r["eta_2d_baseline"], r["s31_accepted"],
            r["s31_deviation_hours"], r["s31_deviation_days"], s31_dir, r["status_2d"],
        ]

        for c, v in enumerate(values, 1):
            cell = ws.cell(i, c, v)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Color 2P status
        status_2p_cell = ws.cell(i, 21)
        if r["status_2p"] == "Within ±48h":
            status_2p_cell.fill = green
        elif r["status_2p"] == "Outside ±48h":
            status_2p_cell.fill = red
        elif r["status_2p"] == "In Transit":
            status_2p_cell.fill = yellow
        else:
            status_2p_cell.fill = grey

        # Color 2D status
        status_2d_cell = ws.cell(i, 29)
        if r["status_2d"] == "Within ±48h":
            status_2d_cell.fill = green
        elif r["status_2d"] == "Outside ±48h":
            status_2d_cell.fill = red
        elif r["status_2d"] == "Not Delivered":
            status_2d_cell.fill = yellow
        else:
            status_2d_cell.fill = grey

        # Color accepted columns
        for acc_col in [17, 25]:
            acc_cell = ws.cell(i, acc_col)
            if acc_cell.value == 1:
                acc_cell.fill = green
            elif acc_cell.value == 0:
                acc_cell.fill = red

    # Auto-width
    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1])) + 2
        for row in ws.iter_rows(min_row=3, max_row=min(50, ws.max_row), min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[get_column_letter(col)].width = min(max_len, 25)

    # Freeze panes
    ws.freeze_panes = "C3"

    # Add auto-filter
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{ws.max_row}"

    out_path = f"{BASE}/ETA_Shipment_Detail_CW11-CW13.xlsx"
    wb.save(out_path)
    return out_path


def main():
    all_records = []
    for week in WEEKS:
        print(f"  Extracting {week}...")
        records = extract_week(week)
        print(f"    {len(records)} shipments")
        all_records.append((week, records))

    # Print summary
    print(f"\n{'='*80}")
    print(f"SUMMARY")
    print(f"{'='*80}")
    for week, records in all_records:
        total = len(records)
        s07_measured = sum(1 for r in records if r["s07_accepted"] is not None)
        s07_acc = sum(1 for r in records if r["s07_accepted"] == 1)
        s07_skip = sum(1 for r in records if r["status_2p"] == "In Transit")
        s31_measured = sum(1 for r in records if r["s31_accepted"] is not None)
        s31_acc = sum(1 for r in records if r["s31_accepted"] == 1)
        s31_skip = sum(1 for r in records if r["status_2d"] == "Not Delivered")

        # Effective measured = measured - skipped
        s07_eff = s07_measured - s07_skip
        s31_eff = s31_measured - s31_skip

        print(f"\n  {week}: {total} total shipments")
        print(f"    2P: {s07_acc}/{s07_eff} = {round(s07_acc/s07_eff*100,1) if s07_eff else 'N/A'}%  (in-transit excluded: {s07_skip})")
        print(f"    2D: {s31_acc}/{s31_eff} = {round(s31_acc/s31_eff*100,1) if s31_eff else 'N/A'}%  (undelivered excluded: {s31_skip})")

    # Flatten and generate Excel
    flat = []
    for _, records in all_records:
        flat.extend(records)

    out = generate_excel(flat)
    print(f"\n  Exported {len(flat)} shipments to: {out}")


if __name__ == "__main__":
    main()
