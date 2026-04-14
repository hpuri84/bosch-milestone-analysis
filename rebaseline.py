"""
Rebaseline all weeks (CW01-CW08) from raw data using corrected formulas.

KPI Definitions:
  - Completeness = sum(Available) / sum(Required)
  - Timeliness   = sum(In_Time) / sum(Required)
  - Critical milestones: S00 (SC4 only), S02, S04, S07, S31 — both Actual + Estimated, NO S45
  - All milestones: everything from SC3 + SC4
  - ETA 2P = S07_Accepted count / S07 measured count (SC4 shipments)
  - ETA 2D = S31_Accepted count / S31 measured count (SC4 shipments)
  - Reference Completeness = shipments with CIV or DN / total shipments (SC4)

Per-service-type breakdowns computed from detail sheets (FCL, BCO, LCL).
Outputs JSON for dashboard consumption.
"""

import openpyxl
import os
import json

BASE = "/Users/harsh.puri/Documents/work-maersk/Prototype Playground/Bosch Milestone Analysis"
RAW_DIR = os.path.join(BASE, "Bosch Milestone raw data")

WEEKS = ["CW01", "CW02", "CW03", "CW04", "CW05", "CW06", "CW07", "CW08", "CW09", "CW10", "CW11", "CW12", "CW13"]

SC3_FILES = {f"CW{i:02d}": f"Maersk NGTM SC3_2026_CW{i:02d}.xlsx" for i in range(1, 10)}
SC3_FILES["CW10"] = "Maersk SC3_2026_CW10.xlsx"  # CW10+ has different naming
SC3_FILES["CW11"] = "Maersk SC3_2026_CW11.xlsx"
SC3_FILES["CW12"] = "Maersk SC3_2026_CW12.xlsx"
SC3_FILES["CW13"] = "Maersk SC3_2026_CW13.xlsx"
SC4_FILES = {f"CW{i:02d}": f"Maersk SC4_2026_CW{i:02d}.xlsx" for i in range(1, 14)}

SC3_CRITICAL_CODES = {"S02", "S04", "S07", "S31"}
SC4_CRITICAL_CODES = {"S00", "S02", "S04", "S07", "S31"}

SERVICE_TYPES = ["FCL", "BCO", "LCL"]


def find_sheet(wb, pattern_fn):
    for sn in wb.sheetnames:
        if pattern_fn(sn):
            return sn
    return None


def find_total_sheet(wb):
    return find_sheet(wb, lambda sn: sn.strip().upper().startswith("TOTAL") or sn.strip().upper() == "ALL")


def find_shipments_sheet(wb):
    return find_sheet(wb, lambda sn: sn.strip().lower() == "shipments")


def find_service_sheet(wb, service_type):
    return find_sheet(wb, lambda sn: sn.strip().upper() == service_type.upper())


def parse_milestone_code(name):
    if name and " - " in str(name):
        return str(name).split(" - ")[0].strip()
    return str(name).strip() if name else None


def read_summary_sheet(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = find_total_sheet(wb)
    if not sheet_name:
        print(f"  WARNING: No TOTAL sheet in {os.path.basename(filepath)}")
        wb.close()
        return []

    ws = wb[sheet_name]
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        for cell in row:
            if cell and str(cell).strip().lower().startswith("required status"):
                header_row = i
                break
        if header_row:
            break
    if not header_row:
        header_row = 3

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        milestone_name = row[0]
        if not milestone_name or "required" in str(milestone_name).lower():
            continue
        code = parse_milestone_code(milestone_name)
        status_type = str(row[1]).strip() if row[1] else ""
        required = row[3] if row[3] and isinstance(row[3], (int, float)) else 0
        available = row[4] if row[4] and isinstance(row[4], (int, float)) else 0
        in_time = row[5] if row[5] and isinstance(row[5], (int, float)) else 0

        rows.append({
            "code": code,
            "type": status_type,
            "required": required,
            "available": available,
            "in_time": in_time,
        })
    wb.close()
    return rows


def read_service_detail_as_summary(filepath, service_type):
    """Read per-service-type detail sheet and aggregate into summary-like rows.

    CW01-CW07 detail sheets have individual shipment-milestone rows with 0/1 available flag.
    We aggregate: Required = count of rows per (code, type), Available = sum(available flag).
    Timeliness (in_time) is NOT available from detail sheets — returned as 0.

    SC3 detail cols: LOAD_TO, Status Code, STATUS_TYPE, STATUS_CODE_AVAILABLE
    SC4 detail cols: CONSIGNMENT, HAWB, HBL, MBL, STATUS_CODE_REQUIRED, STATUS_TYPE, STATUS_CODE_AVAILABLE
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = find_service_sheet(wb, service_type)
    if not sheet_name:
        wb.close()
        return []

    ws = wb[sheet_name]

    # Find the detail header row
    detail_start = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        row_strs = [str(c).strip().upper() if c else "" for c in row]
        if any("STATUS_CODE_AVAILABLE" in s for s in row_strs):
            detail_start = i
            break

    if not detail_start:
        wb.close()
        return []

    headers = []
    for cell in next(ws.iter_rows(min_row=detail_start, max_row=detail_start, values_only=True)):
        headers.append(str(cell).strip().upper() if cell else "")

    avl_col = code_col = type_col = None
    for idx, h in enumerate(headers):
        if "STATUS_CODE_AVAILABLE" in h:
            avl_col = idx
        if h in ("STATUS CODE", "STATUS_CODE_REQUIRED"):
            code_col = idx
        if h == "STATUS_TYPE":
            type_col = idx

    if avl_col is None or code_col is None:
        wb.close()
        return []

    # Aggregate by (milestone_code, status_type)
    from collections import defaultdict
    agg = defaultdict(lambda: {"required": 0, "available": 0})

    for row in ws.iter_rows(min_row=detail_start + 1, max_row=ws.max_row, values_only=True):
        vals = list(row)
        if not vals[0]:
            continue
        raw_code = vals[code_col]
        milestone_code = parse_milestone_code(raw_code) if raw_code else None
        if not milestone_code:
            continue
        status_type = str(vals[type_col]).strip() if type_col is not None and vals[type_col] else ""
        avl_val = vals[avl_col] if isinstance(vals[avl_col], (int, float)) else 0

        key = (milestone_code, status_type)
        agg[key]["required"] += 1
        agg[key]["available"] += avl_val

    wb.close()

    rows = []
    for (code, stype), counts in sorted(agg.items()):
        rows.append({
            "code": code,
            "type": stype,
            "required": int(counts["required"]),
            "available": int(counts["available"]),
            "in_time": 0,  # Not available from detail sheets
        })
    return rows


def read_service_summary(filepath, service_type):
    """Try to read a summary (8-col) section from a service type sheet.
    Returns [] if no summary header found (detail-only sheets).
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = find_service_sheet(wb, service_type)
    if not sheet_name:
        wb.close()
        return []

    ws = wb[sheet_name]
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        for cell in row:
            if cell and str(cell).strip().lower().startswith("required status"):
                header_row = i
                break
        if header_row:
            break

    if not header_row:
        wb.close()
        return []

    rows = read_summary_from_row(ws, header_row)
    wb.close()
    return rows


def read_summary_from_row(ws, header_row):
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        milestone_name = row[0]
        if not milestone_name or "required" in str(milestone_name).lower():
            continue
        # Stop if we hit detail section marker
        if any(str(c).strip().upper() == "STATUS_CODE_AVAILABLE" for c in row if c):
            break
        code = parse_milestone_code(milestone_name)
        status_type = str(row[1]).strip() if row[1] else ""
        required = row[3] if len(row) > 3 and row[3] and isinstance(row[3], (int, float)) else 0
        available = row[4] if len(row) > 4 and row[4] and isinstance(row[4], (int, float)) else 0
        in_time = row[5] if len(row) > 5 and row[5] and isinstance(row[5], (int, float)) else 0

        rows.append({
            "code": code,
            "type": status_type,
            "required": required,
            "available": available,
            "in_time": in_time,
        })
    return rows


def count_sc3_shipments(filepath):
    """Count unique shipments in SC3 shipments sheet."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = find_shipments_sheet(wb)
    if not sheet_name:
        wb.close()
        return 0
    ws = wb[sheet_name]
    count = sum(1 for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True) if row[0])
    wb.close()
    return count


def read_sc4_shipments(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = find_shipments_sheet(wb)
    if not sheet_name:
        wb.close()
        return []

    ws = wb[sheet_name]
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        for cell in row:
            if cell and "UNIQUE_SHIPMENT" in str(cell):
                header_row = i
                break
        if header_row:
            break
    if not header_row:
        header_row = 3

    headers = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            col_map[str(h).strip()] = i

    shipments = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        vals = list(row)
        if not vals[0]:
            continue
        shipment = {}
        for name, idx in col_map.items():
            shipment[name] = vals[idx] if idx < len(vals) else None
        shipments.append(shipment)

    wb.close()
    return shipments


def compute_kpis(rows, critical_codes=None):
    """Compute completeness and timeliness from summary rows.
    If critical_codes is given, filter to those codes only.
    """
    if critical_codes is not None:
        rows = [r for r in rows if r["code"] in critical_codes]

    req = sum(r["required"] for r in rows)
    avl = sum(r["available"] for r in rows)
    it = sum(r["in_time"] for r in rows)
    comp = avl / req if req > 0 else 0
    time = it / req if req > 0 else 0

    return {
        "required": req,
        "available": avl,
        "in_time": it,
        "completeness": round(comp, 4),
        "timeliness": round(time, 4),
        "row_count": len(rows),
    }


def compute_eta_and_ref(shipments):
    """Compute ETA 2P, ETA 2D, and Reference Completeness from SC4 shipments."""
    if not shipments:
        return {"eta_2p": None, "eta_2d": None, "ref_comp": None}

    s07_col = s31_col = civ_col = dn_col = None
    ata_col = delivered_col = None
    for key in shipments[0].keys():
        if "S07_Accepted" in key:
            s07_col = key
        if "S31_Accepted" in key:
            s31_col = key
        if "COMMERCIAL_INVOICE" in key:
            civ_col = key
        if "DELIVERY_NOTE" in key:
            dn_col = key
        if key == "ATA_DATE_TIME":
            ata_col = key
        if key == "DELIVERED_DATE_TIME":
            delivered_col = key

    # ETA 2P — exclude in-transit (S07_Accepted=0, no ATA)
    s07_total = 0
    s07_acc = 0
    if s07_col:
        for s in shipments:
            val = s.get(s07_col)
            if val is None:
                continue
            # Skip in-transit: not accepted and no actual arrival
            if val == 0 and (ata_col is None or s.get(ata_col) is None):
                continue
            s07_total += 1
            if val == 1:
                s07_acc += 1
    eta_2p = round(s07_acc / s07_total, 4) if s07_total > 0 else None

    # ETA 2D — exclude undelivered (S31_Accepted=0, no delivered date)
    s31_total = 0
    s31_acc = 0
    if s31_col:
        for s in shipments:
            val = s.get(s31_col)
            if val is None:
                continue
            # Skip undelivered: not accepted and no actual delivery
            if val == 0 and (delivered_col is None or s.get(delivered_col) is None):
                continue
            s31_total += 1
            if val == 1:
                s31_acc += 1
    eta_2d = round(s31_acc / s31_total, 4) if s31_total > 0 else None

    # Reference Completeness
    ref_total = len(shipments)
    ref_complete = 0
    for s in shipments:
        civ = s.get(civ_col) if civ_col else None
        dn = s.get(dn_col) if dn_col else None
        has_civ = civ is not None and str(civ).strip() not in ("", "None", "NA")
        has_dn = dn is not None and str(dn).strip() not in ("", "None", "NA")
        if has_civ or has_dn:
            ref_complete += 1
    ref_comp = round(ref_complete / ref_total, 4) if ref_total > 0 else None

    return {
        "eta_2p": eta_2p, "s07_acc": s07_acc, "s07_total": s07_total,
        "eta_2d": eta_2d, "s31_acc": s31_acc, "s31_total": s31_total,
        "ref_comp": ref_comp, "ref_complete": ref_complete, "ref_total": ref_total,
    }


def process_week(week):
    """Process a single week and return all KPIs."""
    print(f"\n  Processing {week}...")

    sc3_file = os.path.join(RAW_DIR, SC3_FILES[week])
    sc4_file = os.path.join(RAW_DIR, SC4_FILES[week])

    if not os.path.exists(sc3_file):
        print(f"    SC3 file missing: {SC3_FILES[week]}")
        return None
    if not os.path.exists(sc4_file):
        print(f"    SC4 file missing: {SC4_FILES[week]}")
        return None

    # Read TOTAL summary sheets
    sc3_data = read_summary_sheet(sc3_file)
    sc4_data = read_summary_sheet(sc4_file)

    if not sc3_data:
        print(f"    WARNING: No SC3 summary data for {week}")
    if not sc4_data:
        print(f"    WARNING: No SC4 summary data for {week}")

    # Combined KPIs
    combined = sc3_data + sc4_data
    sc3_crit_codes = SC3_CRITICAL_CODES
    sc4_crit_codes = SC4_CRITICAL_CODES

    sc3_crit = [r for r in sc3_data if r["code"] in sc3_crit_codes]
    sc4_crit = [r for r in sc4_data if r["code"] in sc4_crit_codes]
    crit_rows = sc3_crit + sc4_crit

    crit_kpis = compute_kpis(crit_rows)
    all_kpis = compute_kpis(combined)

    # Per-service-type breakdowns
    service_breakdown = {"SC3": {}, "SC4": {}}

    for svc in SERVICE_TYPES:
        # SC3 service type — try summary first, fall back to detail aggregation
        sc3_svc_rows = read_service_summary(sc3_file, svc)
        if not sc3_svc_rows:
            sc3_svc_rows = read_service_detail_as_summary(sc3_file, svc)
        if sc3_svc_rows:
            service_breakdown["SC3"][svc] = {
                "all": compute_kpis(sc3_svc_rows),
                "critical": compute_kpis(sc3_svc_rows, sc3_crit_codes),
            }

        # SC4 service type — try summary first, fall back to detail aggregation
        sc4_svc_rows = read_service_summary(sc4_file, svc)
        if not sc4_svc_rows:
            sc4_svc_rows = read_service_detail_as_summary(sc4_file, svc)
        if sc4_svc_rows:
            service_breakdown["SC4"][svc] = {
                "all": compute_kpis(sc4_svc_rows),
                "critical": compute_kpis(sc4_svc_rows, sc4_crit_codes),
            }

    # SC3 / SC4 totals separately
    sc3_total = compute_kpis(sc3_data)
    sc4_total = compute_kpis(sc4_data)
    sc3_crit_total = compute_kpis(sc3_crit)
    sc4_crit_total = compute_kpis(sc4_crit)

    # ETA & Reference (SC4 only)
    shipments = read_sc4_shipments(sc4_file)
    eta_ref = compute_eta_and_ref(shipments)

    # Shipment counts (unique shipments from Shipments sheets)
    sc3_ship_count = count_sc3_shipments(sc3_file)
    sc4_ship_count = len(shipments)

    return {
        "week": week,
        "critical": crit_kpis,
        "all": all_kpis,
        "sc3_total": sc3_total,
        "sc4_total": sc4_total,
        "sc3_critical": sc3_crit_total,
        "sc4_critical": sc4_crit_total,
        "sc3_shipments": sc3_ship_count,
        "sc4_shipments": sc4_ship_count,
        "service_breakdown": service_breakdown,
        **eta_ref,
    }


def fmt_pct(val):
    if val is None:
        return "  N/A  "
    return f"{val:>7.2%}"


def print_results(all_results):
    weeks = [r["week"] for r in all_results]
    header = f"{'KPI':<45s}" + "".join(f" | {w:>7s}" for w in weeks)
    sep = "─" * len(header)

    print(f"\n{'='*len(header)}")
    print("REBASELINED KPIs — All Weeks from Raw Data")
    print(f"{'='*len(header)}")
    print(header)
    print(sep)

    # Critical
    print(f"{'Completeness (Critical)':<45s}" + "".join(f" | {fmt_pct(r['critical']['completeness'])}" for r in all_results))
    print(f"{'Timeliness (Critical)':<45s}" + "".join(f" | {fmt_pct(r['critical']['timeliness'])}" for r in all_results))
    print(sep)

    # All
    print(f"{'Completeness (All)':<45s}" + "".join(f" | {fmt_pct(r['all']['completeness'])}" for r in all_results))
    print(f"{'Timeliness (All)':<45s}" + "".join(f" | {fmt_pct(r['all']['timeliness'])}" for r in all_results))
    print(sep)

    # SC3 breakdown
    print(f"{'SC3 Completeness (All)':<45s}" + "".join(f" | {fmt_pct(r['sc3_total']['completeness'])}" for r in all_results))
    print(f"{'SC3 Timeliness (All)':<45s}" + "".join(f" | {fmt_pct(r['sc3_total']['timeliness'])}" for r in all_results))
    print(f"{'SC4 Completeness (All)':<45s}" + "".join(f" | {fmt_pct(r['sc4_total']['completeness'])}" for r in all_results))
    print(f"{'SC4 Timeliness (All)':<45s}" + "".join(f" | {fmt_pct(r['sc4_total']['timeliness'])}" for r in all_results))
    print(sep)

    # Per-service-type
    for scenario in ["SC3", "SC4"]:
        for svc in SERVICE_TYPES:
            comp_vals = []
            time_vals = []
            for r in all_results:
                svc_data = r["service_breakdown"].get(scenario, {}).get(svc)
                if svc_data:
                    comp_vals.append(fmt_pct(svc_data["all"]["completeness"]))
                    time_vals.append(fmt_pct(svc_data["all"]["timeliness"]))
                else:
                    comp_vals.append("  N/A  ")
                    time_vals.append("  N/A  ")
            print(f"{f'{scenario} {svc} Completeness':<45s}" + "".join(f" | {v}" for v in comp_vals))
            print(f"{f'{scenario} {svc} Timeliness':<45s}" + "".join(f" | {v}" for v in time_vals))
    print(sep)

    # ETA & Reference
    print(f"{'ETA accuracy 2P (±48h port)':<45s}" + "".join(f" | {fmt_pct(r.get('eta_2p'))}" for r in all_results))
    print(f"{'ETA accuracy 2D (±48h delivery)':<45s}" + "".join(f" | {fmt_pct(r.get('eta_2d'))}" for r in all_results))
    print(f"{'Reference Completeness (CIV/DN)':<45s}" + "".join(f" | {fmt_pct(r.get('ref_comp'))}" for r in all_results))
    print(sep)

    # Raw counts for audit
    print(f"\n{'AUDIT: Raw counts':<45s}")
    print(sep)
    print(f"{'Critical Required':<45s}" + "".join(f" | {r['critical']['required']:>7d}" for r in all_results))
    print(f"{'Critical Available':<45s}" + "".join(f" | {r['critical']['available']:>7d}" for r in all_results))
    print(f"{'Critical In_Time':<45s}" + "".join(f" | {r['critical']['in_time']:>7d}" for r in all_results))
    print(f"{'All Required':<45s}" + "".join(f" | {r['all']['required']:>7d}" for r in all_results))
    print(f"{'All Available':<45s}" + "".join(f" | {r['all']['available']:>7d}" for r in all_results))
    print(f"{'All In_Time':<45s}" + "".join(f" | {r['all']['in_time']:>7d}" for r in all_results))
    print(f"{'S07 accepted/total':<45s}" + "".join(f" | {r.get('s07_acc',0):>3d}/{r.get('s07_total',0):>3d}" for r in all_results))
    print(f"{'S31 accepted/total':<45s}" + "".join(f" | {r.get('s31_acc',0):>3d}/{r.get('s31_total',0):>3d}" for r in all_results))
    print(f"{'Ref complete/total':<45s}" + "".join(f" | {r.get('ref_complete',0):>3d}/{r.get('ref_total',0):>3d}" for r in all_results))


def export_json(all_results):
    output_path = os.path.join(BASE, "kpi_data.json")
    # Make values JSON-serializable
    with open(output_path, "w") as f:
        json.dump(all_results, f, indent=2, default=str)
    print(f"\nExported KPI data to: {output_path}")


def main():
    print("=" * 100)
    print("BOSCH MILESTONE REBASELINE — All Weeks from Raw Data")
    print("  Corrected formulas: No S45 in Critical, All includes everything")
    print("  Completeness = Available / Required")
    print("  Timeliness   = In_Time / Required")
    print("=" * 100)

    all_results = []
    for week in WEEKS:
        result = process_week(week)
        if result:
            all_results.append(result)

    if all_results:
        print_results(all_results)
        export_json(all_results)

    print("\nDone.")


if __name__ == "__main__":
    main()
